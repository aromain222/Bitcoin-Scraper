#!/usr/bin/env python3
"""
Professional M&A Merger Model Builder
Evaluates the impact of Company A acquiring Company B with accretion/dilution analysis

Author: Investment Banking Modeler
Date: 2024

Features:
- Transaction Assumptions & Company Details
- Sources & Uses of Funds Calculation
- Pro Forma Financial Statement Adjustments
- Accretion/Dilution Analysis to EPS
- Sensitivity Analysis (Premium, Synergies, Financing Mix)
- Professional Excel Output with Banker Formatting
- Deal Summary Box with Key Metrics
"""

import numpy as np
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# Professional color scheme
MERGER_MODEL_COLORS = {
    'header_blue': '1F4E79',
    'input_light_blue': 'DDEBF7',
    'calculation_black': '000000',
    'output_green': 'D8E4BC',
    'warning_red': 'F8CECC',
    'border_dark': '2C3E50',
    'text_dark': '000000',
    'text_white': 'FFFFFF',
    'accretion_green': 'C4D79B',
    'dilution_red': 'F2DCDB',
    'neutral_yellow': 'FFFF00'
}

class ProfessionalMergerModel:
    """
    Comprehensive M&A Merger Model with Accretion/Dilution Analysis
    """

    def __init__(self, acquirer_name="Acquirer Corp", acquirer_ticker="ACQ",
                 target_name="Target Corp", target_ticker="TGT"):
        self.acquirer_name = acquirer_name
        self.acquirer_ticker = acquirer_ticker
        self.target_name = target_name
        self.target_ticker = target_ticker
        self.model_date = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Initialize styles
        self.styles = self._create_styles()

    def _create_styles(self):
        """Create professional Excel styles"""
        styles = {}

        # Header style
        styles['header'] = NamedStyle(name='header')
        styles['header'].font = Font(name='Calibri', size=12, bold=True, color=MERGER_MODEL_COLORS['text_white'])
        styles['header'].fill = PatternFill(start_color=MERGER_MODEL_COLORS['header_blue'], end_color=MERGER_MODEL_COLORS['header_blue'], fill_type='solid')
        styles['header'].alignment = Alignment(horizontal='center', vertical='center')

        # Input style
        styles['input'] = NamedStyle(name='input')
        styles['input'].font = Font(name='Calibri', size=10, color=MERGER_MODEL_COLORS['text_dark'])
        styles['input'].fill = PatternFill(start_color=MERGER_MODEL_COLORS['input_light_blue'], end_color=MERGER_MODEL_COLORS['input_light_blue'], fill_type='solid')
        styles['input'].alignment = Alignment(horizontal='right', vertical='center')
        styles['input'].number_format = '#,##0.00'

        # Calculation style
        styles['calculation'] = NamedStyle(name='calculation')
        styles['calculation'].font = Font(name='Calibri', size=10, color=MERGER_MODEL_COLORS['calculation_black'])
        styles['calculation'].alignment = Alignment(horizontal='right', vertical='center')
        styles['calculation'].number_format = '#,##0.00'

        # Output style
        styles['output'] = NamedStyle(name='output')
        styles['output'].font = Font(name='Calibri', size=10, bold=True, color=MERGER_MODEL_COLORS['text_dark'])
        styles['output'].fill = PatternFill(start_color=MERGER_MODEL_COLORS['output_green'], end_color=MERGER_MODEL_COLORS['output_green'], fill_type='solid')
        styles['output'].alignment = Alignment(horizontal='right', vertical='center')
        styles['output'].number_format = '#,##0.00'

        # Accretion style (green)
        styles['accretion'] = NamedStyle(name='accretion')
        styles['accretion'].font = Font(name='Calibri', size=10, bold=True, color=MERGER_MODEL_COLORS['text_dark'])
        styles['accretion'].fill = PatternFill(start_color=MERGER_MODEL_COLORS['accretion_green'], end_color=MERGER_MODEL_COLORS['accretion_green'], fill_type='solid')
        styles['accretion'].alignment = Alignment(horizontal='right', vertical='center')
        styles['accretion'].number_format = '0.0%'

        # Dilution style (red)
        styles['dilution'] = NamedStyle(name='dilution')
        styles['dilution'].font = Font(name='Calibri', size=10, bold=True, color=MERGER_MODEL_COLORS['text_dark'])
        styles['dilution'].fill = PatternFill(start_color=MERGER_MODEL_COLORS['dilution_red'], end_color=MERGER_MODEL_COLORS['dilution_red'], fill_type='solid')
        styles['dilution'].alignment = Alignment(horizontal='right', vertical='center')
        styles['dilution'].number_format = '0.0%'

        # Warning style
        styles['warning'] = NamedStyle(name='warning')
        styles['warning'].font = Font(name='Calibri', size=10, color=MERGER_MODEL_COLORS['text_dark'])
        styles['warning'].fill = PatternFill(start_color=MERGER_MODEL_COLORS['warning_red'], end_color=MERGER_MODEL_COLORS['warning_red'], fill_type='solid')
        styles['warning'].alignment = Alignment(horizontal='center', vertical='center')

        # Bold label style
        styles['label_bold'] = NamedStyle(name='label_bold')
        styles['label_bold'].font = Font(name='Calibri', size=10, bold=True, color=MERGER_MODEL_COLORS['text_dark'])
        styles['label_bold'].alignment = Alignment(horizontal='left', vertical='center')

        # Normal label style
        styles['label'] = NamedStyle(name='label')
        styles['label'].font = Font(name='Calibri', size=10, color=MERGER_MODEL_COLORS['text_dark'])
        styles['label'].alignment = Alignment(horizontal='left', vertical='center')

        # Summary box style
        styles['summary_box'] = NamedStyle(name='summary_box')
        styles['summary_box'].font = Font(name='Calibri', size=11, bold=True, color=MERGER_MODEL_COLORS['text_white'])
        styles['summary_box'].fill = PatternFill(start_color=MERGER_MODEL_COLORS['header_blue'], end_color=MERGER_MODEL_COLORS['header_blue'], fill_type='solid')
        styles['summary_box'].alignment = Alignment(horizontal='center', vertical='center')
        styles['summary_box'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Subtotal style
        styles['subtotal'] = NamedStyle(name='subtotal')
        styles['subtotal'].font = Font(name='Calibri', size=10, bold=True, color=MERGER_MODEL_COLORS['text_dark'])
        styles['subtotal'].fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')  # Light grey
        styles['subtotal'].alignment = Alignment(horizontal='right', vertical='center')
        styles['subtotal'].number_format = '#,##0.00'

        return styles

    def run_merger_model(self,
                        # Acquirer Details
                        acquirer_share_price=50.0,      # Current share price
                        acquirer_shares_outstanding=100.0,  # Million shares
                        acquirer_eps=2.50,              # Current EPS

                        # Target Details
                        target_share_price=30.0,        # Current share price
                        target_shares_outstanding=50.0, # Million shares
                        target_net_debt=200.0,          # $M (if refinancing)
                        target_eps=1.80,                # Current EPS

                        # Transaction Details
                        offer_price_per_share=45.0,     # Offer price per share
                        premium_pct=0.50,               # 50% premium over current price

                        # Consideration Mix (%)
                        cash_pct=0.60,                  # 60% cash
                        stock_pct=0.40,                 # 40% stock
                        debt_pct=0.00,                  # 0% new debt (financing)

                        # Synergy Assumptions ($M)
                        revenue_synergies=100.0,       # Annual revenue synergies
                        cost_synergies_pct=0.05,       # 5% cost savings on combined costs
                        one_time_costs=50.0,           # One-time integration costs

                        # Financing Assumptions
                        new_debt_interest_rate=0.05,   # 5% interest on new debt
                        foregone_cash_yield=0.03,      # 3% foregone interest on cash used
                        transaction_fees_pct=0.015,    # 1.5% transaction fees

                        # Pro Forma Assumptions
                        intangible_amortization_years=10,  # Years to amortize intangibles
                        tax_rate=0.25,                 # 25% tax rate

                        # Financial Statement Assumptions (Combined)
                        combined_revenue=2000.0,       # Combined annual revenue ($M)
                        combined_ebitda=500.0,         # Combined annual EBITDA ($M)
                        combined_depreciation=80.0,    # Combined annual depreciation ($M)
                        combined_tax_rate=0.25):

        """
        Run complete M&A merger model with accretion/dilution analysis
        """

        print(f"🤝 Building Professional M&A Merger Model: {self.acquirer_name} + {self.target_name}")
        print("=" * 90)

        # Step 1: Create Assumptions & Company Details
        assumptions = self._create_assumptions(
            acquirer_share_price, acquirer_shares_outstanding, acquirer_eps,
            target_share_price, target_shares_outstanding, target_net_debt, target_eps,
            offer_price_per_share, premium_pct, cash_pct, stock_pct, debt_pct,
            revenue_synergies, cost_synergies_pct, one_time_costs,
            new_debt_interest_rate, foregone_cash_yield, transaction_fees_pct,
            intangible_amortization_years, tax_rate,
            combined_revenue, combined_ebitda, combined_depreciation, combined_tax_rate
        )

        # Step 2: Calculate Sources & Uses
        sources_uses = self._calculate_sources_uses(assumptions)

        # Step 3: Create Pro Forma Adjustments
        pro_forma_adjustments = self._create_pro_forma_adjustments(assumptions)

        # Step 4: Build Pro Forma Financials
        pro_forma_financials = self._build_pro_forma_financials(assumptions, pro_forma_adjustments)

        # Step 5: Calculate Accretion/Dilution
        accretion_dilution = self._calculate_accretion_dilution(assumptions, pro_forma_financials)

        # Step 6: Create Sensitivity Analysis
        sensitivity_analysis = self._create_sensitivity_analysis(assumptions)

        # Step 7: Generate Summary Metrics
        summary_metrics = self._generate_summary_metrics(assumptions, sources_uses, accretion_dilution)

        # Compile results
        merger_results = {
            'assumptions': assumptions,
            'sources_uses': sources_uses,
            'pro_forma_adjustments': pro_forma_adjustments,
            'pro_forma_financials': pro_forma_financials,
            'accretion_dilution': accretion_dilution,
            'sensitivity_analysis': sensitivity_analysis,
            'summary_metrics': summary_metrics
        }

        # Create Excel output
        excel_file = self._create_excel_output(merger_results)

        print("\n✅ M&A Merger Model Complete!")
        print("📊 Key Deal Metrics:")
        print(f"   • Offer Price: ${offer_price_per_share:.2f} per share ({premium_pct*100:.0f}% premium)")
        print(f"   • Equity Purchase Price: ${assumptions['equity_purchase_price']/1000:.1f}B")
        print(f"   • Enterprise Value: ${assumptions['enterprise_value']/1000:.1f}B")
        print(f"   • Consideration Mix: {cash_pct*100:.0f}% Cash, {stock_pct*100:.0f}% Stock")
        print(f"   • EPS Accretion/Dilution: {accretion_dilution['accretion_dilution_pct']*100:+.1f}%")
        print(f"   • New Shares Issued: {assumptions['new_shares_issued']/1000:.1f}M")
        print(f"📁 Excel Output: {excel_file}")

        return merger_results, excel_file

    def _create_assumptions(self, acquirer_share_price, acquirer_shares_outstanding, acquirer_eps,
                           target_share_price, target_shares_outstanding, target_net_debt, target_eps,
                           offer_price_per_share, premium_pct, cash_pct, stock_pct, debt_pct,
                           revenue_synergies, cost_synergies_pct, one_time_costs,
                           new_debt_interest_rate, foregone_cash_yield, transaction_fees_pct,
                           intangible_amortization_years, tax_rate,
                           combined_revenue, combined_ebitda, combined_depreciation, combined_tax_rate):

        """Create comprehensive assumptions for the merger analysis"""

        # Calculate derived values
        implied_premium = (offer_price_per_share / target_share_price) - 1
        equity_purchase_price = offer_price_per_share * target_shares_outstanding
        enterprise_value = equity_purchase_price + target_net_debt
        transaction_fees = enterprise_value * transaction_fees_pct

        # Consideration breakdown
        cash_portion = equity_purchase_price * cash_pct
        stock_portion = equity_purchase_price * stock_pct
        debt_portion = enterprise_value * debt_pct

        # New shares to be issued (stock consideration)
        new_shares_issued = stock_portion / acquirer_share_price if stock_portion > 0 else 0

        # Pro forma shares outstanding
        pro_forma_shares_outstanding = acquirer_shares_outstanding + new_shares_issued

        # Purchase accounting adjustments
        goodwill = enterprise_value - (combined_revenue * 0.5)  # Simplified goodwill calculation
        intangibles = goodwill * 0.3  # Assume 30% of goodwill is amortizable intangibles
        annual_intangible_amortization = intangibles / intangible_amortization_years

        assumptions = {
            # Acquirer Details
            'acquirer_share_price': acquirer_share_price,
            'acquirer_shares_outstanding': acquirer_shares_outstanding,
            'acquirer_eps': acquirer_eps,
            'acquirer_market_cap': acquirer_share_price * acquirer_shares_outstanding,

            # Target Details
            'target_share_price': target_share_price,
            'target_shares_outstanding': target_shares_outstanding,
            'target_eps': target_eps,
            'target_net_debt': target_net_debt,
            'target_market_cap': target_share_price * target_shares_outstanding,

            # Transaction Details
            'offer_price_per_share': offer_price_per_share,
            'premium_pct': premium_pct,
            'implied_premium': implied_premium,
            'equity_purchase_price': equity_purchase_price,
            'enterprise_value': enterprise_value,
            'transaction_fees': transaction_fees,

            # Consideration Mix
            'cash_pct': cash_pct,
            'stock_pct': stock_pct,
            'debt_pct': debt_pct,
            'cash_portion': cash_portion,
            'stock_portion': stock_portion,
            'debt_portion': debt_portion,
            'new_shares_issued': new_shares_issued,
            'pro_forma_shares_outstanding': pro_forma_shares_outstanding,

            # Synergy Assumptions
            'revenue_synergies': revenue_synergies,
            'cost_synergies_pct': cost_synergies_pct,
            'one_time_costs': one_time_costs,

            # Financing Assumptions
            'new_debt_interest_rate': new_debt_interest_rate,
            'foregone_cash_yield': foregone_cash_yield,

            # Pro Forma Assumptions
            'intangible_amortization_years': intangible_amortization_years,
            'goodwill': goodwill,
            'intangibles': intangibles,
            'annual_intangible_amortization': annual_intangible_amortization,
            'tax_rate': tax_rate,

            # Combined Financials
            'combined_revenue': combined_revenue,
            'combined_ebitda': combined_ebitda,
            'combined_depreciation': combined_depreciation,
            'combined_tax_rate': combined_tax_rate
        }

        print("📋 Assumptions Created:")
        print(f"   • Acquirer: {acquirer_shares_outstanding:.0f}M shares @ ${acquirer_share_price:.2f} = ${assumptions['acquirer_market_cap']/1000:.1f}B market cap")
        print(f"   • Target: {target_shares_outstanding:.0f}M shares @ ${target_share_price:.2f} = ${assumptions['target_market_cap']/1000:.1f}B market cap")
        print(f"   • Offer: ${offer_price_per_share:.2f}/share ({premium_pct*100:.0f}% premium) = ${equity_purchase_price/1000:.1f}B equity value")
        print(f"   • Consideration: {cash_pct*100:.0f}% Cash, {stock_pct*100:.0f}% Stock, {debt_pct*100:.0f}% Debt")
        print(f"   • New Shares: {new_shares_issued/1000:.1f}M ({new_shares_issued/pro_forma_shares_outstanding*100:.1f}% dilution)")

        return assumptions

    def _calculate_sources_uses(self, assumptions):
        """Calculate Sources & Uses of funds"""

        # USES
        uses = {
            'Equity Purchase Price': assumptions['equity_purchase_price'],
            'Refinance Target Net Debt': assumptions['target_net_debt'],
            'Transaction Fees': assumptions['transaction_fees'],
            'One-time Integration Costs': assumptions['one_time_costs']
        }

        uses['Total Uses'] = sum(uses.values())

        # SOURCES
        sources = {
            'Cash Portion': assumptions['cash_portion'],
            'Stock Portion': assumptions['stock_portion'],
            'New Debt Raised': assumptions['debt_portion']
        }

        sources['Total Sources'] = sum(sources.values())

        # Balance check
        sources_uses_balance = abs(sources['Total Sources'] - uses['Total Uses']) < 1.0

        sources_uses = {
            'sources': sources,
            'uses': uses,
            'balance_check': sources_uses_balance
        }

        print("💰 Sources & Uses:")
        print(f"   • Total Sources: ${sources['Total Sources']/1000:.1f}B")
        print(f"   • Total Uses: ${uses['Total Uses']/1000:.1f}B")
        print(f"   • Balance Check: {'✅ Balanced' if sources_uses_balance else '❌ Mismatch'}")

        return sources_uses

    def _create_pro_forma_adjustments(self, assumptions):
        """Create pro forma adjustments for purchase accounting"""

        # Synergy adjustments
        revenue_adjustment = assumptions['revenue_synergies']
        cost_savings = assumptions['combined_ebitda'] * assumptions['cost_synergies_pct']

        # Purchase accounting adjustments
        intangible_amortization = assumptions['annual_intangible_amortization']

        # Financing adjustments
        new_debt_interest = assumptions['debt_portion'] * assumptions['new_debt_interest_rate']
        foregone_interest = assumptions['cash_portion'] * assumptions['foregone_cash_yield']

        # Tax shield from additional interest expense
        tax_shield = (new_debt_interest - foregone_interest) * assumptions['tax_rate']

        pro_forma_adjustments = {
            'revenue_synergies': revenue_adjustment,
            'cost_savings': cost_savings,
            'intangible_amortization': intangible_amortization,
            'new_debt_interest': new_debt_interest,
            'foregone_interest': foregone_interest,
            'tax_shield': tax_shield,
            'net_interest_adjustment': new_debt_interest - foregone_interest,
            'one_time_costs': assumptions['one_time_costs']
        }

        print("🔧 Pro Forma Adjustments:")
        print(f"   • Revenue Synergies: ${revenue_adjustment:.0f}M")
        print(f"   • Cost Savings: ${cost_savings:.0f}M")
        print(f"   • Intangible Amortization: ${intangible_amortization:.0f}M")
        print(f"   • Net Interest Adjustment: ${(new_debt_interest - foregone_interest):+.0f}M")
        print(f"   • Tax Shield: ${tax_shield:+.0f}M")

        return pro_forma_adjustments

    def _build_pro_forma_financials(self, assumptions, pro_forma_adjustments):
        """Build pro forma income statement"""

        # Start with combined financials
        pro_forma_revenue = assumptions['combined_revenue'] + pro_forma_adjustments['revenue_synergies']
        pro_forma_ebitda = assumptions['combined_ebitda'] + pro_forma_adjustments['cost_savings']
        pro_forma_ebit = pro_forma_ebitda - assumptions['combined_depreciation'] - pro_forma_adjustments['intangible_amortization']

        # Interest expense adjustment
        pro_forma_interest = pro_forma_adjustments['new_debt_interest']  # Simplified - assuming no existing interest
        pro_forma_ebt = pro_forma_ebit - pro_forma_interest

        # Taxes
        pro_forma_taxes = pro_forma_ebt * assumptions['tax_rate']

        # Net income before one-time costs
        pro_forma_net_income_pre_ot = pro_forma_ebt - pro_forma_taxes

        # One-time costs (non-recurring)
        pro_forma_net_income = pro_forma_net_income_pre_ot - pro_forma_adjustments['one_time_costs']

        # Pro forma EPS
        pro_forma_eps = pro_forma_net_income / assumptions['pro_forma_shares_outstanding']

        pro_forma_financials = {
            'revenue': pro_forma_revenue,
            'ebitda': pro_forma_ebitda,
            'depreciation': assumptions['combined_depreciation'],
            'intangible_amortization': pro_forma_adjustments['intangible_amortization'],
            'ebit': pro_forma_ebit,
            'interest': pro_forma_interest,
            'ebt': pro_forma_ebt,
            'taxes': pro_forma_taxes,
            'net_income_pre_ot': pro_forma_net_income_pre_ot,
            'one_time_costs': pro_forma_adjustments['one_time_costs'],
            'net_income': pro_forma_net_income,
            'eps': pro_forma_eps,
            'shares_outstanding': assumptions['pro_forma_shares_outstanding']
        }

        print("📊 Pro Forma Financials:")
        print(f"   • Revenue: ${pro_forma_revenue:.0f}M (+${pro_forma_adjustments['revenue_synergies']:.0f}M synergies)")
        print(f"   • EBITDA: ${pro_forma_ebitda:.0f}M (+${pro_forma_adjustments['cost_savings']:.0f}M cost savings)")
        print(f"   • Net Income: ${pro_forma_net_income:.0f}M")
        print(f"   • Pro Forma EPS: ${pro_forma_eps:.2f}")
        print(f"   • Pro Forma Shares: {assumptions['pro_forma_shares_outstanding']/1000:.1f}M")

        return pro_forma_financials

    def _calculate_accretion_dilution(self, assumptions, pro_forma_financials):
        """Calculate accretion/dilution to EPS"""

        # Standalone acquirer EPS (no transaction)
        standalone_acquirer_eps = assumptions['acquirer_eps']

        # Pro forma EPS (with transaction)
        pro_forma_eps = pro_forma_financials['eps']

        # Handle case where standalone EPS is zero (e.g., PE fund)
        if standalone_acquirer_eps == 0:
            # For PE funds or entities with no earnings, show absolute EPS change
            accretion_dilution_pct = float('inf') if pro_forma_eps > 0 else 0
            eps_difference = pro_forma_eps  # Absolute change since baseline is zero
        else:
            # Normal accretion/dilution calculation
            accretion_dilution_pct = (pro_forma_eps / standalone_acquirer_eps) - 1
            eps_difference = pro_forma_eps - standalone_acquirer_eps

        accretion_dilution = {
            'standalone_acquirer_eps': standalone_acquirer_eps,
            'pro_forma_eps': pro_forma_eps,
            'eps_difference': eps_difference,
            'accretion_dilution_pct': accretion_dilution_pct,
            'is_accretive': eps_difference > 0 if standalone_acquirer_eps == 0 else accretion_dilution_pct > 0,
            'accretion_dilution_amount': eps_difference
        }

        print("📈 Accretion/Dilution Analysis:")
        print(f"   • Standalone Acquirer EPS: ${standalone_acquirer_eps:.2f}")
        print(f"   • Pro Forma EPS: ${pro_forma_eps:.2f}")
        print(f"   • EPS Difference: ${eps_difference:+.2f}")
        if standalone_acquirer_eps == 0:
            print(f"   • EPS Change: {'POSITIVE' if eps_difference > 0 else 'NEGATIVE'} (from zero baseline)")
        else:
            print(f"   • Accretion/Dilution: {accretion_dilution_pct*100:+.1f}% {'(ACCRETIVE)' if accretion_dilution_pct > 0 else '(DILUTIVE)'}")

        return accretion_dilution

    def _create_sensitivity_analysis(self, assumptions):
        """Create sensitivity analysis tables"""

        # Sensitivity: Purchase Premium vs Cost Synergies
        premium_range = np.arange(-0.2, 0.81, 0.2)  # -20% to 60% premium
        synergy_range = np.arange(0.02, 0.11, 0.02)  # 2% to 10% cost synergies

        premium_synergy_sensitivity = []
        for synergy_pct in synergy_range:
            row = []
            for premium in premium_range:
                # Calculate pro forma EPS with these assumptions
                offer_price_adj = assumptions['target_share_price'] * (1 + premium)
                equity_value_adj = offer_price_adj * assumptions['target_shares_outstanding']

                # New shares issued (assuming same stock %)
                stock_portion_adj = equity_value_adj * assumptions['stock_pct']
                new_shares_adj = stock_portion_adj / assumptions['acquirer_share_price']
                pro_forma_shares_adj = assumptions['acquirer_shares_outstanding'] + new_shares_adj

                # Cost synergies
                cost_savings_adj = assumptions['combined_ebitda'] * synergy_pct
                pro_forma_ebitda_adj = assumptions['combined_ebitda'] + cost_savings_adj
                pro_forma_ebit_adj = pro_forma_ebitda_adj - assumptions['combined_depreciation'] - assumptions['annual_intangible_amortization']
                pro_forma_net_income_adj = pro_forma_ebit_adj * (1 - assumptions['tax_rate'])
                pro_forma_eps_adj = pro_forma_net_income_adj / pro_forma_shares_adj

                # Accretion/dilution
                accretion_dilution_adj = (pro_forma_eps_adj / assumptions['acquirer_eps']) - 1
                row.append(accretion_dilution_adj * 100)  # Convert to percentage
            premium_synergy_sensitivity.append(row)

        # Sensitivity: Cash vs Stock Mix
        cash_mix_range = np.arange(0.0, 1.1, 0.25)  # 0% to 100% cash

        cash_stock_sensitivity = []
        for cash_pct in cash_mix_range:
            stock_pct = 1.0 - cash_pct

            # Calculate with different cash/stock mix
            cash_portion = assumptions['equity_purchase_price'] * cash_pct
            stock_portion = assumptions['equity_purchase_price'] * stock_pct
            new_shares_mix = stock_portion / assumptions['acquirer_share_price']
            pro_forma_shares_mix = assumptions['acquirer_shares_outstanding'] + new_shares_mix

            # Financing impact (foregone interest on cash vs new debt)
            foregone_interest_mix = cash_portion * assumptions['foregone_cash_yield']
            new_debt_mix = assumptions['debt_portion']  # Assuming same debt
            new_debt_interest_mix = new_debt_mix * assumptions['new_debt_interest_rate']

            # Net interest impact
            net_interest_mix = new_debt_interest_mix - foregone_interest_mix

            # Pro forma calculations
            pro_forma_ebit_mix = assumptions['combined_ebitda'] + assumptions['cost_synergies_pct'] * assumptions['combined_ebitda'] - assumptions['combined_depreciation'] - assumptions['annual_intangible_amortization'] - net_interest_mix
            pro_forma_net_income_mix = pro_forma_ebit_mix * (1 - assumptions['tax_rate'])
            pro_forma_eps_mix = pro_forma_net_income_mix / pro_forma_shares_mix

            accretion_dilution_mix = (pro_forma_eps_mix / assumptions['acquirer_eps']) - 1
            cash_stock_sensitivity.append(accretion_dilution_mix * 100)

        sensitivity_analysis = {
            'premium_range': premium_range,
            'synergy_range': synergy_range,
            'premium_synergy_sensitivity': premium_synergy_sensitivity,
            'cash_mix_range': cash_mix_range,
            'cash_stock_sensitivity': cash_stock_sensitivity
        }

        print("📊 Sensitivity Analysis Created:")
        print("   • Purchase Premium vs Cost Synergies matrix")
        print("   • Cash vs Stock Mix sensitivity")
        return sensitivity_analysis

    def _generate_summary_metrics(self, assumptions, sources_uses, accretion_dilution):
        """Generate summary metrics for the deal"""

        summary_metrics = {
            'deal_overview': {
                'acquirer_market_cap': assumptions['acquirer_market_cap'],
                'target_market_cap': assumptions['target_market_cap'],
                'offer_price_per_share': assumptions['offer_price_per_share'],
                'premium_pct': assumptions['premium_pct'],
                'equity_purchase_price': assumptions['equity_purchase_price'],
                'enterprise_value': assumptions['enterprise_value']
            },
            'consideration_mix': {
                'cash_pct': assumptions['cash_pct'],
                'stock_pct': assumptions['stock_pct'],
                'debt_pct': assumptions['debt_pct'],
                'cash_portion': assumptions['cash_portion'],
                'stock_portion': assumptions['stock_portion'],
                'debt_portion': assumptions['debt_portion']
            },
            'financing_impact': {
                'new_shares_issued': assumptions['new_shares_issued'],
                'share_dilution_pct': assumptions['new_shares_issued'] / assumptions['pro_forma_shares_outstanding'],
                'pro_forma_shares_outstanding': assumptions['pro_forma_shares_outstanding']
            },
            'synergies': {
                'revenue_synergies': assumptions['revenue_synergies'],
                'cost_synergies_pct': assumptions['cost_synergies_pct'],
                'cost_synergies_amount': assumptions['combined_ebitda'] * assumptions['cost_synergies_pct'],
                'one_time_costs': assumptions['one_time_costs']
            },
            'accretion_dilution': {
                'standalone_eps': accretion_dilution['standalone_acquirer_eps'],
                'pro_forma_eps': accretion_dilution['pro_forma_eps'],
                'eps_difference': accretion_dilution['eps_difference'],
                'accretion_dilution_pct': accretion_dilution['accretion_dilution_pct'],
                'is_accretive': accretion_dilution['is_accretive']
            },
            'sources_uses_check': sources_uses['balance_check']
        }

        print("📋 Summary Metrics Generated:")
        print(f"   • Deal EV: ${assumptions['enterprise_value']/1000:.1f}B ({assumptions['enterprise_value']/assumptions['acquirer_market_cap']:.1f}x acquirer market cap)")
        print(f"   • EPS Impact: ${accretion_dilution['eps_difference']:+.2f} ({accretion_dilution['accretion_dilution_pct']*100:+.1f}%)")

        return summary_metrics

    def _create_excel_output(self, merger_results):
        """Create professional Excel output with multiple tabs"""

        wb = Workbook()

        # Create worksheets
        ws_summary = wb.active
        ws_summary.title = "Deal Summary"

        ws_assumptions = wb.create_sheet("Assumptions")
        ws_sources_uses = wb.create_sheet("Sources & Uses")
        ws_pro_forma = wb.create_sheet("Pro Forma IS")
        ws_accretion = wb.create_sheet("Accretion Dilution")
        ws_sensitivity = wb.create_sheet("Sensitivity Analysis")

        # Apply styles to workbook
        for style_name, style in self.styles.items():
            if style_name not in wb.named_styles:
                wb.add_named_style(style)

        # Create each tab
        self._create_summary_tab(ws_summary, merger_results)
        self._create_assumptions_tab(ws_assumptions, merger_results)
        self._create_sources_uses_tab(ws_sources_uses, merger_results)
        self._create_pro_forma_tab(ws_pro_forma, merger_results)
        self._create_accretion_tab(ws_accretion, merger_results)
        self._create_sensitivity_tab(ws_sensitivity, merger_results)

        # Save workbook
        filename = f"Merger_Model_{self.acquirer_ticker}_{self.target_ticker}_{self.model_date}.xlsx"
        wb.save(filename)

        return filename

    def _create_summary_tab(self, ws, results):
        """Create Deal Summary tab with banker-style summary box"""

        # Title
        ws['A1'] = f"M&A Merger Analysis: {self.acquirer_name} Acquires {self.target_name}"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:G1')

        # Deal Summary Box
        ws['A3'] = 'DEAL SUMMARY'
        ws['A3'].style = 'summary_box'
        ws.merge_cells('A3:G3')

        summary_data = [
            ("Offer Price per Share", f"${results['assumptions']['offer_price_per_share']:.2f}"),
            ("Premium over Current Price", f"{results['assumptions']['premium_pct']*100:.0f}%"),
            ("Equity Purchase Price", f"${results['assumptions']['equity_purchase_price']/1000:.1f}B"),
            ("Enterprise Value", f"${results['assumptions']['enterprise_value']/1000:.1f}B"),
            ("Consideration Mix", f"{results['assumptions']['cash_pct']*100:.0f}% Cash / {results['assumptions']['stock_pct']*100:.0f}% Stock / {results['assumptions']['debt_pct']*100:.0f}% Debt"),
            ("New Shares Issued", f"{results['assumptions']['new_shares_issued']/1000:.1f}M"),
            ("Share Dilution", f"{results['assumptions']['new_shares_issued']/results['assumptions']['pro_forma_shares_outstanding']*100:.1f}%"),
            ("EPS Accretion/Dilution", f"{results['accretion_dilution']['accretion_dilution_pct']*100:+.1f}%"),
            ("Revenue Synergies", f"${results['assumptions']['revenue_synergies']:.0f}M"),
            ("Cost Synergies", f"{results['assumptions']['cost_synergies_pct']*100:.0f}%"),
            ("Sources & Uses Check", "✅ Balanced" if results['sources_uses']['balance_check'] else "❌ Mismatch")
        ]

        current_row = 4
        for label, value in summary_data:
            ws.cell(row=current_row, column=1, value=label).style = 'label_bold'
            ws.cell(row=current_row, column=2, value=value).style = 'output'
            current_row += 1

        # Key Multiples
        current_row += 2
        ws.cell(row=current_row, column=1, value="KEY MULTIPLES").style = 'summary_box'
        ws.merge_cells(f'A{current_row}:G{current_row}')

        multiples_data = [
            ("EV/Revenue", f"{results['assumptions']['enterprise_value']/results['assumptions']['combined_revenue']:.1f}x"),
            ("EV/EBITDA", f"{results['assumptions']['enterprise_value']/results['assumptions']['combined_ebitda']:.1f}x"),
            ("Equity Purchase Price/Target EPS", f"{results['assumptions']['equity_purchase_price']/(results['assumptions']['target_eps']*results['assumptions']['target_shares_outstanding']):.1f}x"),
            ("Pro Forma Shares Outstanding", f"{results['assumptions']['pro_forma_shares_outstanding']/1000:.1f}M")
        ]

        current_row += 1
        for label, value in multiples_data:
            ws.cell(row=current_row, column=1, value=label).style = 'label_bold'
            ws.cell(row=current_row, column=2, value=value).style = 'calculation'
            current_row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25

    def _create_assumptions_tab(self, ws, results):
        """Create Assumptions tab"""

        assumptions = results['assumptions']

        # Title
        ws['A1'] = f"Transaction Assumptions: {self.acquirer_name} + {self.target_name}"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:G1')

        current_row = 3

        # Acquirer Details
        ws[f'A{current_row}'] = "ACQUIRER DETAILS"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        acquirer_data = [
            ("Share Price ($)", assumptions['acquirer_share_price']),
            ("Shares Outstanding (M)", assumptions['acquirer_shares_outstanding']),
            ("Market Cap ($M)", assumptions['acquirer_market_cap']),
            ("EPS ($)", assumptions['acquirer_eps'])
        ]

        for label, value in acquirer_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].style = 'input'
            current_row += 1

        current_row += 2

        # Target Details
        ws[f'A{current_row}'] = "TARGET DETAILS"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        target_data = [
            ("Share Price ($)", assumptions['target_share_price']),
            ("Shares Outstanding (M)", assumptions['target_shares_outstanding']),
            ("Market Cap ($M)", assumptions['target_market_cap']),
            ("Net Debt ($M)", assumptions['target_net_debt']),
            ("EPS ($)", assumptions['target_eps'])
        ]

        for label, value in target_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].style = 'input'
            current_row += 1

        current_row += 2

        # Transaction Details
        ws[f'A{current_row}'] = "TRANSACTION DETAILS"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        transaction_data = [
            ("Offer Price per Share ($)", assumptions['offer_price_per_share']),
            ("Premium (%)", assumptions['premium_pct'] * 100),
            ("Equity Purchase Price ($M)", assumptions['equity_purchase_price']),
            ("Enterprise Value ($M)", assumptions['enterprise_value']),
            ("Transaction Fees ($M)", assumptions['transaction_fees'])
        ]

        for label, value in transaction_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].style = 'input'
            current_row += 1

        current_row += 2

        # Consideration Mix
        ws[f'A{current_row}'] = "CONSIDERATION MIX"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        consideration_data = [
            ("Cash Portion (%)", assumptions['cash_pct'] * 100),
            ("Stock Portion (%)", assumptions['stock_pct'] * 100),
            ("Debt Portion (%)", assumptions['debt_pct'] * 100),
            ("Cash Amount ($M)", assumptions['cash_portion']),
            ("Stock Amount ($M)", assumptions['stock_portion']),
            ("New Shares Issued (M)", assumptions['new_shares_issued']),
            ("Pro Forma Shares (M)", assumptions['pro_forma_shares_outstanding'])
        ]

        for label, value in consideration_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].style = 'input'
            current_row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20

    def _create_sources_uses_tab(self, ws, results):
        """Create Sources & Uses tab"""

        sources_uses = results['sources_uses']

        # Title
        ws['A1'] = f"Sources & Uses: {self.acquirer_name} Acquires {self.target_name}"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:D1')

        # Sources section
        ws['A3'] = "SOURCES"
        ws['A3'].style = 'header'
        ws.merge_cells('A3:B3')

        current_row = 4
        for source, amount in sources_uses['sources'].items():
            ws[f'A{current_row}'] = source
            ws[f'A{current_row}'].style = 'label'
            ws[f'B{current_row}'] = amount
            ws[f'B{current_row}'].style = 'input'
            ws[f'B{current_row}'].number_format = '#,##0'
            current_row += 1

        # Add total with underline
        from openpyxl.styles import Border, Side
        double_border = Border(bottom=Side(style='double'))

        ws[f'A{current_row}'] = "Total Sources"
        ws[f'A{current_row}'].style = 'label_bold'
        ws[f'B{current_row}'] = sources_uses['sources']['Total Sources']
        ws[f'B{current_row}'].style = 'output'
        ws[f'B{current_row}'].number_format = '#,##0'
        ws[f'A{current_row}'].border = double_border
        ws[f'B{current_row}'].border = double_border

        # Uses section
        ws['D3'] = "USES"
        ws['D3'].style = 'header'
        ws.merge_cells('D3:E3')

        current_row = 4
        for use, amount in sources_uses['uses'].items():
            ws[f'D{current_row}'] = use
            ws[f'D{current_row}'].style = 'label'
            ws[f'E{current_row}'] = amount
            ws[f'E{current_row}'].style = 'input'
            ws[f'E{current_row}'].number_format = '#,##0'
            current_row += 1

        # Add total with underline
        ws[f'D{current_row}'] = "Total Uses"
        ws[f'D{current_row}'].style = 'label_bold'
        ws[f'E{current_row}'] = sources_uses['uses']['Total Uses']
        ws[f'E{current_row}'].style = 'output'
        ws[f'E{current_row}'].number_format = '#,##0'
        ws[f'D{current_row}'].border = double_border
        ws[f'E{current_row}'].border = double_border

        # Balance check
        ws['A10'] = "Balance Check:"
        ws['A10'].style = 'label_bold'
        balance_status = "✅ Sources = Uses" if sources_uses['balance_check'] else "❌ MISMATCH"
        ws['B10'] = balance_status
        ws['B10'].style = 'output' if sources_uses['balance_check'] else 'warning'

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15

    def _create_pro_forma_tab(self, ws, results):
        """Create Pro Forma Income Statement tab"""

        pro_forma = results['pro_forma_financials']
        adjustments = results['pro_forma_adjustments']

        # Title
        ws['A1'] = f"Pro Forma Income Statement: Combined {self.acquirer_name} + {self.target_name}"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:F1')

        # Headers
        headers = ['Line Item', 'Amount ($M)', 'Per Share ($)']

        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header).style = 'header'

        # Income Statement data
        is_data = [
            ("Revenue", pro_forma['revenue'], pro_forma['revenue'] / pro_forma['shares_outstanding']),
            ("EBITDA", pro_forma['ebitda'], pro_forma['ebitda'] / pro_forma['shares_outstanding']),
            ("Depreciation", -pro_forma['depreciation'], -pro_forma['depreciation'] / pro_forma['shares_outstanding']),
            ("Intangible Amortization", -pro_forma['intangible_amortization'], -pro_forma['intangible_amortization'] / pro_forma['shares_outstanding']),
            ("EBIT", pro_forma['ebit'], pro_forma['ebit'] / pro_forma['shares_outstanding']),
            ("Interest Expense", -pro_forma['interest'], -pro_forma['interest'] / pro_forma['shares_outstanding']),
            ("EBT", pro_forma['ebt'], pro_forma['ebt'] / pro_forma['shares_outstanding']),
            ("Taxes", -pro_forma['taxes'], -pro_forma['taxes'] / pro_forma['shares_outstanding']),
            ("Net Income (Pre-One-Time)", pro_forma['net_income_pre_ot'], pro_forma['net_income_pre_ot'] / pro_forma['shares_outstanding']),
            ("One-Time Costs", -pro_forma['one_time_costs'], -pro_forma['one_time_costs'] / pro_forma['shares_outstanding']),
            ("Net Income", pro_forma['net_income'], pro_forma['net_income'] / pro_forma['shares_outstanding'])
        ]

        current_row = 4
        for label, amount, per_share in is_data:
            ws.cell(row=current_row, column=1, value=label).style = 'label_bold'
            ws.cell(row=current_row, column=2, value=amount).style = 'calculation'
            ws.cell(row=current_row, column=3, value=per_share).style = 'calculation'

            # Format key lines
            if label in ['EBITDA', 'EBIT', 'Net Income']:
                ws.cell(row=current_row, column=2).style = 'subtotal'
                ws.cell(row=current_row, column=3).style = 'subtotal'

            if label == 'Net Income':
                ws.cell(row=current_row, column=2).style = 'output'
                ws.cell(row=current_row, column=3).style = 'output'

            current_row += 1

        # Add double underline to Net Income
        from openpyxl.styles import Border, Side
        double_border = Border(bottom=Side(style='double'))

        for col in range(1, 4):
            ws.cell(row=current_row-1, column=col).border = double_border

        # Shares outstanding
        current_row += 2
        ws.cell(row=current_row, column=1, value="Shares Outstanding (M)").style = 'label_bold'
        ws.cell(row=current_row, column=2, value=pro_forma['shares_outstanding']).style = 'calculation'

        # EPS
        current_row += 1
        ws.cell(row=current_row, column=1, value="EPS ($)").style = 'label_bold'
        ws.cell(row=current_row, column=2, value=pro_forma['eps']).style = 'output'

        # Set column widths and number formats
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

        for row in range(4, current_row + 1):
            for col in [2, 3]:
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    cell.number_format = '#,##0.00'

    def _create_accretion_tab(self, ws, results):
        """Create Accretion/Dilution Analysis tab"""

        accretion = results['accretion_dilution']
        assumptions = results['assumptions']

        # Title
        ws['A1'] = f"EPS Accretion/Dilution Analysis: {self.acquirer_name} Acquires {self.target_name}"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:F1')

        current_row = 3

        # EPS Comparison
        ws[f'A{current_row}'] = "EPS COMPARISON"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:F{current_row}')
        current_row += 2

        eps_data = [
            ("Standalone Acquirer EPS", accretion['standalone_acquirer_eps']),
            ("Pro Forma EPS", accretion['pro_forma_eps']),
            ("EPS Difference", accretion['eps_difference']),
            ("Accretion/Dilution %", accretion['accretion_dilution_pct'] * 100)
        ]

        for label, value in eps_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            if label == "Accretion/Dilution %":
                ws[f'B{current_row}'].style = 'accretion' if accretion['is_accretive'] else 'dilution'
                ws[f'B{current_row}'].number_format = '0.0%'
            else:
                ws[f'B{current_row}'].style = 'calculation'
                ws[f'B{current_row}'].number_format = '#,##0.00'
            current_row += 1

        current_row += 2

        # Deal Metrics
        ws[f'A{current_row}'] = "DEAL METRICS"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:F{current_row}')
        current_row += 2

        deal_metrics = [
            ("New Shares Issued (M)", assumptions['new_shares_issued']),
            ("Pro Forma Shares (M)", assumptions['pro_forma_shares_outstanding']),
            ("Share Dilution (%)", assumptions['new_shares_issued'] / assumptions['pro_forma_shares_outstanding'] * 100),
            ("Equity Purchase Price ($M)", assumptions['equity_purchase_price']),
            ("Enterprise Value ($M)", assumptions['enterprise_value'])
        ]

        for label, value in deal_metrics:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].style = 'calculation'
            if "%" in label:
                ws[f'B{current_row}'].number_format = '0.0%'
            elif "Shares" in label:
                ws[f'B{current_row}'].number_format = '#,##0'
            else:
                ws[f'B{current_row}'].number_format = '#,##0'
            current_row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

    def _create_sensitivity_tab(self, ws, results):
        """Create Sensitivity Analysis tab"""

        sensitivity = results['sensitivity_analysis']

        # Title
        ws['A1'] = f"Sensitivity Analysis: {self.acquirer_name} + {self.target_name}"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:K1')

        # Premium vs Cost Synergies Sensitivity
        ws['A3'] = 'SENSITIVITY: Purchase Premium vs Cost Synergies (EPS Accretion/Dilution %)'
        ws['A3'].style = 'header'
        ws.merge_cells('A3:K3')

        # Headers
        ws['A5'] = 'Cost Synergies \\ Premium'
        ws['A5'].style = 'header'

        for i, premium in enumerate(sensitivity['premium_range']):
            ws.cell(row=5, column=i+2, value=f'{premium*100:.0f}%').style = 'header'

        # Data
        for i, synergy_pct in enumerate(sensitivity['synergy_range']):
            row = i + 6
            ws.cell(row=row, column=1, value=f'{synergy_pct*100:.0f}%').style = 'label_bold'

            for j, accretion_pct in enumerate(sensitivity['premium_synergy_sensitivity'][i]):
                cell = ws.cell(row=row, column=j+2, value=accretion_pct)
                if accretion_pct > 0:
                    cell.style = 'accretion'
                else:
                    cell.style = 'dilution'
                cell.number_format = '0.0%'

        # Cash vs Stock Mix Sensitivity
        current_row = len(sensitivity['synergy_range']) + 10

        ws[f'A{current_row}'] = 'SENSITIVITY: Cash vs Stock Mix (EPS Accretion/Dilution %)'
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:K{current_row}')

        # Headers
        ws.cell(row=current_row+2, column=1, value='Cash %').style = 'header'
        ws.cell(row=current_row+2, column=2, value='Stock %').style = 'header'
        ws.cell(row=current_row+2, column=3, value='EPS Accretion/Dilution %').style = 'header'

        # Data
        for i, cash_pct in enumerate(sensitivity['cash_mix_range']):
            row = current_row + 3 + i
            stock_pct = 1.0 - cash_pct

            ws.cell(row=row, column=1, value=f'{cash_pct*100:.0f}%').style = 'input'
            ws.cell(row=row, column=2, value=f'{stock_pct*100:.0f}%').style = 'input'

            accretion_pct = sensitivity['cash_stock_sensitivity'][i]
            cell = ws.cell(row=row, column=3, value=accretion_pct)
            if accretion_pct > 0:
                cell.style = 'accretion'
            else:
                cell.style = 'dilution'
            cell.number_format = '0.0%'

        # Set column widths
        for col in range(1, 12):
            ws.column_dimensions[get_column_letter(col)].width = 12


def run_sample_merger_model():
    """Run a sample M&A merger model"""

    print("🤝 Running Professional M&A Merger Model Sample")
    print("=" * 70)

    # Create merger model instance
    merger_model = ProfessionalMergerModel("TechCorp Inc.", "TECH", "TargetSoft Ltd.", "TSFT")

    # Run the model with sample assumptions
    results, excel_file = merger_model.run_merger_model(
        # Acquirer Details
        acquirer_share_price=50.0,      # $50/share
        acquirer_shares_outstanding=100.0,  # 100M shares
        acquirer_eps=2.50,              # $2.50 EPS

        # Target Details
        target_share_price=30.0,        # $30/share
        target_shares_outstanding=50.0, # 50M shares
        target_net_debt=200.0,          # $200M net debt
        target_eps=1.80,                # $1.80 EPS

        # Transaction Details
        offer_price_per_share=45.0,     # $45/share offer
        premium_pct=0.50,               # 50% premium

        # Consideration Mix (%)
        cash_pct=0.60,                  # 60% cash
        stock_pct=0.40,                 # 40% stock
        debt_pct=0.00,                  # 0% new debt

        # Synergy Assumptions ($M)
        revenue_synergies=100.0,       # $100M revenue synergies
        cost_synergies_pct=0.05,       # 5% cost savings
        one_time_costs=50.0,           # $50M integration costs

        # Financing Assumptions
        new_debt_interest_rate=0.05,   # 5%
        foregone_cash_yield=0.03,      # 3%
        transaction_fees_pct=0.015,    # 1.5%

        # Pro Forma Assumptions
        intangible_amortization_years=10,
        tax_rate=0.25,
        combined_revenue=2000.0,       # $2B combined revenue
        combined_ebitda=500.0,         # $500M combined EBITDA
        combined_depreciation=80.0,    # $80M combined depreciation
        combined_tax_rate=0.25
    )

    return results, excel_file


if __name__ == "__main__":
    # Run sample merger model
    results, excel_file = run_sample_merger_model()

    print("\n📋 Model Complete!")
    print(f"📁 Excel file saved as: {excel_file}")

    # Display key results
    print("\n🔑 Key Results Summary:")
    print("=" * 40)

    summary = results['summary_metrics']
    deal = summary['deal_overview']
    accretion = summary['accretion_dilution']

    print(f"Offer Price: ${deal['offer_price_per_share']:.2f} per share ({deal['premium_pct']*100:.0f}% premium)")
    print(f"Enterprise Value: ${deal['enterprise_value']/1000:.1f}B")
    print(f"Consideration Mix: {summary['consideration_mix']['cash_pct']*100:.0f}% Cash, {summary['consideration_mix']['stock_pct']*100:.0f}% Stock")
    print(f"New Shares Issued: {summary['financing_impact']['new_shares_issued']/1000:.1f}M")
    print(f"EPS Impact: ${accretion['eps_difference']:+.2f} ({accretion['accretion_dilution_pct']*100:+.1f}%)")
    print(f"Sources & Uses: {'Balanced' if summary['sources_uses_check'] else 'Mismatch'}")
