#!/usr/bin/env python3
"""
Unified DCF Model Builder
Combines all functionality: Excel output, Google Sheets, multiple models, auto-installation
"""

import sys
import subprocess
import os
import re
import requests
import pandas as pd
from bs4 import BeautifulSoup, Tag
from datetime import datetime
from openpyxl import Workbook
from urllib.parse import urljoin

def install_and_import(package, pip_name=None):
    """Auto-install missing packages."""
    pip_name = pip_name or package
    try:
        __import__(package)
    except ImportError:
        print(f"Installing missing package: {pip_name}")
        subprocess.check_call([sys.executable, "-m", "pip", "install", pip_name])
        __import__(package)

# Install all required packages
REQUIRED_MODULES = [
    ("pandas", "pandas"),
    ("openpyxl", "openpyxl"),
    ("requests", "requests"),
    ("bs4", "beautifulsoup4"),
    ("yfinance", "yfinance"),

    ("openai", "openai"),
    ("gspread", "gspread"),
    ("google.auth", "google-auth"),
    ("dotenv", "python-dotenv"),
    ("tiktoken", "tiktoken"),
    ("tldextract", "tldextract"),
    ("lxml", "lxml"),
]

for mod, pip_name in REQUIRED_MODULES:
    install_and_import(mod, pip_name)

# Optional imports with error handling
try:
    import openai
except ImportError:
    openai = None
try:
    import yfinance as yf
except ImportError:
    yf = None
try:
    import tldextract
except ImportError:
    tldextract = None
try:
    import gspread
    from google.oauth2.service_account import Credentials
except ImportError:
    gspread = None
    Credentials = None

# Load environment variables
from dotenv import load_dotenv
load_dotenv()
OPENAI_API_KEY = os.getenv('OPENAI_API_KEY')
if openai and OPENAI_API_KEY:
    openai.api_key = OPENAI_API_KEY

# Configuration
YEARS = 5
DISCOUNT_RATE = 0.10
TERMINAL_GROWTH = 0.03

# Correct project information
CORRECT_PROJECT_ID = "modeo-466403"
CORRECT_SERVICE_ACCOUNT = "service-account@modeo-466403.iam.gserviceaccount.com"

def check_credentials():
    """Check if credentials are correct."""
    creds_path = os.getenv('GOOGLE_SHEETS_CREDENTIALS') or 'credentials/google_sheets_credentials.json'
    
    try:
        import json
        with open(creds_path, 'r') as f:
            creds_data = json.load(f)
        
        project_id = creds_data.get('project_id', '')
        client_email = creds_data.get('client_email', '')
        
        print(f"Current credentials:")
        print(f"  Project ID: {project_id}")
        print(f"  Service Account: {client_email}")
        print()
        
        if project_id != CORRECT_PROJECT_ID:
            print("‚ùå WRONG PROJECT ID!")
            print(f"Expected: {CORRECT_PROJECT_ID}")
            print(f"Found: {project_id}")
            print()
            print("SETUP INSTRUCTIONS:")
            print("1. Go to https://console.cloud.google.com/")
            print("2. Select project: modeo-466403")
            print("3. Go to 'IAM & Admin' ‚Üí 'Service Accounts'")
            print("4. Find: service-account@modeo-466403.iam.gserviceaccount.com")
            print("5. Create/download JSON key")
            print("6. Replace credentials/google_sheets_credentials.json")
            print()
            return False
        
        print("‚úÖ Correct project ID found!")
        return True
        
    except Exception as e:
        print(f"‚ùå Error reading credentials: {e}")
        return False

def get_yfinance_financials(ticker):
    """Get financial data from yfinance for public companies."""
    if not yf:
        print("yfinance not available")
        return {}
    try:
        ticker_obj = yf.Ticker(ticker)
        fin = ticker_obj.financials.T
        years = fin.index[-YEARS:]
        data = {}
        if 'Total Revenue' in fin.columns:
            data['Revenue'] = [float(fin.loc[y, 'Total Revenue']) for y in years]
        if 'EBITDA' in fin.columns:
            data['EBITDA'] = [float(fin.loc[y, 'EBITDA']) for y in years]
        if 'Operating Income' in fin.columns:
            data['EBIT'] = [float(fin.loc[y, 'Operating Income']) for y in years]
        return data
    except Exception as e:
        print(f"yfinance extraction failed: {e}")
        return {}

def extract_financials_with_llm(company_name):
    """Extract financials using OpenAI."""
    if not openai or not OPENAI_API_KEY:
        print("OpenAI not available")
        return {}
    prompt = f"""
Extract the most recent annual financials for {company_name}. Return a JSON object with keys: Revenue, EBITDA, EBIT, Net Income, Depreciation, CapEx, Working Capital. Use 'Unknown' if not found.
"""
    try:
        response = openai.chat.completions.create(
            model="gpt-3.5-turbo-16k",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.0,
            max_tokens=600,
        )
        import json
        content = response.choices[0].message.content
        if content:
            return json.loads(content)
        else:
            return {}
    except Exception as e:
        print(f"LLM extraction failed: {e}")
        return {}

def crawl_company_website(base_url, max_depth=2, max_pages=20):
    """Web crawler for company financial data."""
    if not tldextract:
        print("tldextract is not installed. Please install it to use this feature.")
        return set()
    visited = set()
    to_visit = [(base_url, 0)]
    relevant_urls = set()
    keywords = ['investor', 'financial', 'annual', 'report', 'press', 'news', 'sec', 'filing', 'earnings']
    domain = tldextract.extract(base_url).registered_domain
    while to_visit and len(visited) < max_pages:
        url, depth = to_visit.pop(0)
        if url in visited or depth > max_depth:
            continue
        try:
            resp = requests.get(url, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
            if resp.status_code != 200:
                continue
            visited.add(url)
            soup = BeautifulSoup(resp.text, 'lxml')
            if any(kw in url.lower() for kw in keywords) or any(kw in soup.get_text().lower() for kw in keywords):
                relevant_urls.add(url)
            for link in soup.find_all('a', href=True):
                if not isinstance(link, Tag):
                    continue
                href = link.get('href', None)
                if not isinstance(href, str):
                    continue
                href_str = str(href)
                if not href_str.startswith('http'):
                    href_str = str(urljoin(url, href_str))
                link_domain = tldextract.extract(href_str).registered_domain
                if link_domain == domain and href_str not in visited:
                    to_visit.append((href_str, depth + 1))
        except Exception:
            continue
    return relevant_urls

def scrape_financials_from_url(url):
    """Scrape financial data from a URL."""
    try:
        resp = requests.get(url, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
        if resp.status_code != 200:
            return {}
        soup = BeautifulSoup(resp.text, 'lxml')
        text = soup.get_text(" ", strip=True)
        patterns = {
            'Revenue': r'(?:revenue|total sales)[^\$\d]{0,20}\$?([\d,.]+) ?(million|billion|bn|m|b)?',
            'EBITDA': r'(?:ebitda)[^\$\d]{0,20}\$?([\d,.]+) ?(million|billion|bn|m|b)?',
            'EBIT': r'(?:ebit|operating income)[^\$\d]{0,20}\$?([\d,.]+) ?(million|billion|bn|m|b)?',
        }
        found = {}
        for field, pat in patterns.items():
            m = re.search(pat, text, re.IGNORECASE)
            if m:
                val = m.group(1).replace(',', '')
                mult = m.group(2) or ''
                try:
                    val = float(val)
                    if mult.lower() in ['billion', 'bn', 'b']:
                        val *= 1e9
                    elif mult.lower() in ['million', 'm']:
                        val *= 1e6
                    found[field] = val
                except Exception:
                    continue
        if len(found) < 2:
            llm_fields = extract_financials_with_llm(text[:4000])
            found.update({k: v for k, v in llm_fields.items() if k not in found})
        return found
    except Exception:
        return {}

def project_financials(base, growth=0.08, years=YEARS):
    """Project financials into the future."""
    return [round(base * ((1 + growth) ** i), 2) for i in range(years)]

def calculate_fcf(revenue, ebitda, ebit, taxes, depreciation, capex, delta_wc):
    """Calculate Free Cash Flow."""
    tax_rate = 0.25
    fcf = []
    for i in range(len(revenue)):
        EBIT = ebit[i] if ebit else revenue[i] * 0.15
        Taxes = EBIT * tax_rate
        Dep = depreciation[i] if depreciation else revenue[i] * 0.05
        Cap = capex[i] if capex else revenue[i] * 0.06
        DWC = delta_wc[i] if delta_wc else revenue[i] * 0.02
        fcf.append(round(EBIT - Taxes + Dep - Cap - DWC, 2))
    return fcf

def calculate_terminal_value(last_fcf, discount_rate=DISCOUNT_RATE, growth=TERMINAL_GROWTH):
    """Calculate terminal value."""
    return last_fcf * (1 + growth) / (discount_rate - growth)

def discount_cash_flows(fcfs, terminal_value, discount_rate=DISCOUNT_RATE):
    """Discount cash flows to present value."""
    discounted = [fcf / ((1 + discount_rate) ** (i + 1)) for i, fcf in enumerate(fcfs)]
    tv_discounted = terminal_value / ((1 + discount_rate) ** len(fcfs))
    return discounted, tv_discounted

def write_to_excel(company, years, revenue, ebitda, ebit, fcf, terminal_value, enterprise_value, filename):
    """Write DCF model to Excel file."""
    wb = Workbook()
    ws = wb.active
    if ws:
        ws.title = "DCF Model"
        
        # Add data
        ws.append(["Company", company])
        ws.append([])
        ws.append(["Year"] + years)
        ws.append(["Revenue"] + revenue)
        ws.append(["EBITDA"] + ebitda)
        ws.append(["EBIT"] + ebit)
        ws.append(["FCF"] + fcf)
        ws.append([])
        ws.append(["Terminal Value", terminal_value])
        ws.append(["Enterprise Value", enterprise_value])
        
        # Add formatting
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Format headers
        for cell in ws[3]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Save file
        wb.save(filename)
        print(f"‚úÖ Saved DCF model to {filename}")
    else:
        print("‚ùå Could not create Excel worksheet")

def write_to_google_sheets(company, years, revenue, ebitda, ebit, fcf, terminal_value, enterprise_value, sheet_name="DCF", worksheet_name="Microsoft"):
    """Write DCF model to Google Sheets."""
    if not gspread or not Credentials:
        print("Google Sheets not available")
        return False
    
    creds_path = os.getenv('GOOGLE_SHEETS_CREDENTIALS') or 'credentials/google_sheets_credentials.json'
    
    try:
        print(f"Connecting to Google Sheets...")
        
        # Use both Sheets and Drive API scopes for full access
        scopes = [
            'https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/drive'
        ]
        creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
        gc = gspread.authorize(creds)
        
        # Open the sheet
        sh = gc.open(sheet_name)
        print(f"‚úÖ Found '{sheet_name}' sheet")
        
        # Get the worksheet
        try:
            ws = sh.worksheet(worksheet_name)
            print(f"‚úÖ Found '{worksheet_name}' tab")
        except gspread.WorksheetNotFound:
            print(f"‚ùå '{worksheet_name}' tab not found in '{sheet_name}' sheet")
            print(f"Please create a tab named '{worksheet_name}' in your {sheet_name} sheet")
            return False
        
        # Clear existing data and add new DCF model
        ws.clear()
        
        # Prepare data
        data = []
        data.append(["Company", company])
        data.append([])
        data.append(["Year"] + years)
        data.append(["Revenue"] + revenue)
        data.append(["EBITDA"] + ebitda)
        data.append(["EBIT"] + ebit)
        data.append(["FCF"] + fcf)
        data.append([])
        data.append(["Terminal Value", terminal_value])
        data.append(["Enterprise Value", enterprise_value])
        
        # Update the sheet
        ws.update('A1', data)
        
        # Try to format headers
        try:
            fmt = {"textFormat": {"bold": True}}
            ws.format("A3:Z3", fmt)
        except Exception:
            pass
        
        print(f"‚úÖ DCF model written to {worksheet_name} tab")
        print(f"üìä Sheet URL: {sh.url}")
        return True
        
    except Exception as e:
        print(f"‚ùå Error writing to Google Sheets: {e}")
        return False

def build_dcf_model(company_name, ticker=None, website=None, output="both", sheet_name="DCF", worksheet_name="Microsoft"):
    """Build DCF model for a company."""
    print(f"Building DCF model for {company_name}")
    
    # Get financial data
    data = get_yfinance_financials(ticker) if ticker else {}
    if not data or not data.get('Revenue'):
        print("Falling back to LLM extraction...")
        data = extract_financials_with_llm(company_name)
    if (not data or not data.get('Revenue')) and website:
        print("Trying web crawler and scraper...")
        urls = crawl_company_website(website)
        for url in urls:
            scraped = scrape_financials_from_url(url)
            if scraped.get('Revenue'):
                data.update(scraped)
                break
    
    # Use default values if no data found
    if not data or not data.get('Revenue'):
        print("Using default financial assumptions...")
        base_revenue = 1000000  # $1M default
    else:
        base_revenue = float(str(data.get('Revenue', [0])[0] if isinstance(data.get('Revenue'), list) else data.get('Revenue', 0)).replace(',', '').replace('$', ''))
    
    # Project financials
    revenue_proj = project_financials(base_revenue)
    ebitda_proj = [r * 0.25 for r in revenue_proj]
    ebit_proj = [r * 0.15 for r in revenue_proj]
    fcf_proj = calculate_fcf(revenue_proj, ebitda_proj, ebit_proj, None, None, None, None)
    
    # Calculate terminal value and enterprise value
    terminal_value = calculate_terminal_value(fcf_proj[-1])
    discounted_fcfs, discounted_tv = discount_cash_flows(fcf_proj, terminal_value)
    enterprise_value = round(sum(discounted_fcfs) + discounted_tv, 2)
    
    # Create years
    years = [datetime.now().year + i for i in range(YEARS)]
    
    # Output results
    success = True
    
    if output in ["excel", "both"]:
        filename = f"dcf_{company_name.replace(' ', '_')}.xlsx"
        write_to_excel(company_name, years, revenue_proj, ebitda_proj, ebit_proj, fcf_proj, terminal_value, enterprise_value, filename)
    
    if output in ["gsheet", "both"]:
        gsheet_success = write_to_google_sheets(company_name, years, revenue_proj, ebitda_proj, ebit_proj, fcf_proj, terminal_value, enterprise_value, sheet_name, worksheet_name)
        if not gsheet_success:
            success = False
    
    if success:
        # Print summary
        print(f"\nüìä DCF Model Summary for {company_name}")
        print(f"Base Revenue: ${base_revenue:,.0f}")
        print(f"Terminal Value: ${terminal_value:,.0f}")
        print(f"Enterprise Value: ${enterprise_value:,.0f}")
        print(f"Discount Rate: {DISCOUNT_RATE*100}%")
        print(f"Terminal Growth: {TERMINAL_GROWTH*100}%")
        print(f"‚úÖ Model successfully created!")
    else:
        print("‚ùå Some outputs failed")
    
    return success

def main_menu():
    """Main interactive menu."""
    print("üöÄ Unified DCF Model Builder")
    print("=" * 50)
    print("This tool combines all DCF functionality:")
    print("‚Ä¢ Excel output")
    print("‚Ä¢ Google Sheets integration")
    print("‚Ä¢ yfinance data extraction")
    print("‚Ä¢ LLM fallback")
    print("‚Ä¢ Web crawling and scraping")
    print("‚Ä¢ Auto-installation of dependencies")
    print()
    
    # Check credentials if Google Sheets will be used
    use_gsheets = input("Use Google Sheets output? (y/N): ").strip().lower() == 'y'
    if use_gsheets:
        if not check_credentials():
            print("Google Sheets disabled due to credential issues.")
            use_gsheets = False
    
    # Get user inputs
    company_name = input("Enter company name: ").strip()
    ticker = input("Enter ticker symbol (optional): ").strip() or None
    website = input("Enter company website (optional): ").strip() or None
    
    # Determine output type
    if use_gsheets:
        output = input("Output type (excel/gsheet/both) [both]: ").strip() or "both"
    else:
        output = "excel"
    
    # Get Google Sheets details if needed
    sheet_name = "DCF"
    worksheet_name = "Microsoft"
    if output in ["gsheet", "both"]:
        sheet_name = input(f"Google Sheet name [{sheet_name}]: ").strip() or sheet_name
        worksheet_name = input(f"Worksheet/tab name [{worksheet_name}]: ").strip() or worksheet_name
    
    # Build the model
    build_dcf_model(company_name, ticker, website, output, sheet_name, worksheet_name)

if __name__ == "__main__":
    main_menu() 