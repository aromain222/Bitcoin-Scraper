#!/usr/bin/env python3
"""
Professional Trading Comparables (Comps) Model
Values a target company relative to its peer group using valuation multiples

Author: Investment Banking Modeler
Date: 2024

Features:
- Peer group analysis with financial metrics
- Valuation multiples calculation (EV/Revenue, EV/EBITDA, P/E)
- Statistical analysis (mean, median, percentiles)
- Implied valuation ranges for target company
- Professional Excel Output with Banker Formatting
- Summary table with valuation ranges
"""

import numpy as np
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# Professional color scheme
COMPS_MODEL_COLORS = {
    'header_blue': '1F4E79',
    'input_light_blue': 'DDEBF7',
    'calculation_black': '000000',
    'output_green': 'D8E4BC',
    'warning_red': 'F8CECC',
    'border_dark': '2C3E50',
    'text_dark': '000000',
    'text_white': 'FFFFFF',
    'peer_data_yellow': 'FFF2CC',
    'target_highlight': 'E6F3FF'
}

class ProfessionalTradingCompsModel:
    """
    Comprehensive Trading Comparables Model with Professional Formatting
    """

    def __init__(self, target_company="Target Company", target_ticker="TARGET"):
        self.target_company = target_company
        self.target_ticker = target_ticker
        self.model_date = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Initialize styles
        self.styles = self._create_styles()

    def _create_styles(self):
        """Create professional Excel styles"""
        styles = {}

        # Header style
        styles['header'] = NamedStyle(name='header')
        styles['header'].font = Font(name='Calibri', size=12, bold=True, color=COMPS_MODEL_COLORS['text_white'])
        styles['header'].fill = PatternFill(start_color=COMPS_MODEL_COLORS['header_blue'], end_color=COMPS_MODEL_COLORS['header_blue'], fill_type='solid')
        styles['header'].alignment = Alignment(horizontal='center', vertical='center')

        # Input style
        styles['input'] = NamedStyle(name='input')
        styles['input'].font = Font(name='Calibri', size=10, color=COMPS_MODEL_COLORS['text_dark'])
        styles['input'].fill = PatternFill(start_color=COMPS_MODEL_COLORS['input_light_blue'], end_color=COMPS_MODEL_COLORS['input_light_blue'], fill_type='solid')
        styles['input'].alignment = Alignment(horizontal='right', vertical='center')
        styles['input'].number_format = '#,##0.00'

        # Peer data style
        styles['peer_data'] = NamedStyle(name='peer_data')
        styles['peer_data'].font = Font(name='Calibri', size=10, color=COMPS_MODEL_COLORS['text_dark'])
        styles['peer_data'].fill = PatternFill(start_color=COMPS_MODEL_COLORS['peer_data_yellow'], end_color=COMPS_MODEL_COLORS['peer_data_yellow'], fill_type='solid')
        styles['peer_data'].alignment = Alignment(horizontal='right', vertical='center')
        styles['peer_data'].number_format = '#,##0.00'

        # Target highlight style
        styles['target_highlight'] = NamedStyle(name='target_highlight')
        styles['target_highlight'].font = Font(name='Calibri', size=10, bold=True, color=COMPS_MODEL_COLORS['text_dark'])
        styles['target_highlight'].fill = PatternFill(start_color=COMPS_MODEL_COLORS['target_highlight'], end_color=COMPS_MODEL_COLORS['target_highlight'], fill_type='solid')
        styles['target_highlight'].alignment = Alignment(horizontal='right', vertical='center')
        styles['target_highlight'].number_format = '#,##0.00'

        # Calculation style
        styles['calculation'] = NamedStyle(name='calculation')
        styles['calculation'].font = Font(name='Calibri', size=10, color=COMPS_MODEL_COLORS['calculation_black'])
        styles['calculation'].alignment = Alignment(horizontal='right', vertical='center')
        styles['calculation'].number_format = '#,##0.00'

        # Output style
        styles['output'] = NamedStyle(name='output')
        styles['output'].font = Font(name='Calibri', size=10, bold=True, color=COMPS_MODEL_COLORS['text_dark'])
        styles['output'].fill = PatternFill(start_color=COMPS_MODEL_COLORS['output_green'], end_color=COMPS_MODEL_COLORS['output_green'], fill_type='solid')
        styles['output'].alignment = Alignment(horizontal='right', vertical='center')
        styles['output'].number_format = '#,##0.00'

        # Statistics style
        styles['statistics'] = NamedStyle(name='statistics')
        styles['statistics'].font = Font(name='Calibri', size=10, bold=True, color=COMPS_MODEL_COLORS['text_dark'])
        styles['statistics'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        styles['statistics'].alignment = Alignment(horizontal='right', vertical='center')
        styles['statistics'].number_format = '#,##0.00'

        # Warning style
        styles['warning'] = NamedStyle(name='warning')
        styles['warning'].font = Font(name='Calibri', size=10, color=COMPS_MODEL_COLORS['text_dark'])
        styles['warning'].fill = PatternFill(start_color=COMPS_MODEL_COLORS['warning_red'], end_color=COMPS_MODEL_COLORS['warning_red'], fill_type='solid')
        styles['warning'].alignment = Alignment(horizontal='center', vertical='center')

        # Bold label style
        styles['label_bold'] = NamedStyle(name='label_bold')
        styles['label_bold'].font = Font(name='Calibri', size=10, bold=True, color=COMPS_MODEL_COLORS['text_dark'])
        styles['label_bold'].alignment = Alignment(horizontal='left', vertical='center')

        # Normal label style
        styles['label'] = NamedStyle(name='label')
        styles['label'].font = Font(name='Calibri', size=10, color=COMPS_MODEL_COLORS['text_dark'])
        styles['label'].alignment = Alignment(horizontal='left', vertical='center')

        # Summary box style
        styles['summary_box'] = NamedStyle(name='summary_box')
        styles['summary_box'].font = Font(name='Calibri', size=11, bold=True, color=COMPS_MODEL_COLORS['text_white'])
        styles['summary_box'].fill = PatternFill(start_color=COMPS_MODEL_COLORS['header_blue'], end_color=COMPS_MODEL_COLORS['header_blue'], fill_type='solid')
        styles['summary_box'].alignment = Alignment(horizontal='center', vertical='center')
        styles['summary_box'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        return styles

    def run_trading_comps_model(self,
                               # Target Company Data
                               target_revenue=1000.0,      # $M
                               target_ebitda=250.0,        # $M
                               target_net_income=150.0,    # $M
                               target_eps=2.50,            # $
                               target_net_debt=200.0,      # $M
                               target_shares_outstanding=80.0,  # Million shares

                               # Peer Group Data (as lists)
                               peer_names=None,
                               peer_tickers=None,
                               peer_revenues=None,          # $M
                               peer_ebitdas=None,          # $M
                               peer_net_incomes=None,      # $M
                               peer_epss=None,             # $
                               peer_net_debts=None,        # $M
                               peer_shares_outstanding=None,  # Million shares
                               peer_share_prices=None):    # $

        """
        Run complete trading comparables model with peer group analysis
        """

        print(f"📊 Building Professional Trading Comps Model for {self.target_company} ({self.target_ticker})")
        print("=" * 90)

        # Set default peer group if not provided
        if peer_names is None:
            peer_names = ["Peer A Corp", "Peer B Inc", "Peer C Ltd", "Peer D Corp", "Peer E LLC"]
            peer_tickers = ["PEERA", "PEERB", "PEERC", "PEERD", "PEERE"]
            peer_revenues = [800.0, 1200.0, 600.0, 1500.0, 900.0]
            peer_ebitdas = [180.0, 300.0, 120.0, 375.0, 225.0]
            peer_net_incomes = [120.0, 180.0, 80.0, 225.0, 135.0]
            peer_epss = [2.40, 3.60, 1.60, 4.50, 2.70]
            peer_net_debts = [150.0, 250.0, 100.0, 300.0, 180.0]
            peer_shares_outstanding = [50.0, 50.0, 50.0, 50.0, 50.0]
            peer_share_prices = [40.0, 60.0, 32.0, 75.0, 45.0]

        # Step 1: Create Assumptions & Peer Group Data
        assumptions = self._create_assumptions(
            target_revenue, target_ebitda, target_net_income, target_eps,
            target_net_debt, target_shares_outstanding,
            peer_names, peer_tickers, peer_revenues, peer_ebitdas,
            peer_net_incomes, peer_epss, peer_net_debts,
            peer_shares_outstanding, peer_share_prices
        )

        # Step 2: Calculate Peer Valuation Multiples
        peer_multiples = self._calculate_peer_multiples(assumptions)

        # Step 3: Compute Summary Statistics
        summary_stats = self._compute_summary_statistics(peer_multiples)

        # Step 4: Generate Target Valuation
        target_valuation = self._generate_target_valuation(assumptions, summary_stats)

        # Step 5: Create Valuation Ranges
        valuation_ranges = self._create_valuation_ranges(target_valuation)

        # Compile results
        comps_results = {
            'assumptions': assumptions,
            'peer_multiples': peer_multiples,
            'summary_stats': summary_stats,
            'target_valuation': target_valuation,
            'valuation_ranges': valuation_ranges
        }

        # Create Excel output
        excel_file = self._create_excel_output(comps_results)

        print("\n✅ Trading Comps Model Complete!")
        print("📊 Key Valuation Metrics:")
        print(f"   • Target Revenue: ${target_revenue:.0f}M | EBITDA: ${target_ebitda:.0f}M")
        print(f"   • Peer Group: {len(peer_names)} companies")
        print(f"   • Median EV/Revenue: {summary_stats['ev_revenue']['median']:.1f}x")
        print(f"   • Median EV/EBITDA: {summary_stats['ev_ebitda']['median']:.1f}x")
        print(f"   • Median P/E: {summary_stats['pe']['median']:.1f}x")
        print(f"   • Implied EV Range: ${valuation_ranges['ev_range']['low']/1000:.1f}B - ${valuation_ranges['ev_range']['high']/1000:.1f}B")
        print(f"   • Implied Share Price Range: ${valuation_ranges['share_price_range']['low']:.2f} - ${valuation_ranges['share_price_range']['high']:.2f}")
        print(f"📁 Excel Output: {excel_file}")

        return comps_results, excel_file

    def _create_assumptions(self, target_revenue, target_ebitda, target_net_income, target_eps,
                           target_net_debt, target_shares_outstanding,
                           peer_names, peer_tickers, peer_revenues, peer_ebitdas,
                           peer_net_incomes, peer_epss, peer_net_debts,
                           peer_shares_outstanding, peer_share_prices):

        """Create comprehensive assumptions and peer group data"""

        # Target company data
        target_data = {
            'company_name': self.target_company,
            'ticker': self.target_ticker,
            'revenue': target_revenue,
            'ebitda': target_ebitda,
            'net_income': target_net_income,
            'eps': target_eps,
            'net_debt': target_net_debt,
            'shares_outstanding': target_shares_outstanding,
            'market_cap': target_eps * target_shares_outstanding if target_eps > 0 else 0,  # Placeholder
            'enterprise_value': (target_eps * target_shares_outstanding) + target_net_debt if target_eps > 0 else target_net_debt
        }

        # Peer group data
        peers_data = []
        for i in range(len(peer_names)):
            peer = {
                'name': peer_names[i],
                'ticker': peer_tickers[i],
                'revenue': peer_revenues[i],
                'ebitda': peer_ebitdas[i],
                'net_income': peer_net_incomes[i],
                'eps': peer_epss[i],
                'net_debt': peer_net_debts[i],
                'shares_outstanding': peer_shares_outstanding[i],
                'share_price': peer_share_prices[i],
                'market_cap': peer_share_prices[i] * peer_shares_outstanding[i],
                'enterprise_value': (peer_share_prices[i] * peer_shares_outstanding[i]) + peer_net_debts[i]
            }
            peers_data.append(peer)

        assumptions = {
            'target': target_data,
            'peers': peers_data,
            'num_peers': len(peers_data)
        }

        print("📋 Assumptions Created:")
        print(f"   • Target: {self.target_company} (${target_revenue:.0f}M revenue, ${target_ebitda:.0f}M EBITDA)")
        print(f"   • Peer Group: {len(peers_data)} companies")
        print(f"   • Peer Revenue Range: ${min(peer_revenues):.0f}M - ${max(peer_revenues):.0f}M")
        print(f"   • Peer EBITDA Range: ${min(peer_ebitdas):.0f}M - ${max(peer_ebitdas):.0f}M")

        return assumptions

    def _calculate_peer_multiples(self, assumptions):
        """Calculate valuation multiples for each peer"""

        peer_multiples = []

        for peer in assumptions['peers']:
            multiples = {
                'name': peer['name'],
                'ticker': peer['ticker'],
                'ev_revenue': peer['enterprise_value'] / peer['revenue'] if peer['revenue'] > 0 else 0,
                'ev_ebitda': peer['enterprise_value'] / peer['ebitda'] if peer['ebitda'] > 0 else 0,
                'pe': peer['market_cap'] / peer['net_income'] if peer['net_income'] > 0 else 0,
                'price_to_eps': peer['share_price'] / peer['eps'] if peer['eps'] > 0 else 0,
                'market_cap': peer['market_cap'],
                'enterprise_value': peer['enterprise_value'],
                'revenue': peer['revenue'],
                'ebitda': peer['ebitda'],
                'net_income': peer['net_income'],
                'eps': peer['eps']
            }
            peer_multiples.append(multiples)

        print("📊 Peer Multiples Calculated:")
        ev_revenue_ratios = [p['ev_revenue'] for p in peer_multiples]
        ev_ebitda_ratios = [p['ev_ebitda'] for p in peer_multiples]
        pe_ratios = [p['pe'] for p in peer_multiples]

        print(f"   • EV/Revenue Range: {min(ev_revenue_ratios):.1f}x - {max(ev_revenue_ratios):.1f}x")
        print(f"   • EV/EBITDA Range: {min(ev_ebitda_ratios):.1f}x - {max(ev_ebitda_ratios):.1f}x")
        print(f"   • P/E Range: {min(pe_ratios):.1f}x - {max(pe_ratios):.1f}x")

        return peer_multiples

    def _compute_summary_statistics(self, peer_multiples):
        """Compute summary statistics for peer multiples"""

        # Extract multiple arrays
        ev_revenue_ratios = [p['ev_revenue'] for p in peer_multiples]
        ev_ebitda_ratios = [p['ev_ebitda'] for p in peer_multiples]
        pe_ratios = [p['pe'] for p in peer_multiples]
        price_to_eps_ratios = [p['price_to_eps'] for p in peer_multiples]

        def calculate_stats(ratios):
            """Calculate mean, median, and percentiles for a list of ratios"""
            if not ratios or all(r == 0 for r in ratios):
                return {'mean': 0, 'median': 0, 'p25': 0, 'p75': 0, 'min': 0, 'max': 0}

            valid_ratios = [r for r in ratios if r > 0]  # Filter out zero ratios
            if not valid_ratios:
                return {'mean': 0, 'median': 0, 'p25': 0, 'p75': 0, 'min': 0, 'max': 0}

            return {
                'mean': np.mean(valid_ratios),
                'median': np.median(valid_ratios),
                'p25': np.percentile(valid_ratios, 25),
                'p75': np.percentile(valid_ratios, 75),
                'min': min(valid_ratios),
                'max': max(valid_ratios)
            }

        summary_stats = {
            'ev_revenue': calculate_stats(ev_revenue_ratios),
            'ev_ebitda': calculate_stats(ev_ebitda_ratios),
            'pe': calculate_stats(pe_ratios),
            'price_to_eps': calculate_stats(price_to_eps_ratios)
        }

        print("📈 Summary Statistics Computed:")
        print(f"   • EV/Revenue: Mean {summary_stats['ev_revenue']['mean']:.1f}x, Median {summary_stats['ev_revenue']['median']:.1f}x")
        print(f"   • EV/EBITDA: Mean {summary_stats['ev_ebitda']['mean']:.1f}x, Median {summary_stats['ev_ebitda']['median']:.1f}x")
        print(f"   • P/E: Mean {summary_stats['pe']['mean']:.1f}x, Median {summary_stats['pe']['median']:.1f}x")

        return summary_stats

    def _generate_target_valuation(self, assumptions, summary_stats):
        """Generate implied valuation for target using peer multiples"""

        target = assumptions['target']

        # Apply multiples to target financials
        target_valuation = {
            'ev_revenue': {
                'p25_multiple': summary_stats['ev_revenue']['p25'],
                'median_multiple': summary_stats['ev_revenue']['median'],
                'p75_multiple': summary_stats['ev_revenue']['p75'],
                'p25_ev': target['revenue'] * summary_stats['ev_revenue']['p25'],
                'median_ev': target['revenue'] * summary_stats['ev_revenue']['median'],
                'p75_ev': target['revenue'] * summary_stats['ev_revenue']['p75']
            },
            'ev_ebitda': {
                'p25_multiple': summary_stats['ev_ebitda']['p25'],
                'median_multiple': summary_stats['ev_ebitda']['median'],
                'p75_multiple': summary_stats['ev_ebitda']['p75'],
                'p25_ev': target['ebitda'] * summary_stats['ev_ebitda']['p25'],
                'median_ev': target['ebitda'] * summary_stats['ev_ebitda']['median'],
                'p75_ev': target['ebitda'] * summary_stats['ev_ebitda']['p75']
            },
            'pe': {
                'p25_multiple': summary_stats['pe']['p25'],
                'median_multiple': summary_stats['pe']['median'],
                'p75_multiple': summary_stats['pe']['p75'],
                'p25_market_cap': target['net_income'] * summary_stats['pe']['p25'],
                'median_market_cap': target['net_income'] * summary_stats['pe']['median'],
                'p75_market_cap': target['net_income'] * summary_stats['pe']['p75']
            },
            'price_to_eps': {
                'p25_multiple': summary_stats['price_to_eps']['p25'],
                'median_multiple': summary_stats['price_to_eps']['median'],
                'p75_multiple': summary_stats['price_to_eps']['p75'],
                'p25_share_price': target['eps'] * summary_stats['price_to_eps']['p25'],
                'median_share_price': target['eps'] * summary_stats['price_to_eps']['median'],
                'p75_share_price': target['eps'] * summary_stats['price_to_eps']['p75']
            }
        }

        print("🎯 Target Valuation Generated:")
        print(f"   • EV/Revenue: ${target_valuation['ev_revenue']['median_ev']/1000:.1f}B (median)")
        print(f"   • EV/EBITDA: ${target_valuation['ev_ebitda']['median_ev']/1000:.1f}B (median)")
        print(f"   • P/E: ${target_valuation['pe']['median_market_cap']/1000:.1f}B market cap (median)")

        return target_valuation

    def _create_valuation_ranges(self, target_valuation):
        """Create valuation ranges from target valuation data"""

        # Use EV/EBITDA as primary valuation method (most common for comps)
        ev_range = {
            'low': target_valuation['ev_ebitda']['p25_ev'],
            'median': target_valuation['ev_ebitda']['median_ev'],
            'high': target_valuation['ev_ebitda']['p75_ev']
        }

        # Calculate equity value ranges (EV - Net Debt)
        equity_range = {
            'low': ev_range['low'] - target_valuation['target_net_debt'] if 'target_net_debt' in target_valuation else ev_range['low'],
            'median': ev_range['median'] - target_valuation['target_net_debt'] if 'target_net_debt' in target_valuation else ev_range['median'],
            'high': ev_range['high'] - target_valuation['target_net_debt'] if 'target_net_debt' in target_valuation else ev_range['high']
        }

        # Calculate share price ranges (Market Cap / Shares Outstanding)
        target_shares = 80.0  # Default, should be passed in
        share_price_range = {
            'low': equity_range['low'] / target_shares,
            'median': equity_range['median'] / target_shares,
            'high': equity_range['high'] / target_shares
        }

        valuation_ranges = {
            'ev_range': ev_range,
            'equity_range': equity_range,
            'share_price_range': share_price_range,
            'primary_multiple': 'EV/EBITDA',
            'range_type': '25th - 75th Percentile'
        }

        print("📊 Valuation Ranges Created:")
        print(f"   • Enterprise Value: ${ev_range['low']/1000:.1f}B - ${ev_range['high']/1000:.1f}B")
        print(f"   • Equity Value: ${equity_range['low']/1000:.1f}B - ${equity_range['high']/1000:.1f}B")
        print(f"   • Share Price: ${share_price_range['low']:.2f} - ${share_price_range['high']:.2f}")

        return valuation_ranges

    def _create_excel_output(self, comps_results):
        """Create professional Excel output with multiple tabs"""

        wb = Workbook()

        # Create worksheets
        ws_summary = wb.active
        ws_summary.title = "Valuation Summary"

        ws_assumptions = wb.create_sheet("Assumptions & Peers")
        ws_peer_multiples = wb.create_sheet("Peer Multiples")
        ws_target_valuation = wb.create_sheet("Target Valuation")

        # Apply styles to workbook
        for style_name, style in self.styles.items():
            if style_name not in wb.named_styles:
                wb.add_named_style(style)

        # Create each tab
        self._create_summary_tab(ws_summary, comps_results)
        self._create_assumptions_tab(ws_assumptions, comps_results)
        self._create_peer_multiples_tab(ws_peer_multiples, comps_results)
        self._create_target_valuation_tab(ws_target_valuation, comps_results)

        # Save workbook
        filename = f"Trading_Comps_{self.target_ticker}_{self.model_date}.xlsx"
        wb.save(filename)

        return filename

    def _create_summary_tab(self, ws, results):
        """Create Valuation Summary tab with banker-style summary boxes"""

        # Title
        ws['A1'] = f"Trading Comparables Analysis: {self.target_company} ({self.target_ticker})"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:G1')

        current_row = 3

        # Trading Comps Summary Box
        ws[f'A{current_row}'] = 'TRADING COMPARABLES SUMMARY'
        ws[f'A{current_row}'].style = 'summary_box'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        # Multiples summary
        summary_stats = results['summary_stats']
        multiples_data = [
            ("EV/Revenue", summary_stats['ev_revenue']['p25'], summary_stats['ev_revenue']['median'], summary_stats['ev_revenue']['p75']),
            ("EV/EBITDA", summary_stats['ev_ebitda']['p25'], summary_stats['ev_ebitda']['median'], summary_stats['ev_ebitda']['p75']),
            ("P/E", summary_stats['pe']['p25'], summary_stats['pe']['median'], summary_stats['pe']['p75'])
        ]

        # Headers
        ws.cell(row=current_row, column=1, value="Multiple").style = 'header'
        ws.cell(row=current_row, column=2, value="25th Percentile").style = 'header'
        ws.cell(row=current_row, column=3, value="Median").style = 'header'
        ws.cell(row=current_row, column=4, value="75th Percentile").style = 'header'
        current_row += 1

        # Data
        for multiple, p25, median, p75 in multiples_data:
            ws.cell(row=current_row, column=1, value=multiple).style = 'label_bold'
            ws.cell(row=current_row, column=2, value=p25).style = 'output'
            ws.cell(row=current_row, column=3, value=median).style = 'output'
            ws.cell(row=current_row, column=4, value=p75).style = 'output'
            current_row += 1

        current_row += 2

        # Target Valuation Summary Box
        ws[f'A{current_row}'] = 'TARGET VALUATION SUMMARY'
        ws[f'A{current_row}'].style = 'summary_box'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        # Valuation ranges
        valuation_ranges = results['valuation_ranges']
        valuation_data = [
            ("Implied Enterprise Value", f"${valuation_ranges['ev_range']['low']/1000:.1f}B", f"${valuation_ranges['ev_range']['median']/1000:.1f}B", f"${valuation_ranges['ev_range']['high']/1000:.1f}B"),
            ("Implied Equity Value", f"${valuation_ranges['equity_range']['low']/1000:.1f}B", f"${valuation_ranges['equity_range']['median']/1000:.1f}B", f"${valuation_ranges['equity_range']['high']/1000:.1f}B"),
            ("Implied Share Price", f"${valuation_ranges['share_price_range']['low']:.2f}", f"${valuation_ranges['share_price_range']['median']:.2f}", f"${valuation_ranges['share_price_range']['high']:.2f}")
        ]

        # Headers
        ws.cell(row=current_row, column=1, value="Metric").style = 'header'
        ws.cell(row=current_row, column=2, value="Low").style = 'header'
        ws.cell(row=current_row, column=3, value="Median").style = 'header'
        ws.cell(row=current_row, column=4, value="High").style = 'header'
        current_row += 1

        # Data
        for metric, low, median, high in valuation_data:
            ws.cell(row=current_row, column=1, value=metric).style = 'label_bold'
            ws.cell(row=current_row, column=2, value=low).style = 'output'
            ws.cell(row=current_row, column=3, value=median).style = 'output'
            ws.cell(row=current_row, column=4, value=high).style = 'output'
            current_row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    def _create_assumptions_tab(self, ws, results):
        """Create Assumptions & Peers tab"""

        assumptions = results['assumptions']

        # Title
        ws['A1'] = f"Assumptions & Peer Group: {self.target_company} vs Peers"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:K1')

        current_row = 3

        # Target Company Data
        ws[f'A{current_row}'] = "TARGET COMPANY DATA"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:K{current_row}')
        current_row += 2

        target_data = [
            ("Company Name", assumptions['target']['company_name']),
            ("Ticker", assumptions['target']['ticker']),
            ("Revenue ($M)", assumptions['target']['revenue']),
            ("EBITDA ($M)", assumptions['target']['ebitda']),
            ("Net Income ($M)", assumptions['target']['net_income']),
            ("EPS ($)", assumptions['target']['eps']),
            ("Net Debt ($M)", assumptions['target']['net_debt']),
            ("Shares Outstanding (M)", assumptions['target']['shares_outstanding'])
        ]

        for label, value in target_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            if label == "Company Name" or label == "Ticker":
                ws[f'B{current_row}'].style = 'input'
            else:
                ws[f'B{current_row}'].style = 'target_highlight'
            current_row += 1

        current_row += 2

        # Peer Group Data
        ws[f'A{current_row}'] = "PEER GROUP DATA"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:K{current_row}')
        current_row += 2

        # Headers
        headers = ['Company', 'Ticker', 'Revenue ($M)', 'EBITDA ($M)', 'Net Income ($M)',
                  'EPS ($)', 'Net Debt ($M)', 'Shares (M)', 'Share Price ($)', 'Market Cap ($M)', 'EV ($M)']

        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header).style = 'header'

        current_row += 1

        # Peer data
        for peer in assumptions['peers']:
            ws.cell(row=current_row, column=1, value=peer['name']).style = 'peer_data'
            ws.cell(row=current_row, column=2, value=peer['ticker']).style = 'peer_data'
            ws.cell(row=current_row, column=3, value=peer['revenue']).style = 'peer_data'
            ws.cell(row=current_row, column=4, value=peer['ebitda']).style = 'peer_data'
            ws.cell(row=current_row, column=5, value=peer['net_income']).style = 'peer_data'
            ws.cell(row=current_row, column=6, value=peer['eps']).style = 'peer_data'
            ws.cell(row=current_row, column=7, value=peer['net_debt']).style = 'peer_data'
            ws.cell(row=current_row, column=8, value=peer['shares_outstanding']).style = 'peer_data'
            ws.cell(row=current_row, column=9, value=peer['share_price']).style = 'peer_data'
            ws.cell(row=current_row, column=10, value=peer['market_cap']).style = 'peer_data'
            ws.cell(row=current_row, column=11, value=peer['enterprise_value']).style = 'peer_data'
            current_row += 1

        # Set column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_peer_multiples_tab(self, ws, results):
        """Create Peer Multiples tab"""

        peer_multiples = results['peer_multiples']
        summary_stats = results['summary_stats']

        # Title
        ws['A1'] = f"Peer Valuation Multiples: {len(peer_multiples)} Companies"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:I1')

        current_row = 3

        # Peer Multiples Table
        ws[f'A{current_row}'] = "PEER VALUATION MULTIPLES"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:I{current_row}')
        current_row += 2

        # Headers
        headers = ['Company', 'EV/Revenue', 'EV/EBITDA', 'P/E', 'Price/EPS', 'Market Cap ($M)', 'EV ($M)']

        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header).style = 'header'

        current_row += 1

        # Peer data
        for peer in peer_multiples:
            ws.cell(row=current_row, column=1, value=peer['ticker']).style = 'peer_data'
            ws.cell(row=current_row, column=2, value=peer['ev_revenue']).style = 'peer_data'
            ws.cell(row=current_row, column=3, value=peer['ev_ebitda']).style = 'peer_data'
            ws.cell(row=current_row, column=4, value=peer['pe']).style = 'peer_data'
            ws.cell(row=current_row, column=5, value=peer['price_to_eps']).style = 'peer_data'
            ws.cell(row=current_row, column=6, value=peer['market_cap']).style = 'peer_data'
            ws.cell(row=current_row, column=7, value=peer['enterprise_value']).style = 'peer_data'
            current_row += 1

        # Add summary statistics
        current_row += 2
        ws[f'A{current_row}'] = "SUMMARY STATISTICS"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:I{current_row}')
        current_row += 2

        # Statistics headers
        ws.cell(row=current_row, column=1, value="Statistic").style = 'header'
        ws.cell(row=current_row, column=2, value="EV/Revenue").style = 'header'
        ws.cell(row=current_row, column=3, value="EV/EBITDA").style = 'header'
        ws.cell(row=current_row, column=4, value="P/E").style = 'header'
        current_row += 1

        # Statistics data
        stats_labels = ['Mean', 'Median', '25th Percentile', '75th Percentile', 'Min', 'Max']
        stats_keys = ['mean', 'median', 'p25', 'p75', 'min', 'max']

        for i, (label, key) in enumerate(zip(stats_labels, stats_keys)):
            ws.cell(row=current_row, column=1, value=label).style = 'statistics'
            ws.cell(row=current_row, column=2, value=summary_stats['ev_revenue'][key]).style = 'statistics'
            ws.cell(row=current_row, column=3, value=summary_stats['ev_ebitda'][key]).style = 'statistics'
            ws.cell(row=current_row, column=4, value=summary_stats['pe'][key]).style = 'statistics'
            current_row += 1

        # Set column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_target_valuation_tab(self, ws, results):
        """Create Target Valuation tab"""

        target_valuation = results['target_valuation']
        valuation_ranges = results['valuation_ranges']

        # Title
        ws['A1'] = f"Target Valuation: {self.target_company} ({self.target_ticker})"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:G1')

        current_row = 3

        # EV/Revenue Valuation
        ws[f'A{current_row}'] = "EV/REVENUE VALUATION"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        ev_rev_data = [
            ("Multiple", target_valuation['ev_revenue']['p25_multiple'], target_valuation['ev_revenue']['median_multiple'], target_valuation['ev_revenue']['p75_multiple']),
            ("Enterprise Value ($M)", target_valuation['ev_revenue']['p25_ev'], target_valuation['ev_revenue']['median_ev'], target_valuation['ev_revenue']['p75_ev'])
        ]

        for label, p25, median, p75 in ev_rev_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = p25
            ws[f'C{current_row}'] = median
            ws[f'D{current_row}'] = p75
            ws[f'B{current_row}'].style = 'calculation'
            ws[f'C{current_row}'].style = 'calculation'
            ws[f'D{current_row}'].style = 'calculation'
            current_row += 1

        current_row += 2

        # EV/EBITDA Valuation
        ws[f'A{current_row}'] = "EV/EBITDA VALUATION"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        ev_ebitda_data = [
            ("Multiple", target_valuation['ev_ebitda']['p25_multiple'], target_valuation['ev_ebitda']['median_multiple'], target_valuation['ev_ebitda']['p75_multiple']),
            ("Enterprise Value ($M)", target_valuation['ev_ebitda']['p25_ev'], target_valuation['ev_ebitda']['median_ev'], target_valuation['ev_ebitda']['p75_ev'])
        ]

        for label, p25, median, p75 in ev_ebitda_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = p25
            ws[f'C{current_row}'] = median
            ws[f'D{current_row}'] = p75
            if label == "Enterprise Value ($M)":
                ws[f'B{current_row}'].style = 'output'
                ws[f'C{current_row}'].style = 'output'
                ws[f'D{current_row}'].style = 'output'
            else:
                ws[f'B{current_row}'].style = 'calculation'
                ws[f'C{current_row}'].style = 'calculation'
                ws[f'D{current_row}'].style = 'calculation'
            current_row += 1

        current_row += 2

        # P/E Valuation
        ws[f'A{current_row}'] = "P/E VALUATION"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        pe_data = [
            ("Multiple", target_valuation['pe']['p25_multiple'], target_valuation['pe']['median_multiple'], target_valuation['pe']['p75_multiple']),
            ("Market Cap ($M)", target_valuation['pe']['p25_market_cap'], target_valuation['pe']['median_market_cap'], target_valuation['pe']['p75_market_cap'])
        ]

        for label, p25, median, p75 in pe_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = p25
            ws[f'C{current_row}'] = median
            ws[f'D{current_row}'] = p75
            if label == "Market Cap ($M)":
                ws[f'B{current_row}'].style = 'output'
                ws[f'C{current_row}'].style = 'output'
                ws[f'D{current_row}'].style = 'output'
            else:
                ws[f'B{current_row}'].style = 'calculation'
                ws[f'C{current_row}'].style = 'calculation'
                ws[f'D{current_row}'].style = 'calculation'
            current_row += 1

        current_row += 2

        # Valuation Summary
        ws[f'A{current_row}'] = "VALUATION SUMMARY (PRIMARY: EV/EBITDA)"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        valuation_summary = [
            ("Enterprise Value Range", f"${valuation_ranges['ev_range']['low']/1000:.1f}B", f"${valuation_ranges['ev_range']['median']/1000:.1f}B", f"${valuation_ranges['ev_range']['high']/1000:.1f}B"),
            ("Equity Value Range", f"${valuation_ranges['equity_range']['low']/1000:.1f}B", f"${valuation_ranges['equity_range']['median']/1000:.1f}B", f"${valuation_ranges['equity_range']['high']/1000:.1f}B"),
            ("Implied Share Price", f"${valuation_ranges['share_price_range']['low']:.2f}", f"${valuation_ranges['share_price_range']['median']:.2f}", f"${valuation_ranges['share_price_range']['high']:.2f}")
        ]

        for label, low, median, high in valuation_summary:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = low
            ws[f'C{current_row}'] = median
            ws[f'D{current_row}'] = high
            ws[f'B{current_row}'].style = 'output'
            ws[f'C{current_row}'].style = 'output'
            ws[f'D{current_row}'].style = 'output'
            current_row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15


def run_sample_trading_comps_model():
    """Run a sample trading comparables model"""

    print("📊 Running Professional Trading Comps Model Sample")
    print("=" * 70)

    # Create model instance
    model = ProfessionalTradingCompsModel("TechTarget Inc.", "TECHTARGET")

    # Run the model with sample data
    results, excel_file = model.run_trading_comps_model(
        # Target Company Data
        target_revenue=1200.0,      # $1.2B revenue
        target_ebitda=360.0,        # $360M EBITDA
        target_net_income=216.0,    # $216M net income
        target_eps=3.60,            # $3.60 EPS
        target_net_debt=240.0,      # $240M net debt
        target_shares_outstanding=60.0  # 60M shares
    )

    return results, excel_file


if __name__ == "__main__":
    # Run sample trading comps model
    results, excel_file = run_sample_trading_comps_model()

    print("\n📋 Model Complete!")
    print(f"📁 Excel file saved as: {excel_file}")

    # Display key results
    print("\n🔑 Key Results Summary:")
    print("=" * 40)

    valuation_ranges = results['valuation_ranges']
    summary_stats = results['summary_stats']

    print(f"Median EV/Revenue: {summary_stats['ev_revenue']['median']:.1f}x")
    print(f"Median EV/EBITDA: {summary_stats['ev_ebitda']['median']:.1f}x")
    print(f"Median P/E: {summary_stats['pe']['median']:.1f}x")
    print(f"Implied EV Range: ${valuation_ranges['ev_range']['low']/1000:.1f}B - ${valuation_ranges['ev_range']['high']/1000:.1f}B")
    print(f"Implied Share Price Range: ${valuation_ranges['share_price_range']['low']:.2f} - ${valuation_ranges['share_price_range']['high']:.2f}")
