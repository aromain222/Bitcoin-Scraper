#!/usr/bin/env python3
"""
FinModAI Excel Generation Engine
Creates professional Excel files with advanced formatting and calculations.
"""

import os
import logging
from typing import Dict, List, Any, Optional, Union
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.drawing.image import Image
from openpyxl.chart import LineChart, BarChart, ScatterChart, Reference

logger = logging.getLogger('FinModAI.ExcelEngine')

class ExcelGenerationEngine:
    """Professional Excel file generation engine."""

    def __init__(self, config):
        self.config = config
        self.templates_dir = Path(config.model_templates_dir)
        self.output_dir = Path(config.output_dir)
        self.output_dir.mkdir(exist_ok=True)

        # Display currency values in USD millions throughout workbooks
        self.currency_scale = 1_000_000  # USD millions
        self.currency_note = "All figures in USD millions (MM) unless noted"

        # Define professional styling
        self._setup_styles()

        logger.info("📊 Excel Generation Engine initialized")

    def _setup_styles(self):
        """Setup professional styling templates."""
        # Create named styles that can be applied to cells
        self.header_style = NamedStyle(name='header')
        self.header_style.font = Font(size=12, bold=True, color='FFFFFF')
        self.header_style.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        self.header_style.alignment = Alignment(horizontal='center', vertical='center')
        self.header_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        self.subheader_style = NamedStyle(name='subheader')
        self.subheader_style.font = Font(size=11, bold=True, color='000000')
        self.subheader_style.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        self.subheader_style.alignment = Alignment(horizontal='left', vertical='center')
        self.subheader_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        self.input_style = NamedStyle(name='input_cell')
        self.input_style.font = Font(color='1F4E79')
        self.input_style.alignment = Alignment(horizontal='right')
        self.input_style.number_format = '#,##0.0'

        self.calculation_style = NamedStyle(name='calculation_cell')
        self.calculation_style.font = Font(color='000000')
        self.calculation_style.alignment = Alignment(horizontal='right')
        self.calculation_style.number_format = '#,##0.00'

        self.output_style = NamedStyle(name='output_cell')
        self.output_style.font = Font(color='2E7D32', bold=True)
        self.output_style.fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
        self.output_style.alignment = Alignment(horizontal='right')
        self.output_style.number_format = '#,##0.00'

        # Store styles dictionary for easy reference
        self.styles = {
            'header': self.header_style,
            'subheader': self.subheader_style,
            'input_cell': self.input_style,
            'calculation_cell': self.calculation_style,
            'output_cell': self.output_style,
            'currency': {'number_format': '"$"#,##0;[Red]-"$"#,##0'},
            'currency_decimals': {'number_format': '"$"#,##0.0;[Red]-"$"#,##0.0'},
            'percent': {'number_format': '0.0%'},
            'multiple': {'number_format': '0.0"x"'},
            'price': {'number_format': '"$"#,##0.00'}
        }

    def generate_output(
        self,
        model_spec: Any,
        output_format: str = "excel"
    ) -> List[str]:
        """
        Generate output files from model specification.

        Args:
            model_spec: ModelSpecification object
            output_format: Output format (excel, pdf, json)

        Returns:
            List of generated file paths
        """

        logger.info(f"📄 Generating {output_format.upper()} output for {model_spec.model_type}")
        print(f"🔧 DEBUG: Excel engine generate_output called with model_type: {model_spec.model_type}")

        if output_format.lower() == "excel":
            files = self._generate_excel_output(model_spec)
            print(f"🔧 DEBUG: Excel engine generated {len(files)} files: {files}")
            return files
        elif output_format.lower() == "json":
            return self._generate_json_output(model_spec)
        else:
            raise ValueError(f"Unsupported output format: {output_format}")

    def _generate_excel_output(self, model_spec: Any) -> List[str]:
        """Generate professional Excel file."""
        wb = Workbook()

        # Register named styles with the workbook
        if 'header' not in wb.named_styles:
            wb.add_named_style(self.header_style)
        if 'subheader' not in wb.named_styles:
            wb.add_named_style(self.subheader_style)
        if 'input_cell' not in wb.named_styles:
            wb.add_named_style(self.input_style)
        if 'calculation_cell' not in wb.named_styles:
            wb.add_named_style(self.calculation_style)
        if 'output_cell' not in wb.named_styles:
            wb.add_named_style(self.output_style)

        # Remove default sheet
        wb.remove(wb.active)

        # Create sheets based on model type
        if model_spec.model_type == 'dcf':
            self._create_dcf_excel(wb, model_spec)
        elif model_spec.model_type == 'lbo':
            self._create_lbo_excel(wb, model_spec)
        elif model_spec.model_type == 'comps':
            self._create_comps_excel(wb, model_spec)
        elif model_spec.model_type == 'three_statement':
            self._create_three_statement_excel(wb, model_spec)

        # Add summary sheet
        self._create_summary_sheet(wb, model_spec)

        # Add sensitivity analysis sheet
        if model_spec.sensitivity_analysis:
            self._create_sensitivity_sheet(wb, model_spec)

        # Add dashboard sheet
        if model_spec.visualization_config:
            self._create_dashboard_sheet(wb, model_spec)

        # Setup workbook properties
        company_name = getattr(model_spec.company_data, 'company_name', 'Unknown')
        wb.properties.title = f"{model_spec.model_type.upper()} Model - {company_name}"
        wb.properties.creator = "FinModAI"
        wb.properties.created = datetime.now()

        # Generate filename
        safe_name = "".join(c for c in company_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"FinModAI_{model_spec.model_type.upper()}_{safe_name}_{timestamp}.xlsx"
        filepath = self.output_dir / filename

        # Save workbook
        wb.save(filepath)

        logger.info(f"✅ Excel file generated: {filepath}")
        return [str(filepath)]

    def _create_dcf_excel(self, wb: Workbook, model_spec: Any):
        """Create DCF-specific Excel sheets."""
        # Assumptions sheet
        ws_assumptions = wb.create_sheet("Assumptions")
        self._populate_assumptions_sheet(ws_assumptions, model_spec)

        # Forecast sheet
        ws_forecast = wb.create_sheet("Forecast_FCF")
        self._populate_forecast_sheet(ws_forecast, model_spec)

        # DCF Valuation sheet
        ws_dcf = wb.create_sheet("DCF_Valuation")
        self._populate_dcf_sheet(ws_dcf, model_spec)

        # Append currency scale note
        for sheet in (ws_assumptions, ws_forecast, ws_dcf):
            note_row = sheet.max_row + 2
            sheet[f'A{note_row}'] = self.currency_note
            sheet[f'A{note_row}'].font = Font(italic=True, size=9, color='666666')
            sheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)

    def _populate_assumptions_sheet(self, ws, model_spec: Any):
        """Populate assumptions sheet with professional formatting."""
        assumptions = model_spec.assumptions
        company_data = model_spec.company_data

        # Handle FinancialData object attributes
        company_name = getattr(company_data, 'company_name', 'Company')
        ticker = getattr(company_data, 'ticker', 'N/A')
        sector = getattr(company_data, 'sector', 'N/A')
        revenue = getattr(company_data, 'revenue', 0)
        beta = getattr(company_data, 'beta', 1.2)

        # Title
        ws['A1'] = f"{company_name} - DCF Assumptions"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:E1')

        # Company info section
        ws['A3'] = "Company Information"
        ws['A3'].style = 'subheader'
        ws.merge_cells('A3:B3')

        row = 4
        ws[f'A{row}'] = "Company Name:"
        ws[f'B{row}'] = company_name

        row += 1
        ws[f'A{row}'] = "Ticker:"
        ws[f'B{row}'] = ticker

        row += 1
        ws[f'A{row}'] = "Sector:"
        ws[f'B{row}'] = sector

        # Timing section
        row += 2
        ws[f'A{row}'] = "Timing & Horizon"
        ws[f'A{row}'].style = 'subheader'
        ws.merge_cells(f'A{row}:B{row}')

        row += 1
        ws[f'A{row}'] = "Base Year:"
        ws[f'B{row}'] = assumptions.get('base_year', datetime.now().year)

        row += 1
        ws[f'A{row}'] = "Forecast Years:"
        ws[f'B{row}'] = assumptions.get('forecast_years', 6)

        # Financial assumptions
        row += 2
        ws[f'A{row}'] = "Financial Assumptions"
        ws[f'A{row}'].style = 'subheader'
        ws.merge_cells(f'A{row}:B{row}')

        base_revenue_value = assumptions.get('base_revenue', revenue * 1_000_000_000)
        base_ebitda_value = assumptions.get('base_ebitda', base_revenue_value * assumptions.get('ebitda_margin', 0.25))

        assumptions_data = [
            ("Starting Revenue ($MM)", base_revenue_value / self.currency_scale),
            ("Revenue Growth Rate", assumptions.get('growth_rate', 0.08)),
            ("EBITDA Margin", assumptions.get('ebitda_margin', 0.25)),
            ("Base EBITDA ($MM)", base_ebitda_value / self.currency_scale),
            ("D&A % of Revenue", assumptions.get('da_pct', 0.05)),
            ("CapEx % of Revenue", assumptions.get('capex_pct', 0.06)),
            ("NWC % of Revenue", assumptions.get('nwc_pct', 0.02)),
            ("Cash Tax Rate", assumptions.get('tax_rate', 0.25))
        ]

        for label, value in assumptions_data:
            row += 1
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value

            # Apply formatting
            if '%' in label or 'Rate' in label or 'Margin' in label:
                ws[f'B{row}'].number_format = '0.0%'
                ws[f'B{row}'].font = Font(color='1F4E79')  # Input blue
            elif 'Revenue' in label or 'EBITDA' in label:
                ws[f'B{row}'].number_format = '"$"#,##0;[Red]-"$"#,##0'
                ws[f'B{row}'].font = Font(color='1F4E79')  # Input blue
            else:
                ws[f'B{row}'].font = Font(color='1F4E79')  # Input blue

        # WACC section
        row += 2
        ws[f'A{row}'] = "WACC Calculation"
        ws[f'A{row}'].style = 'subheader'
        ws.merge_cells(f'A{row}:B{row}')

        wacc_data = [
            ("Risk-Free Rate", assumptions.get('risk_free_rate', 0.045)),
            ("Beta", beta),
            ("Market Risk Premium", assumptions.get('market_risk_premium', 0.06)),
            ("Cost of Equity", assumptions.get('cost_of_equity', 0.105)),
            ("Cost of Debt", assumptions.get('cost_of_debt', 0.055)),
            ("Debt Ratio", assumptions.get('debt_ratio', 0.40)),
            ("WACC", model_spec.calculations.get('wacc', 0.092))
        ]

        for label, value in wacc_data:
            row += 1
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = '0.0%'
            if label == "WACC":
                ws[f'B{row}'].font = Font(color='2E7D32', bold=True)  # Output green
            else:
                ws[f'B{row}'].font = Font(color='1F4E79')  # Input blue

        # Add units note
        note_row = row + 3
        ws[f'A{note_row}'] = self.currency_note
        ws[f'A{note_row}'].font = Font(italic=True, size=9, color='666666')
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)

        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15

        # Print setup
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.fitToWidth = 1

    def _populate_forecast_sheet(self, ws, model_spec: Any):
        """Populate forecast sheet with projections."""
        calculations = model_spec.calculations
        assumptions = model_spec.assumptions

        # Title
        ws['A1'] = "Revenue & Cash Flow Forecast"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:I1')

        # Header row
        ws['A3'] = "Line Item"
        ws['A3'].style = 'header'
        ws['B3'] = "Base Year"
        ws['B3'].style = 'header'

        # Year headers
        forecast_years = assumptions.get('forecast_years', 6)
        base_year = assumptions.get('base_year', datetime.now().year)

        for i in range(forecast_years):
            col = get_column_letter(i + 3)
            ws[f'{col}3'] = base_year + i + 1
            ws[f'{col}3'].style = 'header'

        # Revenue section
        row = 5
        ws[f'A{row}'] = "Revenue ($MM)"
        ws[f'A{row}'].style = 'subheader'
        ws.merge_cells(f'A{row}:B{row}')

        # Revenue data
        revenue_proj = calculations.get('revenue_projection', [])

        # Handle both dict and FinancialData object
        if hasattr(model_spec.company_data, 'revenue'):
            base_revenue = model_spec.company_data.revenue
            base_ebitda = model_spec.company_data.ebitda if hasattr(model_spec.company_data, 'ebitda') else base_revenue * 0.25
        else:
            base_revenue = getattr(model_spec.company_data, 'revenue', 1000)
            base_ebitda = getattr(model_spec.company_data, 'ebitda', base_revenue * 0.25)

        base_revenue_value = model_spec.assumptions.get('base_revenue', base_revenue * 1_000_000_000)
        ws[f'B{row + 1}'] = base_revenue_value / self.currency_scale
        ws[f'B{row + 1}'].number_format = '"$"#,##0;[Red]-"$"#,##0'

        for i, rev in enumerate(revenue_proj):
            col = get_column_letter(i + 3)
            ws[f'{col}{row + 1}'] = rev / self.currency_scale
            ws[f'{col}{row + 1}'].number_format = '"$"#,##0;[Red]-"$"#,##0'

        # EBITDA section
        row += 3
        ws[f'A{row}'] = "EBITDA ($MM)"
        ws[f'A{row}'].style = 'subheader'
        ws.merge_cells(f'A{row}:B{row}')

        ebitda_proj = calculations.get('ebitda_projection', [])

        base_ebitda_value = model_spec.assumptions.get('base_ebitda', base_ebitda * 1_000_000_000 if base_ebitda else 0)
        ws[f'B{row + 1}'] = base_ebitda_value / self.currency_scale
        ws[f'B{row + 1}'].number_format = '"$"#,##0;[Red]-"$"#,##0'

        for i, ebitda in enumerate(ebitda_proj):
            col = get_column_letter(i + 3)
            ws[f'{col}{row + 1}'] = ebitda / self.currency_scale
            ws[f'{col}{row + 1}'].number_format = '"$"#,##0;[Red]-"$"#,##0'

        # UFCF section
        row += 3
        ws[f'A{row}'] = "Unlevered Free Cash Flow ($MM)"
        ws[f'A{row}'].style = 'subheader'
        ws.merge_cells(f'A{row}:B{row}')

        ufcf_proj = calculations.get('ufcf_projection', [])
        tax_rate = assumptions.get('tax_rate', 0.25)
        capex_pct = assumptions.get('capex_pct', 0.06)
        nwc_pct = assumptions.get('nwc_pct', 0.02)

        capex_amount = base_revenue_value * capex_pct
        base_nwc = base_revenue_value * nwc_pct
        base_ufcf = base_ebitda_value * (1 - tax_rate) - capex_amount - base_nwc

        ws[f'B{row + 1}'] = base_ufcf / self.currency_scale
        ws[f'B{row + 1}'].number_format = '"$"#,##0;[Red]-"$"#,##0'
        ws[f'B{row + 1}'].font = Font(color='2E7D32', bold=True)  # Output green

        current_nwc = base_nwc
        for i, ufcf in enumerate(ufcf_proj):
            col = get_column_letter(i + 3)
            # Adjust for projected NWC amounts already embedded in model calculations
            projected_nwc = calc_nwc = model_spec.calculations['nwc_projection'][i] if 'nwc_projection' in model_spec.calculations else 0
            nwc_change = projected_nwc - current_nwc
            current_nwc = projected_nwc

            projected_capex = model_spec.calculations['capex_projection'][i] if 'capex_projection' in model_spec.calculations else capex_amount

            # Recreate UFCF using EBITDA projections to ensure alignment
            projected_ebitda = model_spec.calculations['ebitda_projection'][i] if 'ebitda_projection' in model_spec.calculations else base_ebitda_value
            projected_nopat = projected_ebitda * (1 - tax_rate)
            reconstructed_ufcf = projected_nopat - projected_capex - nwc_change

            ws[f'{col}{row + 1}'] = reconstructed_ufcf / self.currency_scale
            ws[f'{col}{row + 1}'].number_format = '"$"#,##0;[Red]-"$"#,##0'
            ws[f'{col}{row + 1}'].font = Font(color='2E7D32', bold=True)  # Output green

        # Set column widths
        ws.column_dimensions['A'].width = 30
        for i in range(forecast_years + 2):
            col = get_column_letter(i + 1)
            ws.column_dimensions[col].width = 15

        # Print setup
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1

    def _populate_dcf_sheet(self, ws, model_spec: Any):
        """Populate DCF valuation sheet."""
        calculations = model_spec.calculations
        outputs = model_spec.outputs

        # Title
        ws['A1'] = "DCF Valuation Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:I1')

        # Discounting section
        row = 3
        ws[f'A{row}'] = "Discounted Cash Flow Analysis"
        ws[f'A{row}'].style = 'subheader'
        ws.merge_cells(f'A{row}:B{row}')

        # Headers
        row += 2
        ws[f'A{row}'] = "Year"
        ws[f'A{row}'].style = 'header'
        ws[f'B{row}'] = "UFCF ($MM)"
        ws[f'B{row}'].style = 'header'
        ws[f'C{row}'] = "Discount Factor"
        ws[f'C{row}'].style = 'header'
        ws[f'D{row}'] = "PV of UFCF ($MM)"
        ws[f'D{row}'].style = 'header'

        # Data
        ufcf_proj = calculations.get('ufcf_projection', [])
        discount_factors = calculations.get('discount_factors', [])
        pv_ufcf = calculations.get('pv_ufcf', [])

        forecast_years = len(ufcf_proj)
        base_ufcf = calculations.get('base_ufcf', 0)

        for i in range(forecast_years):
            row += 1
            ws[f'A{row}'] = f"Year {i + 1}"
            ws[f'B{row}'] = (ufcf_proj[i] / self.currency_scale) if i < len(ufcf_proj) else 0
            ws[f'B{row}'].number_format = '"$"#,##0;[Red]-"$"#,##0'

            ws[f'C{row}'] = discount_factors[i] if i < len(discount_factors) else 0
            ws[f'C{row}'].number_format = '0.0000'

            ws[f'D{row}'] = (pv_ufcf[i] / self.currency_scale) if i < len(pv_ufcf) else 0
            ws[f'D{row}'].number_format = '"$"#,##0;[Red]-"$"#,##0'

        # Terminal value
        row += 2
        ws[f'A{row}'] = "Terminal Value"
        ws[f'B{row}'] = calculations.get('terminal_value', 0) / self.currency_scale
        ws[f'B{row}'].number_format = '"$"#,##0;[Red]-"$"#,##0'
        ws[f'B{row}'].font = Font(color='2E7D32', bold=True)

        ws[f'C{row}'] = calculations.get('terminal_discount_factor', discount_factors[-1] if discount_factors else 1)
        ws[f'C{row}'].number_format = '0.0000'

        ws[f'D{row}'] = calculations.get('pv_terminal', 0) / self.currency_scale
        ws[f'D{row}'].number_format = '"$"#,##0;[Red]-"$"#,##0'
        ws[f'D{row}'].font = Font(color='2E7D32', bold=True)

        # Enterprise value
        row += 2
        ws[f'A{row}'] = "Enterprise Value ($MM)"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        ws[f'B{row}'] = outputs.get('enterprise_value', 0) / self.currency_scale
        ws[f'B{row}'].number_format = '"$"#,##0;[Red]-"$"#,##0'
        ws[f'B{row}'].font = Font(color='2E7D32', bold=True, size=12)

        # Equity value
        row += 1
        ws[f'A{row}'] = "Equity Value ($MM)"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        ws[f'B{row}'] = outputs.get('equity_value', 0) / self.currency_scale
        ws[f'B{row}'].number_format = '"$"#,##0;[Red]-"$"#,##0'
        ws[f'B{row}'].font = Font(color='2E7D32', bold=True, size=12)

        # Implied price
        row += 1
        ws[f'A{row}'] = "Implied Price per Share"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        ws[f'B{row}'] = outputs.get('implied_price', 0)
        ws[f'B{row}'].number_format = '"$"#,##0.00'
        ws[f'B{row}'].font = Font(color='2E7D32', bold=True, size=12)

        # Add scale note
        note_row = row + 2
        ws[f'A{note_row}'] = self.currency_note
        ws[f'A{note_row}'].font = Font(italic=True, size=9, color='666666')
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)

        # Set column widths
        ws.column_dimensions['A'].width = 25
        for col in ['B', 'C', 'D']:
            ws.column_dimensions[col].width = 18

        # Print setup
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.fitToWidth = 1

    def _create_summary_sheet(self, wb: Workbook, model_spec: Any):
        """Create executive summary sheet."""
        ws = wb.create_sheet("Summary")

        # Title
        ws['A1'] = f"{model_spec.model_type.upper()} Model Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')

        company_name = getattr(model_spec.company_data, 'company_name', 'Unknown')
        ws['A2'] = f"Company: {company_name}"
        ws['A2'].font = Font(size=12, bold=True)
        ws.merge_cells('A2:D2')

        # Key outputs box
        ws['A4'] = "Key Valuation Outputs"
        ws['A4'].style = 'subheader'
        ws.merge_cells('A4:D4')

        outputs = model_spec.outputs
        summary_data = [
            ("Enterprise Value (MM)", outputs.get('enterprise_value', 0) / self.currency_scale, '#,##0'),
            ("Equity Value (MM)", outputs.get('equity_value', 0) / self.currency_scale, '#,##0'),
            ("Implied Price per Share", outputs.get('implied_price', 0), '"$"#,##0.00'),
            ("WACC", outputs.get('wacc', 0), '0.0%')
        ]

        row = 5
        for label, value, format_str in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = format_str
            ws[f'B{row}'].font = Font(color='2E7D32', bold=True, size=12)
            row += 1

        # Key assumptions
        row += 1
        ws[f'A{row}'] = "Key Assumptions"
        ws[f'A{row}'].style = 'subheader'
        ws.merge_cells(f'A{row}:D{row}')

        assumptions = model_spec.assumptions
        assumptions_data = [
            ("Revenue Growth Rate", assumptions.get('growth_rate', 0), '0.0%'),
            ("EBITDA Margin", assumptions.get('ebitda_margin', 0), '0.0%'),
            ("Terminal Growth Rate", assumptions.get('terminal_growth', 0), '0.0%'),
            ("Beta", getattr(model_spec.company_data, 'beta', 1.2), '0.00')
        ]

        row += 1
        for label, value, format_str in assumptions_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = format_str
            ws[f'B{row}'].font = Font(color='1F4E79')
            row += 1

        # Model info
        row += 1
        ws[f'A{row}'] = "Model Information"
        ws[f'A{row}'].style = 'subheader'
        ws.merge_cells(f'A{row}:D{row}')

        model_info = [
            ("Model Type", model_spec.model_type.upper()),
            ("Data Source", getattr(model_spec.company_data, 'data_source', 'Unknown')),
            ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("Platform", "FinModAI v1.0.0")
        ]

        row += 1
        for label, value in model_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

        # Print setup
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.fitToWidth = 1

    def _create_sensitivity_sheet(self, wb: Workbook, model_spec: Any):
        """Create sensitivity analysis sheet."""
        if not model_spec.sensitivity_analysis:
            return

        ws = wb.create_sheet("Sensitivity")

        # Title
        ws['A1'] = "Sensitivity Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:F1')

        # WACC vs Terminal Growth sensitivity table
        ws['A3'] = "WACC vs Terminal Growth Sensitivity"
        ws['A3'].style = 'subheader'
        ws.merge_cells('A3:F3')

        # This is a simplified version - in production, would calculate actual sensitivities
        sensitivity = model_spec.sensitivity_analysis

        # Headers
        ws['A5'] = "WACC \\ Terminal Growth"
        ws['A5'].style = 'header'

        # Terminal growth headers
        growth_range = sensitivity.get('terminal_growth_sensitivity', [0.01, 0.02, 0.03])
        for i, growth in enumerate(growth_range):
            ws[f'{get_column_letter(i + 2)}5'] = ".1%"
            ws[f'{get_column_letter(i + 2)}5'].style = 'header'

        # WACC rows
        wacc_range = sensitivity.get('wacc_sensitivity', [0.08, 0.10, 0.12])
        base_ev = model_spec.outputs.get('enterprise_value', 1000)

        for i, wacc in enumerate(wacc_range):
            row = i + 6
            ws[f'A{row}'] = ".1%"
            ws[f'A{row}'].font = Font(bold=True)

            for j, growth in enumerate(growth_range):
                col = get_column_letter(j + 2)
                # Simplified sensitivity calculation
                sensitivity_factor = (0.10 - wacc) / 0.10 + (0.025 - growth) / 0.025
                ev_sensitivity = base_ev * (1 + sensitivity_factor * 0.5)

                ws[f'{col}{row}'] = ev_sensitivity
                ws[f'{col}{row}'].number_format = '"$"#,##0;[Red]-"$"#,##0'

                # Highlight base case
                if abs(wacc - 0.10) < 0.001 and abs(growth - 0.025) < 0.001:
                    ws[f'{col}{row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # Set column widths
        ws.column_dimensions['A'].width = 20
        for i in range(len(growth_range) + 1):
            col = get_column_letter(i + 1)
            ws.column_dimensions[col].width = 15

    def _create_dashboard_sheet(self, wb: Workbook, model_spec: Any):
        """Create dashboard sheet with charts and visualizations."""
        if not model_spec.visualization_config:
            return

        ws = wb.create_sheet("Dashboard")

        # Title
        ws['A1'] = "Financial Model Dashboard"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:G1')

        # Key metrics section
        ws['A3'] = "Key Metrics"
        ws['A3'].style = 'subheader'
        ws.merge_cells('A3:B3')

        outputs = model_spec.outputs
        metrics = [
            ("Enterprise Value (MM)", outputs.get('enterprise_value', 0) / self.currency_scale),
            ("Equity Value (MM)", outputs.get('equity_value', 0) / self.currency_scale),
            ("Implied Price", outputs.get('implied_price', 0)),
            ("WACC", outputs.get('wacc', 0))
        ]

        row = 4
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            if '%' in label.lower():
                ws[f'B{row}'].number_format = '0.0%'
            elif 'price' in label.lower():
                ws[f'B{row}'].number_format = '"$"#,##0.00'
            else:
                ws[f'B{row}'].number_format = '"$"#,##0;[Red]-"$"#,##0'
            ws[f'B{row}'].font = Font(color='2E7D32', bold=True, size=12)
            row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18

    def _create_lbo_excel(self, wb: Workbook, model_spec: Any):
        """Create LBO-specific Excel sheets."""
        # Simplified LBO structure
        ws_assumptions = wb.create_sheet("LBO_Assumptions")
        ws_model = wb.create_sheet("LBO_Model")
        ws_returns = wb.create_sheet("LBO_Returns")

        # Populate with LBO-specific content
        self._populate_lbo_assumptions(ws_assumptions, model_spec)
        self._populate_lbo_model(ws_model, model_spec)
        self._populate_lbo_returns(ws_returns, model_spec)

    def _create_comps_excel(self, wb: Workbook, model_spec: Any):
        """Create trading comps Excel sheets."""
        ws_comps = wb.create_sheet("Trading_Comps")
        self._populate_comps_sheet(ws_comps, model_spec)

    def _create_three_statement_excel(self, wb: Workbook, model_spec: Any):
        """Create three-statement Excel sheets."""
        ws_income = wb.create_sheet("Income_Statement")
        ws_balance = wb.create_sheet("Balance_Sheet")
        ws_cashflow = wb.create_sheet("Cash_Flow")

        # Populate three-statement sheets
        self._populate_income_statement(ws_income, model_spec)
        self._populate_balance_sheet(ws_balance, model_spec)
        self._populate_cash_flow(ws_cashflow, model_spec)

    # Placeholder methods for other model types
    def _populate_lbo_assumptions(self, ws, model_spec):
        ws['A1'] = "LBO Assumptions - Coming Soon"
        ws['A1'].font = Font(size=14, bold=True)

    def _populate_lbo_model(self, ws, model_spec):
        ws['A1'] = "LBO Model - Coming Soon"
        ws['A1'].font = Font(size=14, bold=True)

    def _populate_lbo_returns(self, ws, model_spec):
        ws['A1'] = "LBO Returns - Coming Soon"
        ws['A1'].font = Font(size=14, bold=True)

    def _populate_comps_sheet(self, ws, model_spec):
        ws['A1'] = "Trading Comparables - Coming Soon"
        ws['A1'].font = Font(size=14, bold=True)

    def _populate_income_statement(self, ws, model_spec):
        ws['A1'] = "Income Statement - Coming Soon"
        ws['A1'].font = Font(size=14, bold=True)

    def _populate_balance_sheet(self, ws, model_spec):
        ws['A1'] = "Balance Sheet - Coming Soon"
        ws['A1'].font = Font(size=14, bold=True)

    def _populate_cash_flow(self, ws, model_spec):
        ws['A1'] = "Cash Flow Statement - Coming Soon"
        ws['A1'].font = Font(size=14, bold=True)

    def _generate_json_output(self, model_spec: Any) -> List[str]:
        """Generate JSON output for API consumption."""
        import json

        output_data = {
            "model_type": model_spec.model_type,
            "company": getattr(model_spec.company_data, 'company_name', 'Unknown'),
            "generated_at": datetime.now().isoformat(),
            "assumptions": model_spec.assumptions,
            "outputs": model_spec.outputs,
            "calculations": model_spec.calculations
        }

        filename = f"model_output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        filepath = self.output_dir / filename

        with open(filepath, 'w') as f:
            json.dump(output_data, f, indent=2, default=str)

        logger.info(f"✅ JSON output generated: {filepath}")
        return [str(filepath)]
