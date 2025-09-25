#!/usr/bin/env python3
"""
Professional Precedent Transactions (M&A Comps) Model
Values a target company based on multiples paid in comparable M&A transactions

Author: Investment Banking Modeler
Date: 2024

Features:
- Historical M&A transaction analysis
- Valuation multiples from actual deal prices
- Statistical analysis (mean, median, percentiles)
- Implied valuation ranges for target company
- Professional Excel Output with Banker Formatting
- Deal summary and valuation ranges
"""

import numpy as np
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# Professional color scheme
PRECEDENT_MODEL_COLORS = {
    'header_blue': '1F4E79',
    'input_light_blue': 'DDEBF7',
    'calculation_black': '000000',
    'output_green': 'D8E4BC',
    'warning_red': 'F8CECC',
    'border_dark': '2C3E50',
    'text_dark': '000000',
    'text_white': 'FFFFFF',
    'deal_data_yellow': 'FFF2CC',
    'target_highlight': 'E6F3FF'
}

class ProfessionalPrecedentTransactionsModel:
    """
    Comprehensive Precedent Transactions Model with Professional Formatting
    """

    def __init__(self, target_company="Target Company", target_ticker="TARGET"):
        self.target_company = target_company
        self.target_ticker = target_ticker
        self.model_date = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Initialize styles
        self.styles = self._create_styles()

    def _create_styles(self):
        """Create professional Excel styles"""
        styles = {}

        # Header style
        styles['header'] = NamedStyle(name='header')
        styles['header'].font = Font(name='Calibri', size=12, bold=True, color=PRECEDENT_MODEL_COLORS['text_white'])
        styles['header'].fill = PatternFill(start_color=PRECEDENT_MODEL_COLORS['header_blue'], end_color=PRECEDENT_MODEL_COLORS['header_blue'], fill_type='solid')
        styles['header'].alignment = Alignment(horizontal='center', vertical='center')

        # Input style
        styles['input'] = NamedStyle(name='input')
        styles['input'].font = Font(name='Calibri', size=10, color=PRECEDENT_MODEL_COLORS['text_dark'])
        styles['input'].fill = PatternFill(start_color=PRECEDENT_MODEL_COLORS['input_light_blue'], end_color=PRECEDENT_MODEL_COLORS['input_light_blue'], fill_type='solid')
        styles['input'].alignment = Alignment(horizontal='right', vertical='center')
        styles['input'].number_format = '#,##0.00'

        # Deal data style
        styles['deal_data'] = NamedStyle(name='deal_data')
        styles['deal_data'].font = Font(name='Calibri', size=10, color=PRECEDENT_MODEL_COLORS['text_dark'])
        styles['deal_data'].fill = PatternFill(start_color=PRECEDENT_MODEL_COLORS['deal_data_yellow'], end_color=PRECEDENT_MODEL_COLORS['deal_data_yellow'], fill_type='solid')
        styles['deal_data'].alignment = Alignment(horizontal='right', vertical='center')
        styles['deal_data'].number_format = '#,##0.00'

        # Target highlight style
        styles['target_highlight'] = NamedStyle(name='target_highlight')
        styles['target_highlight'].font = Font(name='Calibri', size=10, bold=True, color=PRECEDENT_MODEL_COLORS['text_dark'])
        styles['target_highlight'].fill = PatternFill(start_color=PRECEDENT_MODEL_COLORS['target_highlight'], end_color=PRECEDENT_MODEL_COLORS['target_highlight'], fill_type='solid')
        styles['target_highlight'].alignment = Alignment(horizontal='right', vertical='center')
        styles['target_highlight'].number_format = '#,##0.00'

        # Calculation style
        styles['calculation'] = NamedStyle(name='calculation')
        styles['calculation'].font = Font(name='Calibri', size=10, color=PRECEDENT_MODEL_COLORS['calculation_black'])
        styles['calculation'].alignment = Alignment(horizontal='right', vertical='center')
        styles['calculation'].number_format = '#,##0.00'

        # Output style
        styles['output'] = NamedStyle(name='output')
        styles['output'].font = Font(name='Calibri', size=10, bold=True, color=PRECEDENT_MODEL_COLORS['text_dark'])
        styles['output'].fill = PatternFill(start_color=PRECEDENT_MODEL_COLORS['output_green'], end_color=PRECEDENT_MODEL_COLORS['output_green'], fill_type='solid')
        styles['output'].alignment = Alignment(horizontal='right', vertical='center')
        styles['output'].number_format = '#,##0.00'

        # Statistics style
        styles['statistics'] = NamedStyle(name='statistics')
        styles['statistics'].font = Font(name='Calibri', size=10, bold=True, color=PRECEDENT_MODEL_COLORS['text_dark'])
        styles['statistics'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        styles['statistics'].alignment = Alignment(horizontal='right', vertical='center')
        styles['statistics'].number_format = '#,##0.00'

        # Warning style
        styles['warning'] = NamedStyle(name='warning')
        styles['warning'].font = Font(name='Calibri', size=10, color=PRECEDENT_MODEL_COLORS['text_dark'])
        styles['warning'].fill = PatternFill(start_color=PRECEDENT_MODEL_COLORS['warning_red'], end_color=PRECEDENT_MODEL_COLORS['warning_red'], fill_type='solid')
        styles['warning'].alignment = Alignment(horizontal='center', vertical='center')

        # Bold label style
        styles['label_bold'] = NamedStyle(name='label_bold')
        styles['label_bold'].font = Font(name='Calibri', size=10, bold=True, color=PRECEDENT_MODEL_COLORS['text_dark'])
        styles['label_bold'].alignment = Alignment(horizontal='left', vertical='center')

        # Normal label style
        styles['label'] = NamedStyle(name='label')
        styles['label'].font = Font(name='Calibri', size=10, color=PRECEDENT_MODEL_COLORS['text_dark'])
        styles['label'].alignment = Alignment(horizontal='left', vertical='center')

        # Summary box style
        styles['summary_box'] = NamedStyle(name='summary_box')
        styles['summary_box'].font = Font(name='Calibri', size=11, bold=True, color=PRECEDENT_MODEL_COLORS['text_white'])
        styles['summary_box'].fill = PatternFill(start_color=PRECEDENT_MODEL_COLORS['header_blue'], end_color=PRECEDENT_MODEL_COLORS['header_blue'], fill_type='solid')
        styles['summary_box'].alignment = Alignment(horizontal='center', vertical='center')
        styles['summary_box'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        return styles

    def run_precedent_transactions_model(self,
                                        # Target Company Data
                                        target_revenue=1000.0,      # $M
                                        target_ebitda=250.0,        # $M
                                        target_net_income=150.0,    # $M
                                        target_eps=2.50,            # $
                                        target_net_debt=200.0,      # $M
                                        target_shares_outstanding=80.0,  # Million shares

                                        # Precedent Transaction Data (as lists)
                                        deal_dates=None,
                                        acquirers=None,
                                        targets=None,
                                        equity_values=None,         # $M
                                        enterprise_values=None,     # $M
                                        deal_revenues=None,         # $M
                                        deal_ebitdas=None,          # $M
                                        deal_net_incomes=None):     # $M

        """
        Run complete precedent transactions model with deal analysis
        """

        print(f"🤝 Building Professional Precedent Transactions Model for {self.target_company} ({self.target_ticker})")
        print("=" * 90)

        # Set default precedent deals if not provided
        if deal_dates is None:
            deal_dates = ["2023-01-15", "2023-03-22", "2023-06-10", "2023-09-05", "2023-11-28"]
            acquirers = ["BigTech Corp", "Global Inc", "Mega Corp", "TechGiant Ltd", "IndustCo Inc"]
            targets = ["CloudStart Inc", "DataFirm Corp", "SoftTech Ltd", "AppWorks Inc", "Analytics Plus"]
            equity_values = [800.0, 1200.0, 600.0, 1500.0, 900.0]      # $M
            enterprise_values = [950.0, 1400.0, 700.0, 1700.0, 1050.0] # $M
            deal_revenues = [400.0, 600.0, 300.0, 750.0, 450.0]        # $M
            deal_ebitdas = [120.0, 180.0, 90.0, 225.0, 135.0]          # $M
            deal_net_incomes = [80.0, 120.0, 60.0, 150.0, 90.0]        # $M

        # Step 1: Create Assumptions & Deal Data
        assumptions = self._create_assumptions(
            target_revenue, target_ebitda, target_net_income, target_eps,
            target_net_debt, target_shares_outstanding,
            deal_dates, acquirers, targets, equity_values, enterprise_values,
            deal_revenues, deal_ebitdas, deal_net_incomes
        )

        # Step 2: Calculate Deal Valuation Multiples
        deal_multiples = self._calculate_deal_multiples(assumptions)

        # Step 3: Compute Summary Statistics
        summary_stats = self._compute_summary_statistics(deal_multiples)

        # Step 4: Generate Target Valuation
        target_valuation = self._generate_target_valuation(assumptions, summary_stats)

        # Step 5: Create Valuation Ranges
        valuation_ranges = self._create_valuation_ranges(target_valuation)

        # Compile results
        precedent_results = {
            'assumptions': assumptions,
            'deal_multiples': deal_multiples,
            'summary_stats': summary_stats,
            'target_valuation': target_valuation,
            'valuation_ranges': valuation_ranges
        }

        # Create Excel output
        excel_file = self._create_excel_output(precedent_results)

        print("\n✅ Precedent Transactions Model Complete!")
        print("📊 Key Valuation Metrics:")
        print(f"   • Target Revenue: ${target_revenue:.0f}M | EBITDA: ${target_ebitda:.0f}M")
        print(f"   • Precedent Deals: {len(deal_dates)} transactions")
        print(f"   • Median EV/Revenue: {summary_stats['ev_revenue']['median']:.1f}x")
        print(f"   • Median EV/EBITDA: {summary_stats['ev_ebitda']['median']:.1f}x")
        print(f"   • Median Equity Value/Net Income: {summary_stats['equity_value_net_income']['median']:.1f}x")
        print(f"   • Implied EV Range: ${valuation_ranges['ev_range']['low']/1000:.1f}B - ${valuation_ranges['ev_range']['high']/1000:.1f}B")
        print(f"   • Implied Share Price Range: ${valuation_ranges['share_price_range']['low']:.2f} - ${valuation_ranges['share_price_range']['high']:.2f}")
        print(f"📁 Excel Output: {excel_file}")

        return precedent_results, excel_file

    def _create_assumptions(self, target_revenue, target_ebitda, target_net_income, target_eps,
                           target_net_debt, target_shares_outstanding,
                           deal_dates, acquirers, targets, equity_values, enterprise_values,
                           deal_revenues, deal_ebitdas, deal_net_incomes):

        """Create comprehensive assumptions and precedent deal data"""

        # Target company data
        target_data = {
            'company_name': self.target_company,
            'ticker': self.target_ticker,
            'revenue': target_revenue,
            'ebitda': target_ebitda,
            'net_income': target_net_income,
            'eps': target_eps,
            'net_debt': target_net_debt,
            'shares_outstanding': target_shares_outstanding,
            'market_cap': target_eps * target_shares_outstanding if target_eps > 0 else 0,  # Placeholder
            'enterprise_value': (target_eps * target_shares_outstanding) + target_net_debt if target_eps > 0 else target_net_debt
        }

        # Precedent deal data
        deals_data = []
        for i in range(len(deal_dates)):
            deal = {
                'date': deal_dates[i],
                'acquirer': acquirers[i],
                'target': targets[i],
                'equity_value': equity_values[i],
                'enterprise_value': enterprise_values[i],
                'revenue': deal_revenues[i],
                'ebitda': deal_ebitdas[i],
                'net_income': deal_net_incomes[i]
            }
            deals_data.append(deal)

        assumptions = {
            'target': target_data,
            'deals': deals_data,
            'num_deals': len(deals_data)
        }

        print("📋 Assumptions Created:")
        print(f"   • Target: {self.target_company} (${target_revenue:.0f}M revenue, ${target_ebitda:.0f}M EBITDA)")
        print(f"   • Precedent Deals: {len(deals_data)} transactions")
        deal_sizes = [d['enterprise_value'] for d in deals_data]
        print(f"   • Deal Size Range: ${min(deal_sizes):.0f}M - ${max(deal_sizes):.0f}M")
        deal_dates_range = sorted(deal_dates)
        print(f"   • Deal Date Range: {deal_dates_range[0]} - {deal_dates_range[-1]}")

        return assumptions

    def _calculate_deal_multiples(self, assumptions):
        """Calculate valuation multiples from precedent deals"""

        deal_multiples = []

        for deal in assumptions['deals']:
            multiples = {
                'date': deal['date'],
                'acquirer': deal['acquirer'],
                'target': deal['target'],
                'ev_revenue': deal['enterprise_value'] / deal['revenue'] if deal['revenue'] > 0 else 0,
                'ev_ebitda': deal['enterprise_value'] / deal['ebitda'] if deal['ebitda'] > 0 else 0,
                'equity_value_net_income': deal['equity_value'] / deal['net_income'] if deal['net_income'] > 0 else 0,
                'equity_value': deal['equity_value'],
                'enterprise_value': deal['enterprise_value'],
                'revenue': deal['revenue'],
                'ebitda': deal['ebitda'],
                'net_income': deal['net_income']
            }
            deal_multiples.append(multiples)

        print("📊 Deal Multiples Calculated:")
        ev_revenue_ratios = [d['ev_revenue'] for d in deal_multiples]
        ev_ebitda_ratios = [d['ev_ebitda'] for d in deal_multiples]
        equity_value_net_income_ratios = [d['equity_value_net_income'] for d in deal_multiples]

        print(f"   • EV/Revenue Range: {min(ev_revenue_ratios):.1f}x - {max(ev_revenue_ratios):.1f}x")
        print(f"   • EV/EBITDA Range: {min(ev_ebitda_ratios):.1f}x - {max(ev_ebitda_ratios):.1f}x")
        print(f"   • Equity Value/Net Income Range: {min(equity_value_net_income_ratios):.1f}x - {max(equity_value_net_income_ratios):.1f}x")

        return deal_multiples

    def _compute_summary_statistics(self, deal_multiples):
        """Compute summary statistics for deal multiples"""

        # Extract multiple arrays
        ev_revenue_ratios = [d['ev_revenue'] for d in deal_multiples]
        ev_ebitda_ratios = [d['ev_ebitda'] for d in deal_multiples]
        equity_value_net_income_ratios = [d['equity_value_net_income'] for d in deal_multiples]

        def calculate_stats(ratios):
            """Calculate mean, median, and percentiles for a list of ratios"""
            if not ratios or all(r == 0 for r in ratios):
                return {'mean': 0, 'median': 0, 'p25': 0, 'p75': 0, 'min': 0, 'max': 0}

            valid_ratios = [r for r in ratios if r > 0]  # Filter out zero ratios
            if not valid_ratios:
                return {'mean': 0, 'median': 0, 'p25': 0, 'p75': 0, 'min': 0, 'max': 0}

            return {
                'mean': np.mean(valid_ratios),
                'median': np.median(valid_ratios),
                'p25': np.percentile(valid_ratios, 25),
                'p75': np.percentile(valid_ratios, 75),
                'min': min(valid_ratios),
                'max': max(valid_ratios)
            }

        summary_stats = {
            'ev_revenue': calculate_stats(ev_revenue_ratios),
            'ev_ebitda': calculate_stats(ev_ebitda_ratios),
            'equity_value_net_income': calculate_stats(equity_value_net_income_ratios)
        }

        print("📈 Summary Statistics Computed:")
        print(f"   • EV/Revenue: Mean {summary_stats['ev_revenue']['mean']:.1f}x, Median {summary_stats['ev_revenue']['median']:.1f}x")
        print(f"   • EV/EBITDA: Mean {summary_stats['ev_ebitda']['mean']:.1f}x, Median {summary_stats['ev_ebitda']['median']:.1f}x")
        print(f"   • Equity Value/Net Income: Mean {summary_stats['equity_value_net_income']['mean']:.1f}x, Median {summary_stats['equity_value_net_income']['median']:.1f}x")

        return summary_stats

    def _generate_target_valuation(self, assumptions, summary_stats):
        """Generate implied valuation for target using precedent multiples"""

        target = assumptions['target']

        # Apply multiples to target financials
        target_valuation = {
            'ev_revenue': {
                'p25_multiple': summary_stats['ev_revenue']['p25'],
                'median_multiple': summary_stats['ev_revenue']['median'],
                'p75_multiple': summary_stats['ev_revenue']['p75'],
                'p25_ev': target['revenue'] * summary_stats['ev_revenue']['p25'],
                'median_ev': target['revenue'] * summary_stats['ev_revenue']['median'],
                'p75_ev': target['revenue'] * summary_stats['ev_revenue']['p75']
            },
            'ev_ebitda': {
                'p25_multiple': summary_stats['ev_ebitda']['p25'],
                'median_multiple': summary_stats['ev_ebitda']['median'],
                'p75_multiple': summary_stats['ev_ebitda']['p75'],
                'p25_ev': target['ebitda'] * summary_stats['ev_ebitda']['p25'],
                'median_ev': target['ebitda'] * summary_stats['ev_ebitda']['median'],
                'p75_ev': target['ebitda'] * summary_stats['ev_ebitda']['p75']
            },
            'equity_value_net_income': {
                'p25_multiple': summary_stats['equity_value_net_income']['p25'],
                'median_multiple': summary_stats['equity_value_net_income']['median'],
                'p75_multiple': summary_stats['equity_value_net_income']['p75'],
                'p25_equity_value': target['net_income'] * summary_stats['equity_value_net_income']['p25'],
                'median_equity_value': target['net_income'] * summary_stats['equity_value_net_income']['median'],
                'p75_equity_value': target['net_income'] * summary_stats['equity_value_net_income']['p75']
            }
        }

        print("🎯 Target Valuation Generated:")
        print(f"   • EV/Revenue: ${target_valuation['ev_revenue']['median_ev']/1000:.1f}B (median)")
        print(f"   • EV/EBITDA: ${target_valuation['ev_ebitda']['median_ev']/1000:.1f}B (median)")
        print(f"   • Equity Value/Net Income: ${target_valuation['equity_value_net_income']['median_equity_value']/1000:.1f}B market cap (median)")

        return target_valuation

    def _create_valuation_ranges(self, target_valuation):
        """Create valuation ranges from target valuation data"""

        # Use EV/EBITDA as primary valuation method (most common for precedent transactions)
        ev_range = {
            'low': target_valuation['ev_ebitda']['p25_ev'],
            'median': target_valuation['ev_ebitda']['median_ev'],
            'high': target_valuation['ev_ebitda']['p75_ev']
        }

        # Calculate equity value ranges (EV - Net Debt)
        target_net_debt = 200.0  # Default, should be passed in
        equity_range = {
            'low': ev_range['low'] - target_net_debt,
            'median': ev_range['median'] - target_net_debt,
            'high': ev_range['high'] - target_net_debt
        }

        # Calculate share price ranges (Market Cap / Shares Outstanding)
        target_shares = 80.0  # Default, should be passed in
        share_price_range = {
            'low': equity_range['low'] / target_shares,
            'median': equity_range['median'] / target_shares,
            'high': equity_range['high'] / target_shares
        }

        valuation_ranges = {
            'ev_range': ev_range,
            'equity_range': equity_range,
            'share_price_range': share_price_range,
            'primary_multiple': 'EV/EBITDA',
            'range_type': '25th - 75th Percentile of Precedent Deals'
        }

        print("📊 Valuation Ranges Created:")
        print(f"   • Enterprise Value: ${ev_range['low']/1000:.1f}B - ${ev_range['high']/1000:.1f}B")
        print(f"   • Equity Value: ${equity_range['low']/1000:.1f}B - ${equity_range['high']/1000:.1f}B")
        print(f"   • Share Price: ${share_price_range['low']:.2f} - ${share_price_range['high']:.2f}")

        return valuation_ranges

    def _create_excel_output(self, precedent_results):
        """Create professional Excel output with multiple tabs"""

        wb = Workbook()

        # Create worksheets
        ws_summary = wb.active
        ws_summary.title = "Valuation Summary"

        ws_deal_data = wb.create_sheet("Deal Data & Inputs")
        ws_deal_multiples = wb.create_sheet("Deal Multiples")
        ws_target_valuation = wb.create_sheet("Target Valuation")

        # Apply styles to workbook
        for style_name, style in self.styles.items():
            if style_name not in wb.named_styles:
                wb.add_named_style(style)

        # Create each tab
        self._create_summary_tab(ws_summary, precedent_results)
        self._create_deal_data_tab(ws_deal_data, precedent_results)
        self._create_deal_multiples_tab(ws_deal_multiples, precedent_results)
        self._create_target_valuation_tab(ws_target_valuation, precedent_results)

        # Save workbook
        filename = f"Precedent_Transactions_{self.target_ticker}_{self.model_date}.xlsx"
        wb.save(filename)

        return filename

    def _create_summary_tab(self, ws, results):
        """Create Valuation Summary tab with banker-style summary boxes"""

        # Title
        ws['A1'] = f"Precedent Transactions Analysis: {self.target_company} ({self.target_ticker})"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:G1')

        current_row = 3

        # Precedent Transactions Summary Box
        ws[f'A{current_row}'] = 'PRECEDENT TRANSACTIONS SUMMARY'
        ws[f'A{current_row}'].style = 'summary_box'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        # Multiples summary
        summary_stats = results['summary_stats']
        multiples_data = [
            ("EV/Revenue", summary_stats['ev_revenue']['p25'], summary_stats['ev_revenue']['median'], summary_stats['ev_revenue']['p75']),
            ("EV/EBITDA", summary_stats['ev_ebitda']['p25'], summary_stats['ev_ebitda']['median'], summary_stats['ev_ebitda']['p75']),
            ("Equity Value/Net Income", summary_stats['equity_value_net_income']['p25'], summary_stats['equity_value_net_income']['median'], summary_stats['equity_value_net_income']['p75'])
        ]

        # Headers
        ws.cell(row=current_row, column=1, value="Multiple").style = 'header'
        ws.cell(row=current_row, column=2, value="25th Percentile").style = 'header'
        ws.cell(row=current_row, column=3, value="Median").style = 'header'
        ws.cell(row=current_row, column=4, value="75th Percentile").style = 'header'
        current_row += 1

        # Data
        for multiple, p25, median, p75 in multiples_data:
            ws.cell(row=current_row, column=1, value=multiple).style = 'label_bold'
            ws.cell(row=current_row, column=2, value=p25).style = 'output'
            ws.cell(row=current_row, column=3, value=median).style = 'output'
            ws.cell(row=current_row, column=4, value=p75).style = 'output'
            current_row += 1

        current_row += 2

        # Target Valuation Summary Box
        ws[f'A{current_row}'] = 'TARGET VALUATION SUMMARY'
        ws[f'A{current_row}'].style = 'summary_box'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        # Valuation ranges
        valuation_ranges = results['valuation_ranges']
        valuation_data = [
            ("Implied Enterprise Value", f"${valuation_ranges['ev_range']['low']/1000:.1f}B", f"${valuation_ranges['ev_range']['median']/1000:.1f}B", f"${valuation_ranges['ev_range']['high']/1000:.1f}B"),
            ("Implied Equity Value", f"${valuation_ranges['equity_range']['low']/1000:.1f}B", f"${valuation_ranges['equity_range']['median']/1000:.1f}B", f"${valuation_ranges['equity_range']['high']/1000:.1f}B"),
            ("Implied Share Price", f"${valuation_ranges['share_price_range']['low']:.2f}", f"${valuation_ranges['share_price_range']['median']:.2f}", f"${valuation_ranges['share_price_range']['high']:.2f}")
        ]

        # Headers
        ws.cell(row=current_row, column=1, value="Metric").style = 'header'
        ws.cell(row=current_row, column=2, value="Low").style = 'header'
        ws.cell(row=current_row, column=3, value="Median").style = 'header'
        ws.cell(row=current_row, column=4, value="High").style = 'header'
        current_row += 1

        # Data
        for metric, low, median, high in valuation_data:
            ws.cell(row=current_row, column=1, value=metric).style = 'label_bold'
            ws.cell(row=current_row, column=2, value=low).style = 'output'
            ws.cell(row=current_row, column=3, value=median).style = 'output'
            ws.cell(row=current_row, column=4, value=high).style = 'output'
            current_row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    def _create_deal_data_tab(self, ws, results):
        """Create Deal Data & Inputs tab"""

        assumptions = results['assumptions']

        # Title
        ws['A1'] = f"Precedent Transaction Data: {self.target_company} vs Historical Deals"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:K1')

        current_row = 3

        # Target Company Data
        ws[f'A{current_row}'] = "TARGET COMPANY DATA"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:K{current_row}')
        current_row += 2

        target_data = [
            ("Company Name", assumptions['target']['company_name']),
            ("Ticker", assumptions['target']['ticker']),
            ("Revenue ($M)", assumptions['target']['revenue']),
            ("EBITDA ($M)", assumptions['target']['ebitda']),
            ("Net Income ($M)", assumptions['target']['net_income']),
            ("EPS ($)", assumptions['target']['eps']),
            ("Net Debt ($M)", assumptions['target']['net_debt']),
            ("Shares Outstanding (M)", assumptions['target']['shares_outstanding'])
        ]

        for label, value in target_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            if label == "Company Name" or label == "Ticker":
                ws[f'B{current_row}'].style = 'input'
            else:
                ws[f'B{current_row}'].style = 'target_highlight'
            current_row += 1

        current_row += 2

        # Precedent Deal Data
        ws[f'A{current_row}'] = "PRECEDENT TRANSACTION DATA"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:K{current_row}')
        current_row += 2

        # Headers
        headers = ['Date', 'Acquirer', 'Target', 'Equity Value ($M)', 'Enterprise Value ($M)',
                  'Revenue ($M)', 'EBITDA ($M)', 'Net Income ($M)']

        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header).style = 'header'

        current_row += 1

        # Deal data
        for deal in assumptions['deals']:
            ws.cell(row=current_row, column=1, value=deal['date']).style = 'deal_data'
            ws.cell(row=current_row, column=2, value=deal['acquirer']).style = 'deal_data'
            ws.cell(row=current_row, column=3, value=deal['target']).style = 'deal_data'
            ws.cell(row=current_row, column=4, value=deal['equity_value']).style = 'deal_data'
            ws.cell(row=current_row, column=5, value=deal['enterprise_value']).style = 'deal_data'
            ws.cell(row=current_row, column=6, value=deal['revenue']).style = 'deal_data'
            ws.cell(row=current_row, column=7, value=deal['ebitda']).style = 'deal_data'
            ws.cell(row=current_row, column=8, value=deal['net_income']).style = 'deal_data'
            current_row += 1

        # Set column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_deal_multiples_tab(self, ws, results):
        """Create Deal Multiples tab"""

        deal_multiples = results['deal_multiples']
        summary_stats = results['summary_stats']

        # Title
        ws['A1'] = f"Precedent Transaction Multiples: {len(deal_multiples)} Deals"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:I1')

        current_row = 3

        # Deal Multiples Table
        ws[f'A{current_row}'] = "PRECEDENT DEAL MULTIPLES"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:I{current_row}')
        current_row += 2

        # Headers
        headers = ['Date', 'Acquirer', 'Target', 'EV/Revenue', 'EV/EBITDA', 'Equity Value/Net Income', 'Equity Value ($M)', 'EV ($M)']

        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header).style = 'header'

        current_row += 1

        # Deal data
        for deal in deal_multiples:
            ws.cell(row=current_row, column=1, value=deal['date']).style = 'deal_data'
            ws.cell(row=current_row, column=2, value=deal['acquirer']).style = 'deal_data'
            ws.cell(row=current_row, column=3, value=deal['target']).style = 'deal_data'
            ws.cell(row=current_row, column=4, value=deal['ev_revenue']).style = 'deal_data'
            ws.cell(row=current_row, column=5, value=deal['ev_ebitda']).style = 'deal_data'
            ws.cell(row=current_row, column=6, value=deal['equity_value_net_income']).style = 'deal_data'
            ws.cell(row=current_row, column=7, value=deal['equity_value']).style = 'deal_data'
            ws.cell(row=current_row, column=8, value=deal['enterprise_value']).style = 'deal_data'
            current_row += 1

        # Add summary statistics
        current_row += 2
        ws[f'A{current_row}'] = "SUMMARY STATISTICS"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:I{current_row}')
        current_row += 2

        # Statistics headers
        ws.cell(row=current_row, column=1, value="Statistic").style = 'header'
        ws.cell(row=current_row, column=2, value="EV/Revenue").style = 'header'
        ws.cell(row=current_row, column=3, value="EV/EBITDA").style = 'header'
        ws.cell(row=current_row, column=4, value="Equity Value/Net Income").style = 'header'
        current_row += 1

        # Statistics data
        stats_labels = ['Mean', 'Median', '25th Percentile', '75th Percentile', 'Min', 'Max']
        stats_keys = ['mean', 'median', 'p25', 'p75', 'min', 'max']

        for i, (label, key) in enumerate(zip(stats_labels, stats_keys)):
            ws.cell(row=current_row, column=1, value=label).style = 'statistics'
            ws.cell(row=current_row, column=2, value=summary_stats['ev_revenue'][key]).style = 'statistics'
            ws.cell(row=current_row, column=3, value=summary_stats['ev_ebitda'][key]).style = 'statistics'
            ws.cell(row=current_row, column=4, value=summary_stats['equity_value_net_income'][key]).style = 'statistics'
            current_row += 1

        # Set column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_target_valuation_tab(self, ws, results):
        """Create Target Valuation tab"""

        target_valuation = results['target_valuation']
        valuation_ranges = results['valuation_ranges']

        # Title
        ws['A1'] = f"Target Valuation: {self.target_company} ({self.target_ticker})"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:G1')

        current_row = 3

        # EV/Revenue Valuation
        ws[f'A{current_row}'] = "EV/REVENUE VALUATION"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        ev_rev_data = [
            ("Multiple", target_valuation['ev_revenue']['p25_multiple'], target_valuation['ev_revenue']['median_multiple'], target_valuation['ev_revenue']['p75_multiple']),
            ("Enterprise Value ($M)", target_valuation['ev_revenue']['p25_ev'], target_valuation['ev_revenue']['median_ev'], target_valuation['ev_revenue']['p75_ev'])
        ]

        for label, p25, median, p75 in ev_rev_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = p25
            ws[f'C{current_row}'] = median
            ws[f'D{current_row}'] = p75
            ws[f'B{current_row}'].style = 'calculation'
            ws[f'C{current_row}'].style = 'calculation'
            ws[f'D{current_row}'].style = 'calculation'
            current_row += 1

        current_row += 2

        # EV/EBITDA Valuation
        ws[f'A{current_row}'] = "EV/EBITDA VALUATION"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        ev_ebitda_data = [
            ("Multiple", target_valuation['ev_ebitda']['p25_multiple'], target_valuation['ev_ebitda']['median_multiple'], target_valuation['ev_ebitda']['p75_multiple']),
            ("Enterprise Value ($M)", target_valuation['ev_ebitda']['p25_ev'], target_valuation['ev_ebitda']['median_ev'], target_valuation['ev_ebitda']['p75_ev'])
        ]

        for label, p25, median, p75 in ev_ebitda_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = p25
            ws[f'C{current_row}'] = median
            ws[f'D{current_row}'] = p75
            if label == "Enterprise Value ($M)":
                ws[f'B{current_row}'].style = 'output'
                ws[f'C{current_row}'].style = 'output'
                ws[f'D{current_row}'].style = 'output'
            else:
                ws[f'B{current_row}'].style = 'calculation'
                ws[f'C{current_row}'].style = 'calculation'
                ws[f'D{current_row}'].style = 'calculation'
            current_row += 1

        current_row += 2

        # Equity Value/Net Income Valuation
        ws[f'A{current_row}'] = "EQUITY VALUE/NET INCOME VALUATION"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        equity_value_data = [
            ("Multiple", target_valuation['equity_value_net_income']['p25_multiple'], target_valuation['equity_value_net_income']['median_multiple'], target_valuation['equity_value_net_income']['p75_multiple']),
            ("Equity Value ($M)", target_valuation['equity_value_net_income']['p25_equity_value'], target_valuation['equity_value_net_income']['median_equity_value'], target_valuation['equity_value_net_income']['p75_equity_value'])
        ]

        for label, p25, median, p75 in equity_value_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = p25
            ws[f'C{current_row}'] = median
            ws[f'D{current_row}'] = p75
            if label == "Equity Value ($M)":
                ws[f'B{current_row}'].style = 'output'
                ws[f'C{current_row}'].style = 'output'
                ws[f'D{current_row}'].style = 'output'
            else:
                ws[f'B{current_row}'].style = 'calculation'
                ws[f'C{current_row}'].style = 'calculation'
                ws[f'D{current_row}'].style = 'calculation'
            current_row += 1

        current_row += 2

        # Valuation Summary
        ws[f'A{current_row}'] = "VALUATION SUMMARY (PRIMARY: EV/EBITDA)"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        valuation_summary = [
            ("Enterprise Value Range", f"${valuation_ranges['ev_range']['low']/1000:.1f}B", f"${valuation_ranges['ev_range']['median']/1000:.1f}B", f"${valuation_ranges['ev_range']['high']/1000:.1f}B"),
            ("Equity Value Range", f"${valuation_ranges['equity_range']['low']/1000:.1f}B", f"${valuation_ranges['equity_range']['median']/1000:.1f}B", f"${valuation_ranges['equity_range']['high']/1000:.1f}B"),
            ("Implied Share Price", f"${valuation_ranges['share_price_range']['low']:.2f}", f"${valuation_ranges['share_price_range']['median']:.2f}", f"${valuation_ranges['share_price_range']['high']:.2f}")
        ]

        for label, low, median, high in valuation_summary:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = low
            ws[f'C{current_row}'] = median
            ws[f'D{current_row}'] = high
            ws[f'B{current_row}'].style = 'output'
            ws[f'C{current_row}'].style = 'output'
            ws[f'D{current_row}'].style = 'output'
            current_row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15


def run_sample_precedent_transactions_model():
    """Run a sample precedent transactions model"""

    print("🤝 Running Professional Precedent Transactions Model Sample")
    print("=" * 70)

    # Create model instance
    model = ProfessionalPrecedentTransactionsModel("TechTarget Inc.", "TECHTARGET")

    # Run the model with sample data
    results, excel_file = model.run_precedent_transactions_model(
        # Target Company Data
        target_revenue=1200.0,      # $1.2B revenue
        target_ebitda=360.0,        # $360M EBITDA
        target_net_income=216.0,    # $216M net income
        target_eps=3.60,            # $3.60 EPS
        target_net_debt=240.0,      # $240M net debt
        target_shares_outstanding=60.0  # 60M shares
    )

    return results, excel_file


if __name__ == "__main__":
    # Run sample precedent transactions model
    results, excel_file = run_sample_precedent_transactions_model()

    print("\n📋 Model Complete!")
    print(f"📁 Excel file saved as: {excel_file}")

    # Display key results
    print("\n🔑 Key Results Summary:")
    print("=" * 40)

    valuation_ranges = results['valuation_ranges']
    summary_stats = results['summary_stats']

    print(f"Median EV/Revenue: {summary_stats['ev_revenue']['median']:.1f}x")
    print(f"Median EV/EBITDA: {summary_stats['ev_ebitda']['median']:.1f}x")
    print(f"Median Equity Value/Net Income: {summary_stats['equity_value_net_income']['median']:.1f}x")
    print(f"Implied EV Range: ${valuation_ranges['ev_range']['low']/1000:.1f}B - ${valuation_ranges['ev_range']['high']/1000:.1f}B")
    print(f"Implied Share Price Range: ${valuation_ranges['share_price_range']['low']:.2f} - ${valuation_ranges['share_price_range']['high']:.2f}")
