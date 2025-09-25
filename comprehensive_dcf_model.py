#!/usr/bin/env python3
"""
Comprehensive DCF Model Builder
Includes all essential DCF components: WACC, proper FCF, terminal value, debt adjustments, share price
"""

import sys
import subprocess
import os
import re
import requests
import pandas as pd
from bs4 import BeautifulSoup, Tag
from datetime import datetime
from openpyxl import Workbook
from urllib.parse import urljoin

def install_and_import(package, pip_name=None):
    """Auto-install missing packages."""
    pip_name = pip_name or package
    try:
        __import__(package)
    except ImportError:
        print(f"Installing missing package: {pip_name}")
        subprocess.check_call([sys.executable, "-m", "pip", "install", pip_name])
        __import__(package)

# Install required packages
REQUIRED_MODULES = [
    ("pandas", "pandas"),
    ("openpyxl", "openpyxl"),
    ("requests", "requests"),
    ("bs4", "beautifulsoup4"),
    ("yfinance", "yfinance"),
    ("openai", "openai"),
    ("gspread", "gspread"),
    ("google.auth", "google-auth"),
    ("dotenv", "python-dotenv"),
]

for mod, pip_name in REQUIRED_MODULES:
    install_and_import(mod, pip_name)

# Optional imports
try:
    import openai
except ImportError:
    openai = None
try:
    import yfinance as yf
except ImportError:
    yf = None
try:
    import gspread
    from google.oauth2.service_account import Credentials
except ImportError:
    gspread = None
    Credentials = None

# Load environment variables
from dotenv import load_dotenv
load_dotenv()
OPENAI_API_KEY = os.getenv('OPENAI_API_KEY')
if openai and OPENAI_API_KEY:
    openai.api_key = OPENAI_API_KEY

# Default Configuration
DEFAULT_YEARS = 5
TERMINAL_GROWTH = 0.03  # 3% perpetual growth
RISK_FREE_RATE = 0.04   # 4% risk-free rate
MARKET_RISK_PREMIUM = 0.06  # 6% market risk premium
BETA = 1.1  # Default beta
COST_OF_DEBT = 0.05  # 5% cost of debt
TAX_RATE = 0.25  # 25% effective tax rate

def get_comprehensive_financials(ticker, years):
    """Get comprehensive financial data from yfinance."""
    if not yf:
        print("yfinance not available")
        return {}
    try:
        ticker_obj = yf.Ticker(ticker)
        
        # Get financial statements
        fin = ticker_obj.financials.T
        bal = ticker_obj.balance_sheet.T
        
        # Get market data
        info = ticker_obj.info
        
        years_data = fin.index[-years:]
        data = {}
        
        # Income Statement Items
        if 'Total Revenue' in fin.columns:
            data['Revenue'] = [float(fin.loc[y, 'Total Revenue']) for y in years_data]
        if 'EBITDA' in fin.columns:
            data['EBITDA'] = [float(fin.loc[y, 'EBITDA']) for y in years_data]
        if 'Operating Income' in fin.columns:
            data['EBIT'] = [float(fin.loc[y, 'Operating Income']) for y in years_data]
        if 'Net Income' in fin.columns:
            data['Net Income'] = [float(fin.loc[y, 'Net Income']) for y in years_data]
        if 'Depreciation' in fin.columns:
            data['Depreciation'] = [float(fin.loc[y, 'Depreciation']) for y in years_data]
        if 'Capital Expenditure' in fin.columns:
            data['CapEx'] = [float(fin.loc[y, 'Capital Expenditure']) for y in years_data]
        
        # Balance Sheet Items
        if 'Total Current Assets' in bal.columns and 'Total Current Liabilities' in bal.columns:
            data['Current Assets'] = [float(bal.loc[y, 'Total Current Assets']) for y in years_data]
            data['Current Liabilities'] = [float(bal.loc[y, 'Total Current Liabilities']) for y in years_data]
        
        # Market Data
        data['Market Cap'] = info.get('marketCap', 0)
        data['Total Debt'] = info.get('totalDebt', 0)
        data['Cash'] = info.get('cash', 0)
        data['Shares Outstanding'] = info.get('sharesOutstanding', 0)
        data['Beta'] = info.get('beta', BETA)
        
        return data
    except Exception as e:
        print(f"yfinance extraction failed: {e}")
        return {}

def extract_financials_with_llm(company_name):
    """Extract comprehensive financials using OpenAI."""
    if not openai or not OPENAI_API_KEY:
        print("OpenAI not available")
        return {}
    prompt = f"""
Extract comprehensive financial data for {company_name}. Return a JSON object with these keys:
Revenue, EBITDA, EBIT, Net Income, Depreciation, CapEx, Current Assets, Current Liabilities, 
Total Debt, Cash, Shares Outstanding, Beta. Use 'Unknown' if not found.
"""
    try:
        response = openai.chat.completions.create(
            model="gpt-3.5-turbo-16k",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.0,
            max_tokens=800,
        )
        import json
        content = response.choices[0].message.content
        if content:
            return json.loads(content)
        else:
            return {}
    except Exception as e:
        print(f"LLM extraction failed: {e}")
        return {}

def calculate_wacc(beta, market_cap, total_debt, cost_of_debt=COST_OF_DEBT, tax_rate=TAX_RATE):
    """Calculate Weighted Average Cost of Capital."""
    # Cost of Equity (CAPM)
    cost_of_equity = RISK_FREE_RATE + beta * MARKET_RISK_PREMIUM
    
    # Capital structure
    total_capital = market_cap + total_debt
    equity_weight = market_cap / total_capital if total_capital > 0 else 0.7
    debt_weight = total_debt / total_capital if total_capital > 0 else 0.3
    
    # WACC calculation
    wacc = (equity_weight * cost_of_equity) + (debt_weight * cost_of_debt * (1 - tax_rate))
    
    return wacc, cost_of_equity, equity_weight, debt_weight

def calculate_working_capital_change(current_assets, current_liabilities, years):
    """Calculate change in net working capital."""
    if len(current_assets) < 2 or len(current_liabilities) < 2:
        return [0] * years
    
    nwc_changes = []
    for i in range(1, len(current_assets)):
        nwc_current = current_assets[i] - current_liabilities[i]
        nwc_previous = current_assets[i-1] - current_liabilities[i-1]
        nwc_changes.append(nwc_current - nwc_previous)
    
    # Pad with zeros if needed
    while len(nwc_changes) < years:
        nwc_changes.append(0)
    
    return nwc_changes[:years]

def calculate_unlevered_fcf(revenue, ebit, depreciation, capex, nwc_change, tax_rate=TAX_RATE):
    """Calculate Unlevered Free Cash Flow."""
    fcf = []
    for i in range(len(revenue)):
        # NOPAT = EBIT * (1 - Tax Rate)
        nopat = ebit[i] * (1 - tax_rate)
        
        # FCF = NOPAT + D&A - CapEx - Î”NWC
        fcf_val = nopat + depreciation[i] - capex[i] - nwc_change[i]
        fcf.append(round(fcf_val, 2))
    
    return fcf

def calculate_terminal_value(final_fcf, wacc, growth_rate=TERMINAL_GROWTH):
    """Calculate terminal value using Gordon Growth Model."""
    return final_fcf * (1 + growth_rate) / (wacc - growth_rate)

def discount_cash_flows(fcfs, terminal_value, wacc):
    """Discount cash flows to present value."""
    discounted_fcfs = []
    for i, fcf in enumerate(fcfs):
        pv = fcf / ((1 + wacc) ** (i + 1))
        discounted_fcfs.append(round(pv, 2))
    
    # Discount terminal value
    pv_terminal = terminal_value / ((1 + wacc) ** len(fcfs))
    
    return discounted_fcfs, round(pv_terminal, 2)

def calculate_equity_value(enterprise_value, total_debt, cash):
    """Calculate equity value from enterprise value."""
    net_debt = total_debt - cash
    equity_value = enterprise_value - net_debt
    return equity_value, net_debt

def calculate_share_price(equity_value, shares_outstanding):
    """Calculate intrinsic share price."""
    if shares_outstanding > 0:
        return equity_value / shares_outstanding
    return 0

def project_financials(base_revenue, years, growth_rate=0.08):
    """Project financials into the future."""
    revenue = [round(base_revenue * ((1 + growth_rate) ** i), 2) for i in range(years)]
    
    # Project other items as percentages of revenue
    ebitda = [r * 0.25 for r in revenue]  # 25% EBITDA margin
    ebit = [r * 0.15 for r in revenue]    # 15% EBIT margin
    depreciation = [r * 0.05 for r in revenue]  # 5% of revenue
    capex = [r * 0.06 for r in revenue]   # 6% of revenue
    
    return revenue, ebitda, ebit, depreciation, capex

def write_comprehensive_excel(company, years, revenue, ebitda, ebit, depreciation, capex, nwc_change, fcf, 
                            discounted_fcfs, terminal_value, pv_terminal, enterprise_value, equity_value, 
                            share_price, wacc, cost_of_equity, filename):
    """Write comprehensive DCF model to Excel."""
    wb = Workbook()
    ws = wb.active
    if ws:
        ws.title = "Comprehensive DCF Model"
        
        # Header
        ws.append(["COMPREHENSIVE DCF MODEL"])
        ws.append(["Company:", company])
        ws.append([])
        
        # WACC Summary
        ws.append(["WACC CALCULATION"])
        ws.append(["Cost of Equity:", f"{cost_of_equity:.1%}"])
        ws.append(["WACC:", f"{wacc:.1%}"])
        ws.append([])
        
        # Financial Projections
        ws.append(["FINANCIAL PROJECTIONS"])
        ws.append(["Year"] + [str(y) for y in years])
        ws.append(["Revenue"] + [f"${val:,.0f}" for val in revenue])
        ws.append(["EBITDA"] + [f"${val:,.0f}" for val in ebitda])
        ws.append(["EBIT"] + [f"${val:,.0f}" for val in ebit])
        ws.append(["Depreciation"] + [f"${val:,.0f}" for val in depreciation])
        ws.append(["CapEx"] + [f"${val:,.0f}" for val in capex])
        ws.append(["Î”NWC"] + [f"${val:,.0f}" for val in nwc_change])
        ws.append([])
        
        # Free Cash Flow
        ws.append(["FREE CASH FLOW"])
        ws.append(["Unlevered FCF"] + [f"${val:,.0f}" for val in fcf])
        ws.append([])
        
        # Valuation
        ws.append(["VALUATION"])
        ws.append(["Discounted FCFs"] + [f"${val:,.0f}" for val in discounted_fcfs])
        ws.append(["Terminal Value", f"${terminal_value:,.0f}"])
        ws.append(["PV of Terminal Value", f"${pv_terminal:,.0f}"])
        ws.append([])
        ws.append(["Enterprise Value", f"${enterprise_value:,.0f}"])
        ws.append(["Equity Value", f"${equity_value:,.0f}"])
        ws.append(["Share Price", f"${share_price:.2f}"])
        
        # Formatting
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Format headers
        for row in [1, 5, 15, 22]:
            for cell in ws[row]:
                cell.font = header_font
                cell.fill = header_fill
        
        # Save file
        wb.save(filename)
        print(f"âœ… Saved comprehensive DCF model to {filename}")
    else:
        print("âŒ Could not create Excel worksheet")

def write_comprehensive_gsheets(company, years, revenue, ebitda, ebit, depreciation, capex, nwc_change, fcf, 
                              discounted_fcfs, terminal_value, pv_terminal, enterprise_value, equity_value, 
                              share_price, wacc, cost_of_equity, sheet_name, worksheet_name):
    """Write comprehensive DCF model to Google Sheets."""
    if not gspread or not Credentials:
        print("Google Sheets not available")
        return False
    
    creds_path = os.getenv('GOOGLE_SHEETS_CREDENTIALS') or 'credentials/google_sheets_credentials.json'
    
    try:
        scopes = ['https://www.googleapis.com/auth/spreadsheets']
        creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
        gc = gspread.authorize(creds)
        
        # Open or create sheet
        try:
            sh = gc.open(sheet_name)
        except gspread.SpreadsheetNotFound:
            sh = gc.create(sheet_name)
        
        # Get or create worksheet
        try:
            ws = sh.worksheet(worksheet_name)
        except gspread.WorksheetNotFound:
            ws = sh.add_worksheet(title=worksheet_name, rows=50, cols=20)
        
        # Clear and update data
        ws.clear()
        
        # Prepare comprehensive data
        data = []
        data.append(["COMPREHENSIVE DCF MODEL"])
        data.append(["Company:", company])
        data.append([])
        
        # WACC Summary
        data.append(["WACC CALCULATION"])
        data.append(["Cost of Equity:", f"{cost_of_equity:.1%}"])
        data.append(["WACC:", f"{wacc:.1%}"])
        data.append([])
        
        # Financial Projections
        data.append(["FINANCIAL PROJECTIONS"])
        data.append(["Year"] + [str(y) for y in years])
        data.append(["Revenue"] + [f"${val:,.0f}" for val in revenue])
        data.append(["EBITDA"] + [f"${val:,.0f}" for val in ebitda])
        data.append(["EBIT"] + [f"${val:,.0f}" for val in ebit])
        data.append(["Depreciation"] + [f"${val:,.0f}" for val in depreciation])
        data.append(["CapEx"] + [f"${val:,.0f}" for val in capex])
        data.append(["Î”NWC"] + [f"${val:,.0f}" for val in nwc_change])
        data.append([])
        
        # Free Cash Flow
        data.append(["FREE CASH FLOW"])
        data.append(["Unlevered FCF"] + [f"${val:,.0f}" for val in fcf])
        data.append([])
        
        # Valuation
        data.append(["VALUATION"])
        data.append(["Discounted FCFs"] + [f"${val:,.0f}" for val in discounted_fcfs])
        data.append(["Terminal Value", f"${terminal_value:,.0f}"])
        data.append(["PV of Terminal Value", f"${pv_terminal:,.0f}"])
        data.append([])
        data.append(["Enterprise Value", f"${enterprise_value:,.0f}"])
        data.append(["Equity Value", f"${equity_value:,.0f}"])
        data.append(["Share Price", f"${share_price:.2f}"])
        
        # Update the sheet
        ws.update('A1', data)
        
        # Try formatting
        try:
            fmt = {"textFormat": {"bold": True}}
            ws.format("A1:A1", fmt)
            ws.format("A5:A5", fmt)
            ws.format("A15:A15", fmt)
            ws.format("A22:A22", fmt)
        except Exception:
            pass
        
        print(f"âœ… Comprehensive DCF model written to {worksheet_name} tab")
        print(f"ðŸ“Š Sheet URL: {sh.url}")
        return True
        
    except Exception as e:
        print(f"âŒ Error writing to Google Sheets: {e}")
        return False

def build_comprehensive_dcf_model(company_name, ticker=None, years=DEFAULT_YEARS, output="both", sheet_name="DCF", worksheet_name="Comprehensive"):
    """Build comprehensive DCF model."""
    print(f"Building comprehensive DCF model for {company_name}")
    print(f"Forecast period: {years} years")
    
    # Get comprehensive financial data
    data = get_comprehensive_financials(ticker, years) if ticker else {}
    if not data or not data.get('Revenue'):
        print("Falling back to LLM extraction...")
        data = extract_financials_with_llm(company_name)
    
    # Use default values if no data found
    if not data or not data.get('Revenue'):
        print("Using default financial assumptions...")
        base_revenue = 1000000  # $1M default
        market_cap = 5000000    # $5M default
        total_debt = 1000000    # $1M default
        cash = 500000           # $500K default
        shares_outstanding = 1000000  # 1M shares
        beta = BETA
    else:
        base_revenue = float(str(data.get('Revenue', [0])[0] if isinstance(data.get('Revenue'), list) else data.get('Revenue', 0)).replace(',', '').replace('$', ''))
        market_cap = data.get('Market Cap', 5000000)
        total_debt = data.get('Total Debt', 1000000)
        cash = data.get('Cash', 500000)
        shares_outstanding = data.get('Shares Outstanding', 1000000)
        beta = data.get('Beta', BETA)
    
    # Calculate WACC
    wacc, cost_of_equity, equity_weight, debt_weight = calculate_wacc(beta, market_cap, total_debt)
    
    # Project financials
    revenue, ebitda, ebit, depreciation, capex = project_financials(base_revenue, years)
    
    # Calculate working capital changes
    current_assets = data.get('Current Assets', [base_revenue * 0.3] * years)
    current_liabilities = data.get('Current Liabilities', [base_revenue * 0.2] * years)
    nwc_change = calculate_working_capital_change(current_assets, current_liabilities, years)
    
    # Calculate Free Cash Flow
    fcf = calculate_unlevered_fcf(revenue, ebit, depreciation, capex, nwc_change)
    
    # Calculate Terminal Value
    terminal_value = calculate_terminal_value(fcf[-1], wacc)
    
    # Discount cash flows
    discounted_fcfs, pv_terminal = discount_cash_flows(fcf, terminal_value, wacc)
    
    # Calculate Enterprise Value
    enterprise_value = sum(discounted_fcfs) + pv_terminal
    
    # Calculate Equity Value and Share Price
    equity_value, net_debt = calculate_equity_value(enterprise_value, total_debt, cash)
    share_price = calculate_share_price(equity_value, shares_outstanding)
    
    # Create years
    forecast_years = [datetime.now().year + i for i in range(years)]
    
    # Output results
    success = True
    
    if output in ["excel", "both"]:
        filename = f"comprehensive_dcf_{company_name.replace(' ', '_')}.xlsx"
        write_comprehensive_excel(company_name, forecast_years, revenue, ebitda, ebit, depreciation, capex, 
                                nwc_change, fcf, discounted_fcfs, terminal_value, pv_terminal, 
                                enterprise_value, equity_value, share_price, wacc, cost_of_equity, filename)
    
    if output in ["gsheet", "both"]:
        gsheet_success = write_comprehensive_gsheets(company_name, forecast_years, revenue, ebitda, ebit, depreciation, capex, 
                                                   nwc_change, fcf, discounted_fcfs, terminal_value, pv_terminal, 
                                                   enterprise_value, equity_value, share_price, wacc, cost_of_equity, 
                                                   sheet_name, worksheet_name)
        if not gsheet_success:
            success = False
    
    if success:
        # Print comprehensive summary
        print(f"\nðŸ“Š COMPREHENSIVE DCF MODEL SUMMARY for {company_name}")
        print("=" * 60)
        print(f"Forecast Period: {years} years")
        print(f"WACC Calculation:")
        print(f"  Cost of Equity: {cost_of_equity:.1%}")
        print(f"  WACC: {wacc:.1%}")
        print(f"  Equity Weight: {equity_weight:.1%}")
        print(f"  Debt Weight: {debt_weight:.1%}")
        print()
        print(f"Valuation:")
        print(f"  Enterprise Value: ${enterprise_value:,.0f}")
        print(f"  Net Debt: ${net_debt:,.0f}")
        print(f"  Equity Value: ${equity_value:,.0f}")
        print(f"  Share Price: ${share_price:.2f}")
        print()
        print(f"Key Assumptions:")
        print(f"  Terminal Growth: {TERMINAL_GROWTH:.1%}")
        print(f"  Tax Rate: {TAX_RATE:.1%}")
        print(f"  Risk-Free Rate: {RISK_FREE_RATE:.1%}")
        print(f"  Market Risk Premium: {MARKET_RISK_PREMIUM:.1%}")
        print(f"  Beta: {beta:.2f}")
        print("=" * 60)
        print(f"âœ… Comprehensive DCF model successfully created!")
    else:
        print("âŒ Some outputs failed")

def main_menu():
    """Main interactive menu."""
    print("ðŸš€ Comprehensive DCF Model Builder")
    print("=" * 60)
    print("This tool creates professional DCF models with:")
    print("â€¢ WACC calculation (CAPM)")
    print("â€¢ Proper FCF calculation (NOPAT + D&A - CapEx - Î”NWC)")
    print("â€¢ Terminal value (Gordon Growth Model)")
    print("â€¢ Debt/cash adjustments")
    print("â€¢ Share price calculation")
    print("â€¢ Excel and Google Sheets output")
    print()
    
    # Get user inputs
    company_name = input("Enter company name: ").strip()
    ticker = input("Enter ticker symbol (optional): ").strip() or None
    
    # Get forecast years
    while True:
        try:
            years_input = input(f"Enter forecast period in years [{DEFAULT_YEARS}]: ").strip()
            years = int(years_input) if years_input else DEFAULT_YEARS
            if years > 0 and years <= 20:
                break
            else:
                print("Please enter a number between 1 and 20")
        except ValueError:
            print("Please enter a valid number")
    
    # Choose output
    output = input("Output type (excel/gsheet/both) [both]: ").strip() or "both"
    
    # Get Google Sheets details if needed
    sheet_name = "DCF"
    worksheet_name = "Comprehensive"
    if output in ["gsheet", "both"]:
        sheet_name = input("Enter Google Sheet name [DCF]: ").strip() or "DCF"
        worksheet_name = input("Enter worksheet/tab name [Comprehensive]: ").strip() or "Comprehensive"
    
    # Build the comprehensive model
    build_comprehensive_dcf_model(company_name, ticker, years, output, sheet_name, worksheet_name)

if __name__ == "__main__":
    main_menu() 