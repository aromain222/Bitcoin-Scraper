#!/usr/bin/env python3
"""
Google Drive Setup Script for Excel Sharing System
Helps configure Google Drive integration for sharing financial models
"""

import os
import json
from pathlib import Path

def setup_google_drive():
    """Interactive setup for Google Drive integration."""
    print("🚀 Google Drive Setup for Excel Sharing System")
    print("=" * 50)
    
    print("\n📋 Step 1: Google Cloud Console Setup")
    print("1. Go to https://console.cloud.google.com/")
    print("2. Create a new project or select existing one")
    print("3. Enable Google Drive API:")
    print("   - Go to APIs & Services > Library")
    print("   - Search for 'Google Drive API'")
    print("   - Click Enable")
    
    print("\n🔑 Step 2: Create Service Account")
    print("1. Go to APIs & Services > Credentials")
    print("2. Click 'Create Credentials' > 'Service Account'")
    print("3. Fill in service account details")
    print("4. Add 'Editor' or 'Owner' role")
    print("5. Create and download JSON key file")
    
    print("\n📁 Step 3: Configure Credentials")
    credentials_path = input("\nEnter path to your Google Drive credentials JSON file: ").strip()
    
    if not credentials_path:
        credentials_path = "google_drive_credentials.json"
    
    if not os.path.exists(credentials_path):
        print(f"❌ Credentials file not found: {credentials_path}")
        print("Please download the JSON file from Google Cloud Console first.")
        return False
    
    # Validate JSON file
    try:
        with open(credentials_path, 'r') as f:
            creds = json.load(f)
        
        required_fields = ['type', 'project_id', 'private_key_id', 'private_key', 'client_email', 'client_id']
        missing_fields = [field for field in required_fields if field not in creds]
        
        if missing_fields:
            print(f"❌ Invalid credentials file. Missing fields: {missing_fields}")
            return False
        
        print("✅ Credentials file validated successfully")
        
    except json.JSONDecodeError:
        print("❌ Invalid JSON file")
        return False
    except Exception as e:
        print(f"❌ Error reading credentials: {e}")
        return False
    
    # Ask for folder ID (optional)
    folder_id = input("\nEnter Google Drive folder ID (optional, press Enter to skip): ").strip()
    
    # Create configuration
    config = {
        "provider": "google_drive",
        "credentials_path": credentials_path,
        "folder_id": folder_id if folder_id else None,
        "link_expiration_days": 30,
        "enable_access_logging": True,
        "preserve_formatting": True
    }
    
    # Save configuration
    config_path = "google_drive_config.json"
    with open(config_path, 'w') as f:
        json.dump(config, f, indent=2)
    
    print(f"\n✅ Configuration saved to {config_path}")
    
    # Test the setup
    print("\n🧪 Testing Google Drive Integration...")
    try:
        from excel_sharing_system import ExcelSharingSystem, SharingConfig
        
        # Load configuration
        sharing_config = SharingConfig(
            provider="google_drive",
            credentials_path=credentials_path,
            folder_id=folder_id if folder_id else None,
            link_expiration_days=30
        )
        
        sharing_system = ExcelSharingSystem(sharing_config)
        
        # Test pre-flight check
        test_file = "test_excel.xlsx"
        if not os.path.exists(test_file):
            print(f"Creating test file: {test_file}")
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Test"
            ws['A1'] = "Test Excel File"
            ws['A2'] = "Created for Google Drive setup"
            wb.save(test_file)
        
        validation = sharing_system.pre_flight_check(test_file, "test")
        
        if validation.get('file_exists') and validation.get('excel_format_valid'):
            print("✅ Pre-flight check passed!")
            print("🎉 Google Drive setup complete!")
            print(f"\n📝 Next steps:")
            print(f"1. Use the sharing system with: python excel_sharing_demo.py")
            print(f"2. Or integrate into your app with the config in {config_path}")
            
            # Clean up test file
            os.remove(test_file)
            return True
        else:
            print("❌ Pre-flight check failed")
            print(f"Validation results: {validation}")
            return False
            
    except Exception as e:
        print(f"❌ Test failed: {e}")
        return False

if __name__ == "__main__":
    setup_google_drive()
