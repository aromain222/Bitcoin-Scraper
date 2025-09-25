#!/usr/bin/env python3
"""
Professional Leveraged Buyout (LBO) Model Builder
Creates comprehensive LBO analysis with professional formatting and Excel output

Author: Investment Banking Modeler
Date: 2024

Features:
- Transaction Assumptions & Financing Structure
- Sources & Uses Table
- 5-7 Year Operating Forecast
- Comprehensive Debt Schedule with Multiple Tranches
- Cash Flow Waterfall
- Exit Analysis with Multiple Scenarios
- IRR & MOIC Calculations (Base, Bull, Bear)
- Sensitivity Analysis
- Professional Excel Output with Formatting
"""

import numpy as np
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, ScatterChart, Series
# Optional matplotlib import for charting
try:
    import matplotlib.pyplot as plt
    HAS_MATPLOTLIB = True
except ImportError:
    HAS_MATPLOTLIB = False
# from scipy.optimize import newton  # Not currently used
import warnings
warnings.filterwarnings('ignore')

# Professional color scheme
LBO_COLORS = {
    'header_blue': '1F4E79',
    'input_light_blue': 'DDEBF7',
    'calculation_yellow': 'FFF2CC',
    'result_green': 'D8E4BC',
    'warning_orange': 'FCE4D6',
    'border_dark': '2C3E50',
    'text_dark': '000000',
    'text_white': 'FFFFFF'
}

class ProfessionalLBOModel:
    """
    Comprehensive LBO Model with Professional Formatting
    """

    def __init__(self, company_name="Sample Company", ticker="TICK"):
        self.company_name = company_name
        self.ticker = ticker
        self.model_date = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Initialize styles
        self.styles = self._create_styles()

    def _create_styles(self):
        """Create professional Excel styles"""
        styles = {}

        # Header style
        styles['header'] = NamedStyle(name='header')
        styles['header'].font = Font(name='Calibri', size=12, bold=True, color=LBO_COLORS['text_white'])
        styles['header'].fill = PatternFill(start_color=LBO_COLORS['header_blue'], end_color=LBO_COLORS['header_blue'], fill_type='solid')
        styles['header'].alignment = Alignment(horizontal='center', vertical='center')

        # Input style
        styles['input'] = NamedStyle(name='input')
        styles['input'].font = Font(name='Calibri', size=10, color=LBO_COLORS['text_dark'])
        styles['input'].fill = PatternFill(start_color=LBO_COLORS['input_light_blue'], end_color=LBO_COLORS['input_light_blue'], fill_type='solid')
        styles['input'].alignment = Alignment(horizontal='right', vertical='center')
        styles['input'].number_format = '#,##0'

        # Calculation style
        styles['calculation'] = NamedStyle(name='calculation')
        styles['calculation'].font = Font(name='Calibri', size=10, color=LBO_COLORS['text_dark'])
        styles['calculation'].fill = PatternFill(start_color=LBO_COLORS['calculation_yellow'], end_color=LBO_COLORS['calculation_yellow'], fill_type='solid')
        styles['calculation'].alignment = Alignment(horizontal='right', vertical='center')
        styles['calculation'].number_format = '#,##0'

        # Result style
        styles['result'] = NamedStyle(name='result')
        styles['result'].font = Font(name='Calibri', size=10, bold=True, color=LBO_COLORS['text_dark'])
        styles['result'].fill = PatternFill(start_color=LBO_COLORS['result_green'], end_color=LBO_COLORS['result_green'], fill_type='solid')
        styles['result'].alignment = Alignment(horizontal='right', vertical='center')
        styles['result'].number_format = '#,##0'

        # Warning style
        styles['warning'] = NamedStyle(name='warning')
        styles['warning'].font = Font(name='Calibri', size=10, color=LBO_COLORS['text_dark'])
        styles['warning'].fill = PatternFill(start_color=LBO_COLORS['warning_orange'], end_color=LBO_COLORS['warning_orange'], fill_type='solid')
        styles['warning'].alignment = Alignment(horizontal='center', vertical='center')

        # Bold label style
        styles['label_bold'] = NamedStyle(name='label_bold')
        styles['label_bold'].font = Font(name='Calibri', size=10, bold=True, color=LBO_COLORS['text_dark'])
        styles['label_bold'].alignment = Alignment(horizontal='left', vertical='center')

        # Normal label style
        styles['label'] = NamedStyle(name='label')
        styles['label'].font = Font(name='Calibri', size=10, color=LBO_COLORS['text_dark'])
        styles['label'].alignment = Alignment(horizontal='left', vertical='center')

        return styles

    def run_lbo_model(self,
                     # Transaction Assumptions
                     entry_ebitda=200.0,  # $M
                     entry_multiple=10.0,
                     exit_multiple_base=11.0,
                     exit_multiple_bull=12.5,
                     exit_multiple_bear=9.5,

                     # Financing Structure
                     senior_debt_pct=0.50,  # % of purchase price
                     mezzanine_pct=0.10,
                     equity_pct=0.40,
                     fees_pct=0.02,  # Transaction fees as % of purchase price

                     # Senior Debt Terms
                     senior_rate=0.065,  # 6.5%
                     senior_term=7,  # years
                     senior_amort_pct=0.10,  # 10% annual mandatory amortization

                     # Mezzanine Debt Terms
                     mezz_rate=0.125,  # 12.5%
                     mezz_term=8,  # years
                     mezz_amort_pct=0.05,  # 5% annual mandatory amortization

                     # Operating Assumptions
                     revenue_growth=[0.08, 0.06, 0.05, 0.04, 0.03, 0.02],  # Year 1-6
                     ebitda_margin=0.28,
                     capex_pct=0.06,  # % of revenue
                     nwc_pct=0.03,  # % of revenue
                     tax_rate=0.25,
                     da_pct=0.08,  # D&A as % of revenue

                     # Optional Prepayments
                     optional_prepay_senior=0.0,  # $M per year
                     optional_prepay_mezz=0.0,  # $M per year

                     forecast_years=6):

        """
        Run complete LBO model with all components
        """

        print(f"🏦 Building Professional LBO Model for {self.company_name} ({self.ticker})")
        print("=" * 80)

        # Step 1: Transaction Assumptions
        assumptions = self._create_transaction_assumptions(
            entry_ebitda, entry_multiple, exit_multiple_base, exit_multiple_bull, exit_multiple_bear,
            senior_debt_pct, mezzanine_pct, equity_pct, fees_pct,
            senior_rate, senior_term, senior_amort_pct,
            mezz_rate, mezz_term, mezz_amort_pct,
            revenue_growth, ebitda_margin, capex_pct, nwc_pct, tax_rate, da_pct,
            optional_prepay_senior, optional_prepay_mezz, forecast_years
        )

        # Step 2: Sources & Uses
        sources_uses = self._create_sources_uses(assumptions)

        # Step 3: Operating Forecast
        operating_forecast = self._create_operating_forecast(assumptions)

        # Step 4: Debt Schedule
        debt_schedule = self._create_debt_schedule(assumptions, operating_forecast)

        # Step 5: Cash Flow Waterfall
        cash_flow_waterfall = self._create_cash_flow_waterfall(assumptions, operating_forecast, debt_schedule)

        # Step 6: Exit Analysis
        exit_analysis = self._create_exit_analysis(assumptions, operating_forecast, debt_schedule)

        # Step 7: Returns Analysis
        returns_analysis = self._calculate_returns(assumptions, exit_analysis)

        # Step 8: Sensitivity Analysis
        sensitivity_analysis = self._create_sensitivity_analysis(assumptions)

        # Compile results
        lbo_results = {
            'assumptions': assumptions,
            'sources_uses': sources_uses,
            'operating_forecast': operating_forecast,
            'debt_schedule': debt_schedule,
            'cash_flow_waterfall': cash_flow_waterfall,
            'exit_analysis': exit_analysis,
            'returns_analysis': returns_analysis,
            'sensitivity_analysis': sensitivity_analysis
        }

        # Create Excel output
        excel_file = self._create_excel_output(lbo_results)

        print("\n✅ LBO Model Complete!")
        print("📊 Key Metrics:")
        print(f"   • Purchase Price: ${assumptions['purchase_price']/1000:.0f}B")
        print(f"   • Total Debt: ${assumptions['total_debt']/1000:.0f}B ({assumptions['total_debt']/assumptions['equity_investment']:.1f}x equity)")
        print(f"   • IRR: {returns_analysis['irr_base']:.1%} (Base), {returns_analysis['irr_bull']:.1%} (Bull), {returns_analysis['irr_bear']:.1%} (Bear)")
        print(f"   • MOIC: {returns_analysis['moic_base']:.1f}x (Base), {returns_analysis['moic_bull']:.1f}x (Bull), {returns_analysis['moic_bear']:.1f}x (Bear)")
        print(f"📁 Excel Output: {excel_file}")

        return lbo_results, excel_file

    def _create_transaction_assumptions(self, entry_ebitda, entry_multiple, exit_multiple_base,
                                       exit_multiple_bull, exit_multiple_bear, senior_debt_pct,
                                       mezzanine_pct, equity_pct, fees_pct, senior_rate,
                                       senior_term, senior_amort_pct, mezz_rate, mezz_term,
                                       mezz_amort_pct, revenue_growth, ebitda_margin, capex_pct,
                                       nwc_pct, tax_rate, da_pct, optional_prepay_senior,
                                       optional_prepay_mezz, forecast_years):

        """Create comprehensive transaction assumptions"""

        assumptions = {
            # Entry Assumptions
            'entry_ebitda': entry_ebitda,
            'entry_multiple': entry_multiple,
            'purchase_price': entry_ebitda * entry_multiple,

            # Exit Assumptions
            'exit_multiple_base': exit_multiple_base,
            'exit_multiple_bull': exit_multiple_bull,
            'exit_multiple_bear': exit_multiple_bear,

            # Financing Structure (% of purchase price)
            'senior_debt_pct': senior_debt_pct,
            'mezzanine_pct': mezzanine_pct,
            'equity_pct': equity_pct,
            'fees_pct': fees_pct,

            # Financing Amounts
            'senior_debt': entry_ebitda * entry_multiple * senior_debt_pct,
            'mezzanine_debt': entry_ebitda * entry_multiple * mezzanine_pct,
            'equity_investment': entry_ebitda * entry_multiple * equity_pct,
            'transaction_fees': entry_ebitda * entry_multiple * fees_pct,
            'total_debt': entry_ebitda * entry_multiple * (senior_debt_pct + mezzanine_pct),

            # Senior Debt Terms
            'senior_rate': senior_rate,
            'senior_term': senior_term,
            'senior_amort_pct': senior_amort_pct,

            # Mezzanine Debt Terms
            'mezz_rate': mezz_rate,
            'mezz_term': mezz_term,
            'mezz_amort_pct': mezz_amort_pct,

            # Operating Assumptions
            'revenue_growth': revenue_growth,
            'ebitda_margin': ebitda_margin,
            'capex_pct': capex_pct,
            'nwc_pct': nwc_pct,
            'tax_rate': tax_rate,
            'da_pct': da_pct,

            # Optional Prepayments
            'optional_prepay_senior': optional_prepay_senior,
            'optional_prepay_mezz': optional_prepay_mezz,

            # Model Parameters
            'forecast_years': forecast_years,
            'years': list(range(2025, 2025 + forecast_years))
        }

        print("📋 Transaction Assumptions:")
        print(f"   • Entry EBITDA: ${entry_ebitda:.0f}M at {entry_multiple:.1f}x = ${assumptions['purchase_price']:.0f}M purchase price")
        print(f"   • Financing: {senior_debt_pct:.0%} Senior, {mezzanine_pct:.0%} Mezzanine, {equity_pct:.0%} Equity")
        print(f"   • Total Debt: ${assumptions['total_debt']:.0f}M | Equity: ${assumptions['equity_investment']:.0f}M")
        print(f"   • Debt/Equity Ratio: {assumptions['total_debt']/assumptions['equity_investment']:.1f}x")

        return assumptions

    def _create_sources_uses(self, assumptions):
        """Create Sources & Uses table"""

        sources = {
            'Senior Debt': assumptions['senior_debt'],
            'Mezzanine Debt': assumptions['mezzanine_debt'],
            'Equity Investment': assumptions['equity_investment'],
            'Total Sources': assumptions['senior_debt'] + assumptions['mezzanine_debt'] + assumptions['equity_investment']
        }

        uses = {
            'Purchase Price': assumptions['purchase_price'],
            'Transaction Fees': assumptions['transaction_fees'],
            'Total Uses': assumptions['purchase_price'] + assumptions['transaction_fees']
        }

        sources_uses = {
            'sources': sources,
            'uses': uses,
            'check': abs(sources['Total Sources'] - uses['Total Uses']) < 0.01
        }

        print("💰 Sources & Uses:")
        print(f"   • Sources: ${sources['Total Sources']:.0f}M | Uses: ${uses['Total Uses']:.0f}M")
        print(f"   • Balance Check: {'✅ Balanced' if sources_uses['check'] else '❌ Mismatch'}")

        return sources_uses

    def _create_operating_forecast(self, assumptions):
        """Create 5-7 year operating forecast"""

        years = assumptions['years']
        forecast_years = assumptions['forecast_years']

        # Start with entry year revenue (reverse from EBITDA)
        entry_revenue = assumptions['entry_ebitda'] / assumptions['ebitda_margin']
        revenue = [entry_revenue]

        # Project revenue growth
        for i in range(1, forecast_years):
            if i-1 < len(assumptions['revenue_growth']):
                growth = assumptions['revenue_growth'][i-1]
            else:
                growth = assumptions['revenue_growth'][-1]  # Use terminal growth
            revenue.append(revenue[-1] * (1 + growth))

        # Calculate other line items
        ebitda = [rev * assumptions['ebitda_margin'] for rev in revenue]
        da = [rev * assumptions['da_pct'] for rev in revenue]
        ebit = [ebitda[i] - da[i] for i in range(len(ebitda))]
        capex = [rev * assumptions['capex_pct'] for rev in revenue]
        nwc_change = [revenue[0] * assumptions['nwc_pct']]  # Initial NWC investment
        for i in range(1, len(revenue)):
            nwc_change.append((revenue[i] - revenue[i-1]) * assumptions['nwc_pct'])

        # Unlevered Free Cash Flow (UFCF)
        nopat = [ebit[i] * (1 - assumptions['tax_rate']) for i in range(len(ebit))]
        ufcf = [nopat[i] + da[i] - capex[i] - nwc_change[i] for i in range(len(nopat))]

        operating_forecast = {
            'years': years,
            'revenue': revenue,
            'ebitda': ebitda,
            'da': da,
            'ebit': ebit,
            'capex': capex,
            'nwc_change': nwc_change,
            'nopat': nopat,
            'ufcf': ufcf
        }

        print("📈 Operating Forecast:")
        print(f"   • Year 1 Revenue: ${revenue[1]/1000:.1f}B | EBITDA: ${ebitda[1]:.0f}M")
        print(f"   • Year {forecast_years} Revenue: ${revenue[-1]/1000:.1f}B | EBITDA: ${ebitda[-1]:.0f}M")
        print(f"   • CAGR: {((revenue[-1]/revenue[1])**(1/(forecast_years-1))-1)*100:.1f}%")

        return operating_forecast

    def _create_debt_schedule(self, assumptions, operating_forecast):
        """Create comprehensive debt schedule with multiple tranches"""

        years = assumptions['forecast_years']

        # Senior Debt Schedule
        senior_balance = [assumptions['senior_debt']]
        senior_interest = []
        senior_mandatory = []
        senior_optional = []
        senior_total_paydown = []

        # Mezzanine Debt Schedule
        mezz_balance = [assumptions['mezzanine_debt']]
        mezz_interest = []
        mezz_mandatory = []
        mezz_optional = []
        mezz_total_paydown = []

        for year in range(years):
            # Senior Debt
            interest_senior = senior_balance[-1] * assumptions['senior_rate']
            mandatory_senior = min(assumptions['senior_amort_pct'] * assumptions['senior_debt'], senior_balance[-1])
            optional_senior = min(assumptions['optional_prepay_senior'], senior_balance[-1] - mandatory_senior)
            total_paydown_senior = mandatory_senior + optional_senior
            new_senior_balance = senior_balance[-1] - total_paydown_senior

            senior_interest.append(interest_senior)
            senior_mandatory.append(mandatory_senior)
            senior_optional.append(optional_senior)
            senior_total_paydown.append(total_paydown_senior)
            senior_balance.append(new_senior_balance)

            # Mezzanine Debt
            interest_mezz = mezz_balance[-1] * assumptions['mezz_rate']
            mandatory_mezz = min(assumptions['mezz_amort_pct'] * assumptions['mezzanine_debt'], mezz_balance[-1])
            optional_mezz = min(assumptions['optional_prepay_mezz'], mezz_balance[-1] - mandatory_mezz)
            total_paydown_mezz = mandatory_mezz + optional_mezz
            new_mezz_balance = mezz_balance[-1] - total_paydown_mezz

            mezz_interest.append(interest_mezz)
            mezz_mandatory.append(mandatory_mezz)
            mezz_optional.append(optional_mezz)
            mezz_total_paydown.append(total_paydown_mezz)
            mezz_balance.append(new_mezz_balance)

        # Remove the last balance (which is beyond the forecast period)
        senior_balance = senior_balance[:-1]
        mezz_balance = mezz_balance[:-1]

        debt_schedule = {
            'years': operating_forecast['years'],
            'senior_balance': senior_balance,
            'senior_interest': senior_interest,
            'senior_mandatory': senior_mandatory,
            'senior_optional': senior_optional,
            'senior_total_paydown': senior_total_paydown,
            'mezz_balance': mezz_balance,
            'mezz_interest': mezz_interest,
            'mezz_mandatory': mezz_mandatory,
            'mezz_optional': mezz_optional,
            'mezz_total_paydown': mezz_total_paydown,
            'total_debt_balance': [senior_balance[i] + mezz_balance[i] for i in range(len(senior_balance))],
            'total_interest': [senior_interest[i] + mezz_interest[i] for i in range(len(senior_interest))],
            'total_debt_paydown': [senior_total_paydown[i] + mezz_total_paydown[i] for i in range(len(senior_total_paydown))]
        }

        print("🏦 Debt Schedule:")
        print(f"   • Year 1 Senior Balance: ${senior_balance[0]:.0f}M | Interest: ${senior_interest[0]:.0f}M")
        print(f"   • Year 1 Mezz Balance: ${mezz_balance[0]:.0f}M | Interest: ${mezz_interest[0]:.0f}M")
        print(f"   • Final Debt Balance: ${debt_schedule['total_debt_balance'][-1]:.0f}M")

        return debt_schedule

    def _create_cash_flow_waterfall(self, assumptions, operating_forecast, debt_schedule):
        """Create cash flow waterfall showing how FCF is used"""

        ufcf = operating_forecast['ufcf'][1:]  # Exclude entry year
        interest_expense = debt_schedule['total_interest']
        tax_rate = assumptions['tax_rate']

        # Calculate levered cash flows
        ebitda = operating_forecast['ebitda'][1:]
        da = operating_forecast['da'][1:]
        capex = operating_forecast['capex'][1:]
        nwc_change = operating_forecast['nwc_change'][1:]

        # EBIT = EBITDA - D&A
        ebit = [ebitda[i] - da[i] for i in range(len(ebitda))]

        # EBT = EBIT - Interest
        ebt = [ebit[i] - interest_expense[i] for i in range(len(ebit))]

        # Taxes = EBT * tax_rate
        taxes = [ebt[i] * tax_rate for i in range(len(ebt))]

        # NOPAT = EBT - Taxes
        nopat = [ebt[i] - taxes[i] for i in range(len(ebt))]

        # Operating Cash Flow = NOPAT + D&A
        operating_cf = [nopat[i] + da[i] for i in range(len(nopat))]

        # Free Cash Flow = Operating CF - CapEx - ΔNWC
        free_cf = [operating_cf[i] - capex[i] - nwc_change[i] for i in range(len(operating_cf))]

        # Cash Available for Debt Paydown
        cash_available = free_cf.copy()

        # Debt Paydown (from debt schedule)
        debt_paydown = debt_schedule['total_debt_paydown']

        # Excess Cash Flow (after debt paydown)
        excess_cf = [cash_available[i] - debt_paydown[i] for i in range(len(cash_available))]

        cash_flow_waterfall = {
            'years': operating_forecast['years'][1:],  # Exclude entry year
            'ebitda': ebitda,
            'da': da,
            'ebit': ebit,
            'interest_expense': interest_expense,
            'ebt': ebt,
            'taxes': taxes,
            'nopat': nopat,
            'operating_cf': operating_cf,
            'capex': capex,
            'nwc_change': nwc_change,
            'free_cf': free_cf,
            'cash_available': cash_available,
            'debt_paydown': debt_paydown,
            'excess_cf': excess_cf
        }

        print("💧 Cash Flow Waterfall:")
        print(f"   • Average Annual FCF: ${np.mean(free_cf):.0f}M")
        print(f"   • Average Annual Debt Paydown: ${np.mean(debt_paydown):.0f}M")
        print(f"   • Average Annual Excess Cash: ${np.mean(excess_cf):.0f}M")

        return cash_flow_waterfall

    def _create_exit_analysis(self, assumptions, operating_forecast, debt_schedule):
        """Create exit analysis with multiple scenarios"""

        exit_year = assumptions['forecast_years']

        # Exit EBITDA (from operating forecast)
        exit_ebitda_base = operating_forecast['ebitda'][-1]  # Last year EBITDA
        exit_ebitda_bull = exit_ebitda_base * 1.15  # 15% upside
        exit_ebitda_bear = exit_ebitda_base * 0.85  # 15% downside

        # Exit Enterprise Value
        exit_ev_base = exit_ebitda_base * assumptions['exit_multiple_base']
        exit_ev_bull = exit_ebitda_bull * assumptions['exit_multiple_bull']
        exit_ev_bear = exit_ebitda_bear * assumptions['exit_multiple_bear']

        # Net Debt at Exit
        exit_debt_base = debt_schedule['total_debt_balance'][-1]
        exit_debt_bull = exit_debt_base * 0.8  # Assume some debt paydown
        exit_debt_bear = exit_debt_base * 1.1  # Assume less paydown

        # Exit Equity Value
        exit_equity_base = exit_ev_base - exit_debt_base
        exit_equity_bull = exit_ev_bull - exit_debt_bull
        exit_equity_bear = exit_ev_bear - exit_debt_bear

        exit_analysis = {
            'exit_year': exit_year,
            'exit_ebitda_base': exit_ebitda_base,
            'exit_ebitda_bull': exit_ebitda_bull,
            'exit_ebitda_bear': exit_ebitda_bear,
            'exit_multiple_base': assumptions['exit_multiple_base'],
            'exit_multiple_bull': assumptions['exit_multiple_bull'],
            'exit_multiple_bear': assumptions['exit_multiple_bear'],
            'exit_ev_base': exit_ev_base,
            'exit_ev_bull': exit_ev_bull,
            'exit_ev_bear': exit_ev_bear,
            'exit_debt_base': exit_debt_base,
            'exit_debt_bull': exit_debt_bull,
            'exit_debt_bear': exit_debt_bear,
            'exit_equity_base': exit_equity_base,
            'exit_equity_bull': exit_equity_bull,
            'exit_equity_bear': exit_equity_bear
        }

        print("🚪 Exit Analysis:")
        print(f"   • Base Case: ${exit_ev_base:.0f}M EV | ${exit_equity_base:.0f}M Equity")
        print(f"   • Bull Case: ${exit_ev_bull:.0f}M EV | ${exit_equity_bull:.0f}M Equity")
        print(f"   • Bear Case: ${exit_ev_bear:.0f}M EV | ${exit_equity_bear:.0f}M Equity")

        return exit_analysis

    def _calculate_returns(self, assumptions, exit_analysis):
        """Calculate IRR and MOIC for all scenarios"""

        equity_investment = assumptions['equity_investment']
        forecast_years = assumptions['forecast_years']

        # Calculate IRR using cash flows
        def calculate_irr(exit_value, years):
            # Create cash flow stream: -investment + annual cash flows + exit value
            cash_flows = [-equity_investment]  # Initial investment

            # Add annual excess cash flows (distributions to equity)
            annual_cf = 0  # Assuming no annual distributions for conservative case

            for year in range(1, years):
                cash_flows.append(annual_cf)

            # Add exit value
            cash_flows.append(exit_value)

            # Calculate IRR
            try:
                # Use numpy's irr function
                irr = np.irr(cash_flows)
                return max(irr, -0.99)  # Cap at -99%
            except (Exception,):
                # Fallback calculation
                if exit_value > equity_investment:
                    return (exit_value / equity_investment) ** (1/years) - 1
                else:
                    return -0.5

        # Base Case
        irr_base = calculate_irr(exit_analysis['exit_equity_base'], forecast_years)
        moic_base = exit_analysis['exit_equity_base'] / equity_investment

        # Bull Case
        irr_bull = calculate_irr(exit_analysis['exit_equity_bull'], forecast_years)
        moic_bull = exit_analysis['exit_equity_bull'] / equity_investment

        # Bear Case
        irr_bear = calculate_irr(exit_analysis['exit_equity_bear'], forecast_years)
        moic_bear = exit_analysis['exit_equity_bear'] / equity_investment

        returns_analysis = {
            'equity_investment': equity_investment,
            'irr_base': irr_base,
            'irr_bull': irr_bull,
            'irr_bear': irr_bear,
            'moic_base': moic_base,
            'moic_bull': moic_bull,
            'moic_bear': moic_bear,
            'investment_period': forecast_years
        }

        print("💎 Returns Analysis:")
        print(f"   • Base Case: {irr_base:.1%} IRR | {moic_base:.1f}x MOIC")
        print(f"   • Bull Case: {irr_bull:.1%} IRR | {moic_bull:.1f}x MOIC")
        print(f"   • Bear Case: {irr_bear:.1%} IRR | {moic_bear:.1f}x MOIC")

        return returns_analysis

    def _create_sensitivity_analysis(self, assumptions):
        """Create sensitivity analysis for key variables"""

        base_entry_multiple = assumptions['entry_multiple']
        base_exit_multiple = assumptions['exit_multiple_base']
        base_leverage = (assumptions['senior_debt_pct'] + assumptions['mezzanine_pct']) / assumptions['equity_pct']

        # Sensitivity ranges
        entry_multiples = np.arange(8.0, 13.0, 0.5)
        exit_multiples = np.arange(9.0, 14.0, 0.5)
        leverage_ratios = np.arange(2.0, 6.0, 0.5)

        # IRR sensitivity to entry vs exit multiple
        irr_sensitivity = []
        for entry_mult in entry_multiples:
            row = []
            for exit_mult in exit_multiples:
                # Recalculate with new multiples
                purchase_price = assumptions['entry_ebitda'] * entry_mult
                equity_investment = purchase_price * assumptions['equity_pct']
                exit_ebitda = assumptions['entry_ebitda'] * (1 + sum(assumptions['revenue_growth'][:assumptions['forecast_years']-1])/assumptions['forecast_years']) ** (assumptions['forecast_years']-1)
                exit_value = exit_ebitda * exit_mult

                # Simple IRR calculation
                irr = (exit_value / equity_investment) ** (1/assumptions['forecast_years']) - 1
                row.append(irr)
            irr_sensitivity.append(row)

        # MOIC sensitivity to leverage
        moic_sensitivity = []
        for lev_ratio in leverage_ratios:
            row = []
            for exit_mult in exit_multiples:
                # Recalculate with new leverage
                debt_pct = lev_ratio / (1 + lev_ratio)
                equity_pct = 1 / (1 + lev_ratio)

                purchase_price = assumptions['entry_ebitda'] * base_entry_multiple
                equity_investment = purchase_price * equity_pct
                exit_ebitda = assumptions['entry_ebitda'] * (1 + sum(assumptions['revenue_growth'][:assumptions['forecast_years']-1])/assumptions['forecast_years']) ** (assumptions['forecast_years']-1)
                exit_value = exit_ebitda * exit_mult

                moic = exit_value / equity_investment
                row.append(moic)
            moic_sensitivity.append(row)

        sensitivity_analysis = {
            'entry_multiples': entry_multiples,
            'exit_multiples': exit_multiples,
            'leverage_ratios': leverage_ratios,
            'irr_sensitivity': irr_sensitivity,
            'moic_sensitivity': moic_sensitivity
        }

        print("📊 Sensitivity Analysis:")
        print("   • IRR sensitivity table created")
        print("   • MOIC sensitivity table created")
        return sensitivity_analysis

    def _create_excel_output(self, lbo_results):
        """Create professional Excel output with multiple tabs"""

        wb = Workbook()

        # Create worksheets
        ws_assumptions = wb.active
        ws_assumptions.title = "Assumptions"

        ws_sources_uses = wb.create_sheet("Sources & Uses")
        ws_forecast = wb.create_sheet("Operating Forecast")
        ws_debt = wb.create_sheet("Debt Schedule")
        ws_waterfall = wb.create_sheet("Cash Flow Waterfall")
        ws_exit = wb.create_sheet("Exit & Returns")
        ws_sensitivity = wb.create_sheet("Sensitivity")
        ws_summary = wb.create_sheet("Summary")

        # Apply styles to workbook
        for style_name, style in self.styles.items():
            if style_name not in wb.named_styles:
                wb.add_named_style(style)

        # Create each tab
        self._create_assumptions_tab(ws_assumptions, lbo_results)
        self._create_sources_uses_tab(ws_sources_uses, lbo_results)
        self._create_forecast_tab(ws_forecast, lbo_results)
        self._create_debt_tab(ws_debt, lbo_results)
        self._create_waterfall_tab(ws_waterfall, lbo_results)
        self._create_exit_tab(ws_exit, lbo_results)
        self._create_sensitivity_tab(ws_sensitivity, lbo_results)
        self._create_summary_tab(ws_summary, lbo_results)

        # Save workbook
        filename = f"LBO_Model_{self.ticker}_{self.model_date}.xlsx"
        wb.save(filename)

        return filename

    def _create_assumptions_tab(self, ws, lbo_results):
        """Create Assumptions tab"""

        assumptions = lbo_results['assumptions']

        # Title
        ws['A1'] = f"{self.company_name} ({self.ticker}) - LBO Model Assumptions"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:G1')

        current_row = 3

        # Transaction Assumptions
        ws[f'A{current_row}'] = "TRANSACTION ASSUMPTIONS"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        transaction_data = [
            ("Entry EBITDA ($M)", assumptions['entry_ebitda']),
            ("Entry Multiple (x)", assumptions['entry_multiple']),
            ("Purchase Price ($M)", assumptions['purchase_price']),
            ("Exit Multiple - Base (x)", assumptions['exit_multiple_base']),
            ("Exit Multiple - Bull (x)", assumptions['exit_multiple_bull']),
            ("Exit Multiple - Bear (x)", assumptions['exit_multiple_bear']),
        ]

        for label, value in transaction_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].style = 'input'
            current_row += 1

        current_row += 2

        # Financing Structure
        ws[f'A{current_row}'] = "FINANCING STRUCTURE"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        financing_data = [
            ("Senior Debt (% of Purchase Price)", f"{assumptions['senior_debt_pct']:.1%}"),
            ("Mezzanine Debt (% of Purchase Price)", f"{assumptions['mezzanine_pct']:.1%}"),
            ("Equity (% of Purchase Price)", f"{assumptions['equity_pct']:.1%}"),
            ("Transaction Fees (% of Purchase Price)", f"{assumptions['fees_pct']:.1%}"),
            ("Senior Debt ($M)", assumptions['senior_debt']),
            ("Mezzanine Debt ($M)", assumptions['mezzanine_debt']),
            ("Equity Investment ($M)", assumptions['equity_investment']),
            ("Transaction Fees ($M)", assumptions['transaction_fees']),
            ("Total Debt ($M)", assumptions['total_debt']),
            ("Debt/Equity Ratio", f"{assumptions['total_debt']/assumptions['equity_investment']:.1f}x"),
        ]

        for label, value in financing_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].style = 'input'
            current_row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20

    def _create_sources_uses_tab(self, ws, lbo_results):
        """Create Sources & Uses tab with professional formatting"""

        sources_uses = lbo_results['sources_uses']

        # Title
        ws['A1'] = f"{self.company_name} - Sources & Uses"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:D1')

        # Sources section
        ws['A3'] = "SOURCES"
        ws['A3'].style = 'header'
        ws.merge_cells('A3:B3')

        current_row = 4
        for source, amount in sources_uses['sources'].items():
            ws[f'A{current_row}'] = source
            ws[f'A{current_row}'].style = 'label'
            ws[f'B{current_row}'] = amount
            ws[f'B{current_row}'].style = 'input'
            ws[f'B{current_row}'].number_format = '#,##0'
            current_row += 1

        # Add total with underline
        ws[f'A{current_row}'] = "Total Sources"
        ws[f'A{current_row}'].style = 'label_bold'
        ws[f'B{current_row}'] = sources_uses['sources']['Total Sources']
        ws[f'B{current_row}'].style = 'result'
        ws[f'B{current_row}'].number_format = '#,##0'

        # Add border
        from openpyxl.styles import Border, Side
        thin_border = Border(bottom=Side(style='thin'))
        ws[f'A{current_row}'].border = thin_border
        ws[f'B{current_row}'].border = thin_border

        # Uses section
        ws['D3'] = "USES"
        ws['D3'].style = 'header'
        ws.merge_cells('D3:E3')

        current_row = 4
        for use, amount in sources_uses['uses'].items():
            ws[f'D{current_row}'] = use
            ws[f'D{current_row}'].style = 'label'
            ws[f'E{current_row}'] = amount
            ws[f'E{current_row}'].style = 'input'
            ws[f'E{current_row}'].number_format = '#,##0'
            current_row += 1

        # Add total with underline
        ws[f'D{current_row}'] = "Total Uses"
        ws[f'D{current_row}'].style = 'label_bold'
        ws[f'E{current_row}'] = sources_uses['uses']['Total Uses']
        ws[f'E{current_row}'].style = 'result'
        ws[f'E{current_row}'].number_format = '#,##0'

        # Add border
        ws[f'D{current_row}'].border = thin_border
        ws[f'E{current_row}'].border = thin_border

        # Balance check
        ws['A8'] = "Balance Check:"
        ws['A8'].style = 'label_bold'
        ws['B8'] = "Sources = Uses" if sources_uses['check'] else "MISMATCH"
        ws['B8'].style = 'result' if sources_uses['check'] else 'warning'

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15

    def _create_forecast_tab(self, ws, lbo_results):
        """Create Operating Forecast tab"""

        forecast = lbo_results['operating_forecast']

        # Title
        ws['A1'] = f"{self.company_name} - Operating Forecast"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:I1')

        # Headers
        headers = ['Year', 'Revenue ($M)', 'EBITDA ($M)', 'D&A ($M)', 'EBIT ($M)',
                  'CapEx ($M)', 'ΔNWC ($M)', 'NOPAT ($M)', 'UFCF ($M)']

        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header).style = 'header'

        # Data
        for i, year in enumerate(forecast['years']):
            row = i + 4
            ws.cell(row=row, column=1, value=year).style = 'label'

            values = [
                forecast['revenue'][i],
                forecast['ebitda'][i],
                forecast['da'][i],
                forecast['ebit'][i],
                forecast['capex'][i],
                forecast['nwc_change'][i],
                forecast['nopat'][i],
                forecast['ufcf'][i]
            ]

            for col, value in enumerate(values, 2):
                cell = ws.cell(row=row, column=col, value=value)
                cell.style = 'calculation'
                cell.number_format = '#,##0'

        # Set column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_debt_tab(self, ws, lbo_results):
        """Create Debt Schedule tab"""

        debt = lbo_results['debt_schedule']

        # Title
        ws['A1'] = f"{self.company_name} - Debt Schedule"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:L1')

        # Headers
        headers = ['Year', 'Senior Balance', 'Senior Interest', 'Senior Mandatory', 'Senior Optional',
                  'Senior Paydown', 'Mezz Balance', 'Mezz Interest', 'Mezz Mandatory',
                  'Mezz Optional', 'Mezz Paydown', 'Total Debt']

        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header).style = 'header'

        # Data
        for i, year in enumerate(debt['years']):
            row = i + 4
            ws.cell(row=row, column=1, value=year).style = 'label'

            values = [
                debt['senior_balance'][i],
                debt['senior_interest'][i],
                debt['senior_mandatory'][i],
                debt['senior_optional'][i],
                debt['senior_total_paydown'][i],
                debt['mezz_balance'][i],
                debt['mezz_interest'][i],
                debt['mezz_mandatory'][i],
                debt['mezz_optional'][i],
                debt['mezz_total_paydown'][i],
                debt['total_debt_balance'][i]
            ]

            for col, value in enumerate(values, 2):
                cell = ws.cell(row=row, column=col, value=value)
                cell.style = 'calculation'
                cell.number_format = '#,##0'

        # Set column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_waterfall_tab(self, ws, lbo_results):
        """Create Cash Flow Waterfall tab"""

        waterfall = lbo_results['cash_flow_waterfall']

        # Title
        ws['A1'] = f"{self.company_name} - Cash Flow Waterfall"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:K1')

        # Headers
        headers = ['Year', 'EBITDA', 'D&A', 'EBIT', 'Interest', 'EBT', 'Taxes',
                  'NOPAT', 'Operating CF', 'CapEx', 'ΔNWC', 'Free CF',
                  'Cash Available', 'Debt Paydown', 'Excess CF']

        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header).style = 'header'

        # Data
        for i, year in enumerate(waterfall['years']):
            row = i + 4
            ws.cell(row=row, column=1, value=year).style = 'label'

            values = [
                waterfall['ebitda'][i],
                waterfall['da'][i],
                waterfall['ebit'][i],
                waterfall['interest_expense'][i],
                waterfall['ebt'][i],
                waterfall['taxes'][i],
                waterfall['nopat'][i],
                waterfall['operating_cf'][i],
                waterfall['capex'][i],
                waterfall['nwc_change'][i],
                waterfall['free_cf'][i],
                waterfall['cash_available'][i],
                waterfall['debt_paydown'][i],
                waterfall['excess_cf'][i]
            ]

            for col, value in enumerate(values, 2):
                cell = ws.cell(row=row, column=col, value=value)
                cell.style = 'calculation'
                cell.number_format = '#,##0'

        # Set column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12

    def _create_exit_tab(self, ws, lbo_results):
        """Create Exit & Returns tab"""

        exit_analysis = lbo_results['exit_analysis']
        returns = lbo_results['returns_analysis']

        # Title
        ws['A1'] = f"{self.company_name} - Exit & Returns Analysis"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:F1')

        current_row = 3

        # Exit Analysis
        ws[f'A{current_row}'] = "EXIT ANALYSIS"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:F{current_row}')
        current_row += 2

        scenarios = ['Base Case', 'Bull Case', 'Bear Case']
        for i, scenario in enumerate(scenarios):
            ws[f'A{current_row}'] = scenario
            ws[f'A{current_row}'].style = 'label_bold'
            current_row += 1

            scenario_key = scenario.lower().split()[0]
            exit_data = [
                ("Exit EBITDA ($M)", exit_analysis[f'exit_ebitda_{scenario_key}']),
                ("Exit Multiple (x)", exit_analysis[f'exit_multiple_{scenario_key}']),
                ("Exit EV ($M)", exit_analysis[f'exit_ev_{scenario_key}']),
                ("Less: Net Debt ($M)", exit_analysis[f'exit_debt_{scenario_key}']),
                ("Exit Equity Value ($M)", exit_analysis[f'exit_equity_{scenario_key}'])
            ]

            for label, value in exit_data:
                ws[f'A{current_row}'] = f"  {label}"
                ws[f'A{current_row}'].style = 'label'
                ws[f'B{current_row}'] = value
                ws[f'B{current_row}'].style = 'calculation'
                ws[f'B{current_row}'].number_format = '#,##0'
                current_row += 1

            current_row += 1

        # Returns Analysis
        current_row += 2
        ws[f'A{current_row}'] = "RETURNS ANALYSIS"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:F{current_row}')
        current_row += 2

        returns_data = [
            ("Initial Equity Investment ($M)", returns['equity_investment']),
            ("Investment Period (Years)", returns['investment_period']),
            ("", ""),
            ("BASE CASE RETURNS", ""),
            ("IRR", f"{returns['irr_base']:.1%}"),
            ("MOIC", f"{returns['moic_base']:.1f}x"),
            ("", ""),
            ("BULL CASE RETURNS", ""),
            ("IRR", f"{returns['irr_bull']:.1%}"),
            ("MOIC", f"{returns['moic_bull']:.1f}x"),
            ("", ""),
            ("BEAR CASE RETURNS", ""),
            ("IRR", f"{returns['irr_bear']:.1%}"),
            ("MOIC", f"{returns['moic_bear']:.1f}x")
        ]

        for label, value in returns_data:
            if label == "":
                current_row += 1
                continue
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold' if 'CASE' in label or label in ['IRR', 'MOIC'] else 'label'
            if value != "":
                ws[f'B{current_row}'] = value
                ws[f'B{current_row}'].style = 'result'
                if '%' in str(value):
                    ws[f'B{current_row}'].number_format = '0.0%'
                elif 'x' in str(value):
                    ws[f'B{current_row}'].number_format = '0.0"x"'
                else:
                    ws[f'B{current_row}'].number_format = '#,##0'
            current_row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

    def _create_sensitivity_tab(self, ws, lbo_results):
        """Create Sensitivity Analysis tab"""

        sensitivity = lbo_results['sensitivity_analysis']

        # Title
        ws['A1'] = f"{self.company_name} - Sensitivity Analysis"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:K1')

        # IRR Sensitivity Table
        ws['A3'] = "IRR SENSITIVITY: Entry Multiple vs Exit Multiple"
        ws['A3'].style = 'header'
        ws.merge_cells('A3:K3')

        # IRR table headers
        ws['A5'] = "Entry\\Exit"
        ws['A5'].style = 'label_bold'

        for i, exit_mult in enumerate(sensitivity['exit_multiples']):
            ws.cell(row=5, column=i+2, value=f"{exit_mult}x").style = 'header'

        for i, entry_mult in enumerate(sensitivity['entry_multiples']):
            ws.cell(row=i+6, column=1, value=f"{entry_mult}x").style = 'label_bold'

            for j, irr_value in enumerate(sensitivity['irr_sensitivity'][i]):
                cell = ws.cell(row=i+6, column=j+2, value=irr_value)
                cell.style = 'calculation'
                cell.number_format = '0.0%'

        # MOIC Sensitivity Table
        current_row = len(sensitivity['entry_multiples']) + 10

        ws[f'A{current_row}'] = "MOIC SENSITIVITY: Leverage Ratio vs Exit Multiple"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:K{current_row}')

        # MOIC table headers
        ws[f'A{current_row+2}'] = "Leverage\\Exit"
        ws[f'A{current_row+2}'].style = 'label_bold'

        for i, exit_mult in enumerate(sensitivity['exit_multiples']):
            ws.cell(row=current_row+2, column=i+2, value=f"{exit_mult}x").style = 'header'

        for i, lev_ratio in enumerate(sensitivity['leverage_ratios']):
            ws.cell(row=current_row+3+i, column=1, value=f"{lev_ratio}x").style = 'label_bold'

            for j, moic_value in enumerate(sensitivity['moic_sensitivity'][i]):
                cell = ws.cell(row=current_row+3+i, column=j+2, value=moic_value)
                cell.style = 'calculation'
                cell.number_format = '0.0"x"'

        # Set column widths
        for col in range(1, len(sensitivity['exit_multiples']) + 3):
            ws.column_dimensions[get_column_letter(col)].width = 12

    def _create_summary_tab(self, ws, lbo_results):
        """Create Summary tab with key metrics"""

        assumptions = lbo_results['assumptions']
        returns = lbo_results['returns_analysis']
        sources_uses = lbo_results['sources_uses']

        # Title
        ws['A1'] = f"{self.company_name} ({self.ticker}) - LBO Investment Summary"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:E1')

        current_row = 3

        # Transaction Overview
        ws[f'A{current_row}'] = "TRANSACTION OVERVIEW"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 2

        transaction_summary = [
            ("Purchase Price", f"${assumptions['purchase_price']/1000:.1f}B"),
            ("Entry EBITDA", f"${assumptions['entry_ebitda']:.0f}M"),
            ("Entry Multiple", f"{assumptions['entry_multiple']:.1f}x"),
            ("Total Debt", f"${assumptions['total_debt']/1000:.1f}B"),
            ("Equity Investment", f"${assumptions['equity_investment']/1000:.1f}B"),
            ("Debt/Equity Ratio", f"{assumptions['total_debt']/assumptions['equity_investment']:.1f}x")
        ]

        for label, value in transaction_summary:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].style = 'input'
            current_row += 1

        current_row += 2

        # Sources & Uses Summary
        ws[f'A{current_row}'] = "SOURCES & USES SUMMARY"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 2

        sources_summary = [
            ("Senior Debt", f"${assumptions['senior_debt']/1000:.1f}B"),
            ("Mezzanine Debt", f"${assumptions['mezzanine_debt']/1000:.1f}B"),
            ("Equity", f"${assumptions['equity_investment']/1000:.1f}B"),
            ("Total Sources", f"${sources_uses['sources']['Total Sources']/1000:.1f}B"),
            ("", ""),
            ("Purchase Price", f"${assumptions['purchase_price']/1000:.1f}B"),
            ("Fees", f"${assumptions['transaction_fees']/1000:.1f}B"),
            ("Total Uses", f"${sources_uses['uses']['Total Uses']/1000:.1f}B")
        ]

        for label, value in sources_summary:
            if label == "":
                current_row += 1
                continue
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label'
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].style = 'calculation'
            current_row += 1

        current_row += 2

        # Returns Summary
        ws[f'A{current_row}'] = "INVESTMENT RETURNS"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 2

        returns_summary = [
            ("Investment Period", f"{returns['investment_period']} years"),
            ("", ""),
            ("BASE CASE", ""),
            ("IRR", f"{returns['irr_base']:.1%}"),
            ("MOIC", f"{returns['moic_base']:.1f}x"),
            ("", ""),
            ("BULL CASE", ""),
            ("IRR", f"{returns['irr_bull']:.1%}"),
            ("MOIC", f"{returns['moic_bull']:.1f}x"),
            ("", ""),
            ("BEAR CASE", ""),
            ("IRR", f"{returns['irr_bear']:.1%}"),
            ("MOIC", f"{returns['moic_bear']:.1f}x")
        ]

        for label, value in returns_summary:
            if label == "":
                current_row += 1
                continue
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold' if 'CASE' in label or label in ['IRR', 'MOIC'] else 'label'
            if value != "":
                ws[f'B{current_row}'] = value
                ws[f'B{current_row}'].style = 'result'
            current_row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20


def run_sample_lbo():
    """Run a sample LBO model with realistic assumptions"""

    print("🏦 Running Professional LBO Model Sample")
    print("=" * 60)

    # Create LBO model instance
    lbo_model = ProfessionalLBOModel("TechCorp Inc.", "TECH")

    # Run the model with sample assumptions
    lbo_results, excel_file = lbo_model.run_lbo_model(
        # Transaction Assumptions
        entry_ebitda=300.0,  # $300M EBITDA
        entry_multiple=12.0,  # 12x EBITDA
        exit_multiple_base=13.0,
        exit_multiple_bull=14.5,
        exit_multiple_bear=11.5,

        # Financing Structure
        senior_debt_pct=0.55,
        mezzanine_pct=0.10,
        equity_pct=0.35,
        fees_pct=0.015,

        # Senior Debt Terms
        senior_rate=0.06,
        senior_term=7,
        senior_amort_pct=0.08,

        # Mezzanine Debt Terms
        mezz_rate=0.11,
        mezz_term=8,
        mezz_amort_pct=0.05,

        # Operating Assumptions
        revenue_growth=[0.12, 0.10, 0.08, 0.06, 0.04, 0.03],
        ebitda_margin=0.32,
        capex_pct=0.055,
        nwc_pct=0.025,
        tax_rate=0.24,
        da_pct=0.075,

        # Optional Prepayments
        optional_prepay_senior=10.0,
        optional_prepay_mezz=5.0,

        forecast_years=6
    )

    return lbo_results, excel_file


if __name__ == "__main__":
    # Run sample LBO model
    results, excel_file = run_sample_lbo()

    print("\n📋 Model Complete!")
    print(f"📁 Excel file saved as: {excel_file}")

    # Display key results
    print("\n🔑 Key Results Summary:")
    print("=" * 40)
    assumptions = results['assumptions']
    returns = results['returns_analysis']

    print(f"Purchase Price: ${assumptions['purchase_price']/1000:.1f}B")
    print(f"Total Financing: ${assumptions['total_debt']/1000:.1f}B Debt + ${assumptions['equity_investment']/1000:.1f}B Equity")
    print(f"Debt/Equity Ratio: {assumptions['total_debt']/assumptions['equity_investment']:.1f}x")
    print()
    print("Base Case Returns:")
    print(f"  IRR: {returns['irr_base']:.1%}")
    print(f"  MOIC: {returns['moic_base']:.1f}x")
    print()
    print("Bull Case Returns:")
    print(f"  IRR: {returns['irr_bull']:.1%}")
    print(f"  MOIC: {returns['moic_bull']:.1f}x")
    print()
    print("Bear Case Returns:")
    print(f"  IRR: {returns['irr_bear']:.1%}")
    print(f"  MOIC: {returns['moic_bear']:.1f}x")
