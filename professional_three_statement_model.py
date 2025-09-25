#!/usr/bin/env python3
"""
Professional Three-Statement Financial Model Builder
Creates fully linked Income Statement, Balance Sheet, and Cash Flow Statement with professional formatting

Author: Investment Banking Modeler
Date: 2024

Features:
- Scenario Analysis (Base, Bull, Bear cases)
- Fully Linked Financial Statements
- Balance Sheet Balancing Checks
- Professional Excel Output with Multiple Tabs
- Investment Banker-Style Formatting
"""

import numpy as np
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# Professional color scheme
FINANCIAL_MODEL_COLORS = {
    'header_blue': '1F4E79',
    'input_light_blue': 'DDEBF7',
    'calculation_yellow': 'FFF2CC',
    'result_green': 'D8E4BC',
    'warning_red': 'F8CECC',
    'border_dark': '2C3E50',
    'text_dark': '000000',
    'text_white': 'FFFFFF',
    'scenario_blue': 'B8CCE4',
    'scenario_green': 'C4D79B',
    'scenario_red': 'F2DCDB'
}

class ProfessionalThreeStatementModel:
    """
    Comprehensive Three-Statement Financial Model with Professional Formatting
    """

    def __init__(self, company_name="Sample Company", ticker="TICK"):
        self.company_name = company_name
        self.ticker = ticker
        self.model_date = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Initialize styles
        self.styles = self._create_styles()

    def _create_styles(self):
        """Create professional Excel styles"""
        styles = {}

        # Header style
        styles['header'] = NamedStyle(name='header')
        styles['header'].font = Font(name='Calibri', size=12, bold=True, color=FINANCIAL_MODEL_COLORS['text_white'])
        styles['header'].fill = PatternFill(start_color=FINANCIAL_MODEL_COLORS['header_blue'], end_color=FINANCIAL_MODEL_COLORS['header_blue'], fill_type='solid')
        styles['header'].alignment = Alignment(horizontal='center', vertical='center')

        # Input style
        styles['input'] = NamedStyle(name='input')
        styles['input'].font = Font(name='Calibri', size=10, color=FINANCIAL_MODEL_COLORS['text_dark'])
        styles['input'].fill = PatternFill(start_color=FINANCIAL_MODEL_COLORS['input_light_blue'], end_color=FINANCIAL_MODEL_COLORS['input_light_blue'], fill_type='solid')
        styles['input'].alignment = Alignment(horizontal='right', vertical='center')
        styles['input'].number_format = '#,##0'

        # Calculation style
        styles['calculation'] = NamedStyle(name='calculation')
        styles['calculation'].font = Font(name='Calibri', size=10, color=FINANCIAL_MODEL_COLORS['text_dark'])
        styles['calculation'].fill = PatternFill(start_color=FINANCIAL_MODEL_COLORS['calculation_yellow'], end_color=FINANCIAL_MODEL_COLORS['calculation_yellow'], fill_type='solid')
        styles['calculation'].alignment = Alignment(horizontal='right', vertical='center')
        styles['calculation'].number_format = '#,##0'

        # Result style
        styles['result'] = NamedStyle(name='result')
        styles['result'].font = Font(name='Calibri', size=10, bold=True, color=FINANCIAL_MODEL_COLORS['text_dark'])
        styles['result'].fill = PatternFill(start_color=FINANCIAL_MODEL_COLORS['result_green'], end_color=FINANCIAL_MODEL_COLORS['result_green'], fill_type='solid')
        styles['result'].alignment = Alignment(horizontal='right', vertical='center')
        styles['result'].number_format = '#,##0'

        # Warning style
        styles['warning'] = NamedStyle(name='warning')
        styles['warning'].font = Font(name='Calibri', size=10, color=FINANCIAL_MODEL_COLORS['text_dark'])
        styles['warning'].fill = PatternFill(start_color=FINANCIAL_MODEL_COLORS['warning_red'], end_color=FINANCIAL_MODEL_COLORS['warning_red'], fill_type='solid')
        styles['warning'].alignment = Alignment(horizontal='center', vertical='center')

        # Scenario styles
        styles['scenario_base'] = NamedStyle(name='scenario_base')
        styles['scenario_base'].font = Font(name='Calibri', size=10, color=FINANCIAL_MODEL_COLORS['text_dark'])
        styles['scenario_base'].fill = PatternFill(start_color=FINANCIAL_MODEL_COLORS['scenario_blue'], end_color=FINANCIAL_MODEL_COLORS['scenario_blue'], fill_type='solid')

        styles['scenario_bull'] = NamedStyle(name='scenario_bull')
        styles['scenario_bull'].font = Font(name='Calibri', size=10, color=FINANCIAL_MODEL_COLORS['text_dark'])
        styles['scenario_bull'].fill = PatternFill(start_color=FINANCIAL_MODEL_COLORS['scenario_green'], end_color=FINANCIAL_MODEL_COLORS['scenario_green'], fill_type='solid')

        styles['scenario_bear'] = NamedStyle(name='scenario_bear')
        styles['scenario_bear'].font = Font(name='Calibri', size=10, color=FINANCIAL_MODEL_COLORS['text_dark'])
        styles['scenario_bear'].fill = PatternFill(start_color=FINANCIAL_MODEL_COLORS['scenario_red'], end_color=FINANCIAL_MODEL_COLORS['scenario_red'], fill_type='solid')

        # Bold label style
        styles['label_bold'] = NamedStyle(name='label_bold')
        styles['label_bold'].font = Font(name='Calibri', size=10, bold=True, color=FINANCIAL_MODEL_COLORS['text_dark'])
        styles['label_bold'].alignment = Alignment(horizontal='left', vertical='center')

        # Normal label style
        styles['label'] = NamedStyle(name='label')
        styles['label'].font = Font(name='Calibri', size=10, color=FINANCIAL_MODEL_COLORS['text_dark'])
        styles['label'].alignment = Alignment(horizontal='left', vertical='center')

        # Subtotal style
        styles['subtotal'] = NamedStyle(name='subtotal')
        styles['subtotal'].font = Font(name='Calibri', size=10, bold=True, color=FINANCIAL_MODEL_COLORS['text_dark'])
        styles['subtotal'].fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')  # Light grey
        styles['subtotal'].alignment = Alignment(horizontal='right', vertical='center')
        styles['subtotal'].number_format = '#,##0'

        return styles

    def run_three_statement_model(self,
                                 # Company Basics
                                 starting_revenue=1000.0,  # $M

                                 # Revenue Growth Scenarios (%)
                                 growth_base=[0.08, 0.07, 0.06, 0.05, 0.04],  # Year 1-5
                                 growth_bull=[0.12, 0.10, 0.08, 0.06, 0.05],
                                 growth_bear=[0.04, 0.03, 0.02, 0.02, 0.01],

                                 # Operating Margins (%)
                                 gross_margin=0.65,  # Revenue - COGS
                                 ebitda_margin=0.25,  # EBITDA / Revenue
                                 ebit_margin=0.20,    # EBIT / Revenue

                                 # Depreciation & CapEx (% of revenue)
                                 depreciation_pct=0.05,
                                 capex_pct=0.04,

                                 # Working Capital (% of revenue)
                                 ar_pct=0.15,         # Accounts Receivable
                                 inventory_pct=0.10,  # Inventory
                                 ap_pct=0.08,         # Accounts Payable
                                 other_current_liab_pct=0.05,

                                 # Debt Schedule
                                 opening_debt=300.0,  # $M
                                 interest_rate=0.06,  # 6%
                                 annual_amortization=50.0,  # $M per year

                                 # Other Assumptions
                                 tax_rate=0.25,
                                 dividend_payout=0.30,  # % of net income
                                 starting_cash=100.0,   # $M
                                 forecast_years=5):

        """
        Run complete three-statement financial model
        """

        print(f"🏢 Building Professional Three-Statement Model for {self.company_name} ({self.ticker})")
        print("=" * 80)

        # Step 1: Create Assumptions & Drivers
        assumptions = self._create_assumptions(
            starting_revenue, growth_base, growth_bull, growth_bear,
            gross_margin, ebitda_margin, ebit_margin, depreciation_pct, capex_pct,
            ar_pct, inventory_pct, ap_pct, other_current_liab_pct,
            opening_debt, interest_rate, annual_amortization,
            tax_rate, dividend_payout, starting_cash, forecast_years
        )

        # Step 2: Generate Income Statements for all scenarios
        income_statements = self._create_income_statements(assumptions)

        # Step 3: Generate Balance Sheets for all scenarios
        balance_sheets = self._create_balance_sheets(assumptions, income_statements)

        # Step 4: Generate Cash Flow Statements for all scenarios
        cash_flows = self._create_cash_flow_statements(assumptions, income_statements, balance_sheets)

        # Step 5: Perform Balance Checks
        balance_checks = self._perform_balance_checks(balance_sheets)

        # Step 6: Create Supporting Schedules
        supporting_schedules = self._create_supporting_schedules(assumptions, income_statements, balance_sheets, cash_flows)

        # Compile results
        three_statement_results = {
            'assumptions': assumptions,
            'income_statements': income_statements,
            'balance_sheets': balance_sheets,
            'cash_flows': cash_flows,
            'balance_checks': balance_checks,
            'supporting_schedules': supporting_schedules
        }

        # Create Excel output
        excel_file = self._create_excel_output(three_statement_results)

        print("\n✅ Three-Statement Model Complete!")
        print("📊 Key Metrics (Base Case):")
        print(f"   • Year 5 Revenue: ${income_statements['base']['revenue'][-1]/1000:.1f}B")
        print(f"   • Year 5 EBITDA: ${income_statements['base']['ebitda'][-1]:.0f}M")
        print(f"   • Year 5 Net Income: ${income_statements['base']['net_income'][-1]:.0f}M")
        print(f"   • Year 5 Cash Balance: ${balance_sheets['base']['cash'][-1]:.0f}M")
        print(f"   • Balance Sheet Check: {'✅ All Years Balanced' if all(balance_checks['base']) else '❌ Balance Issues'}")
        print(f"📁 Excel Output: {excel_file}")

        return three_statement_results, excel_file

    def _create_assumptions(self, starting_revenue, growth_base, growth_bull, growth_bear,
                           gross_margin, ebitda_margin, ebit_margin, depreciation_pct, capex_pct,
                           ar_pct, inventory_pct, ap_pct, other_current_liab_pct,
                           opening_debt, interest_rate, annual_amortization,
                           tax_rate, dividend_payout, starting_cash, forecast_years):

        """Create comprehensive assumptions for all scenarios"""

        assumptions = {
            # Company Basics
            'starting_revenue': starting_revenue,
            'forecast_years': forecast_years,
            'years': list(range(2025, 2025 + forecast_years)),

            # Growth Scenarios (%)
            'growth_base': growth_base,
            'growth_bull': growth_bull,
            'growth_bear': growth_bear,

            # Operating Margins (%)
            'gross_margin': gross_margin,
            'ebitda_margin': ebitda_margin,
            'ebit_margin': ebit_margin,

            # Depreciation & CapEx (% of revenue)
            'depreciation_pct': depreciation_pct,
            'capex_pct': capex_pct,

            # Working Capital (% of revenue)
            'ar_pct': ar_pct,                           # Accounts Receivable
            'inventory_pct': inventory_pct,             # Inventory
            'ap_pct': ap_pct,                          # Accounts Payable
            'other_current_liab_pct': other_current_liab_pct,

            # Debt Schedule
            'opening_debt': opening_debt,
            'interest_rate': interest_rate,
            'annual_amortization': annual_amortization,

            # Other Assumptions
            'tax_rate': tax_rate,
            'dividend_payout': dividend_payout,
            'starting_cash': starting_cash
        }

        print("📋 Assumptions Created:")
        print(f"   • Starting Revenue: ${starting_revenue:.0f}M")
        print(f"   • Forecast Period: {forecast_years} years")
        print(f"   • Base Case Growth: {growth_base[0]*100:.1f}% → {growth_base[-1]*100:.1f}%")
        print(f"   • EBITDA Margin: {ebitda_margin*100:.1f}%")
        print(f"   • Debt: ${opening_debt:.0f}M at {interest_rate*100:.1f}%")

        return assumptions

    def _create_income_statements(self, assumptions):
        """Create income statements for all scenarios"""

        income_statements = {}

        for scenario in ['base', 'bull', 'bear']:
            growth_rates = assumptions[f'growth_{scenario}']

            # Calculate revenue
            revenue = [assumptions['starting_revenue']]
            for i in range(assumptions['forecast_years'] - 1):
                if i < len(growth_rates):
                    growth = growth_rates[i]
                else:
                    growth = growth_rates[-1]  # Use terminal growth
                revenue.append(revenue[-1] * (1 + growth))

            # Cost of Goods Sold
            cogs = [rev * (1 - assumptions['gross_margin']) for rev in revenue]
            gross_profit = [rev - cogs[i] for i, rev in enumerate(revenue)]

            # Operating Expenses (back into EBITDA and EBIT margins)
            ebitda = [rev * assumptions['ebitda_margin'] for rev in revenue]
            opex = [gross_profit[i] - ebitda[i] for i in range(len(gross_profit))]

            # Depreciation & Amortization
            depreciation = [rev * assumptions['depreciation_pct'] for rev in revenue]

            # EBIT
            ebit = [ebitda[i] - depreciation[i] for i in range(len(ebitda))]

            # Interest Expense (from debt schedule)
            debt_balance = assumptions['opening_debt']
            interest_expense = []

            for year in range(assumptions['forecast_years']):
                interest = debt_balance * assumptions['interest_rate']
                interest_expense.append(interest)
                debt_balance = max(debt_balance - assumptions['annual_amortization'], 0)

            # EBT
            ebt = [ebit[i] - interest_expense[i] for i in range(len(ebit))]

            # Taxes
            taxes = [ebt[i] * assumptions['tax_rate'] for i in range(len(ebt))]

            # Net Income
            net_income = [ebt[i] - taxes[i] for i in range(len(ebt))]

            # Dividends
            dividends = [ni * assumptions['dividend_payout'] for ni in net_income]

            income_statements[scenario] = {
                'years': assumptions['years'],
                'revenue': revenue,
                'cogs': cogs,
                'gross_profit': gross_profit,
                'opex': opex,
                'ebitda': ebitda,
                'depreciation': depreciation,
                'ebit': ebit,
                'interest_expense': interest_expense,
                'ebt': ebt,
                'taxes': taxes,
                'net_income': net_income,
                'dividends': dividends
            }

        print("💰 Income Statements Created:")
        print(f"   • Base Case Year 5 Revenue: ${income_statements['base']['revenue'][-1]/1000:.1f}B")
        print(f"   • Base Case Year 5 EBITDA: ${income_statements['base']['ebitda'][-1]:.0f}M")
        print(f"   • Base Case Year 5 Net Income: ${income_statements['base']['net_income'][-1]:.0f}M")

        return income_statements

    def _create_balance_sheets(self, assumptions, income_statements):
        """Create balance sheets for all scenarios"""

        balance_sheets = {}

        for scenario in ['base', 'bull', 'bear']:
            income = income_statements[scenario]

            # ASSETS
            # Cash (will be calculated from cash flow statement)
            cash = [assumptions['starting_cash']]

            # Accounts Receivable
            ar = [rev * assumptions['ar_pct'] for rev in income['revenue']]

            # Inventory
            inventory = [rev * assumptions['inventory_pct'] for rev in income['revenue']]

            # PP&E (accumulated CapEx - accumulated Depreciation)
            capex = [rev * assumptions['capex_pct'] for rev in income['revenue']]
            accumulated_capex = [sum(capex[:i+1]) for i in range(len(capex))]
            accumulated_depreciation = [sum(income['depreciation'][:i+1]) for i in range(len(income['depreciation']))]
            ppe = [accumulated_capex[i] - accumulated_depreciation[i] for i in range(len(accumulated_capex))]

            # Total Assets (will be updated after cash flow calculation)
            total_assets = [cash[0] + ar[0] + inventory[0] + ppe[0]]

            # LIABILITIES
            # Debt
            debt_balance = assumptions['opening_debt']
            debt = [debt_balance]
            for year in range(1, assumptions['forecast_years']):
                debt_balance = max(debt_balance - assumptions['annual_amortization'], 0)
                debt.append(debt_balance)

            # Accounts Payable
            ap = [rev * assumptions['ap_pct'] for rev in income['revenue']]

            # Other Current Liabilities
            other_current_liab = [rev * assumptions['other_current_liab_pct'] for rev in income['revenue']]

            # Total Current Liabilities
            current_liabilities = [ap[i] + other_current_liab[i] for i in range(len(ap))]

            # EQUITY
            # Opening Equity (assumed)
            opening_equity = 200.0  # $M

            # Retained Earnings
            retained_earnings = [opening_equity]
            for i in range(1, len(income['net_income'])):
                new_retained = retained_earnings[-1] + income['net_income'][i] - income['dividends'][i]
                retained_earnings.append(new_retained)

            # Total Equity
            total_equity = retained_earnings.copy()

            # Total Liabilities & Equity
            total_liabilities_equity = [debt[i] + current_liabilities[i] + total_equity[i] for i in range(len(debt))]

            balance_sheets[scenario] = {
                'years': assumptions['years'],
                # Assets
                'cash': cash,
                'ar': ar,
                'inventory': inventory,
                'ppe': ppe,
                'total_assets': total_assets,
                # Liabilities
                'debt': debt,
                'ap': ap,
                'other_current_liab': other_current_liab,
                'current_liabilities': current_liabilities,
                'total_liabilities': [debt[i] + current_liabilities[i] for i in range(len(debt))],
                # Equity
                'retained_earnings': retained_earnings,
                'total_equity': total_equity,
                'total_liabilities_equity': total_liabilities_equity
            }

        print("📊 Balance Sheets Created:")
        print(f"   • Base Case Year 5 Total Assets: ${balance_sheets['base']['total_assets'][-1]:.0f}M")

        return balance_sheets

    def _create_cash_flow_statements(self, assumptions, income_statements, balance_sheets):
        """Create cash flow statements for all scenarios"""

        cash_flows = {}

        for scenario in ['base', 'bull', 'bear']:
            income = income_statements[scenario]
            balance = balance_sheets[scenario]

            # CASH FROM OPERATIONS (CFO)
            # Net Income
            net_income = income['net_income']

            # Add back Depreciation
            depreciation = income['depreciation']

            # Changes in Working Capital
            ar_change = [0]  # First year has no prior year
            inventory_change = [0]
            ap_change = [0]
            other_liab_change = [0]

            for i in range(1, len(balance['ar'])):
                ar_change.append(balance['ar'][i-1] - balance['ar'][i])
                inventory_change.append(balance['inventory'][i-1] - balance['inventory'][i])
                ap_change.append(balance['ap'][i] - balance['ap'][i-1])  # AP increase is use of cash
                other_liab_change.append(balance['other_current_liab'][i] - balance['other_current_liab'][i-1])

            # CFO = Net Income + Depreciation - ∆NWC
            delta_nwc = [ar_change[i] + inventory_change[i] - ap_change[i] - other_liab_change[i] for i in range(len(ar_change))]
            cfo = [net_income[i] + depreciation[i] - delta_nwc[i] for i in range(len(net_income))]

            # CASH FROM INVESTING (CFI)
            capex = [rev * assumptions['capex_pct'] for rev in income['revenue']]
            cfi = [-capex[i] for i in range(len(capex))]

            # CASH FROM FINANCING (CFF)
            # Debt changes
            debt_change = [0]
            for i in range(1, len(balance['debt'])):
                debt_change.append(balance['debt'][i-1] - balance['debt'][i])  # Positive = debt raised

            # Dividends
            dividends = income['dividends']
            dividends_paid = [-dividends[i] for i in range(len(dividends))]

            cff = [debt_change[i] + dividends_paid[i] for i in range(len(debt_change))]

            # NET CHANGE IN CASH
            net_change_cash = [cfo[i] + cfi[i] + cff[i] for i in range(len(cfo))]

            # ENDING CASH BALANCE
            cash_balance = [balance['cash'][0]]  # Starting cash
            for i in range(1, len(net_change_cash) + 1):
                new_cash = cash_balance[-1] + net_change_cash[i-1]
                cash_balance.append(new_cash)

            # Update balance sheet with correct cash balances
            balance_sheets[scenario]['cash'] = cash_balance[:-1]  # Remove the last element to match forecast years
            balance_sheets[scenario]['total_assets'] = [
                balance_sheets[scenario]['cash'][i] +
                balance_sheets[scenario]['ar'][i] +
                balance_sheets[scenario]['inventory'][i] +
                balance_sheets[scenario]['ppe'][i]
                for i in range(len(balance_sheets[scenario]['cash']))
            ]

            cash_flows[scenario] = {
                'years': assumptions['years'],
                # CFO
                'net_income': net_income,
                'depreciation': depreciation,
                'delta_nwc': delta_nwc,
                'cfo': cfo,
                # CFI
                'capex': capex,
                'cfi': cfi,
                # CFF
                'debt_change': debt_change,
                'dividends_paid': dividends_paid,
                'cff': cff,
                # Summary
                'net_change_cash': net_change_cash,
                'ending_cash': cash_balance[1:]  # Remove starting cash
            }

        print("💸 Cash Flow Statements Created:")
        print(f"   • Base Case Avg. CFO: ${np.mean(cash_flows['base']['cfo']):.0f}M")
        print(f"   • Base Case Avg. CapEx: ${np.mean(cash_flows['base']['capex']):.0f}M")

        return cash_flows

    def _perform_balance_checks(self, balance_sheets):
        """Perform balance checks to ensure Assets = Liabilities + Equity"""

        balance_checks = {}

        for scenario in ['base', 'bull', 'bear']:
            balance = balance_sheets[scenario]
            checks = []

            for i in range(len(balance['total_assets'])):
                assets = balance['total_assets'][i]
                liabilities_equity = balance['total_liabilities_equity'][i]
                difference = abs(assets - liabilities_equity)
                checks.append(difference < 0.01)  # Allow for small rounding differences

            balance_checks[scenario] = checks

        # Summary statistics
        all_balanced = all([all(checks) for checks in balance_checks.values()])

        print("⚖️  Balance Checks:")
        if all_balanced:
            print("   • ✅ All balance sheets are properly balanced")
        else:
            print("   • ❌ Balance sheet imbalances detected")

        return balance_checks

    def _create_supporting_schedules(self, assumptions, income_statements, balance_sheets, cash_flows):
        """Create supporting schedules (Debt, PP&E, Working Capital)"""

        supporting_schedules = {}

        for scenario in ['base', 'bull', 'bear']:
            income = income_statements[scenario]
            balance = balance_sheets[scenario]
            cash_flow = cash_flows[scenario]

            # Debt Schedule
            debt_schedule = {
                'years': assumptions['years'],
                'opening_debt': [assumptions['opening_debt']] + balance['debt'][:-1],
                'interest_expense': income['interest_expense'],
                'amortization': [assumptions['annual_amortization']] * assumptions['forecast_years'],
                'ending_debt': balance['debt']
            }

            # PP&E Schedule
            ppe_schedule = {
                'years': assumptions['years'],
                'opening_ppe': [0] + balance['ppe'][:-1],  # Assume starting PP&E = 0 for simplicity
                'capex': cash_flow['capex'],
                'depreciation': income['depreciation'],
                'ending_ppe': balance['ppe']
            }

            # Working Capital Schedule
            working_capital = {
                'years': assumptions['years'],
                'ar': balance['ar'],
                'inventory': balance['inventory'],
                'ap': balance['ap'],
                'other_current_liab': balance['other_current_liab'],
                'net_working_capital': [balance['ar'][i] + balance['inventory'][i] - balance['ap'][i] - balance['other_current_liab'][i] for i in range(len(balance['ar']))]
            }

            supporting_schedules[scenario] = {
                'debt_schedule': debt_schedule,
                'ppe_schedule': ppe_schedule,
                'working_capital': working_capital
            }

        print("📋 Supporting Schedules Created")
        return supporting_schedules

    def _create_excel_output(self, three_statement_results):
        """Create professional Excel output with multiple tabs"""

        wb = Workbook()

        # Create worksheets
        ws_assumptions = wb.active
        ws_assumptions.title = "Assumptions"

        ws_income_base = wb.create_sheet("Income Statement - Base")
        ws_income_bull = wb.create_sheet("Income Statement - Bull")
        ws_income_bear = wb.create_sheet("Income Statement - Bear")

        ws_balance_base = wb.create_sheet("Balance Sheet - Base")
        ws_balance_bull = wb.create_sheet("Balance Sheet - Bull")
        ws_balance_bear = wb.create_sheet("Balance Sheet - Bear")

        ws_cashflow_base = wb.create_sheet("Cash Flow - Base")
        ws_cashflow_bull = wb.create_sheet("Cash Flow - Bull")
        ws_cashflow_bear = wb.create_sheet("Cash Flow - Bear")

        ws_supporting = wb.create_sheet("Supporting Schedules")
        ws_summary = wb.create_sheet("Summary & Checks")

        # Apply styles to workbook
        for style_name, style in self.styles.items():
            if style_name not in wb.named_styles:
                wb.add_named_style(style)

        # Create each tab
        self._create_assumptions_tab(ws_assumptions, three_statement_results)
        self._create_income_statement_tabs(ws_income_base, ws_income_bull, ws_income_bear, three_statement_results)
        self._create_balance_sheet_tabs(ws_balance_base, ws_balance_bull, ws_balance_bear, three_statement_results)
        self._create_cash_flow_tabs(ws_cashflow_base, ws_cashflow_bull, ws_cashflow_bear, three_statement_results)
        self._create_supporting_tab(ws_supporting, three_statement_results)
        self._create_summary_tab(ws_summary, three_statement_results)

        # Save workbook
        filename = f"Three_Statement_Model_{self.ticker}_{self.model_date}.xlsx"
        wb.save(filename)

        return filename

    def _create_assumptions_tab(self, ws, results):
        """Create Assumptions tab"""

        assumptions = results['assumptions']

        # Title
        ws['A1'] = f"{self.company_name} ({self.ticker}) - Assumptions & Drivers"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:G1')

        current_row = 3

        # Company Basics
        ws[f'A{current_row}'] = "COMPANY BASICS"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        basics_data = [
            ("Starting Revenue ($M)", assumptions['starting_revenue']),
            ("Forecast Years", assumptions['forecast_years'])
        ]

        for label, value in basics_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].style = 'input'
            current_row += 1

        current_row += 2

        # Growth Scenarios
        ws[f'A{current_row}'] = "REVENUE GROWTH SCENARIOS (%)"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        growth_headers = ['Year'] + ['Base Case', 'Bull Case', 'Bear Case']

        for col, header in enumerate(growth_headers, 1):
            ws.cell(row=current_row, column=col, value=header).style = 'header'

        current_row += 1

        for i in range(max(len(assumptions['growth_base']), len(assumptions['growth_bull']), len(assumptions['growth_bear']))):
            ws.cell(row=current_row, column=1, value=f'Year {i+1}').style = 'label'

            if i < len(assumptions['growth_base']):
                ws.cell(row=current_row, column=2, value=f"{assumptions['growth_base'][i]*100:.1f}%").style = 'input'
            if i < len(assumptions['growth_bull']):
                ws.cell(row=current_row, column=3, value=f"{assumptions['growth_bull'][i]*100:.1f}%").style = 'input'
            if i < len(assumptions['growth_bear']):
                ws.cell(row=current_row, column=4, value=f"{assumptions['growth_bear'][i]*100:.1f}%").style = 'input'

            current_row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    def _create_income_statement_tabs(self, ws_base, ws_bull, ws_bear, results):
        """Create Income Statement tabs for all scenarios"""

        scenarios = [('base', ws_base), ('bull', ws_bull), ('bear', ws_bear)]

        for scenario, ws in scenarios:
            income = results['income_statements'][scenario]

            # Title
            ws['A1'] = f"{self.company_name} - Income Statement ({scenario.title()} Case)"
            ws['A1'].style = 'header'
            ws.merge_cells('A1:G1')

            # Headers
            headers = ['Year', 'Revenue', 'COGS', 'Gross Profit', 'OpEx', 'EBITDA', 'D&A', 'EBIT', 'Interest', 'EBT', 'Taxes', 'Net Income']

            for col, header in enumerate(headers, 1):
                ws.cell(row=3, column=col, value=header).style = 'header'

            # Data
            for i, year in enumerate(income['years']):
                row = i + 4
                ws.cell(row=row, column=1, value=year).style = 'label'

                values = [
                    income['revenue'][i],
                    income['cogs'][i],
                    income['gross_profit'][i],
                    income['opex'][i],
                    income['ebitda'][i],
                    income['depreciation'][i],
                    income['ebit'][i],
                    income['interest_expense'][i],
                    income['ebt'][i],
                    income['taxes'][i],
                    income['net_income'][i]
                ]

                for col, value in enumerate(values, 2):
                    cell = ws.cell(row=row, column=col, value=value)
                    if col in [6, 8, 12]:  # EBITDA, EBIT, Net Income
                        cell.style = 'subtotal'
                    else:
                        cell.style = 'calculation'
                    cell.number_format = '#,##0'

            # Set column widths
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 12

    def _create_balance_sheet_tabs(self, ws_base, ws_bull, ws_bear, results):
        """Create Balance Sheet tabs for all scenarios"""

        scenarios = [('base', ws_base), ('bull', ws_bull), ('bear', ws_bear)]

        for scenario, ws in scenarios:
            balance = results['balance_sheets'][scenario]

            # Title
            ws['A1'] = f"{self.company_name} - Balance Sheet ({scenario.title()} Case)"
            ws['A1'].style = 'header'
            ws.merge_cells('A1:L1')

            # ASSETS section
            ws['A3'] = 'ASSETS'
            ws['A3'].style = 'header'
            ws.merge_cells('A3:D3')

            # Headers
            headers = ['Year', 'Cash', 'AR', 'Inventory', 'PP&E', 'Total Assets', '', 'Debt', 'AP', 'Other CL', 'Total CL', 'Retained Earnings', 'Total Equity', 'Total L&E']

            for col, header in enumerate(headers, 1):
                if header == '':
                    continue
                ws.cell(row=5, column=col, value=header).style = 'header'

            # LIABILITIES & EQUITY section
            ws['H3'] = 'LIABILITIES & EQUITY'
            ws['H3'].style = 'header'
            ws.merge_cells('H3:L3')

            # Data
            for i, year in enumerate(balance['years']):
                row = i + 6
                ws.cell(row=row, column=1, value=year).style = 'label'

                # Assets
                asset_values = [
                    balance['cash'][i],
                    balance['ar'][i],
                    balance['inventory'][i],
                    balance['ppe'][i],
                    balance['total_assets'][i]
                ]

                for col, value in enumerate(asset_values, 2):
                    cell = ws.cell(row=row, column=col, value=value)
                    if col == 6:  # Total Assets
                        cell.style = 'subtotal'
                    else:
                        cell.style = 'calculation'
                    cell.number_format = '#,##0'

                # Empty column
                ws.cell(row=row, column=7, value='')

                # Liabilities & Equity
                liability_equity_values = [
                    balance['debt'][i],
                    balance['ap'][i],
                    balance['other_current_liab'][i],
                    balance['current_liabilities'][i],
                    balance['retained_earnings'][i],
                    balance['total_equity'][i],
                    balance['total_liabilities_equity'][i]
                ]

                for col, value in enumerate(liability_equity_values, 8):
                    cell = ws.cell(row=row, column=col, value=value)
                    if col in [11, 13, 14]:  # Total CL, Total Equity, Total L&E
                        cell.style = 'subtotal'
                    else:
                        cell.style = 'calculation'
                    cell.number_format = '#,##0'

            # Set column widths
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 12

    def _create_cash_flow_tabs(self, ws_base, ws_bull, ws_bear, results):
        """Create Cash Flow tabs for all scenarios"""

        scenarios = [('base', ws_base), ('bull', ws_bull), ('bear', ws_bear)]

        for scenario, ws in scenarios:
            cash_flow = results['cash_flows'][scenario]

            # Title
            ws['A1'] = f"{self.company_name} - Cash Flow Statement ({scenario.title()} Case)"
            ws['A1'].style = 'header'
            ws.merge_cells('A1:J1')

            # Headers
            headers = ['Year', 'Net Income', '+ D&A', '- ∆NWC', '= CFO', '- CapEx', '= CFI', 'Debt Change', 'Dividends', '= CFF', 'Net Change', 'Ending Cash']

            for col, header in enumerate(headers, 1):
                ws.cell(row=3, column=col, value=header).style = 'header'

            # Data
            for i, year in enumerate(cash_flow['years']):
                row = i + 4
                ws.cell(row=row, column=1, value=year).style = 'label'

                values = [
                    cash_flow['net_income'][i],
                    cash_flow['depreciation'][i],
                    -cash_flow['delta_nwc'][i],  # Show as positive for CFO calculation
                    cash_flow['cfo'][i],
                    -cash_flow['capex'][i],
                    cash_flow['cfi'][i],
                    cash_flow['debt_change'][i],
                    cash_flow['dividends_paid'][i],
                    cash_flow['cff'][i],
                    cash_flow['net_change_cash'][i],
                    cash_flow['ending_cash'][i]
                ]

                for col, value in enumerate(values, 2):
                    cell = ws.cell(row=row, column=col, value=value)
                    if col in [5, 7, 10]:  # CFO, CFI, CFF
                        cell.style = 'subtotal'
                    else:
                        cell.style = 'calculation'
                    cell.number_format = '#,##0'

            # Set column widths
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 12

    def _create_supporting_tab(self, ws, results):
        """Create Supporting Schedules tab"""

        # Title
        ws['A1'] = f"{self.company_name} - Supporting Schedules (Base Case)"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:K1')

        supporting = results['supporting_schedules']['base']

        # Debt Schedule
        ws['A3'] = 'DEBT SCHEDULE'
        ws['A3'].style = 'header'
        ws.merge_cells('A3:E3')

        debt_headers = ['Year', 'Opening Debt', 'Interest', 'Amortization', 'Ending Debt']
        for col, header in enumerate(debt_headers, 1):
            ws.cell(row=5, column=col, value=header).style = 'header'

        for i, year in enumerate(supporting['debt_schedule']['years']):
            row = i + 6
            ws.cell(row=row, column=1, value=year).style = 'label'
            values = [
                supporting['debt_schedule']['opening_debt'][i],
                supporting['debt_schedule']['interest_expense'][i],
                supporting['debt_schedule']['amortization'][i],
                supporting['debt_schedule']['ending_debt'][i]
            ]
            for col, value in enumerate(values, 2):
                ws.cell(row=row, column=col, value=value).style = 'calculation'
                ws.cell(row=row, column=col).number_format = '#,##0'

        # PP&E Schedule
        ws['G3'] = 'PP&E SCHEDULE'
        ws['G3'].style = 'header'
        ws.merge_cells('G3:K3')

        ppe_headers = ['Year', 'Opening PP&E', 'CapEx', 'Depreciation', 'Ending PP&E']
        for col, header in enumerate(ppe_headers, 7):
            ws.cell(row=5, column=col, value=header).style = 'header'

        for i, year in enumerate(supporting['ppe_schedule']['years']):
            row = i + 6
            values = [
                supporting['ppe_schedule']['opening_ppe'][i],
                supporting['ppe_schedule']['capex'][i],
                supporting['ppe_schedule']['depreciation'][i],
                supporting['ppe_schedule']['ending_ppe'][i]
            ]
            for col, value in enumerate(values, 8):
                ws.cell(row=row, column=col, value=value).style = 'calculation'
                ws.cell(row=row, column=col).number_format = '#,##0'

        # Set column widths
        for col in range(1, 12):
            ws.column_dimensions[get_column_letter(col)].width = 12

    def _create_summary_tab(self, ws, results):
        """Create Summary & Checks tab"""

        # Title
        ws['A1'] = f"{self.company_name} - Summary & Balance Checks"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:F1')

        current_row = 3

        # Key Metrics Summary
        ws[f'A{current_row}'] = "KEY METRICS SUMMARY (Base Case)"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:F{current_row}')
        current_row += 2

        base_income = results['income_statements']['base']
        base_balance = results['balance_sheets']['base']
        base_cash_flow = results['cash_flows']['base']

        summary_data = [
            ("Starting Revenue", f"${base_income['revenue'][0]:.0f}M"),
            ("Year 5 Revenue", f"${base_income['revenue'][-1]/1000:.1f}B"),
            ("Revenue CAGR", f"{((base_income['revenue'][-1]/base_income['revenue'][0])**(1/(len(base_income['revenue'])-1))-1)*100:.1f}%"),
            ("Average EBITDA Margin", f"{np.mean([ebitda/rev for ebitda, rev in zip(base_income['ebitda'], base_income['revenue'])])*100:.1f}%"),
            ("Year 5 Cash Balance", f"${base_balance['cash'][-1]:.0f}M"),
            ("Average CapEx", f"${np.mean(base_cash_flow['capex']):.0f}M"),
            ("Debt Paydown", f"${results['assumptions']['annual_amortization']:.0f}M/year")
        ]

        for label, value in summary_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].style = 'result'
            current_row += 1

        current_row += 2

        # Balance Check
        ws[f'A{current_row}'] = "BALANCE SHEET CHECKS"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:F{current_row}')
        current_row += 2

        # Headers
        ws.cell(row=current_row, column=1, value="Year").style = 'header'
        ws.cell(row=current_row, column=2, value="Assets").style = 'header'
        ws.cell(row=current_row, column=3, value="Liabilities + Equity").style = 'header'
        ws.cell(row=current_row, column=4, value="Difference").style = 'header'
        ws.cell(row=current_row, column=5, value="Status").style = 'header'
        current_row += 1

        # Check data
        checks = results['balance_checks']['base']
        balance = results['balance_sheets']['base']

        for i, year in enumerate(balance['years']):
            ws.cell(row=current_row, column=1, value=year).style = 'label'
            ws.cell(row=current_row, column=2, value=balance['total_assets'][i]).style = 'calculation'
            ws.cell(row=current_row, column=3, value=balance['total_liabilities_equity'][i]).style = 'calculation'

            difference = balance['total_assets'][i] - balance['total_liabilities_equity'][i]
            ws.cell(row=current_row, column=4, value=difference).style = 'calculation'

            status = "✅ Balanced" if abs(difference) < 0.01 else "❌ Imbalance"
            ws.cell(row=current_row, column=5, value=status).style = 'result' if abs(difference) < 0.01 else 'warning'

            current_row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12


def run_sample_three_statement_model():
    """Run a sample three-statement financial model"""

    print("🏢 Running Professional Three-Statement Model Sample")
    print("=" * 60)

    # Create model instance
    model = ProfessionalThreeStatementModel("TechFlow Inc.", "TECHFLOW")

    # Run the model with sample assumptions
    results, excel_file = model.run_three_statement_model(
        # Company Basics
        starting_revenue=1000.0,  # $1B starting revenue

        # Revenue Growth Scenarios (%)
        growth_base=[0.08, 0.07, 0.06, 0.05, 0.04],  # Year 1-5
        growth_bull=[0.12, 0.10, 0.08, 0.06, 0.05],
        growth_bear=[0.04, 0.03, 0.02, 0.02, 0.01],

        # Operating Margins (%)
        gross_margin=0.65,     # 65% gross margin
        ebitda_margin=0.25,    # 25% EBITDA margin
        ebit_margin=0.20,      # 20% EBIT margin

        # Depreciation & CapEx (% of revenue)
        depreciation_pct=0.05,  # 5% of revenue
        capex_pct=0.04,         # 4% of revenue

        # Working Capital (% of revenue)
        ar_pct=0.15,            # 15% AR
        inventory_pct=0.10,     # 10% Inventory
        ap_pct=0.08,            # 8% AP
        other_current_liab_pct=0.05,

        # Debt Schedule
        opening_debt=300.0,     # $300M debt
        interest_rate=0.06,     # 6%
        annual_amortization=50.0,  # $50M/year

        # Other Assumptions
        tax_rate=0.25,
        dividend_payout=0.30,
        starting_cash=100.0,
        forecast_years=5
    )

    return results, excel_file


if __name__ == "__main__":
    # Run sample three-statement model
    results, excel_file = run_sample_three_statement_model()

    print("\n📋 Model Complete!")
    print(f"📁 Excel file saved as: {excel_file}")

    # Display key results
    print("\n🔑 Key Results Summary:")
    print("=" * 40)
    assumptions = results['assumptions']
    base_income = results['income_statements']['base']
    base_balance = results['balance_sheets']['base']

    print(f"Starting Revenue: ${base_income['revenue'][0]:.0f}M")
    print(f"Year 5 Revenue: ${base_income['revenue'][-1]/1000:.1f}B")
    print(f"Year 5 EBITDA: ${base_income['ebitda'][-1]:.0f}M")
    print(f"Year 5 Net Income: ${base_income['net_income'][-1]:.0f}M")
    print(f"Year 5 Cash Balance: ${base_balance['cash'][-1]:.0f}M")
    print(f"Balance Sheet Status: {'✅ All Years Balanced' if all(results['balance_checks']['base']) else '❌ Imbalances Detected'}")
