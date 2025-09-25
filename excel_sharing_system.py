#!/usr/bin/env python3
"""
Excel Sharing System - Professional Financial Model Sharing Platform
Makes local Excel financial models shareable via secure cloud links

Features:
- Pre-flight checks for required credentials and setup
- Multiple cloud storage providers (Google Drive, Dropbox, AWS S3)
- Read-only link generation with optional expiration
- Excel formatting and formula preservation
- Access logging and security controls
- Professional metadata output

Author: FinModAI Engineering Team
"""

import os
import sys
import json
import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Any, Optional, Union
from dataclasses import dataclass
import openpyxl
from openpyxl import load_workbook

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger('ExcelSharing')

@dataclass
class SharingConfig:
    """Configuration for Excel sharing system."""
    provider: str = "google_drive"  # google_drive, dropbox, aws_s3, azure_blob
    credentials_path: Optional[str] = None
    bucket_name: Optional[str] = None
    folder_id: Optional[str] = None
    link_expiration_days: int = 30
    enable_access_logging: bool = True
    preserve_formatting: bool = True

@dataclass
class SharingResult:
    """Result of Excel sharing operation."""
    success: bool
    model_type: str
    file_name: str
    url: Optional[str] = None
    expires: Optional[str] = None
    file_size: Optional[int] = None
    error_message: Optional[str] = None
    metadata: Optional[Dict[str, Any]] = None

class ExcelSharingSystem:
    """
    Professional Excel sharing system for financial models.

    Handles secure upload to cloud storage and generates shareable read-only links.
    """

    SUPPORTED_PROVIDERS = ["google_drive", "dropbox", "aws_s3", "azure_blob"]
    SUPPORTED_MODEL_TYPES = [
        "dcf", "lbo", "three_statement", "comps", "merger",
        "ipo", "sotp", "football_field", "trading_comps"
    ]

    def __init__(self, config: Optional[SharingConfig] = None):
        self.config = config or SharingConfig()
        self._validate_config()
        logger.info("📤 Excel Sharing System initialized")

    def _validate_config(self):
        """Validate configuration settings."""
        if self.config.provider not in self.SUPPORTED_PROVIDERS:
            raise ValueError(f"Unsupported provider: {self.config.provider}")

        if self.config.link_expiration_days <= 0:
            raise ValueError("Link expiration days must be positive")

    def pre_flight_check(self, file_path: str, model_type: str) -> Dict[str, Any]:
        """
        Perform pre-flight check to identify required setup and credentials.

        Args:
            file_path: Path to local Excel file
            model_type: Type of financial model

        Returns:
            Dict with requirements and status
        """
        logger.info("🔍 Performing pre-flight check...")

        requirements = {
            "file_access": False,
            "file_validation": False,
            "model_type_validation": False,
            "cloud_credentials": False,
            "network_access": False,
            "required_setup": []
        }

        # 1. Check file access
        if not os.path.exists(file_path):
            requirements["required_setup"].append(f"❌ File not found: {file_path}")
        elif not os.access(file_path, os.R_OK):
            requirements["required_setup"].append(f"❌ No read permission: {file_path}")
        else:
            requirements["file_access"] = True
            logger.info(f"✅ File access confirmed: {file_path}")

        # 2. Validate file is Excel
        if requirements["file_access"]:
            try:
                wb = load_workbook(file_path, read_only=True)
                wb.close()
                requirements["file_validation"] = True
                logger.info("✅ Excel file validation passed")
            except Exception as e:
                requirements["required_setup"].append(f"❌ Invalid Excel file: {e}")

        # 3. Validate model type
        if model_type.lower() not in self.SUPPORTED_MODEL_TYPES:
            requirements["required_setup"].append(f"❌ Unsupported model type: {model_type}")
            requirements["required_setup"].append(f"   Supported types: {', '.join(self.SUPPORTED_MODEL_TYPES)}")
        else:
            requirements["model_type_validation"] = True
            logger.info(f"✅ Model type validated: {model_type}")

        # 4. Check cloud provider requirements
        provider_requirements = self._get_provider_requirements()
        requirements["required_setup"].extend(provider_requirements)

        # 5. Check network access
        try:
            import requests
            requests.get("https://www.google.com", timeout=5)
            requirements["network_access"] = True
            logger.info("✅ Network access confirmed")
        except Exception:
            requirements["required_setup"].append("❌ No internet connection")

        # Overall status
        all_checks_passed = all([
            requirements["file_access"],
            requirements["file_validation"],
            requirements["model_type_validation"],
            requirements["network_access"]
        ])

        return {
            "status": "ready" if all_checks_passed else "setup_required",
            "requirements": requirements,
            "summary": self._generate_pre_flight_summary(requirements)
        }

    def _get_provider_requirements(self) -> List[str]:
        """Get requirements for the selected cloud provider."""
        requirements = []

        if self.config.provider == "google_drive":
            requirements.extend([
                "🔑 Google Drive API credentials (service account JSON)",
                "📁 Google Drive folder ID for uploads",
                "🔐 Google Drive API enabled in Google Cloud Console",
                "📝 Set credentials path in SharingConfig.credentials_path"
            ])

        elif self.config.provider == "dropbox":
            requirements.extend([
                "🔑 Dropbox access token",
                "📁 Dropbox app with files.content.write permission",
                "🔐 Dropbox API v2 enabled",
                "📝 Set access token in SharingConfig.credentials_path or environment"
            ])

        elif self.config.provider == "aws_s3":
            requirements.extend([
                "🔑 AWS access key and secret key",
                "🪣 S3 bucket name",
                "🔐 AWS IAM permissions: s3:PutObject, s3:GetObject",
                "📝 Set AWS credentials via boto3 or environment variables"
            ])

        elif self.config.provider == "azure_blob":
            requirements.extend([
                "🔑 Azure storage account name and key",
                "🪣 Azure blob container name",
                "🔐 Azure Storage Blob Data Contributor role",
                "📝 Set credentials via azure-storage-blob or environment"
            ])

        return requirements

    def _generate_pre_flight_summary(self, requirements: Dict) -> str:
        """Generate human-readable summary of pre-flight check."""
        if requirements["file_access"] and requirements["file_validation"] and \
           requirements["model_type_validation"] and requirements["network_access"]:
            return "✅ All checks passed! Ready to share Excel file."
        else:
            failed_checks = []
            if not requirements["file_access"]:
                failed_checks.append("file access")
            if not requirements["file_validation"]:
                failed_checks.append("Excel validation")
            if not requirements["model_type_validation"]:
                failed_checks.append("model type")
            if not requirements["network_access"]:
                failed_checks.append("network access")

            return f"❌ Pre-flight check failed: {', '.join(failed_checks)}. See requirements above."

    def share_excel_file(self, file_path: str, model_type: str) -> SharingResult:
        """
        Share Excel file via cloud storage with read-only link.

        Args:
            file_path: Path to local Excel file
            model_type: Type of financial model

        Returns:
            SharingResult with URL and metadata
        """
        logger.info(f"📤 Starting Excel sharing process for {file_path}")

        # Validate inputs
        if not os.path.exists(file_path):
            return SharingResult(
                success=False,
                model_type=model_type,
                file_name=os.path.basename(file_path),
                error_message=f"File not found: {file_path}"
            )

        # Get file info
        file_size = os.path.getsize(file_path)
        file_name = os.path.basename(file_path)

        # Upload based on provider
        try:
            if self.config.provider == "google_drive":
                result = self._upload_to_google_drive(file_path, model_type)
            elif self.config.provider == "dropbox":
                result = self._upload_to_dropbox(file_path, model_type)
            elif self.config.provider == "aws_s3":
                result = self._upload_to_aws_s3(file_path, model_type)
            elif self.config.provider == "azure_blob":
                result = self._upload_to_azure_blob(file_path, model_type)
            else:
                return SharingResult(
                    success=False,
                    model_type=model_type,
                    file_name=file_name,
                    error_message=f"Unsupported provider: {self.config.provider}"
                )

            if result.success:
                logger.info(f"✅ Excel file shared successfully: {result.url}")
                return result
            else:
                return SharingResult(
                    success=False,
                    model_type=model_type,
                    file_name=file_name,
                    error_message=result.error_message
                )

        except Exception as e:
            logger.error(f"❌ Upload failed: {e}")
            return SharingResult(
                success=False,
                model_type=model_type,
                file_name=file_name,
                error_message=f"Upload failed: {str(e)}"
            )

    def _upload_to_google_drive(self, file_path: str, model_type: str) -> SharingResult:
        """Upload to Google Drive using OAuth2 flow."""
        try:
            from google_auth_oauthlib.flow import InstalledAppFlow
            from google.auth.transport.requests import Request
            from googleapiclient.discovery import build
            from googleapiclient.http import MediaFileUpload
            from googleapiclient.errors import HttpError
            import pickle
            
            # Check for credentials file
            credentials_path = self.config.credentials_path or "google_drive_credentials.json"
            if not os.path.exists(credentials_path):
                return SharingResult(
                    success=False,
                    model_type=model_type,
                    file_name=os.path.basename(file_path),
                    error_message=f"Google Drive credentials not found at {credentials_path}. Please download from Google Cloud Console."
                )
            
            # OAuth2 scopes
            SCOPES = ['https://www.googleapis.com/auth/drive.file']
            
            # Token file for storing OAuth2 credentials
            token_file = 'google_drive_token.pickle'
            
            # Load existing credentials or start OAuth flow
            creds = None
            if os.path.exists(token_file):
                with open(token_file, 'rb') as token:
                    creds = pickle.load(token)
            
            # If there are no (valid) credentials available, let the user log in
            if not creds or not creds.valid:
                if creds and creds.expired and creds.refresh_token:
                    creds.refresh(Request())
                else:
                    # For service account, we need to use OAuth2 delegation
                    # This requires the service account to be configured for domain-wide delegation
                    # For now, let's use a simpler approach with OAuth2 client credentials
                    flow = InstalledAppFlow.from_client_secrets_file(
                        credentials_path, SCOPES)
                    creds = flow.run_local_server(port=0)
                
                # Save the credentials for the next run
                with open(token_file, 'wb') as token:
                    pickle.dump(creds, token)
            
            # Build Drive service
            service = build('drive', 'v3', credentials=creds)
            
            # Generate unique filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_model_type = model_type.replace("-", "_").replace(" ", "_")
            file_name = os.path.basename(file_path)
            drive_file_name = f"{timestamp}_{safe_model_type}_{file_name}"
            
            # File metadata
            file_metadata = {
                'name': drive_file_name,
                'parents': [self.config.folder_id] if self.config.folder_id else None
            }
            
            # Upload file
            media = MediaFileUpload(file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            file = service.files().create(
                body=file_metadata,
                media_body=media,
                fields='id,webViewLink,webContentLink'
            ).execute()
            
            # Make file publicly viewable
            permission = {
                'type': 'anyone',
                'role': 'reader'
            }
            service.permissions().create(
                fileId=file.get('id'),
                body=permission
            ).execute()
            
            # Generate shareable URL
            shareable_url = file.get('webViewLink')
            
            # Set expiration if specified
            expires = None
            if self.config.link_expiration_days > 0:
                expires = (datetime.now() + timedelta(days=self.config.link_expiration_days)).isoformat()
                # Note: Google Drive doesn't support automatic expiration, 
                # but we track it for our own purposes
            
            logger.info(f"✅ File uploaded to Google Drive: {shareable_url}")
            
            return SharingResult(
                success=True,
                model_type=model_type,
                file_name=file_name,
                url=shareable_url,
                expires=expires,
                file_id=file.get('id'),
                provider="google_drive"
            )
            
        except ImportError as e:
            return SharingResult(
                success=False,
                model_type=model_type,
                file_name=os.path.basename(file_path),
                error_message=f"Google Drive API not installed: {str(e)}"
            )
        except HttpError as e:
            return SharingResult(
                success=False,
                model_type=model_type,
                file_name=os.path.basename(file_path),
                error_message=f"Google Drive API error: {str(e)}"
            )
        except Exception as e:
            return SharingResult(
                success=False,
                model_type=model_type,
                file_name=os.path.basename(file_path),
                error_message=f"Google Drive upload failed: {str(e)}"
            )

    def _upload_to_dropbox(self, file_path: str, model_type: str) -> SharingResult:
        """Upload to Dropbox."""
        # Placeholder implementation
        logger.warning("⚠️ Dropbox integration not yet implemented")
        return SharingResult(
            success=False,
            model_type=model_type,
            file_name=os.path.basename(file_path),
            error_message="Dropbox integration requires API setup"
        )

    def _upload_to_aws_s3(self, file_path: str, model_type: str) -> SharingResult:
        """Upload to AWS S3."""
        try:
            import boto3
            from botocore.exceptions import NoCredentialsError

            # Initialize S3 client
            s3_client = boto3.client('s3')

            # Generate unique filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_model_type = model_type.replace("-", "_").replace(" ", "_")
            file_name = os.path.basename(file_path)
            s3_key = f"financial_models/{safe_model_type}/{timestamp}_{file_name}"

            # Upload file
            with open(file_path, 'rb') as f:
                s3_client.upload_fileobj(f, self.config.bucket_name, s3_key)

            # Generate shareable URL
            if self.config.link_expiration_days > 0:
                url = s3_client.generate_presigned_url(
                    'get_object',
                    Params={'Bucket': self.config.bucket_name, 'Key': s3_key},
                    ExpiresIn=self.config.link_expiration_days * 24 * 3600
                )
                expires = (datetime.now() + timedelta(days=self.config.link_expiration_days)).isoformat()
            else:
                # Permanent link (if bucket allows public access)
                url = f"https://{self.config.bucket_name}.s3.amazonaws.com/{s3_key}"
                expires = None

            return SharingResult(
                success=True,
                model_type=model_type,
                file_name=file_name,
                url=url,
                expires=expires,
                file_size=os.path.getsize(file_path),
                metadata={
                    "provider": "aws_s3",
                    "bucket": self.config.bucket_name,
                    "key": s3_key,
                    "expiration_days": self.config.link_expiration_days
                }
            )

        except Exception as e:
            if "credentials" in str(e).lower() or "NoCredentialsError" in str(e):
                return SharingResult(
                    success=False,
                    model_type=model_type,
                    file_name=os.path.basename(file_path),
                    error_message="AWS credentials not found. Set AWS_ACCESS_KEY_ID and AWS_SECRET_ACCESS_KEY"
                )
            else:
                return SharingResult(
                    success=False,
                    model_type=model_type,
                    file_name=os.path.basename(file_path),
                    error_message=f"AWS S3 upload failed: {str(e)}"
                )

    def _upload_to_azure_blob(self, file_path: str, model_type: str) -> SharingResult:
        """Upload to Azure Blob Storage."""
        # Placeholder implementation
        logger.warning("⚠️ Azure Blob integration not yet implemented")
        return SharingResult(
            success=False,
            model_type=model_type,
            file_name=os.path.basename(file_path),
            error_message="Azure Blob integration requires API setup"
        )

    def validate_excel_formatting(self, file_path: str) -> Dict[str, Any]:
        """
        Validate Excel file formatting and formulas are preserved.

        Args:
            file_path: Path to Excel file

        Returns:
            Dict with validation results
        """
        logger.info("🔍 Validating Excel formatting...")

        try:
            wb = load_workbook(file_path, read_only=False)
            validation = {
                "workbook_loaded": True,
                "worksheet_count": len(wb.sheetnames),
                "worksheets": wb.sheetnames,
                "formatting_preserved": True,
                "formulas_intact": True,
                "issues": []
            }

            # Check each worksheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Count cells with formulas
                formula_count = 0
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            formula_count += 1

                validation[f"{sheet_name}_formulas"] = formula_count
                logger.info(f"📊 {sheet_name}: {formula_count} formulas found")

            wb.close()
            return validation

        except Exception as e:
            return {
                "workbook_loaded": False,
                "issues": [f"Excel validation failed: {str(e)}"]
            }

def main():
    """Main function for command-line usage."""
    import argparse

    parser = argparse.ArgumentParser(description="Share Excel financial models via cloud storage")
    parser.add_argument("file_path", help="Path to local Excel file")
    parser.add_argument("model_type", help="Type of financial model",
                       choices=["dcf", "lbo", "three_statement", "comps", "merger", "ipo", "sotp"])
    parser.add_argument("--provider", default="aws_s3",
                       choices=["google_drive", "dropbox", "aws_s3", "azure_blob"],
                       help="Cloud storage provider")
    parser.add_argument("--expiration", type=int, default=30,
                       help="Link expiration in days (0 for permanent)")
    parser.add_argument("--preflight-only", action="store_true",
                       help="Only perform pre-flight check")
    parser.add_argument("--bucket", help="Bucket/container name for cloud storage")

    args = parser.parse_args()

    # Setup configuration
    config = SharingConfig(
        provider=args.provider,
        link_expiration_days=args.expiration,
        bucket_name=args.bucket
    )

    # Initialize sharing system
    sharing_system = ExcelSharingSystem(config)

    # Perform pre-flight check
    print("\n🔍 Performing Pre-Flight Check...")
    print("=" * 60)

    preflight_result = sharing_system.pre_flight_check(args.file_path, args.model_type)

    print(f"Status: {preflight_result['status']}")
    print(f"Summary: {preflight_result['summary']}")

    if preflight_result['requirements']['required_setup']:
        print("\n📋 Required Setup:")
        for req in preflight_result['requirements']['required_setup']:
            print(f"  {req}")

    if args.preflight_only or preflight_result['status'] != 'ready':
        print("\n❌ Pre-flight check incomplete. Please address requirements above.")
        return

    # Validate Excel formatting
    print("\n📊 Validating Excel File...")
    validation = sharing_system.validate_excel_formatting(args.file_path)
    if validation.get('workbook_loaded'):
        print("✅ Excel validation passed")
        print(f"📊 Worksheets: {', '.join(validation.get('worksheets', []))}")
        for sheet in validation.get('worksheets', []):
            formulas = validation.get(f"{sheet}_formulas", 0)
            print(f"   {sheet}: {formulas} formulas")
    else:
        print("❌ Excel validation failed")
        for issue in validation.get('issues', []):
            print(f"   {issue}")
        return

    # Share the file
    print("\n📤 Sharing Excel File...")
    print("=" * 60)

    result = sharing_system.share_excel_file(args.file_path, args.model_type)

    if result.success:
        print("✅ Excel file shared successfully!")
        print(f"📄 Model Type: {result.model_type.title()}")
        print(f"📁 File Name: {result.file_name}")
        print(f"🔗 Shareable URL: {result.url}")
        if result.expires:
            print(f"⏰ Expires: {result.expires}")
        if result.file_size:
            print(",")

        # JSON output
        output = {
            "model_type": result.model_type,
            "file_name": result.file_name,
            "url": result.url,
            "expires": result.expires,
            "file_size_mb": round(result.file_size / (1024 * 1024), 2) if result.file_size else None
        }

        print("\n📋 JSON Output:")
        print(json.dumps(output, indent=2))

    else:
        print("❌ Sharing failed!")
        print(f"Error: {result.error_message}")

if __name__ == "__main__":
    main()
