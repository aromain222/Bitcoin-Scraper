import pandas as pd
import sys
sys.path.append('financial-models-app/backend')
from app import get_comprehensive_company_data, create_professional_excel_model

# Generate a DCF model
data = get_comprehensive_company_data('MSFT', 'Microsoft Corporation')
result = create_professional_excel_model(data, 'dcf')

print('=== DETAILED DCF MODEL EXAMINATION ===')
print(f'File: {result}')

try:
    # Read the Excel file with openpyxl to get more details
    from openpyxl import load_workbook
    wb = load_workbook(result)
    ws = wb.active
    
    print(f'Sheet name: {ws.title}')
    print(f'Max row: {ws.max_row}')
    print(f'Max column: {ws.max_column}')
    
    # Check specific rows where projections should be
    print('\n=== CHECKING PROJECTION ROWS (25-33) ===')
    for row in range(25, 34):
        row_data = []
        for col in range(1, 8):  # Check first 7 columns
            cell_value = ws.cell(row=row, column=col).value
            row_data.append(str(cell_value) if cell_value is not None else 'None')
        print(f'Row {row}: {row_data}')
    
    # Check for any non-empty cells in the entire sheet
    print('\n=== ALL NON-EMPTY CELLS ===')
    non_empty_count = 0
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is not None and str(cell_value).strip() != '':
                print(f'Row {row}, Col {col}: "{cell_value}"')
                non_empty_count += 1
    
    print(f'\nTotal non-empty cells: {non_empty_count}')
    
except Exception as e:
    print(f'Error examining Excel file: {e}')
    import traceback
    traceback.print_exc() 