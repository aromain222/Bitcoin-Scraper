import json
import uuid
from datetime import datetime, timedelta
# yfinance removed - now using FMP APIs
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import io
import os
import re
import threading
import time
import requests
from bs4 import BeautifulSoup
from urllib.parse import quote
from typing import Dict, List, Optional, Any
from dataclasses import dataclass, asdict
from enum import Enum
import openai
import anthropic
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file

# Import our FMP data engine
from fmp_data_engine import build_fmp_assumptions

# Create Flask app
app = Flask(__name__)
app.secret_key = 'finmodai_secret_key_2024'

# Simple storage for models
MODEL_STORAGE = {}

# Session Management System
class ClimateType(str, Enum):
    BASE = "base"
    BULL = "bull"
    BEAR = "bear"
    ESG = "esg"

class ModelType(str, Enum):
    DCF = "dcf"
    LBO = "lbo"
    MA = "ma"
    COMPS = "comps"

@dataclass
class Session:
    id: str
    ticker: str
    climate: ClimateType
    model_type: ModelType
    created_at: datetime
    assumption_profile: Dict[str, Any]

class SessionManager:
    def __init__(self):
        self.sessions: Dict[str, Session] = {}
        self.cleanup_thread = threading.Thread(target=self._cleanup_expired_sessions, daemon=True)
        self.cleanup_thread.start()

    def create_session(self, ticker: str, climate: ClimateType, model_type: ModelType, assumption_profile: Dict[str, Any]) -> str:
        session_id = str(uuid.uuid4())
        self.sessions[session_id] = Session(
            id=session_id,
            ticker=ticker,
            climate=climate,
            model_type=model_type,
            created_at=datetime.now(),
            assumption_profile=assumption_profile
        )
        return session_id

    def get_session(self, session_id: str) -> Optional[Session]:
        return self.sessions.get(session_id)

    def _cleanup_expired_sessions(self):
        while True:
            now = datetime.now()
            expired_sessions = [
                session_id for session_id, session in self.sessions.items()
                if (now - session.created_at) > timedelta(hours=2)
            ]
            for session_id in expired_sessions:
                del self.sessions[session_id]
            time.sleep(300)  # Check every 5 minutes

session_manager = SessionManager()

# Function to generate DCF model with company-specific assumptions
def generate_dcf_model(ticker, climate):
    try:
        print(f"Generating DCF model for {ticker}...")
        # Get company-specific assumptions from FMP
        assumptions = build_fmp_assumptions(ticker)
        
        # Check if we got an error
        if "error" in assumptions:
            print(f"FMP Error getting assumptions: {assumptions['message']}")
            return {
                'ticker': ticker,
                'company_name': ticker,
                'data_source': 'ERROR',
                'error': assumptions['message'],
                'error_type': assumptions.get('error', 'FMP_API_ERROR'),
                'insufficient_data': True
            }
        
        print(f"Successfully retrieved assumptions for {ticker}")
        
        # Extract data from FMP assumptions
        company_name = assumptions.get('company_name', ticker)
        
        # Extract key assumptions from FMP data
        fmp_assumptions = assumptions.get('assumptions', {})
        revenue_growth = fmp_assumptions.get('revenue_growth', [0.08, 0.07, 0.06, 0.05, 0.04])
        operating_margin = fmp_assumptions.get('operating_margin', [0.25, 0.25, 0.25, 0.25, 0.25])
        tax_rate = fmp_assumptions.get('tax_rate', 0.21)
        wacc = fmp_assumptions.get('wacc', 0.105)
        terminal_growth = fmp_assumptions.get('terminal_growth', 0.025)
        
        # Get additional metrics
        capex_percent = fmp_assumptions.get('capex_percent_revenue', 0.06)
        da_percent = fmp_assumptions.get('da_percent_revenue', 0.04)
        nwc_percent = fmp_assumptions.get('nwc_percent_revenue', 0.03)
        
        # Get historical data for UI
        historicals = assumptions.get('historicals', {})
        
        # For demo purposes - in real implementation, you'd get this from FMP market data API
        current_price = 25.00  # This should come from FMP market data
        
        # Simple DCF calculation for demonstration
        enterprise_value = 2500000000  # Simplified
        equity_value = 2000000000  # Simplified
        implied_price = 25.00  # Simplified
        upside_downside = ((implied_price / current_price) - 1) * 100 if current_price > 0 else 0
        
        return {
            'ticker': ticker,
            'company_name': company_name,
            'enterprise_value': enterprise_value,
            'equity_value': equity_value,
            'implied_price': implied_price,
            'current_price': current_price,
            'upside_downside': upside_downside,
            'assumptions': {
                'revenue_growth': revenue_growth,
                'operating_margin': operating_margin,
                'wacc': wacc,
                'terminal_growth': terminal_growth,
                'tax_rate': tax_rate,
                'capex_percent_revenue': capex_percent,
                'da_percent_revenue': da_percent,
                'nwc_percent_revenue': nwc_percent
            },
            'raw_assumptions': assumptions,
            'historicals': historicals,
            'provenance': assumptions.get('provenance', {}),
            'data_source': 'FMP'
        }
    except Exception as e:
        print(f"Error in generate_dcf_model: {str(e)}")
        return {
            'ticker': ticker,
            'company_name': ticker,
            'data_source': 'ERROR',
            'error': f"FMP API Error: {str(e)}",
            'error_type': 'FMP_API_EXCEPTION',
            'insufficient_data': True
        }

# Routes
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/generate-model', methods=['GET', 'POST'])
def generate_model():
    if request.method == 'POST':
        ticker = request.form.get('ticker', '').strip().upper()
        model_type = request.form.get('model_type', 'dcf').lower()
        climate = request.form.get('climate', 'base').lower()
        
        if not ticker:
            flash('Please enter a ticker symbol', 'error')
            return redirect(url_for('index'))
        
        try:
            # Generate model
            model_id = str(uuid.uuid4())
            MODEL_STORAGE[model_id] = {
                'ticker': ticker,
                'model_type': model_type,
                'climate': climate,
                'status': 'pending'
            }
            
            # Generate DCF model with company-specific assumptions
            if model_type == 'dcf':
                result = generate_dcf_model(ticker, climate)
                MODEL_STORAGE[model_id]['result'] = result
            else:
                # For other model types, use default values
                MODEL_STORAGE[model_id]['result'] = {
                    'enterprise_value': 2500000000,
                    'equity_value': 2000000000,
                    'implied_price': 25.00,
                    'current_price': 20.00,
                    'upside_downside': 25.0
                }
            
            # Redirect to results page
            return redirect(url_for('model_results', model_id=model_id))
            
        except Exception as e:
            flash(f'Error generating model: {str(e)}', 'error')
            return redirect(url_for('index'))
    
    return render_template('generate_model.html')

@app.route('/model-results/<model_id>')
def model_results(model_id):
    model = MODEL_STORAGE.get(model_id)
    if not model:
        flash('Model not found', 'error')
        return redirect(url_for('index'))
    
    # Get model data
    ticker = model.get('ticker')
    model_type = model.get('model_type')
    climate = model.get('climate')
    result = model.get('result', {})
    
    # Check if we have insufficient data
    if result.get('insufficient_data'):
        error_message = result.get('error', 'Unable to retrieve sufficient historical data')
        error_type = result.get('error_type', 'DATA_ERROR')
        
        return render_template('model_results.html',
            model_id=model_id,
            ticker=ticker,
            model_type=model_type,
            climate=climate,
            error_state=True,
            error_message=error_message,
            error_type=error_type
        )
    
    # Format values for display
    enterprise_value = result.get('enterprise_value', 0)
    equity_value = result.get('equity_value', 0)
    implied_price = result.get('implied_price', 0)
    current_price = result.get('current_price', 0)
    upside_downside = result.get('upside_downside', 0)
    
    # Get assumptions for display
    assumptions = result.get('assumptions', {})
    revenue_growth = assumptions.get('revenue_growth', [0.08, 0.07, 0.06, 0.05, 0.04])
    operating_margin = assumptions.get('operating_margin', [0.25, 0.25, 0.25, 0.25, 0.25])
    wacc = assumptions.get('wacc', 0.105)
    terminal_growth = assumptions.get('terminal_growth', 0.025)
    tax_rate = assumptions.get('tax_rate', 0.21)
    
    # Get additional metrics from FMP
    capex_percent_revenue = assumptions.get('capex_percent_revenue', 0.06)
    da_percent_revenue = assumptions.get('da_percent_revenue', 0.04)
    nwc_percent_revenue = assumptions.get('nwc_percent_revenue', 0.03)
    
    # Get data source and provenance
    data_source = result.get('data_source', 'FMP')
    company_name = result.get('company_name', ticker)
    provenance = result.get('provenance', {})
    historicals = result.get('historicals', {})
    
    # Get detailed assumptions from financial data engine if available
    raw_assumptions = result.get('raw_assumptions', {})
    
    # Helper function to get historical CAGR text
    def get_historical_cagr_text(revenue_values):
        if not revenue_values or len(revenue_values) < 2:
            return "Insufficient historical data"
        years = min(3, len(revenue_values) - 1)
        if years == 0:
            return "Insufficient historical data"
        start_value = revenue_values[years]
        end_value = revenue_values[0]
        if start_value <= 0 or end_value <= 0:
            return "Invalid historical data"
        cagr = (end_value / start_value) ** (1 / years) - 1
        return f"{cagr:.1%}"
    
    # Helper function to get average margin text
    def get_average_margin_text(margin_values):
        if not margin_values or len(margin_values) == 0:
            return "Insufficient historical data"
        years = min(3, len(margin_values))
        avg = sum(margin_values[:years]) / years
        return f"{avg:.1%}"
    
    # Format assumptions HTML with historical context
    assumptions_html = f"""
    <div class="space-y-6">
        <!-- Header -->
        <div class="bg-gradient-to-r from-blue-50 to-blue-100 p-4 rounded-lg border border-blue-200">
            <div class="flex items-center justify-between">
                <div>
                    <div class="text-sm font-medium text-blue-800 mb-1">Data Source: {data_source.upper()}</div>
                    <div class="text-lg font-bold text-blue-900">Company-Specific DCF Assumptions</div>
                    <div class="text-sm text-blue-700">{company_name} ({ticker})</div>
                </div>
                <div class="text-right">
                    <div class="text-xs text-blue-600">Built from</div>
                    <div class="text-sm font-medium text-blue-800">Historical Financials</div>
                </div>
            </div>
        </div>
        
        <!-- Revenue Growth & Operating Margin Table -->
        <div class="bg-white border border-gray-200 rounded-lg overflow-hidden">
            <div class="px-4 py-3 bg-gray-50 border-b border-gray-200">
                <h3 class="text-lg font-semibold text-gray-900">5-Year Forecast Assumptions</h3>
            </div>
            <div class="overflow-x-auto">
                <table class="min-w-full divide-y divide-gray-200">
                    <thead class="bg-gray-50">
                        <tr>
                            <th class="px-6 py-3 text-left text-xs font-medium text-gray-500 uppercase tracking-wider">Metric</th>
                            <th class="px-6 py-3 text-center text-xs font-medium text-gray-500 uppercase tracking-wider">Year 1</th>
                            <th class="px-6 py-3 text-center text-xs font-medium text-gray-500 uppercase tracking-wider">Year 2</th>
                            <th class="px-6 py-3 text-center text-xs font-medium text-gray-500 uppercase tracking-wider">Year 3</th>
                            <th class="px-6 py-3 text-center text-xs font-medium text-gray-500 uppercase tracking-wider">Year 4</th>
                            <th class="px-6 py-3 text-center text-xs font-medium text-gray-500 uppercase tracking-wider">Year 5</th>
                        </tr>
                    </thead>
                    <tbody class="bg-white divide-y divide-gray-200">
                        <tr class="hover:bg-gray-50">
                            <td class="px-6 py-4 whitespace-nowrap text-sm font-medium text-gray-900">Revenue Growth</td>
                            <td class="px-6 py-4 whitespace-nowrap text-center text-sm text-gray-900 font-medium">{revenue_growth[0]:.1%}</td>
                            <td class="px-6 py-4 whitespace-nowrap text-center text-sm text-gray-500">{revenue_growth[1]:.1%}</td>
                            <td class="px-6 py-4 whitespace-nowrap text-center text-sm text-gray-500">{revenue_growth[2]:.1%}</td>
                            <td class="px-6 py-4 whitespace-nowrap text-center text-sm text-gray-500">{revenue_growth[3]:.1%}</td>
                            <td class="px-6 py-4 whitespace-nowrap text-center text-sm text-gray-500">{revenue_growth[4]:.1%}</td>
                        </tr>
                        <tr class="hover:bg-gray-50">
                            <td class="px-6 py-4 whitespace-nowrap text-sm font-medium text-gray-900">Operating Margin</td>
                            <td class="px-6 py-4 whitespace-nowrap text-center text-sm text-gray-900 font-medium">{operating_margin[0]:.1%}</td>
                            <td class="px-6 py-4 whitespace-nowrap text-center text-sm text-gray-500">{operating_margin[1]:.1%}</td>
                            <td class="px-6 py-4 whitespace-nowrap text-center text-sm text-gray-500">{operating_margin[2]:.1%}</td>
                            <td class="px-6 py-4 whitespace-nowrap text-center text-sm text-gray-500">{operating_margin[3]:.1%}</td>
                            <td class="px-6 py-4 whitespace-nowrap text-center text-sm text-gray-500">{operating_margin[4]:.1%}</td>
                        </tr>
                    </tbody>
                </table>
            </div>
        </div>
        
        <!-- Historical Data Context -->
        <div class="bg-gray-50 border border-gray-200 p-4 rounded-lg">
            <h4 class="text-md font-semibold text-gray-900 mb-3">Calculation Methodology</h4>
            <div class="grid grid-cols-1 md:grid-cols-2 gap-4 text-sm text-gray-700">
                <div>
                    <span class="font-medium">Revenue Growth:</span> Based on 3-year historical CAGR ({get_historical_cagr_text(historicals.get('revenue', []))})
                </div>
                <div>
                    <span class="font-medium">Operating Margin:</span> 3-year average ({get_average_margin_text(historicals.get('operating_margins', []))})
                </div>
                <div>
                    <span class="font-medium">CapEx % Revenue:</span> 3-year average ({capex_percent_revenue:.1%})
                </div>
                <div>
                    <span class="font-medium">NWC % Revenue:</span> 3-year average ({nwc_percent_revenue:.1%})
                </div>
            </div>
        </div>
        
        <!-- Key Parameters -->
        <div class="grid grid-cols-1 md:grid-cols-3 gap-4">
            <div class="bg-blue-50 border border-blue-200 p-4 rounded-lg">
                <div class="text-blue-700 font-medium text-sm">WACC</div>
                <div class="text-2xl font-bold text-blue-800">{wacc:.1%}</div>
                <div class="text-xs text-blue-600 mt-1">
                    FMP Historical Data<br>
                    3-year average calculations
                </div>
            </div>
            
            <div class="bg-green-50 border border-green-200 p-4 rounded-lg">
                <div class="text-green-700 font-medium text-sm">Terminal Growth</div>
                <div class="text-2xl font-bold text-green-800">{terminal_growth:.1%}</div>
                <div class="text-xs text-green-600 mt-1">
                    FMP historical analysis<br>
                    Conservative terminal rate
                </div>
            </div>
            
            <div class="bg-purple-50 border border-purple-200 p-4 rounded-lg">
                <div class="text-purple-700 font-medium text-sm">Effective Tax Rate</div>
                <div class="text-2xl font-bold text-purple-800">{tax_rate:.1%}</div>
                <div class="text-xs text-purple-600 mt-1">
                    FMP income statements<br>
                    3-year effective tax average
                </div>
            </div>
        </div>
    </div>
    """
    
    # Format valuation results HTML
    valuation_html = f"""
    <div class="space-y-3">
        <div class="bg-green-50 p-4 rounded-lg">
            <div class="text-green-700 font-medium">Enterprise Value</div>
            <div class="text-2xl font-bold text-green-800">${enterprise_value/1e9:.1f}B</div>
        </div>
        
        <div class="bg-green-50 p-4 rounded-lg">
            <div class="text-green-700 font-medium">Equity Value</div>
            <div class="text-2xl font-bold text-green-800">${equity_value/1e9:.1f}B</div>
        </div>
        
        <div class="bg-blue-50 p-4 rounded-lg">
            <div class="text-blue-700 font-medium">Implied Price</div>
            <div class="text-2xl font-bold text-blue-800">${implied_price:.2f}</div>
        </div>
        
        <div class="bg-blue-50 p-4 rounded-lg">
            <div class="text-blue-700 font-medium">Current Price</div>
            <div class="text-2xl font-bold text-blue-800">${current_price:.2f}</div>
        </div>
        
        <div class="p-4 rounded-lg {upside_downside > 0 and 'bg-green-50' or 'bg-red-50'}">
            <div class="font-medium {upside_downside > 0 and 'text-green-700' or 'text-red-700'}">Upside/Downside</div>
            <div class="text-2xl font-bold {upside_downside > 0 and 'text-green-800' or 'text-red-800'}">{upside_downside:.1f}%</div>
        </div>
    </div>
    """
    
    # Format download section HTML
    download_section_html = """
    <div class="space-y-4">
        <button onclick="window.print()" class="bg-gray-100 text-gray-700 px-4 py-2 rounded-lg font-medium hover:bg-gray-200 transition-colors">
            Print Results
        </button>
        
        <a href="/download-excel" class="block">
            <button class="bg-green-600 text-white px-4 py-2 rounded-lg font-medium hover:bg-green-700 transition-colors">
                Download Excel Model
            </button>
        </a>
    </div>
    """
    
    return render_template(
        'model_results.html',
        model_id=model_id,
        ticker=ticker,
        model_type=model_type,
        climate=climate,
        assumptions_html=assumptions_html,
        valuation_html=valuation_html,
        download_section_html=download_section_html
    )

@app.route('/download-excel')
def download_excel():
    # Get the last model generated (simplified approach)
    last_model_id = list(MODEL_STORAGE.keys())[-1] if MODEL_STORAGE else None
    model = MODEL_STORAGE.get(last_model_id, {})
    result = model.get('result', {})
    
    # Get assumptions
    assumptions = result.get('assumptions', {})
    revenue_growth = assumptions.get('revenue_growth', [0.08, 0.07, 0.06, 0.05, 0.04])
    operating_margin = assumptions.get('operating_margin', [0.25, 0.25, 0.25, 0.25, 0.25])
    wacc = assumptions.get('wacc', 0.105)
    terminal_growth = assumptions.get('terminal_growth', 0.025)
    tax_rate = assumptions.get('tax_rate', 0.21)
    
    # Get additional assumptions from FMP data
    capex_pct_rev = assumptions.get('capex_percent_revenue', 0.06)
    da_pct_rev = assumptions.get('da_percent_revenue', 0.04)
    nwc_pct_rev = assumptions.get('nwc_percent_revenue', 0.03)
    
    # Get valuation results
    enterprise_value = result.get('enterprise_value', 2500000000)
    equity_value = result.get('equity_value', 2000000000)
    implied_price = result.get('implied_price', 25.00)
    current_price = result.get('current_price', 20.00)
    upside_downside = result.get('upside_downside', 25.0)
    
    # Create Excel file
    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    
    # Create Assumptions worksheet
    assumptions_sheet = workbook.active
    assumptions_sheet.title = "Company Assumptions"
    
    # Add headers
    headers = ["Year", "1", "2", "3", "4", "5"]
    for col, header in enumerate(headers, 1):
        cell = assumptions_sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    
    # Add revenue growth
    row_idx = 2
    assumptions_sheet.cell(row=row_idx, column=1).value = "Revenue Growth"
    for i, growth in enumerate(revenue_growth):
        cell = assumptions_sheet.cell(row=row_idx, column=i+2)
        cell.value = growth
        cell.number_format = "0.0%"
        cell.alignment = Alignment(horizontal="right")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    
    # Add operating margin
    row_idx = 3
    assumptions_sheet.cell(row=row_idx, column=1).value = "Operating Margin"
    for i, margin in enumerate(operating_margin):
        cell = assumptions_sheet.cell(row=row_idx, column=i+2)
        cell.value = margin
        cell.number_format = "0.0%"
        cell.alignment = Alignment(horizontal="right")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    
    # Add key parameters
    row_idx = 5
    assumptions_sheet.cell(row=row_idx, column=1).value = "WACC"
    assumptions_sheet.cell(row=row_idx, column=2).value = wacc
    assumptions_sheet.cell(row=row_idx, column=2).number_format = "0.0%"
    
    row_idx = 6
    assumptions_sheet.cell(row=row_idx, column=1).value = "Terminal Growth"
    assumptions_sheet.cell(row=row_idx, column=2).value = terminal_growth
    assumptions_sheet.cell(row=row_idx, column=2).number_format = "0.0%"
    
    row_idx = 7
    assumptions_sheet.cell(row=row_idx, column=1).value = "Tax Rate"
    assumptions_sheet.cell(row=row_idx, column=2).value = tax_rate
    assumptions_sheet.cell(row=row_idx, column=2).number_format = "0.0%"
    
    # Add additional assumptions
    row_idx = 9
    assumptions_sheet.cell(row=row_idx, column=1).value = "Depreciation % Revenue"
    assumptions_sheet.cell(row=row_idx, column=2).value = da_pct_rev
    assumptions_sheet.cell(row=row_idx, column=2).number_format = "0.0%"
    
    row_idx = 10
    assumptions_sheet.cell(row=row_idx, column=1).value = "CapEx % Revenue"
    assumptions_sheet.cell(row=row_idx, column=2).value = capex_pct_rev
    assumptions_sheet.cell(row=row_idx, column=2).number_format = "0.0%"
    
    row_idx = 11
    assumptions_sheet.cell(row=row_idx, column=1).value = "NWC % Revenue"
    assumptions_sheet.cell(row=row_idx, column=2).value = nwc_pct_rev
    assumptions_sheet.cell(row=row_idx, column=2).number_format = "0.0%"
    
    # Create Valuation worksheet
    valuation_sheet = workbook.create_sheet("Valuation Results")
    
    # Add headers
    headers = ["Metric", "Value"]
    for col, header in enumerate(headers, 1):
        cell = valuation_sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    
    # Add valuation data
    data = [
        ["Enterprise Value", enterprise_value],
        ["Equity Value", equity_value],
        ["Implied Price", implied_price],
        ["Current Price", current_price],
        ["Upside/Downside", upside_downside / 100]  # Convert to decimal for Excel
    ]
    
    for row_idx, (metric, value) in enumerate(data, 2):
        # Metric
        cell = valuation_sheet.cell(row=row_idx, column=1)
        cell.value = metric
        cell.alignment = Alignment(horizontal="left")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Value
        cell = valuation_sheet.cell(row=row_idx, column=2)
        cell.value = value
        
        # Format based on metric
        if metric in ["Enterprise Value", "Equity Value"]:
            cell.number_format = '"$"#,##0.0,," B"'
        elif metric in ["Implied Price", "Current Price"]:
            cell.number_format = '"$"#,##0.00'
        elif metric == "Upside/Downside":
            cell.number_format = "0.0%"
        
        cell.alignment = Alignment(horizontal="right")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    
    # Auto-adjust column widths
    for sheet in [assumptions_sheet, valuation_sheet]:
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Save to buffer
    workbook.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="fmp_historical_assumptions_model.xlsx"
    )

if __name__ == '__main__':
    # Use a different port to avoid conflict
    port = 10001
    print(f"Starting app on port {port}")
    app.run(host='0.0.0.0', port=port)