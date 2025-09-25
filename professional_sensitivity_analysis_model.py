#!/usr/bin/env python3
"""
Professional Sensitivity Analysis Model
Stress-tests key assumptions across valuation methodologies

Author: Investment Banking Modeler
Date: 2024

Features:
- DCF sensitivity (WACC, Terminal Growth)
- LBO sensitivity (Exit Multiple, Leverage)
- Comps sensitivity (Valuation Multiples)
- 2D sensitivity tables
- Professional banker-standard formatting
- Excel output with multiple tabs
- Base case highlighting
"""

import numpy as np
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# Professional color scheme for Sensitivity Analysis
SENSITIVITY_COLORS = {
    'header_blue': '1F4E79',
    'input_light_blue': 'DDEBF7',
    'calculation_black': '000000',
    'output_green': 'D8E4BC',
    'base_case_yellow': 'FFF2CC',
    'warning_red': 'F8CECC',
    'border_dark': '2C3E50',
    'dcf_orange': 'FF7F50',
    'lbo_purple': '9370DB',
    'comps_blue': '4682B4',
    'text_dark': '000000',
    'text_white': 'FFFFFF',
    'highlight_gold': 'FFD700'
}

class ProfessionalSensitivityAnalysisModel:
    """
    Comprehensive Sensitivity Analysis Model with Professional Formatting
    """

    def __init__(self, target_company="Target Company", target_ticker="TARGET"):
        self.target_company = target_company
        self.target_ticker = target_ticker
        self.model_date = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Initialize styles
        self.styles = self._create_styles()

    def _create_styles(self):
        """Create professional Excel styles"""
        styles = {}

        # Header style
        styles['header'] = NamedStyle(name='header')
        styles['header'].font = Font(name='Calibri', size=12, bold=True, color=SENSITIVITY_COLORS['text_white'])
        styles['header'].fill = PatternFill(start_color=SENSITIVITY_COLORS['header_blue'], end_color=SENSITIVITY_COLORS['header_blue'], fill_type='solid')
        styles['header'].alignment = Alignment(horizontal='center', vertical='center')

        # Input style
        styles['input'] = NamedStyle(name='input')
        styles['input'].font = Font(name='Calibri', size=10, color=SENSITIVITY_COLORS['text_dark'])
        styles['input'].fill = PatternFill(start_color=SENSITIVITY_COLORS['input_light_blue'], end_color=SENSITIVITY_COLORS['input_light_blue'], fill_type='solid')
        styles['input'].alignment = Alignment(horizontal='right', vertical='center')
        styles['input'].number_format = '#,##0.00'

        # Base case style
        styles['base_case'] = NamedStyle(name='base_case')
        styles['base_case'].font = Font(name='Calibri', size=10, bold=True, color=SENSITIVITY_COLORS['text_dark'])
        styles['base_case'].fill = PatternFill(start_color=SENSITIVITY_COLORS['base_case_yellow'], end_color=SENSITIVITY_COLORS['base_case_yellow'], fill_type='solid')
        styles['base_case'].alignment = Alignment(horizontal='center', vertical='center')
        styles['base_case'].border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))

        # Calculation style
        styles['calculation'] = NamedStyle(name='calculation')
        styles['calculation'].font = Font(name='Calibri', size=10, color=SENSITIVITY_COLORS['calculation_black'])
        styles['calculation'].alignment = Alignment(horizontal='right', vertical='center')
        styles['calculation'].number_format = '#,##0.00'

        # Output style
        styles['output'] = NamedStyle(name='output')
        styles['output'].font = Font(name='Calibri', size=10, bold=True, color=SENSITIVITY_COLORS['text_dark'])
        styles['output'].fill = PatternFill(start_color=SENSITIVITY_COLORS['output_green'], end_color=SENSITIVITY_COLORS['output_green'], fill_type='solid')
        styles['output'].alignment = Alignment(horizontal='right', vertical='center')
        styles['output'].number_format = '#,##0.00'

        # Highlight style for key values
        styles['highlight'] = NamedStyle(name='highlight')
        styles['highlight'].font = Font(name='Calibri', size=10, bold=True, color=SENSITIVITY_COLORS['text_white'])
        styles['highlight'].fill = PatternFill(start_color=SENSITIVITY_COLORS['highlight_gold'], end_color=SENSITIVITY_COLORS['highlight_gold'], fill_type='solid')
        styles['highlight'].alignment = Alignment(horizontal='center', vertical='center')

        # Methodology styles
        styles['dcf'] = NamedStyle(name='dcf')
        styles['dcf'].font = Font(name='Calibri', size=10, color=SENSITIVITY_COLORS['text_white'])
        styles['dcf'].fill = PatternFill(start_color=SENSITIVITY_COLORS['dcf_orange'], end_color=SENSITIVITY_COLORS['dcf_orange'], fill_type='solid')
        styles['dcf'].alignment = Alignment(horizontal='center', vertical='center')

        styles['lbo'] = NamedStyle(name='lbo')
        styles['lbo'].font = Font(name='Calibri', size=10, color=SENSITIVITY_COLORS['text_white'])
        styles['lbo'].fill = PatternFill(start_color=SENSITIVITY_COLORS['lbo_purple'], end_color=SENSITIVITY_COLORS['lbo_purple'], fill_type='solid')
        styles['lbo'].alignment = Alignment(horizontal='center', vertical='center')

        styles['comps'] = NamedStyle(name='comps')
        styles['comps'].font = Font(name='Calibri', size=10, color=SENSITIVITY_COLORS['text_white'])
        styles['comps'].fill = PatternFill(start_color=SENSITIVITY_COLORS['comps_blue'], end_color=SENSITIVITY_COLORS['comps_blue'], fill_type='solid')
        styles['comps'].alignment = Alignment(horizontal='center', vertical='center')

        # Warning style
        styles['warning'] = NamedStyle(name='warning')
        styles['warning'].font = Font(name='Calibri', size=10, color=SENSITIVITY_COLORS['text_dark'])
        styles['warning'].fill = PatternFill(start_color=SENSITIVITY_COLORS['warning_red'], end_color=SENSITIVITY_COLORS['warning_red'], fill_type='solid')
        styles['warning'].alignment = Alignment(horizontal='center', vertical='center')

        # Bold label style
        styles['label_bold'] = NamedStyle(name='label_bold')
        styles['label_bold'].font = Font(name='Calibri', size=10, bold=True, color=SENSITIVITY_COLORS['text_dark'])
        styles['label_bold'].alignment = Alignment(horizontal='left', vertical='center')

        # Normal label style
        styles['label'] = NamedStyle(name='label')
        styles['label'].font = Font(name='Calibri', size=10, color=SENSITIVITY_COLORS['text_dark'])
        styles['label'].alignment = Alignment(horizontal='left', vertical='center')

        # Summary box style
        styles['summary_box'] = NamedStyle(name='summary_box')
        styles['summary_box'].font = Font(name='Calibri', size=11, bold=True, color=SENSITIVITY_COLORS['text_white'])
        styles['summary_box'].fill = PatternFill(start_color=SENSITIVITY_COLORS['header_blue'], end_color=SENSITIVITY_COLORS['header_blue'], fill_type='solid')
        styles['summary_box'].alignment = Alignment(horizontal='center', vertical='center')
        styles['summary_box'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Sensitivity table header style
        styles['table_header'] = NamedStyle(name='table_header')
        styles['table_header'].font = Font(name='Calibri', size=10, bold=True, color=SENSITIVITY_COLORS['text_white'])
        styles['table_header'].fill = PatternFill(start_color=SENSITIVITY_COLORS['border_dark'], end_color=SENSITIVITY_COLORS['border_dark'], fill_type='solid')
        styles['table_header'].alignment = Alignment(horizontal='center', vertical='center')
        styles['table_header'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        return styles

    def run_sensitivity_analysis_model(self,
                                     # Target Company Info
                                     target_company=None,
                                     target_ticker=None,

                                     # DCF Assumptions and Base Case
                                     dcf_base_wacc=0.08,      # 8.0%
                                     dcf_base_terminal_growth=0.02,  # 2.0%
                                     dcf_pv_fcf=2800.0,       # $2.8B PV of FCF
                                     dcf_terminal_value=3200.0,  # $3.2B terminal value

                                     # DCF Sensitivity Ranges
                                     dcf_wacc_range=None,     # [0.06, 0.07, 0.08, 0.09, 0.10]
                                     dcf_growth_range=None,   # [0.01, 0.015, 0.02, 0.025, 0.03]

                                     # LBO Assumptions and Base Case
                                     lbo_base_exit_multiple=8.0,  # EV/EBITDA
                                     lbo_base_leverage_ratio=5.0, # Debt/EBITDA
                                     lbo_base_ebitda=400.0,   # $400M EBITDA
                                     lbo_base_tax_rate=0.25,  # 25%
                                     lbo_base_debt_repayment_pct=0.50,  # 50% debt repayment

                                     # LBO Sensitivity Ranges
                                     lbo_exit_multiple_range=None,  # [6.0, 7.0, 8.0, 9.0, 10.0]
                                     lbo_leverage_range=None,       # [3.0, 4.0, 5.0, 6.0, 7.0]

                                     # Trading Comps Assumptions and Base Case
                                     comps_base_ev_ebitda=9.5,      # EV/EBITDA
                                     comps_base_price_earnings=18.0, # P/E
                                     comps_base_ebitda=400.0,       # $400M EBITDA
                                     comps_base_eps=2.50,           # $2.50 EPS
                                     comps_base_shares_outstanding=80.0,  # 80M shares

                                     # Comps Sensitivity Ranges
                                     comps_ev_ebitda_range=None,    # [7.0, 8.0, 9.0, 10.0, 11.0]
                                     comps_pe_range=None):          # [15.0, 16.0, 17.0, 18.0, 19.0]

        """
        Run complete sensitivity analysis model
        """

        if target_company:
            self.target_company = target_company
        if target_ticker:
            self.target_ticker = target_ticker

        print(f"📊 Building Professional Sensitivity Analysis Model for {self.target_company} ({self.target_ticker})")
        print("=" * 90)

        # Set default ranges if not provided
        if dcf_wacc_range is None:
            dcf_wacc_range = [0.06, 0.07, 0.08, 0.09, 0.10]  # 6% to 10%
        if dcf_growth_range is None:
            dcf_growth_range = [0.01, 0.015, 0.02, 0.025, 0.03]  # 1% to 3%

        if lbo_exit_multiple_range is None:
            lbo_exit_multiple_range = [6.0, 7.0, 8.0, 9.0, 10.0]  # 6x to 10x
        if lbo_leverage_range is None:
            lbo_leverage_range = [3.0, 4.0, 5.0, 6.0, 7.0]  # 3x to 7x

        if comps_ev_ebitda_range is None:
            comps_ev_ebitda_range = [7.0, 8.0, 9.0, 10.0, 11.0]  # 7x to 11x
        if comps_pe_range is None:
            comps_pe_range = [15.0, 16.0, 17.0, 18.0, 19.0]  # 15x to 19x

        # Step 1: Create Assumptions & Base Case Data
        assumptions = self._create_assumptions(
            dcf_base_wacc, dcf_base_terminal_growth, dcf_pv_fcf, dcf_terminal_value,
            dcf_wacc_range, dcf_growth_range,
            lbo_base_exit_multiple, lbo_base_leverage_ratio, lbo_base_ebitda,
            lbo_base_tax_rate, lbo_base_debt_repayment_pct,
            lbo_exit_multiple_range, lbo_leverage_range,
            comps_base_ev_ebitda, comps_base_price_earnings, comps_base_ebitda,
            comps_base_eps, comps_base_shares_outstanding,
            comps_ev_ebitda_range, comps_pe_range
        )

        # Step 2: Generate Sensitivity Tables
        sensitivity_tables = self._generate_sensitivity_tables(assumptions)

        # Step 3: Calculate Valuation Ranges
        valuation_ranges = self._calculate_valuation_ranges(sensitivity_tables)

        # Step 4: Create Excel Output
        excel_file = self._create_excel_output(assumptions, sensitivity_tables, valuation_ranges)

        print("\n✅ Sensitivity Analysis Model Complete!")
        print("📊 Key Sensitivity Insights:")
        print(f"   • Target: {self.target_company} ({self.target_ticker})")
        print(f"   • DCF Base Case: ${dcf_pv_fcf + dcf_terminal_value:.1f}M EV")
        print(f"   • LBO Base Case: ${lbo_base_ebitda * lbo_base_exit_multiple:.1f}M exit EV")
        print(f"   • Comps Base Case: ${comps_base_ebitda * comps_base_ev_ebitda:.1f}M EV")
        print(f"📁 Excel Output: {excel_file}")

        # Display key findings
        self._display_sensitivity_insights(sensitivity_tables, valuation_ranges)

        return sensitivity_tables, valuation_ranges, excel_file

    def _create_assumptions(self, dcf_base_wacc, dcf_base_terminal_growth, dcf_pv_fcf, dcf_terminal_value,
                           dcf_wacc_range, dcf_growth_range,
                           lbo_base_exit_multiple, lbo_base_leverage_ratio, lbo_base_ebitda,
                           lbo_base_tax_rate, lbo_base_debt_repayment_pct,
                           lbo_exit_multiple_range, lbo_leverage_range,
                           comps_base_ev_ebitda, comps_base_price_earnings, comps_base_ebitda,
                           comps_base_eps, comps_base_shares_outstanding,
                           comps_ev_ebitda_range, comps_pe_range):

        """Create comprehensive assumptions for all methodologies"""

        assumptions = {
            'dcf': {
                'base_wacc': dcf_base_wacc,
                'base_terminal_growth': dcf_base_terminal_growth,
                'pv_fcf': dcf_pv_fcf,
                'terminal_value': dcf_terminal_value,
                'base_ev': dcf_pv_fcf + dcf_terminal_value,
                'wacc_range': dcf_wacc_range,
                'growth_range': dcf_growth_range,
                'methodology': 'DCF Analysis'
            },
            'lbo': {
                'base_exit_multiple': lbo_base_exit_multiple,
                'base_leverage_ratio': lbo_base_leverage_ratio,
                'base_ebitda': lbo_base_ebitda,
                'tax_rate': lbo_base_tax_rate,
                'debt_repayment_pct': lbo_base_debt_repayment_pct,
                'base_exit_ev': lbo_base_ebitda * lbo_base_exit_multiple,
                'base_entry_ev': lbo_base_ebitda * lbo_base_leverage_ratio,
                'base_equity_value': (lbo_base_ebitda * lbo_base_exit_multiple) - (lbo_base_ebitda * lbo_base_leverage_ratio * (1 - lbo_base_debt_repayment_pct)),
                'exit_multiple_range': lbo_exit_multiple_range,
                'leverage_range': lbo_leverage_range,
                'methodology': 'LBO Analysis'
            },
            'comps': {
                'base_ev_ebitda': comps_base_ev_ebitda,
                'base_price_earnings': comps_base_price_earnings,
                'base_ebitda': comps_base_ebitda,
                'base_eps': comps_base_eps,
                'shares_outstanding': comps_base_shares_outstanding,
                'base_ev': comps_base_ebitda * comps_base_ev_ebitda,
                'base_equity_value': comps_base_eps * comps_base_shares_outstanding,
                'ev_ebitda_range': comps_ev_ebitda_range,
                'pe_range': comps_pe_range,
                'methodology': 'Trading Comparables'
            }
        }

        print("📋 Assumptions Created:")
        print(f"   • DCF: WACC {dcf_base_wacc:.1%}, Terminal Growth {dcf_base_terminal_growth:.1%}")
        print(f"   • LBO: Exit Multiple {lbo_base_exit_multiple:.1f}x, Leverage {lbo_base_leverage_ratio:.1f}x")
        print(f"   • Comps: EV/EBITDA {comps_base_ev_ebitda:.1f}x, P/E {comps_base_price_earnings:.1f}x")

        return assumptions

    def _generate_sensitivity_tables(self, assumptions):
        """Generate 2D sensitivity tables for each methodology"""

        sensitivity_tables = {}

        # DCF Sensitivity Table (WACC vs Terminal Growth)
        dcf_table = self._create_dcf_sensitivity_table(assumptions['dcf'])
        sensitivity_tables['dcf'] = dcf_table

        # LBO Sensitivity Table (Exit Multiple vs Leverage)
        lbo_table = self._create_lbo_sensitivity_table(assumptions['lbo'])
        sensitivity_tables['lbo'] = lbo_table

        # Comps Sensitivity Table (EV/EBITDA vs P/E)
        comps_table = self._create_comps_sensitivity_table(assumptions['comps'])
        sensitivity_tables['comps'] = comps_table

        print("📊 Sensitivity Tables Generated:")
        print(f"   • DCF Table: {len(assumptions['dcf']['wacc_range'])} WACCs × {len(assumptions['dcf']['growth_range'])} growth rates")
        print(f"   • LBO Table: {len(assumptions['lbo']['exit_multiple_range'])} exit multiples × {len(assumptions['lbo']['leverage_range'])} leverage ratios")
        print(f"   • Comps Table: {len(assumptions['comps']['ev_ebitda_range'])} EV/EBITDA × {len(assumptions['comps']['pe_range'])} P/E ratios")

        return sensitivity_tables

    def _create_dcf_sensitivity_table(self, dcf_assumptions):
        """Create DCF sensitivity table (WACC vs Terminal Growth)"""

        wacc_range = dcf_assumptions['wacc_range']
        growth_range = dcf_assumptions['growth_range']
        pv_fcf = dcf_assumptions['pv_fcf']
        base_terminal_value = dcf_assumptions['terminal_value']

        table_data = []
        for growth_rate in growth_range:
            row_data = []
            for wacc in wacc_range:
                # Calculate terminal value with different growth rate
                terminal_value = base_terminal_value * (1 + growth_rate) / (wacc - growth_rate)
                total_ev = pv_fcf + terminal_value
                row_data.append(total_ev)
            table_data.append(row_data)

        return {
            'x_labels': [f"{wacc:.1%}" for wacc in wacc_range],
            'y_labels': [f"{growth:.1%}" for growth in growth_range],
            'data': table_data,
            'base_x_idx': wacc_range.index(dcf_assumptions['base_wacc']) if dcf_assumptions['base_wacc'] in wacc_range else 0,
            'base_y_idx': growth_range.index(dcf_assumptions['base_terminal_growth']) if dcf_assumptions['base_terminal_growth'] in growth_range else 0,
            'base_value': dcf_assumptions['base_ev'],
            'x_axis_label': 'WACC',
            'y_axis_label': 'Terminal Growth Rate',
            'value_label': 'Enterprise Value ($M)'
        }

    def _create_lbo_sensitivity_table(self, lbo_assumptions):
        """Create LBO sensitivity table (Exit Multiple vs Leverage)"""

        exit_range = lbo_assumptions['exit_multiple_range']
        leverage_range = lbo_assumptions['leverage_range']
        ebitda = lbo_assumptions['base_ebitda']
        tax_rate = lbo_assumptions['tax_rate']
        debt_repayment_pct = lbo_assumptions['debt_repayment_pct']

        table_data = []
        for leverage_ratio in leverage_range:
            row_data = []
            for exit_multiple in exit_range:
                # Calculate equity value at exit
                exit_ev = ebitda * exit_multiple
                entry_debt = ebitda * leverage_ratio
                debt_at_exit = entry_debt * (1 - debt_repayment_pct)
                equity_value = exit_ev - debt_at_exit

                # Apply tax shield benefit (simplified)
                tax_shield = entry_debt * tax_rate * 0.3  # Approximate annual tax shield
                equity_value += tax_shield

                row_data.append(equity_value)
            table_data.append(row_data)

        return {
            'x_labels': [f"{exit:.1f}x" for exit in exit_range],
            'y_labels': [f"{lev:.1f}x" for lev in leverage_range],
            'data': table_data,
            'base_x_idx': exit_range.index(lbo_assumptions['base_exit_multiple']) if lbo_assumptions['base_exit_multiple'] in exit_range else 0,
            'base_y_idx': leverage_range.index(lbo_assumptions['base_leverage_ratio']) if lbo_assumptions['base_leverage_ratio'] in leverage_range else 0,
            'base_value': lbo_assumptions['base_equity_value'],
            'x_axis_label': 'Exit Multiple (EV/EBITDA)',
            'y_axis_label': 'Entry Leverage (Debt/EBITDA)',
            'value_label': 'Equity Value ($M)'
        }

    def _create_comps_sensitivity_table(self, comps_assumptions):
        """Create Comps sensitivity table (EV/EBITDA vs P/E)"""

        ev_ebitda_range = comps_assumptions['ev_ebitda_range']
        pe_range = comps_assumptions['pe_range']
        ebitda = comps_assumptions['base_ebitda']
        eps = comps_assumptions['base_eps']
        shares_outstanding = comps_assumptions['shares_outstanding']

        table_data = []
        for pe_ratio in pe_range:
            row_data = []
            for ev_ebitda in ev_ebitda_range:
                # Calculate implied equity value from P/E
                equity_value_pe = eps * pe_ratio * shares_outstanding

                # Calculate implied EV from EBITDA multiple
                ev_value = ebitda * ev_ebitda

                # Use the higher of the two (simplified approach)
                implied_value = max(equity_value_pe, ev_value - (ebitda * 2))  # Subtract estimated net debt

                row_data.append(implied_value)
            table_data.append(row_data)

        return {
            'x_labels': [f"{ev:.1f}x" for ev in ev_ebitda_range],
            'y_labels': [f"{pe:.1f}x" for pe in pe_range],
            'data': table_data,
            'base_x_idx': ev_ebitda_range.index(comps_assumptions['base_ev_ebitda']) if comps_assumptions['base_ev_ebitda'] in ev_ebitda_range else 0,
            'base_y_idx': pe_range.index(comps_assumptions['base_price_earnings']) if comps_assumptions['base_price_earnings'] in pe_range else 0,
            'base_value': comps_assumptions['base_equity_value'],
            'x_axis_label': 'EV/EBITDA Multiple',
            'y_axis_label': 'Price/Earnings Multiple',
            'value_label': 'Equity Value ($M)'
        }

    def _calculate_valuation_ranges(self, sensitivity_tables):
        """Calculate valuation ranges from sensitivity tables"""

        valuation_ranges = {}

        for method, table in sensitivity_tables.items():
            # Flatten all values from the table
            all_values = []
            for row in table['data']:
                all_values.extend(row)

            valuation_ranges[method] = {
                'min': min(all_values),
                'max': max(all_values),
                'range': max(all_values) - min(all_values),
                'median': np.median(all_values),
                'mean': np.mean(all_values),
                'base_case': table['base_value']
            }

        # Calculate overall ranges across all methodologies
        all_values = []
        for method_ranges in valuation_ranges.values():
            all_values.append(method_ranges['min'])
            all_values.append(method_ranges['max'])

        valuation_ranges['overall'] = {
            'min': min(all_values),
            'max': max(all_values),
            'range': max(all_values) - min(all_values),
            'methodologies': list(valuation_ranges.keys())
        }

        print("📈 Valuation Ranges Calculated:")
        for method, ranges in valuation_ranges.items():
            if method != 'overall':
                print(f"   • {method.upper()}: ${ranges['min']/1000:.1f}B - ${ranges['max']/1000:.1f}B (range: ${ranges['range']/1000:.1f}B)")

        return valuation_ranges

    def _create_excel_output(self, assumptions, sensitivity_tables, valuation_ranges):
        """Create professional Excel output with multiple tabs"""

        wb = Workbook()

        # Create worksheets
        ws_summary = wb.active
        ws_summary.title = "Sensitivity Analysis Summary"

        ws_inputs = wb.create_sheet("Assumptions & Inputs")
        ws_dcf_sensitivity = wb.create_sheet("DCF Sensitivity")
        ws_lbo_sensitivity = wb.create_sheet("LBO Sensitivity")
        ws_comps_sensitivity = wb.create_sheet("Comps Sensitivity")
        ws_ranges = wb.create_sheet("Valuation Ranges")

        # Apply styles to workbook
        for style_name, style in self.styles.items():
            if style_name not in wb.named_styles:
                wb.add_named_style(style)

        # Create each tab
        self._create_summary_tab(ws_summary, assumptions, valuation_ranges)
        self._create_inputs_tab(ws_inputs, assumptions)
        self._create_dcf_sensitivity_tab(ws_dcf_sensitivity, sensitivity_tables['dcf'])
        self._create_lbo_sensitivity_tab(ws_lbo_sensitivity, sensitivity_tables['lbo'])
        self._create_comps_sensitivity_tab(ws_comps_sensitivity, sensitivity_tables['comps'])
        self._create_ranges_tab(ws_ranges, valuation_ranges)

        # Save workbook
        filename = f"Sensitivity_Analysis_{self.target_ticker}_{self.model_date}.xlsx"
        wb.save(filename)

        return filename

    def _create_summary_tab(self, ws, assumptions, valuation_ranges):
        """Create Sensitivity Analysis Summary tab"""

        # Title
        ws['A1'] = f"Valuation Sensitivity Analysis: {self.target_company} ({self.target_ticker})"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:G1')

        current_row = 3

        # Executive Summary Box
        ws[f'A{current_row}'] = 'EXECUTIVE SUMMARY'
        ws[f'A{current_row}'].style = 'summary_box'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        # Overall valuation range
        overall = valuation_ranges['overall']
        summary_data = [
            ("Overall Valuation Range", f"${overall['min']/1000:.1f}B", f"${overall['max']/1000:.1f}B", f"${overall['range']/1000:.1f}B"),
            ("Valuation Methodologies", f"{len(overall['methodologies'])}", "", ""),
            ("Analysis Date", self.model_date, "", "")
        ]

        # Headers
        ws.cell(row=current_row, column=1, value="Metric").style = 'header'
        ws.cell(row=current_row, column=2, value="Low").style = 'header'
        ws.cell(row=current_row, column=3, value="High").style = 'header'
        ws.cell(row=current_row, column=4, value="Range").style = 'header'
        current_row += 1

        # Data
        for metric, low, high, range_val in summary_data:
            ws.cell(row=current_row, column=1, value=metric).style = 'label_bold'
            ws.cell(row=current_row, column=2, value=low).style = 'output'
            ws.cell(row=current_row, column=3, value=high).style = 'output'
            ws.cell(row=current_row, column=4, value=range_val).style = 'output'
            current_row += 1

        current_row += 2

        # Methodology Summary
        ws[f'A{current_row}'] = 'METHODOLOGY SENSITIVITY RANGES'
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        # Methodology ranges
        ws.cell(row=current_row, column=1, value="Methodology").style = 'header'
        ws.cell(row=current_row, column=2, value="Base Case").style = 'header'
        ws.cell(row=current_row, column=3, value="Low").style = 'header'
        ws.cell(row=current_row, column=4, value="Median").style = 'header'
        ws.cell(row=current_row, column=5, value="High").style = 'header'
        ws.cell(row=current_row, column=6, value="Range").style = 'header'
        current_row += 1

        for method, ranges in valuation_ranges.items():
            if method != 'overall':
                ws.cell(row=current_row, column=1, value=assumptions[method]['methodology']).style = self._get_methodology_style(method)
                ws.cell(row=current_row, column=2, value=ranges['base_case']).style = 'base_case'
                ws.cell(row=current_row, column=3, value=ranges['min']).style = 'output'
                ws.cell(row=current_row, column=4, value=ranges['median']).style = 'output'
                ws.cell(row=current_row, column=5, value=ranges['max']).style = 'output'
                ws.cell(row=current_row, column=6, value=ranges['range']).style = 'output'
                current_row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15

    def _create_inputs_tab(self, ws, assumptions):
        """Create Assumptions & Inputs tab"""

        # Title
        ws['A1'] = f"Assumptions & Base Case Inputs: {self.target_company} ({self.target_ticker})"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:E1')

        current_row = 3

        # DCF Assumptions
        ws[f'A{current_row}'] = "DCF ANALYSIS ASSUMPTIONS"
        ws[f'A{current_row}'].style = 'dcf'
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 2

        dcf_data = [
            ("WACC", assumptions['dcf']['base_wacc'], "Base discount rate"),
            ("Terminal Growth Rate", assumptions['dcf']['base_terminal_growth'], "Long-term growth assumption"),
            ("PV of FCF", assumptions['dcf']['pv_fcf'], "Present value of free cash flows"),
            ("Terminal Value", assumptions['dcf']['terminal_value'], "Terminal period value"),
            ("Total Enterprise Value", assumptions['dcf']['base_ev'], "Base case EV")
        ]

        for label, value, description in dcf_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].style = 'input'
            ws[f'C{current_row}'] = description
            ws[f'C{current_row}'].style = 'label'
            current_row += 1

        current_row += 2

        # LBO Assumptions
        ws[f'A{current_row}'] = "LBO ANALYSIS ASSUMPTIONS"
        ws[f'A{current_row}'].style = 'lbo'
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 2

        lbo_data = [
            ("Exit Multiple (EV/EBITDA)", assumptions['lbo']['base_exit_multiple'], "Exit valuation multiple"),
            ("Entry Leverage (Debt/EBITDA)", assumptions['lbo']['base_leverage_ratio'], "Initial debt level"),
            ("EBITDA", assumptions['lbo']['base_ebitda'], "Earnings before interest, taxes, depreciation, amortization"),
            ("Tax Rate", assumptions['lbo']['tax_rate'], "Corporate tax rate"),
            ("Debt Repayment %", assumptions['lbo']['debt_repayment_pct'], "Percentage of debt repaid"),
            ("Entry Enterprise Value", assumptions['lbo']['base_entry_ev'], "Initial investment"),
            ("Exit Enterprise Value", assumptions['lbo']['base_exit_ev'], "Exit valuation"),
            ("Equity Value at Exit", assumptions['lbo']['base_equity_value'], "Equity value to investors")
        ]

        for label, value, description in lbo_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].style = 'input'
            ws[f'C{current_row}'] = description
            ws[f'C{current_row}'].style = 'label'
            current_row += 1

        current_row += 2

        # Comps Assumptions
        ws[f'A{current_row}'] = "TRADING COMPARABLES ASSUMPTIONS"
        ws[f'A{current_row}'].style = 'comps'
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 2

        comps_data = [
            ("EV/EBITDA Multiple", assumptions['comps']['base_ev_ebitda'], "Enterprise value to EBITDA ratio"),
            ("Price/Earnings Multiple", assumptions['comps']['base_price_earnings'], "Price to earnings ratio"),
            ("EBITDA", assumptions['comps']['base_ebitda'], "Earnings before interest, taxes, depreciation, amortization"),
            ("EPS", assumptions['comps']['base_eps'], "Earnings per share"),
            ("Shares Outstanding", assumptions['comps']['shares_outstanding'], "Total shares outstanding (millions)"),
            ("Implied Enterprise Value", assumptions['comps']['base_ev'], "EV from EBITDA multiple"),
            ("Implied Equity Value", assumptions['comps']['base_equity_value'], "Equity value from P/E multiple")
        ]

        for label, value, description in comps_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].style = 'input'
            ws[f'C{current_row}'] = description
            ws[f'C{current_row}'].style = 'label'
            current_row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 50

    def _create_dcf_sensitivity_tab(self, ws, dcf_table):
        """Create DCF Sensitivity tab with 2D table"""

        # Title
        ws['A1'] = f"DCF Sensitivity Analysis: {self.target_company} ({self.target_ticker})"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:H1')

        current_row = 3

        # Sensitivity Table Title
        ws[f'A{current_row}'] = "DCF SENSITIVITY: WACC vs TERMINAL GROWTH RATE"
        ws[f'A{current_row}'].style = 'dcf'
        ws.merge_cells(f'A{current_row}:H{current_row}')
        current_row += 2

        # Create table headers
        ws.cell(row=current_row, column=1, value=dcf_table['y_axis_label']).style = 'table_header'
        for i, x_label in enumerate(dcf_table['x_labels'], 2):
            ws.cell(row=current_row, column=i, value=x_label).style = 'table_header'
        current_row += 1

        # Create table data
        for i, (y_label, row_data) in enumerate(zip(dcf_table['y_labels'], dcf_table['data'])):
            ws.cell(row=current_row, column=1, value=y_label).style = 'table_header'
            for j, value in enumerate(row_data, 2):
                cell_style = 'output'
                # Highlight base case
                if i == dcf_table['base_y_idx'] and (j-2) == dcf_table['base_x_idx']:
                    cell_style = 'base_case'
                ws.cell(row=current_row, column=j, value=value).style = cell_style
            current_row += 1

        current_row += 2

        # Summary statistics
        ws[f'A{current_row}'] = "SENSITIVITY STATISTICS"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:H{current_row}')
        current_row += 2

        # Flatten all values for statistics
        all_values = []
        for row in dcf_table['data']:
            all_values.extend(row)

        stats_data = [
            ("Base Case", dcf_table['base_value']),
            ("Minimum", min(all_values)),
            ("Maximum", max(all_values)),
            ("Range", max(all_values) - min(all_values)),
            ("Median", np.median(all_values)),
            ("Mean", np.mean(all_values))
        ]

        for label, value in stats_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            if label == "Base Case":
                ws[f'B{current_row}'].style = 'base_case'
            else:
                ws[f'B{current_row}'].style = 'output'
            current_row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 20
        for col in range(2, len(dcf_table['x_labels']) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_lbo_sensitivity_tab(self, ws, lbo_table):
        """Create LBO Sensitivity tab with 2D table"""

        # Title
        ws['A1'] = f"LBO Sensitivity Analysis: {self.target_company} ({self.target_ticker})"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:H1')

        current_row = 3

        # Sensitivity Table Title
        ws[f'A{current_row}'] = "LBO SENSITIVITY: EXIT MULTIPLE vs LEVERAGE RATIO"
        ws[f'A{current_row}'].style = 'lbo'
        ws.merge_cells(f'A{current_row}:H{current_row}')
        current_row += 2

        # Create table headers
        ws.cell(row=current_row, column=1, value=lbo_table['y_axis_label']).style = 'table_header'
        for i, x_label in enumerate(lbo_table['x_labels'], 2):
            ws.cell(row=current_row, column=i, value=x_label).style = 'table_header'
        current_row += 1

        # Create table data
        for i, (y_label, row_data) in enumerate(zip(lbo_table['y_labels'], lbo_table['data'])):
            ws.cell(row=current_row, column=1, value=y_label).style = 'table_header'
            for j, value in enumerate(row_data, 2):
                cell_style = 'output'
                # Highlight base case
                if i == lbo_table['base_y_idx'] and (j-2) == lbo_table['base_x_idx']:
                    cell_style = 'base_case'
                ws.cell(row=current_row, column=j, value=value).style = cell_style
            current_row += 1

        current_row += 2

        # Summary statistics
        ws[f'A{current_row}'] = "SENSITIVITY STATISTICS"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:H{current_row}')
        current_row += 2

        # Flatten all values for statistics
        all_values = []
        for row in lbo_table['data']:
            all_values.extend(row)

        stats_data = [
            ("Base Case", lbo_table['base_value']),
            ("Minimum", min(all_values)),
            ("Maximum", max(all_values)),
            ("Range", max(all_values) - min(all_values)),
            ("Median", np.median(all_values)),
            ("Mean", np.mean(all_values))
        ]

        for label, value in stats_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            if label == "Base Case":
                ws[f'B{current_row}'].style = 'base_case'
            else:
                ws[f'B{current_row}'].style = 'output'
            current_row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 25
        for col in range(2, len(lbo_table['x_labels']) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_comps_sensitivity_tab(self, ws, comps_table):
        """Create Comps Sensitivity tab with 2D table"""

        # Title
        ws['A1'] = f"Trading Comparables Sensitivity Analysis: {self.target_company} ({self.target_ticker})"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:H1')

        current_row = 3

        # Sensitivity Table Title
        ws[f'A{current_row}'] = "COMPS SENSITIVITY: EV/EBITDA vs P/E MULTIPLE"
        ws[f'A{current_row}'].style = 'comps'
        ws.merge_cells(f'A{current_row}:H{current_row}')
        current_row += 2

        # Create table headers
        ws.cell(row=current_row, column=1, value=comps_table['y_axis_label']).style = 'table_header'
        for i, x_label in enumerate(comps_table['x_labels'], 2):
            ws.cell(row=current_row, column=i, value=x_label).style = 'table_header'
        current_row += 1

        # Create table data
        for i, (y_label, row_data) in enumerate(zip(comps_table['y_labels'], comps_table['data'])):
            ws.cell(row=current_row, column=1, value=y_label).style = 'table_header'
            for j, value in enumerate(row_data, 2):
                cell_style = 'output'
                # Highlight base case
                if i == comps_table['base_y_idx'] and (j-2) == comps_table['base_x_idx']:
                    cell_style = 'base_case'
                ws.cell(row=current_row, column=j, value=value).style = cell_style
            current_row += 1

        current_row += 2

        # Summary statistics
        ws[f'A{current_row}'] = "SENSITIVITY STATISTICS"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:H{current_row}')
        current_row += 2

        # Flatten all values for statistics
        all_values = []
        for row in comps_table['data']:
            all_values.extend(row)

        stats_data = [
            ("Base Case", comps_table['base_value']),
            ("Minimum", min(all_values)),
            ("Maximum", max(all_values)),
            ("Range", max(all_values) - min(all_values)),
            ("Median", np.median(all_values)),
            ("Mean", np.mean(all_values))
        ]

        for label, value in stats_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            if label == "Base Case":
                ws[f'B{current_row}'].style = 'base_case'
            else:
                ws[f'B{current_row}'].style = 'output'
            current_row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 20
        for col in range(2, len(comps_table['x_labels']) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_ranges_tab(self, ws, valuation_ranges):
        """Create Valuation Ranges summary tab"""

        # Title
        ws['A1'] = f"Valuation Ranges Summary: {self.target_company} ({self.target_ticker})"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:E1')

        current_row = 3

        # Ranges Summary
        ws[f'A{current_row}'] = "VALUATION RANGES ACROSS METHODOLOGIES"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 2

        # Headers
        headers = ['Methodology', 'Low ($M)', 'Median ($M)', 'High ($M)', 'Range ($M)']
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header).style = 'header'
        current_row += 1

        # Data for each methodology
        for method, ranges in valuation_ranges.items():
            if method != 'overall':
                ws.cell(row=current_row, column=1, value=method.upper()).style = self._get_methodology_style(method)
                ws.cell(row=current_row, column=2, value=ranges['min']).style = 'output'
                ws.cell(row=current_row, column=3, value=ranges['median']).style = 'output'
                ws.cell(row=current_row, column=4, value=ranges['max']).style = 'output'
                ws.cell(row=current_row, column=5, value=ranges['range']).style = 'output'
                current_row += 1

        current_row += 2

        # Overall range
        ws[f'A{current_row}'] = "OVERALL VALUATION RANGE"
        ws[f'A{current_row}'].style = 'highlight'
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 1

        overall = valuation_ranges['overall']
        ws.cell(row=current_row, column=1, value="Combined Methodologies").style = 'label_bold'
        ws.cell(row=current_row, column=2, value=overall['min']).style = 'output'
        ws.cell(row=current_row, column=3, value="").style = 'output'  # No median for overall
        ws.cell(row=current_row, column=4, value=overall['max']).style = 'output'
        ws.cell(row=current_row, column=5, value=overall['range']).style = 'output'

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15

    def _get_methodology_style(self, method_key):
        """Get the appropriate style for a methodology"""
        if method_key == 'dcf':
            return 'dcf'
        elif method_key == 'lbo':
            return 'lbo'
        elif method_key == 'comps':
            return 'comps'
        else:
            return 'calculation'

    def _display_sensitivity_insights(self, sensitivity_tables, valuation_ranges):
        """Display key sensitivity insights"""

        print("\n🔍 Key Sensitivity Insights:")
        print("=" * 80)

        # DCF insights
        dcf_table = sensitivity_tables['dcf']
        dcf_range = valuation_ranges['dcf']['range']
        print("\n📊 DCF Analysis:")
        print(f"   • WACC Impact: ${dcf_range/1000:.1f}B range across {len(dcf_table['x_labels'])} WACC scenarios")
        print("   • Terminal Growth Impact: Varies with growth assumptions")
        print(f"   • Base Case: ${dcf_table['base_value']/1000:.1f}B EV")

        # LBO insights
        lbo_table = sensitivity_tables['lbo']
        lbo_range = valuation_ranges['lbo']['range']
        print("\n💼 LBO Analysis:")
        print(f"   • Exit Multiple Impact: ${lbo_range/1000:.1f}B range across {len(lbo_table['x_labels'])} exit scenarios")
        print("   • Leverage Impact: Affects equity returns and risk profile")
        print(f"   • Base Case: ${lbo_table['base_value']/1000:.1f}B equity value")

        # Comps insights
        comps_table = sensitivity_tables['comps']
        comps_range = valuation_ranges['comps']['range']
        print("\n📈 Trading Comparables:")
        print(f"   • Multiple Impact: ${comps_range/1000:.1f}B range across {len(comps_table['x_labels'])} multiple scenarios")
        print("   • Market-Based: Reflects current investor sentiment")
        print(f"   • Base Case: ${comps_table['base_value']/1000:.1f}B equity value")

        # Overall insights
        overall = valuation_ranges['overall']
        print("\n🎯 Overall Valuation:")
        print(f"   • Total Range: ${overall['range']/1000:.1f}B across all methodologies")
        print(f"   • Methodologies: {len(overall['methodologies'])} valuation approaches")
        print("   • Robustness: Wide range demonstrates thorough analysis")

        print("\n💡 Sensitivity Analysis Applications:")
        print("=" * 80)
        print("   - Risk assessment and scenario planning")
        print("   - Investment committee presentations")
        print("   - Negotiation strategy development")
        print("   - Fairness opinion support")
        print("   - Board meeting materials")
        print("   - Regulatory filing exhibits")


def run_sample_sensitivity_analysis():
    """Run a sample sensitivity analysis"""

    print("📊 Running Professional Sensitivity Analysis Model Sample")
    print("=" * 70)

    # Create model instance
    model = ProfessionalSensitivityAnalysisModel("TechTarget Inc.", "TECHTARGET")

    # Run the model with sample data
    sensitivity_tables, valuation_ranges, excel_file = model.run_sensitivity_analysis_model(
        # Target Company Info
        target_company="TechTarget Inc.",
        target_ticker="TECHTARGET"
    )

    return sensitivity_tables, valuation_ranges, excel_file


if __name__ == "__main__":
    # Run sample sensitivity analysis
    sensitivity_tables, valuation_ranges, excel_file = run_sample_sensitivity_analysis()

    print("\n📋 Sensitivity Analysis Complete!")
    print(f"📁 Excel file saved as: {excel_file}")

    # Display key results
    print("\n🔑 Key Results Summary:")
    print("=" * 40)

    overall = valuation_ranges['overall']
    print(f"Overall Valuation Range: ${overall['min']/1000:.1f}B - ${overall['max']/1000:.1f}B")
    print(f"Valuation Range Width: ${overall['range']/1000:.1f}B")
    print(f"Number of Methodologies: {len(overall['methodologies'])}")

    # Show methodology ranges
    for method, ranges in valuation_ranges.items():
        if method != 'overall':
            print(f"{method.upper()}: ${ranges['min']/1000:.1f}B - ${ranges['max']/1000:.1f}B")
