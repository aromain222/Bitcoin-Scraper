#!/usr/bin/env python3
"""
Custom Inputs Module for Financial Models
Allows users to override research-based assumptions with custom values
"""

import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

class CustomInputsManager:
    """Manages custom user inputs for financial models"""
    
    def __init__(self):
        self.custom_input_sections = {
            'growth_assumptions': {
                'title': 'CUSTOM GROWTH ASSUMPTIONS',
                'description': 'Override revenue growth rates and terminal growth',
                'fields': [
                    ('Revenue Growth Year 1', 'percentage', 0.15, 'Custom revenue growth for year 1'),
                    ('Revenue Growth Year 2', 'percentage', 0.12, 'Custom revenue growth for year 2'),
                    ('Revenue Growth Year 3', 'percentage', 0.10, 'Custom revenue growth for year 3'),
                    ('Revenue Growth Year 4', 'percentage', 0.08, 'Custom revenue growth for year 4'),
                    ('Revenue Growth Year 5', 'percentage', 0.06, 'Custom revenue growth for year 5'),
                    ('Terminal Growth Rate', 'percentage', 0.025, 'Long-term growth rate after year 5')
                ]
            },
            'profitability_assumptions': {
                'title': 'CUSTOM PROFITABILITY ASSUMPTIONS',
                'description': 'Override margins and profitability metrics',
                'fields': [
                    ('EBITDA Margin', 'percentage', 0.25, 'Custom EBITDA margin'),
                    ('Gross Margin', 'percentage', 0.60, 'Custom gross margin'),
                    ('R&D as % of Revenue', 'percentage', 0.10, 'Custom R&D spending'),
                    ('SG&A as % of Revenue', 'percentage', 0.20, 'Custom SG&A spending'),
                    ('Tax Rate', 'percentage', 0.21, 'Custom effective tax rate'),
                    ('ROA Target', 'percentage', 0.10, 'Target return on assets'),
                    ('ROE Target', 'percentage', 0.15, 'Target return on equity'),
                    ('ROIC Target', 'percentage', 0.12, 'Target return on invested capital')
                ]
            },
            'capital_structure': {
                'title': 'CUSTOM CAPITAL STRUCTURE',
                'description': 'Override debt and capital structure assumptions',
                'fields': [
                    ('Target Debt/Equity Ratio', 'ratio', 0.40, 'Target debt to equity ratio'),
                    ('Cost of Debt', 'percentage', 0.065, 'Custom cost of debt'),
                    ('Cost of Equity', 'percentage', 0.12, 'Custom cost of equity'),
                    ('Dividend Payout Ratio', 'percentage', 0.25, 'Target dividend payout'),
                    ('Share Repurchase % of FCF', 'percentage', 0.25, 'Share repurchase as % of FCF'),
                    ('Interest Coverage Target', 'ratio', 5.0, 'Target interest coverage ratio'),
                    ('Debt/EBITDA Target', 'ratio', 2.5, 'Target debt to EBITDA ratio')
                ]
            },
            'working_capital': {
                'title': 'CUSTOM WORKING CAPITAL ASSUMPTIONS',
                'description': 'Override working capital and operational metrics',
                'fields': [
                    ('Days Sales in Inventory', 'days', 45, 'Custom inventory turnover'),
                    ('Days Sales Outstanding', 'days', 45, 'Custom receivables turnover'),
                    ('Days Payable Outstanding', 'days', 45, 'Custom payables turnover'),
                    ('Prepaid Expenses % of Revenue', 'percentage', 0.02, 'Prepaid expenses ratio'),
                    ('Other Current Assets % of Revenue', 'percentage', 0.03, 'Other current assets ratio'),
                    ('Accrued Expenses % of Revenue', 'percentage', 0.05, 'Accrued expenses ratio'),
                    ('Other Current Liabilities % of Revenue', 'percentage', 0.03, 'Other current liabilities ratio')
                ]
            },
            'capital_expenditure': {
                'title': 'CUSTOM CAPITAL EXPENDITURE ASSUMPTIONS',
                'description': 'Override CapEx and depreciation assumptions',
                'fields': [
                    ('CapEx as % of Revenue', 'percentage', 0.08, 'Capital expenditure ratio'),
                    ('Depreciation as % of PP&E', 'percentage', 0.10, 'Depreciation rate'),
                    ('Amortization as % of Intangibles', 'percentage', 0.12, 'Amortization rate'),
                    ('Maintenance CapEx % of Revenue', 'percentage', 0.04, 'Maintenance capital expenditure'),
                    ('Growth CapEx % of Revenue', 'percentage', 0.04, 'Growth capital expenditure')
                ]
            },
            'valuation_metrics': {
                'title': 'CUSTOM VALUATION METRICS',
                'description': 'Override valuation multiples and exit assumptions',
                'fields': [
                    ('EV/EBITDA Multiple', 'multiple', 15.0, 'Custom EV/EBITDA multiple'),
                    ('EV/Revenue Multiple', 'multiple', 3.0, 'Custom EV/Revenue multiple'),
                    ('P/E Ratio', 'multiple', 18.0, 'Custom P/E ratio'),
                    ('P/B Ratio', 'multiple', 2.5, 'Custom P/B ratio'),
                    ('P/S Ratio', 'multiple', 2.5, 'Custom P/S ratio'),
                    ('Exit Year', 'integer', 5, 'Year to apply exit multiple'),
                    ('Exit Multiple Type', 'dropdown', 'EV/EBITDA', 'Type of exit multiple to use')
                ]
            },
            'scenario_analysis': {
                'title': 'CUSTOM SCENARIO ANALYSIS',
                'description': 'Define custom scenarios for sensitivity analysis',
                'fields': [
                    ('Base Case Revenue Growth', 'percentage', 0.10, 'Base case revenue growth'),
                    ('Bull Case Revenue Growth', 'percentage', 0.15, 'Bull case revenue growth'),
                    ('Bear Case Revenue Growth', 'percentage', 0.05, 'Bear case revenue growth'),
                    ('Base Case EBITDA Margin', 'percentage', 0.25, 'Base case EBITDA margin'),
                    ('Bull Case EBITDA Margin', 'percentage', 0.30, 'Bull case EBITDA margin'),
                    ('Bear Case EBITDA Margin', 'percentage', 0.20, 'Bear case EBITDA margin'),
                    ('Base Case WACC', 'percentage', 0.10, 'Base case WACC'),
                    ('Bull Case WACC', 'percentage', 0.08, 'Bull case WACC'),
                    ('Bear Case WACC', 'percentage', 0.12, 'Bear case WACC')
                ]
            },
            'private_company': {
                'title': 'PRIVATE COMPANY CUSTOM INPUTS',
                'description': 'Additional inputs for private company modeling',
                'fields': [
                    ('Current Revenue', 'currency', 1000000, 'Current annual revenue'),
                    ('Current EBITDA', 'currency', 200000, 'Current annual EBITDA'),
                    ('Current Net Income', 'currency', 150000, 'Current annual net income'),
                    ('Current Cash Balance', 'currency', 500000, 'Current cash and equivalents'),
                    ('Current Debt Balance', 'currency', 800000, 'Current total debt'),
                    ('Shares Outstanding', 'integer', 1000000, 'Number of shares outstanding'),
                    ('Current Share Price', 'currency', 10.00, 'Current share price (if applicable)'),
                    ('Management Projections Available', 'boolean', False, 'Whether management projections are available'),
                    ('Industry Comparable Companies', 'text', '', 'List of comparable public companies')
                ]
            }
        }
    
    def create_custom_inputs_sheet(self, wb, company_name, research_assumptions=None):
        """Create a comprehensive custom inputs sheet"""
        ws = wb.create_sheet("Custom Inputs")
        
        # Header
        ws['A1'] = f"CUSTOM INPUTS - {company_name.upper()}"
        ws['A1'].font = Font(size=16, bold=True, color='1F4E79')
        ws.merge_cells('A1:H1')
        
        # Instructions
        ws['A3'] = "INSTRUCTIONS:"
        ws['A3'].font = Font(bold=True, size=12)
        ws['A3'].fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        
        instructions = [
            "1. Enter your custom values in the 'Custom Value' column to override research-based assumptions",
            "2. Leave 'Custom Value' blank to use the research-based default",
            "3. All changes will be automatically applied to the financial model",
            "4. Use the 'Notes' column to document your assumptions",
            "5. The 'Research Default' column shows the original research-based values"
        ]
        
        for i, instruction in enumerate(instructions, 4):
            ws[f'A{i}'] = instruction
            ws[f'A{i}'].font = Font(size=10)
        
        current_row = 10
        
        # Create each section
        for section_key, section_data in self.custom_input_sections.items():
            current_row = self._create_section(ws, section_data, current_row, research_assumptions)
            current_row += 2  # Add spacing between sections
        
        # Add summary section
        current_row = self._create_summary_section(ws, current_row)
        
        # Apply formatting
        self._apply_sheet_formatting(ws)
        
        return ws
    
    def _create_section(self, ws, section_data, start_row, research_assumptions):
        """Create a section of custom inputs"""
        # Section header
        ws[f'A{start_row}'] = section_data['title']
        ws[f'A{start_row}'].font = Font(bold=True, size=12, color='FFFFFF')
        ws[f'A{start_row}'].fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        ws.merge_cells(f'A{start_row}:H{start_row}')
        
        # Section description
        ws[f'A{start_row + 1}'] = section_data['description']
        ws[f'A{start_row + 1}'].font = Font(italic=True, size=10, color='666666')
        ws.merge_cells(f'A{start_row + 1}:H{start_row + 1}')
        
        # Column headers
        headers = ['Metric', 'Type', 'Research Default', 'Custom Value', 'Notes', 'Applied Value', 'Source']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row + 3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D5E8D4', end_color='D5E8D4', fill_type='solid')
        
        # Add fields
        current_row = start_row + 4
        for field_name, field_type, default_value, description in section_data['fields']:
            # Get research default if available
            research_value = self._get_research_value(research_assumptions, field_name, default_value)
            
            # Metric name
            ws[f'A{current_row}'] = field_name
            ws[f'A{current_row}'].font = Font(bold=True)
            
            # Field type
            ws[f'B{current_row}'] = field_type
            ws[f'B{current_row}'].font = Font(size=9, color='666666')
            
            # Research default
            ws[f'C{current_row}'] = self._format_value(research_value, field_type)
            ws[f'C{current_row}'].fill = PatternFill(start_color='E1D5E7', end_color='E1D5E7', fill_type='solid')
            
            # Custom value (editable)
            ws[f'D{current_row}'] = ''
            ws[f'D{current_row}'].fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            
            # Notes
            ws[f'E{current_row}'] = description
            ws[f'E{current_row}'].font = Font(size=9, color='666666')
            
            # Applied value (formula)
            ws[f'F{current_row}'] = f'=IF(D{current_row}<>"",D{current_row},C{current_row})'
            ws[f'F{current_row}'].fill = PatternFill(start_color='D5E8D4', end_color='D5E8D4', fill_type='solid')
            
            # Source
            ws[f'G{current_row}'] = 'Custom' if current_row == start_row + 4 else 'Research'
            ws[f'G{current_row}'].font = Font(size=9, color='666666')
            
            current_row += 1
        
        return current_row
    
    def _create_summary_section(self, ws, start_row):
        """Create a summary section showing key customizations"""
        # Summary header
        ws[f'A{start_row}'] = "CUSTOMIZATION SUMMARY"
        ws[f'A{start_row}'].font = Font(bold=True, size=12, color='FFFFFF')
        ws[f'A{start_row}'].fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        ws.merge_cells(f'A{start_row}:H{start_row}')
        
        # Summary metrics
        summary_metrics = [
            ('Total Custom Inputs', '=COUNTIF(D:D,"<>")'),
            ('Growth Assumptions Customized', '=COUNTIF(A:A,"*Growth*")'),
            ('Profitability Assumptions Customized', '=COUNTIF(A:A,"*Margin*")'),
            ('Capital Structure Customized', '=COUNTIF(A:A,"*Debt*")'),
            ('Working Capital Customized', '=COUNTIF(A:A,"*Days*")'),
            ('Valuation Metrics Customized', '=COUNTIF(A:A,"*Multiple*")')
        ]
        
        for i, (metric, formula) in enumerate(summary_metrics, start_row + 2):
            ws[f'A{i}'] = metric
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'] = formula
            ws[f'B{i}'].fill = PatternFill(start_color='D5E8D4', end_color='D5E8D4', fill_type='solid')
        
        return start_row + len(summary_metrics) + 3
    
    def _get_research_value(self, research_assumptions, field_name, default_value):
        """Extract value from research assumptions based on field name"""
        if not research_assumptions:
            return default_value
        
        # Map field names to research assumption paths
        field_mapping = {
            'Revenue Growth Year 1': ['revenue_growth', 0],
            'Revenue Growth Year 2': ['revenue_growth', 1],
            'Revenue Growth Year 3': ['revenue_growth', 2],
            'Revenue Growth Year 4': ['revenue_growth', 3],
            'Revenue Growth Year 5': ['revenue_growth', 4],
            'Terminal Growth Rate': ['terminal_growth'],
            'EBITDA Margin': ['ebitda_margin'],
            'Gross Margin': ['gross_margin'],
            'R&D as % of Revenue': ['rd_pct_revenue'],
            'SG&A as % of Revenue': ['sga_pct_revenue'],
            'Tax Rate': ['tax_rate'],
            'ROA Target': ['roa'],
            'ROE Target': ['roe'],
            'ROIC Target': ['roic'],
            'Target Debt/Equity Ratio': ['debt_to_equity'],
            'Cost of Debt': ['capital_structure', 'cost_of_debt'],
            'Cost of Equity': ['capital_structure', 'cost_of_equity'],
            'Dividend Payout Ratio': ['capital_structure', 'dividend_payout_ratio'],
            'Share Repurchase % of FCF': ['capital_structure', 'share_repurchase_pct_fcf'],
            'Interest Coverage Target': ['interest_coverage'],
            'Debt/EBITDA Target': ['debt_to_ebitda'],
            'Days Sales in Inventory': ['working_capital_days', 'dsi'],
            'Days Sales Outstanding': ['working_capital_days', 'dso'],
            'Days Payable Outstanding': ['working_capital_days', 'dpo'],
            'Prepaid Expenses % of Revenue': ['working_capital_days', 'prepaid_pct_revenue'],
            'Other Current Assets % of Revenue': ['working_capital_days', 'other_current_assets_pct_revenue'],
            'Accrued Expenses % of Revenue': ['working_capital_days', 'accrued_pct_revenue'],
            'Other Current Liabilities % of Revenue': ['working_capital_days', 'other_current_liab_pct_revenue'],
            'CapEx as % of Revenue': ['capex_pct_revenue'],
            'Depreciation as % of PP&E': ['depreciation_pct_ppe'],
            'Amortization as % of Intangibles': ['amortization_pct_intangibles'],
            'EV/EBITDA Multiple': ['valuation_metrics', 'ev_ebitda'],
            'EV/Revenue Multiple': ['valuation_metrics', 'ev_revenue'],
            'P/E Ratio': ['valuation_metrics', 'pe_ratio'],
            'P/B Ratio': ['valuation_metrics', 'pb_ratio'],
            'P/S Ratio': ['valuation_metrics', 'ps_ratio']
        }
        
        if field_name in field_mapping:
            path = field_mapping[field_name]
            value = research_assumptions
            try:
                for key in path:
                    value = value[key]
                return value
            except (KeyError, IndexError, TypeError):
                return default_value
        
        return default_value
    
    def _format_value(self, value, field_type):
        """Format value based on field type"""
        if field_type == 'percentage':
            return f"{value*100:.1f}%"
        elif field_type == 'currency':
            return f"${value:,.0f}"
        elif field_type == 'ratio':
            return f"{value:.2f}"
        elif field_type == 'multiple':
            return f"{value:.1f}x"
        elif field_type == 'days':
            return f"{value:.0f} days"
        elif field_type == 'boolean':
            return "Yes" if value else "No"
        else:
            return str(value)
    
    def _apply_sheet_formatting(self, ws):
        """Apply consistent formatting to the sheet"""
        # Apply borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 12), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def get_custom_inputs(self, ws):
        """Extract custom inputs from the worksheet"""
        custom_inputs = {}
        
        # Find the custom inputs section
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and 'Custom Value' in str(cell.value):
                    # Found the custom inputs section
                    start_row = cell.row + 1
                    break
        
        # Extract custom values
        for row in range(start_row, ws.max_row + 1):
            metric = ws[f'A{row}'].value
            custom_value = ws[f'D{row}'].value
            
            if metric and custom_value and custom_value != '':
                # Clean up the metric name
                clean_metric = metric.strip()
                custom_inputs[clean_metric] = custom_value
        
        return custom_inputs
    
    def apply_custom_inputs(self, research_assumptions, custom_inputs):
        """Apply custom inputs to research assumptions"""
        if not custom_inputs:
            return research_assumptions
        
        # Create a copy of research assumptions
        modified_assumptions = research_assumptions.copy()
        
        # Apply custom inputs
        for metric, value in custom_inputs.items():
            # Parse the value based on the metric type
            parsed_value = self._parse_custom_value(metric, value)
            
            # Apply to the appropriate section
            self._apply_custom_value(modified_assumptions, metric, parsed_value)
        
        return modified_assumptions
    
    def _parse_custom_value(self, metric, value):
        """Parse custom value based on metric type"""
        if isinstance(value, (int, float)):
            return value
        
        # Remove common formatting
        value_str = str(value).replace('%', '').replace('$', '').replace(',', '').replace('x', '').replace(' days', '')
        
        try:
            if '%' in str(value):
                return float(value_str) / 100
            elif 'x' in str(value):
                return float(value_str)
            elif 'days' in str(value):
                return float(value_str)
            elif '$' in str(value):
                return float(value_str)
            else:
                return float(value_str)
        except ValueError:
            return value
    
    def _apply_custom_value(self, assumptions, metric, value):
        """Apply custom value to the appropriate section of assumptions"""
        # Map metrics to assumption paths
        metric_mapping = {
            'Revenue Growth Year 1': ['revenue_growth', 0],
            'Revenue Growth Year 2': ['revenue_growth', 1],
            'Revenue Growth Year 3': ['revenue_growth', 2],
            'Revenue Growth Year 4': ['revenue_growth', 3],
            'Revenue Growth Year 5': ['revenue_growth', 4],
            'Terminal Growth Rate': ['terminal_growth'],
            'EBITDA Margin': ['ebitda_margin'],
            'Gross Margin': ['gross_margin'],
            'R&D as % of Revenue': ['rd_pct_revenue'],
            'SG&A as % of Revenue': ['sga_pct_revenue'],
            'Tax Rate': ['tax_rate'],
            'ROA Target': ['roa'],
            'ROE Target': ['roe'],
            'ROIC Target': ['roic'],
            'Target Debt/Equity Ratio': ['debt_to_equity'],
            'Cost of Debt': ['capital_structure', 'cost_of_debt'],
            'Cost of Equity': ['capital_structure', 'cost_of_equity'],
            'Dividend Payout Ratio': ['capital_structure', 'dividend_payout_ratio'],
            'Share Repurchase % of FCF': ['capital_structure', 'share_repurchase_pct_fcf'],
            'Interest Coverage Target': ['interest_coverage'],
            'Debt/EBITDA Target': ['debt_to_ebitda'],
            'Days Sales in Inventory': ['working_capital_days', 'dsi'],
            'Days Sales Outstanding': ['working_capital_days', 'dso'],
            'Days Payable Outstanding': ['working_capital_days', 'dpo'],
            'Prepaid Expenses % of Revenue': ['working_capital_days', 'prepaid_pct_revenue'],
            'Other Current Assets % of Revenue': ['working_capital_days', 'other_current_assets_pct_revenue'],
            'Accrued Expenses % of Revenue': ['working_capital_days', 'accrued_pct_revenue'],
            'Other Current Liabilities % of Revenue': ['working_capital_days', 'other_current_liab_pct_revenue'],
            'CapEx as % of Revenue': ['capex_pct_revenue'],
            'Depreciation as % of PP&E': ['depreciation_pct_ppe'],
            'Amortization as % of Intangibles': ['amortization_pct_intangibles'],
            'EV/EBITDA Multiple': ['valuation_metrics', 'ev_ebitda'],
            'EV/Revenue Multiple': ['valuation_metrics', 'ev_revenue'],
            'P/E Ratio': ['valuation_metrics', 'pe_ratio'],
            'P/B Ratio': ['valuation_metrics', 'pb_ratio'],
            'P/S Ratio': ['valuation_metrics', 'ps_ratio']
        }
        
        if metric in metric_mapping:
            path = metric_mapping[metric]
            target = assumptions
            for key in path[:-1]:
                if key not in target:
                    target[key] = {}
                target = target[key]
            target[path[-1]] = value

# Global instance
custom_inputs_manager = CustomInputsManager()

def create_custom_inputs_sheet(wb, company_name, research_assumptions=None):
    """Convenience function to create custom inputs sheet"""
    return custom_inputs_manager.create_custom_inputs_sheet(wb, company_name, research_assumptions)

def get_custom_inputs(ws):
    """Convenience function to get custom inputs"""
    return custom_inputs_manager.get_custom_inputs(ws)

def apply_custom_inputs(research_assumptions, custom_inputs):
    """Convenience function to apply custom inputs"""
    return custom_inputs_manager.apply_custom_inputs(research_assumptions, custom_inputs) 