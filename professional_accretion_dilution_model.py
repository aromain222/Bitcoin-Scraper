#!/usr/bin/env python3
"""
Professional Accretion/Dilution Model
Evaluates pro forma EPS impact of an acquisition

Author: Investment Banking Modeler
Date: 2024

Features:
- Buyer and seller financial analysis
- Flexible deal structure (cash, stock, debt)
- Synergies and cost savings modeling
- Pro forma EPS calculations
- Accretion/dilution analysis
- Professional banker-standard formatting
- Excel output with multiple tabs
- Sensitivity analysis
"""

import numpy as np
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
try:
    import matplotlib.pyplot as plt
    HAS_MATPLOTLIB = True
except ImportError:
    HAS_MATPLOTLIB = False
    print("⚠️  Warning: matplotlib not available. Charts will not be generated.")
import warnings
warnings.filterwarnings('ignore')

# Professional color scheme for Accretion/Dilution
ACCRETION_COLORS = {
    'header_blue': '1F4E79',
    'input_light_blue': 'DDEBF7',
    'calculation_black': '000000',
    'output_green': 'D8E4BC',
    'accretion_green': 'C8E6C9',      # Green for accretive deals
    'dilution_red': 'FFCDD2',        # Red for dilutive deals
    'warning_red': 'F8CECC',
    'border_dark': '2C3E50',
    'text_dark': '000000',
    'text_white': 'FFFFFF',
    'chart_background': 'F5F5F5',
    'grid_light': 'E0E0E0',
    'highlight_gold': 'FFD700'
}

class ProfessionalAccretionDilutionModel:
    """
    Comprehensive Accretion/Dilution Model with Professional Formatting
    """

    def __init__(self, buyer_company="Buyer Company", seller_company="Seller Company", buyer_ticker="BUYER", seller_ticker="SELLER"):
        self.buyer_company = buyer_company
        self.seller_company = seller_company
        self.buyer_ticker = buyer_ticker
        self.seller_ticker = seller_ticker
        self.model_date = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Initialize styles
        self.styles = self._create_styles()

    def _create_styles(self):
        """Create professional Excel styles"""
        styles = {}

        # Header style
        styles['header'] = NamedStyle(name='header')
        styles['header'].font = Font(name='Calibri', size=12, bold=True, color=ACCRETION_COLORS['text_white'])
        styles['header'].fill = PatternFill(start_color=ACCRETION_COLORS['header_blue'], end_color=ACCRETION_COLORS['header_blue'], fill_type='solid')
        styles['header'].alignment = Alignment(horizontal='center', vertical='center')

        # Input style
        styles['input'] = NamedStyle(name='input')
        styles['input'].font = Font(name='Calibri', size=10, color=ACCRETION_COLORS['text_dark'])
        styles['input'].fill = PatternFill(start_color=ACCRETION_COLORS['input_light_blue'], end_color=ACCRETION_COLORS['input_light_blue'], fill_type='solid')
        styles['input'].alignment = Alignment(horizontal='right', vertical='center')
        styles['input'].number_format = '#,##0.00'

        # Calculation style
        styles['calculation'] = NamedStyle(name='calculation')
        styles['calculation'].font = Font(name='Calibri', size=10, color=ACCRETION_COLORS['calculation_black'])
        styles['calculation'].alignment = Alignment(horizontal='right', vertical='center')
        styles['calculation'].number_format = '#,##0.00'

        # Output style
        styles['output'] = NamedStyle(name='output')
        styles['output'].font = Font(name='Calibri', size=10, bold=True, color=ACCRETION_COLORS['text_dark'])
        styles['output'].fill = PatternFill(start_color=ACCRETION_COLORS['output_green'], end_color=ACCRETION_COLORS['output_green'], fill_type='solid')
        styles['output'].alignment = Alignment(horizontal='right', vertical='center')
        styles['output'].number_format = '#,##0.00'

        # Accretion style (green)
        styles['accretion'] = NamedStyle(name='accretion')
        styles['accretion'].font = Font(name='Calibri', size=10, bold=True, color=ACCRETION_COLORS['text_dark'])
        styles['accretion'].fill = PatternFill(start_color=ACCRETION_COLORS['accretion_green'], end_color=ACCRETION_COLORS['accretion_green'], fill_type='solid')
        styles['accretion'].alignment = Alignment(horizontal='right', vertical='center')
        styles['accretion'].number_format = '0.00%'

        # Dilution style (red)
        styles['dilution'] = NamedStyle(name='dilution')
        styles['dilution'].font = Font(name='Calibri', size=10, bold=True, color=ACCRETION_COLORS['text_dark'])
        styles['dilution'].fill = PatternFill(start_color=ACCRETION_COLORS['dilution_red'], end_color=ACCRETION_COLORS['dilution_red'], fill_type='solid')
        styles['dilution'].alignment = Alignment(horizontal='right', vertical='center')
        styles['dilution'].number_format = '0.00%'

        # Warning style
        styles['warning'] = NamedStyle(name='warning')
        styles['warning'].font = Font(name='Calibri', size=10, color=ACCRETION_COLORS['text_dark'])
        styles['warning'].fill = PatternFill(start_color=ACCRETION_COLORS['warning_red'], end_color=ACCRETION_COLORS['warning_red'], fill_type='solid')
        styles['warning'].alignment = Alignment(horizontal='center', vertical='center')

        # Bold label style
        styles['label_bold'] = NamedStyle(name='label_bold')
        styles['label_bold'].font = Font(name='Calibri', size=10, bold=True, color=ACCRETION_COLORS['text_dark'])
        styles['label_bold'].alignment = Alignment(horizontal='left', vertical='center')

        # Normal label style
        styles['label'] = NamedStyle(name='label')
        styles['label'].font = Font(name='Calibri', size=10, color=ACCRETION_COLORS['text_dark'])
        styles['label'].alignment = Alignment(horizontal='left', vertical='center')

        # Summary box style
        styles['summary_box'] = NamedStyle(name='summary_box')
        styles['summary_box'].font = Font(name='Calibri', size=11, bold=True, color=ACCRETION_COLORS['text_white'])
        styles['summary_box'].fill = PatternFill(start_color=ACCRETION_COLORS['header_blue'], end_color=ACCRETION_COLORS['header_blue'], fill_type='solid')
        styles['summary_box'].alignment = Alignment(horizontal='center', vertical='center')
        styles['summary_box'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Company header style
        styles['company_header'] = NamedStyle(name='company_header')
        styles['company_header'].font = Font(name='Calibri', size=11, bold=True, color=ACCRETION_COLORS['text_white'])
        styles['company_header'].fill = PatternFill(start_color=ACCRETION_COLORS['border_dark'], end_color=ACCRETION_COLORS['border_dark'], fill_type='solid')
        styles['company_header'].alignment = Alignment(horizontal='center', vertical='center')

        return styles

    def run_accretion_dilution_model(self,
                                   # Buyer Company Financials
                                   buyer_net_income=1000.0,      # $M
                                   buyer_eps=3.50,              # $
                                   buyer_shares_outstanding=285.7,  # Million shares
                                   buyer_pe_ratio=18.0,          # P/E multiple

                                   # Seller Company Financials
                                   seller_net_income=200.0,      # $M
                                   seller_shares_outstanding=57.1,  # Million shares
                                   seller_pe_ratio=15.0,         # P/E multiple

                                   # Deal Structure
                                   purchase_premium_pct=0.25,    # 25% premium over current market value
                                   cash_pct=0.60,                # 60% cash consideration
                                   stock_pct=0.30,               # 30% stock consideration
                                   debt_pct=0.10,                # 10% debt financing

                                   # Financing Assumptions
                                   interest_rate=0.05,           # 5% interest rate on new debt
                                   tax_rate=0.25,                # 25% corporate tax rate

                                   # Synergies and Cost Savings
                                   revenue_synergies=100.0,      # $M annual revenue synergies
                                   cost_synergies=150.0,         # $M annual cost synergies
                                   one_time_costs=50.0,          # $M one-time integration costs

                                   # Sensitivity Ranges
                                   premium_range=None,           # [0.15, 0.20, 0.25, 0.30, 0.35]
                                   synergies_range=None):        # [100, 125, 150, 175, 200]

        """
        Run complete accretion/dilution analysis
        """

        print(f"📊 Building Professional Accretion/Dilution Model for {self.buyer_company} acquiring {self.seller_company}")
        print("=" * 110)

        # Set default ranges if not provided
        if premium_range is None:
            premium_range = [0.15, 0.20, 0.25, 0.30, 0.35]  # 15% to 35%
        if synergies_range is None:
            synergies_range = [100.0, 125.0, 150.0, 175.0, 200.0]  # $100M to $200M

        # Step 1: Create Financial Inputs & Deal Structure
        financial_inputs = self._create_financial_inputs(
            buyer_net_income, buyer_eps, buyer_shares_outstanding, buyer_pe_ratio,
            seller_net_income, seller_shares_outstanding, seller_pe_ratio
        )

        deal_structure = self._create_deal_structure(
            purchase_premium_pct, cash_pct, stock_pct, debt_pct,
            interest_rate, tax_rate
        )

        synergies = self._create_synergies(
            revenue_synergies, cost_synergies, one_time_costs
        )

        # Step 2: Calculate Purchase Price & Financing
        purchase_price_calc = self._calculate_purchase_price(financial_inputs, deal_structure)

        financing_calc = self._calculate_financing(purchase_price_calc, deal_structure, synergies)

        # Step 3: Calculate Pro Forma Financials
        proforma_calc = self._calculate_proforma_financials(
            financial_inputs, purchase_price_calc, financing_calc, synergies
        )

        # Step 4: Calculate Accretion/Dilution
        accretion_dilution = self._calculate_accretion_dilution(
            financial_inputs, proforma_calc
        )

        # Step 5: Generate Sensitivity Analysis
        sensitivity_analysis = self._generate_sensitivity_analysis(
            financial_inputs, deal_structure, synergies,
            premium_range, synergies_range
        )

        # Step 6: Create Charts
        chart_files = self._create_accretion_charts(accretion_dilution, sensitivity_analysis)

        # Step 7: Create Excel Output
        excel_file = self._create_excel_output(
            financial_inputs, deal_structure, synergies,
            purchase_price_calc, financing_calc, proforma_calc,
            accretion_dilution, sensitivity_analysis, chart_files
        )

        print("\n✅ Accretion/Dilution Model Complete!")
        print("📊 Key Deal Metrics:")
        print(f"   • Buyer: {self.buyer_company} ({self.buyer_ticker}) - ${buyer_net_income:.0f}M NI, {buyer_eps:.2f} EPS")
        print(f"   • Seller: {self.seller_company} ({self.seller_ticker}) - ${seller_net_income:.0f}M NI")
        print(f"   • Purchase Premium: {purchase_premium_pct:.1%}")
        print(f"   • Consideration Mix: {cash_pct:.0%} Cash, {stock_pct:.0%} Stock, {debt_pct:.0%} Debt")
        print(f"   • Synergies: ${cost_synergies:.0f}M cost savings")
        print(f"📁 Excel Output: {excel_file}")

        # Display key results
        self._display_accretion_results(accretion_dilution, sensitivity_analysis)

        return {
            'financial_inputs': financial_inputs,
            'deal_structure': deal_structure,
            'synergies': synergies,
            'purchase_price_calc': purchase_price_calc,
            'financing_calc': financing_calc,
            'proforma_calc': proforma_calc,
            'accretion_dilution': accretion_dilution,
            'sensitivity_analysis': sensitivity_analysis,
            'excel_file': excel_file
        }

    def _create_financial_inputs(self, buyer_net_income, buyer_eps, buyer_shares_outstanding, buyer_pe_ratio,
                                seller_net_income, seller_shares_outstanding, seller_pe_ratio):

        """Create comprehensive financial inputs for buyer and seller"""

        # Buyer financials
        buyer_market_cap = buyer_eps * buyer_shares_outstanding
        buyer_enterprise_value = buyer_market_cap  # Simplified, assuming no debt for buyer

        # Seller financials
        seller_eps = seller_net_income / seller_shares_outstanding if seller_shares_outstanding > 0 else 0
        seller_market_cap = seller_eps * seller_shares_outstanding if seller_eps > 0 else seller_net_income * seller_pe_ratio
        seller_enterprise_value = seller_market_cap  # Simplified

        financial_inputs = {
            'buyer': {
                'company': self.buyer_company,
                'ticker': self.buyer_ticker,
                'net_income': buyer_net_income,
                'eps': buyer_eps,
                'shares_outstanding': buyer_shares_outstanding,
                'pe_ratio': buyer_pe_ratio,
                'market_cap': buyer_market_cap,
                'enterprise_value': buyer_enterprise_value
            },
            'seller': {
                'company': self.seller_company,
                'ticker': self.seller_ticker,
                'net_income': seller_net_income,
                'eps': seller_eps,
                'shares_outstanding': seller_shares_outstanding,
                'pe_ratio': seller_pe_ratio,
                'market_cap': seller_market_cap,
                'enterprise_value': seller_enterprise_value
            }
        }

        print("📋 Financial Inputs Created:")
        print(f"   • Buyer Market Cap: ${buyer_market_cap/1000:.1f}B")
        print(f"   • Seller Market Cap: ${seller_market_cap/1000:.1f}B")
        print(f"   • Seller EPS: ${seller_eps:.2f}")

        return financial_inputs

    def _create_deal_structure(self, purchase_premium_pct, cash_pct, stock_pct, debt_pct,
                              interest_rate, tax_rate):

        """Create deal structure and financing assumptions"""

        # Validate consideration mix
        total_pct = cash_pct + stock_pct + debt_pct
        if abs(total_pct - 1.0) > 0.01:
            print(f"⚠️  Warning: Consideration mix sums to {total_pct:.1%}, adjusting to 100%")
            # Normalize to 100%
            scale_factor = 1.0 / total_pct
            cash_pct *= scale_factor
            stock_pct *= scale_factor
            debt_pct *= scale_factor

        deal_structure = {
            'purchase_premium_pct': purchase_premium_pct,
            'consideration_mix': {
                'cash_pct': cash_pct,
                'stock_pct': stock_pct,
                'debt_pct': debt_pct
            },
            'financing': {
                'interest_rate': interest_rate,
                'tax_rate': tax_rate
            }
        }

        print("🤝 Deal Structure Created:")
        print(f"   • Purchase Premium: {purchase_premium_pct:.1%}")
        print(f"   • Cash Consideration: {cash_pct:.1%}")
        print(f"   • Stock Consideration: {stock_pct:.1%}")
        print(f"   • Debt Financing: {debt_pct:.1%}")

        return deal_structure

    def _create_synergies(self, revenue_synergies, cost_synergies, one_time_costs):

        """Create synergies and cost savings assumptions"""

        synergies = {
            'revenue_synergies': revenue_synergies,
            'cost_synergies': cost_synergies,
            'one_time_costs': one_time_costs,
            'total_annual_synergies': revenue_synergies + cost_synergies
        }

        print("💰 Synergies Created:")
        print(f"   • Revenue Synergies: ${revenue_synergies:.0f}M")
        print(f"   • Cost Synergies: ${cost_synergies:.0f}M")
        print(f"   • One-time Costs: ${one_time_costs:.0f}M")
        print(f"   • Net Annual Synergies: ${synergies['total_annual_synergies']:.0f}M")

        return synergies

    def _calculate_purchase_price(self, financial_inputs, deal_structure):

        """Calculate purchase price and consideration breakdown"""

        seller_market_cap = financial_inputs['seller']['market_cap']
        purchase_premium_pct = deal_structure['purchase_premium_pct']

        # Calculate purchase price with premium
        equity_purchase_price = seller_market_cap * (1 + purchase_premium_pct)
        enterprise_purchase_price = equity_purchase_price  # Simplified

        # Breakdown by consideration type
        cash_consideration = equity_purchase_price * deal_structure['consideration_mix']['cash_pct']
        stock_consideration = equity_purchase_price * deal_structure['consideration_mix']['stock_pct']
        debt_consideration = equity_purchase_price * deal_structure['consideration_mix']['debt_pct']

        # Calculate shares to be issued (if stock consideration)
        buyer_pe_ratio = financial_inputs['buyer']['pe_ratio']
        buyer_eps = financial_inputs['buyer']['eps']
        shares_to_issue = stock_consideration / (buyer_eps * buyer_pe_ratio) if buyer_eps > 0 else 0

        purchase_price_calc = {
            'seller_market_cap': seller_market_cap,
            'purchase_premium_pct': purchase_premium_pct,
            'equity_purchase_price': equity_purchase_price,
            'enterprise_purchase_price': enterprise_purchase_price,
            'consideration_breakdown': {
                'cash': cash_consideration,
                'stock': stock_consideration,
                'debt': debt_consideration
            },
            'shares_to_issue': shares_to_issue
        }

        print("💵 Purchase Price Calculated:")
        print(f"   • Seller Market Cap: ${seller_market_cap/1000:.1f}B")
        print(f"   • Purchase Price (with premium): ${equity_purchase_price/1000:.1f}B")
        print(f"   • Shares to Issue: {shares_to_issue:.1f}M")

        return purchase_price_calc

    def _calculate_financing(self, purchase_price_calc, deal_structure, synergies):

        """Calculate financing impacts and interest expenses"""

        cash_consideration = purchase_price_calc['consideration_breakdown']['cash']
        debt_consideration = purchase_price_calc['consideration_breakdown']['debt']

        interest_rate = deal_structure['financing']['interest_rate']
        tax_rate = deal_structure['financing']['tax_rate']

        # Interest expense on new debt
        annual_interest_expense = debt_consideration * interest_rate
        tax_shield = annual_interest_expense * tax_rate
        net_interest_expense = annual_interest_expense - tax_shield

        # Foregone interest income on cash used
        foregone_interest_income = cash_consideration * interest_rate * (1 - tax_rate)

        # Net financing impact
        net_financing_impact = net_interest_expense - foregone_interest_income

        financing_calc = {
            'new_debt_amount': debt_consideration,
            'annual_interest_expense': annual_interest_expense,
            'tax_shield': tax_shield,
            'net_interest_expense': net_interest_expense,
            'cash_used': cash_consideration,
            'foregone_interest_income': foregone_interest_income,
            'net_financing_impact': net_financing_impact
        }

        print("💳 Financing Calculated:")
        print(f"   • New Debt: ${debt_consideration/1000:.1f}B")
        print(f"   • Annual Interest Expense: ${annual_interest_expense:.0f}M")
        print(f"   • Net Financing Impact: ${net_financing_impact:.0f}M")

        return financing_calc

    def _calculate_proforma_financials(self, financial_inputs, purchase_price_calc, financing_calc, synergies):

        """Calculate pro forma financials after acquisition"""

        # Buyer standalone financials
        buyer_net_income = financial_inputs['buyer']['net_income']
        buyer_shares = financial_inputs['buyer']['shares_outstanding']

        # Seller standalone financials
        seller_net_income = financial_inputs['seller']['net_income']

        # Pro forma adjustments
        # 1. Synergies (after-tax)
        tax_rate = financing_calc.get('tax_rate', 0.25)
        synergies_impact = synergies['total_annual_synergies'] * (1 - tax_rate)

        # 2. Financing impact
        financing_impact = financing_calc['net_financing_impact']

        # 3. One-time costs (amortized or immediate impact - simplified as immediate)
        one_time_impact = synergies['one_time_costs'] * (1 - tax_rate)  # Assuming tax-deductible

        # Pro forma net income
        proforma_net_income = buyer_net_income + seller_net_income + synergies_impact - financing_impact - one_time_impact

        # Pro forma shares outstanding
        shares_issued = purchase_price_calc['shares_to_issue']
        proforma_shares = buyer_shares + shares_issued

        # Pro forma EPS
        proforma_eps = proforma_net_income / proforma_shares if proforma_shares > 0 else 0

        proforma_calc = {
            'buyer_standalone_net_income': buyer_net_income,
            'seller_standalone_net_income': seller_net_income,
            'synergies_impact': synergies_impact,
            'financing_impact': financing_impact,
            'one_time_impact': one_time_impact,
            'proforma_net_income': proforma_net_income,
            'buyer_shares': buyer_shares,
            'shares_issued': shares_issued,
            'proforma_shares': proforma_shares,
            'buyer_standalone_eps': financial_inputs['buyer']['eps'],
            'proforma_eps': proforma_eps
        }

        print("📈 Pro Forma Financials Calculated:")
        print(f"   • Buyer Standalone NI: ${buyer_net_income:.0f}M")
        print(f"   • Seller Standalone NI: ${seller_net_income:.0f}M")
        print(f"   • Synergies Impact: ${synergies_impact:.0f}M")
        print(f"   • Pro Forma NI: ${proforma_net_income:.0f}M")
        print(f"   • Pro Forma Shares: {proforma_shares:.1f}M")
        print(f"   • Pro Forma EPS: ${proforma_eps:.2f}")

        return proforma_calc

    def _calculate_accretion_dilution(self, financial_inputs, proforma_calc):

        """Calculate accretion/dilution analysis"""

        buyer_standalone_eps = proforma_calc['buyer_standalone_eps']
        proforma_eps = proforma_calc['proforma_eps']

        # Accretion/Dilution calculation
        eps_impact = proforma_eps - buyer_standalone_eps
        accretion_dilution_pct = eps_impact / buyer_standalone_eps if buyer_standalone_eps != 0 else 0

        # Determine if accretive or dilutive
        is_accretive = accretion_dilution_pct > 0

        accretion_dilution = {
            'buyer_standalone_eps': buyer_standalone_eps,
            'proforma_eps': proforma_eps,
            'eps_impact': eps_impact,
            'accretion_dilution_pct': accretion_dilution_pct,
            'is_accretive': is_accretive,
            'status': "ACCRETIVE" if is_accretive else "DILUTIVE"
        }

        print("📊 Accretion/Dilution Calculated:")
        print(f"   • Buyer Standalone EPS: ${buyer_standalone_eps:.2f}")
        print(f"   • Pro Forma EPS: ${proforma_eps:.2f}")
        print(f"   • EPS Impact: ${eps_impact:.2f}")
        print(f"   • Accretion/(Dilution): {accretion_dilution_pct:.2%}")
        print(f"   • Status: {accretion_dilution['status']}")

        return accretion_dilution

    def _generate_sensitivity_analysis(self, financial_inputs, deal_structure, synergies,
                                     premium_range, synergies_range):

        """Generate sensitivity analysis for key variables"""

        sensitivity_analysis = {}

        # Premium sensitivity
        premium_sensitivity = []
        for premium in premium_range:
            # Recalculate with different premium
            temp_deal = deal_structure.copy()
            temp_deal['purchase_premium_pct'] = premium

            temp_purchase = self._calculate_purchase_price(financial_inputs, temp_deal)
            temp_financing = self._calculate_financing(temp_purchase, temp_deal, synergies)
            temp_proforma = self._calculate_proforma_financials(financial_inputs, temp_purchase, temp_financing, synergies)
            temp_accretion = self._calculate_accretion_dilution(financial_inputs, temp_proforma)

            premium_sensitivity.append({
                'premium_pct': premium,
                'purchase_price': temp_purchase['equity_purchase_price'],
                'proforma_eps': temp_proforma['proforma_eps'],
                'accretion_dilution_pct': temp_accretion['accretion_dilution_pct'],
                'is_accretive': temp_accretion['is_accretive']
            })

        # Synergies sensitivity
        synergies_sensitivity = []
        for synergy_amount in synergies_range:
            # Recalculate with different synergies
            temp_synergies = synergies.copy()
            temp_synergies['cost_synergies'] = synergy_amount
            temp_synergies['total_annual_synergies'] = temp_synergies['revenue_synergies'] + synergy_amount

            temp_financing = self._calculate_financing(self._calculate_purchase_price(financial_inputs, deal_structure), deal_structure, temp_synergies)
            temp_proforma = self._calculate_proforma_financials(financial_inputs, self._calculate_purchase_price(financial_inputs, deal_structure), temp_financing, temp_synergies)
            temp_accretion = self._calculate_accretion_dilution(financial_inputs, temp_proforma)

            synergies_sensitivity.append({
                'synergies_amount': synergy_amount,
                'proforma_eps': temp_proforma['proforma_eps'],
                'accretion_dilution_pct': temp_accretion['accretion_dilution_pct'],
                'is_accretive': temp_accretion['is_accretive']
            })

        sensitivity_analysis = {
            'premium_sensitivity': premium_sensitivity,
            'synergies_sensitivity': synergies_sensitivity
        }

        print("📈 Sensitivity Analysis Generated:")
        print(f"   • Premium Scenarios: {len(premium_range)} cases")
        print(f"   • Synergies Scenarios: {len(synergies_range)} cases")

        return sensitivity_analysis

    def _create_accretion_charts(self, accretion_dilution, sensitivity_analysis):

        """Create professional charts for accretion/dilution analysis"""

        chart_files = {}

        if not HAS_MATPLOTLIB:
            print("📈 Charts Skipped (matplotlib not available)")
            return chart_files

        # Create main accretion/dilution bar chart
        chart_files['main_chart'] = self._create_main_accretion_chart(accretion_dilution)

        # Create sensitivity charts
        if sensitivity_analysis['premium_sensitivity']:
            chart_files['premium_chart'] = self._create_sensitivity_chart(
                sensitivity_analysis['premium_sensitivity'],
                'Purchase Premium Sensitivity',
                'Premium (%)',
                'premium_sensitivity.png'
            )

        if sensitivity_analysis['synergies_sensitivity']:
            chart_files['synergies_chart'] = self._create_sensitivity_chart(
                sensitivity_analysis['synergies_sensitivity'],
                'Synergies Sensitivity',
                'Synergies ($M)',
                'synergies_sensitivity.png'
            )

        print("📊 Accretion Charts Created:")
        for chart_name, chart_file in chart_files.items():
            print(f"   • {chart_name}: {chart_file}")

        return chart_files

    def _create_main_accretion_chart(self, accretion_dilution):

        """Create main accretion/dilution bar chart"""

        fig, ax = plt.subplots(figsize=(10, 6))
        fig.patch.set_facecolor('#' + ACCRETION_COLORS['chart_background'])

        # Data
        categories = ['Buyer Standalone', 'Pro Forma Combined']
        eps_values = [accretion_dilution['buyer_standalone_eps'], accretion_dilution['proforma_eps']]
        colors = ['#' + ACCRETION_COLORS['border_dark'], '#' + ACCRETION_COLORS['accretion_green'] if accretion_dilution['is_accretive'] else '#' + ACCRETION_COLORS['dilution_red']]

        # Create bars
        bars = ax.bar(categories, eps_values, color=colors, alpha=0.8, width=0.6)

        # Add value labels on bars
        for bar, value in zip(bars, eps_values):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height + 0.05,
                   f'${value:.2f}', ha='center', va='bottom', fontweight='bold')

        # Add accretion/dilution annotation
        accretion_pct = accretion_dilution['accretion_dilution_pct']
        color = '#' + ACCRETION_COLORS['accretion_green'] if accretion_dilution['is_accretive'] else '#' + ACCRETION_COLORS['dilution_red']
        ax.text(1.5, max(eps_values) * 0.8,
               f'Accretion/(Dilution): {accretion_pct:.2%}',
               ha='center', va='center', fontsize=12, fontweight='bold',
               bbox=dict(boxstyle='round,pad=0.3', facecolor=color, alpha=0.8))

        # Styling
        ax.set_ylabel('EPS ($)', fontweight='bold', fontsize=11)
        ax.set_title('Accretion/(Dilution) Analysis – Base Case', fontweight='bold', fontsize=14, pad=20)
        ax.grid(True, alpha=0.3, axis='y')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)

        plt.tight_layout()

        # Save chart
        chart_path = "/Users/averyromain/Scraper/accretion_main_chart.png"
        plt.savefig(chart_path, dpi=300, bbox_inches='tight', facecolor=fig.get_facecolor())
        plt.close()

        return chart_path

    def _create_sensitivity_chart(self, sensitivity_data, title, x_label, filename):

        """Create sensitivity analysis chart"""

        fig, ax = plt.subplots(figsize=(12, 6))
        fig.patch.set_facecolor('#' + ACCRETION_COLORS['chart_background'])

        # Extract data
        if 'premium_pct' in sensitivity_data[0]:
            x_values = [item['premium_pct'] * 100 for item in sensitivity_data]  # Convert to percentage
            x_label = 'Premium (%)'
        else:
            x_values = [item['synergies_amount'] for item in sensitivity_data]

        eps_values = [item['proforma_eps'] for item in sensitivity_data]
        accretion_values = [item['accretion_dilution_pct'] * 100 for item in sensitivity_data]  # Convert to percentage

        # Create dual y-axis chart
        ax2 = ax.twinx()

        # Plot EPS line
        line1 = ax.plot(x_values, eps_values, 'b-', linewidth=2, marker='o', markersize=6, label='Pro Forma EPS')
        # Plot accretion/dilution line
        line2 = ax2.plot(x_values, accretion_values, 'r-', linewidth=2, marker='s', markersize=6, label='Accretion/(Dilution) %')

        # Styling
        ax.set_xlabel(x_label, fontweight='bold', fontsize=11)
        ax.set_ylabel('EPS ($)', fontweight='bold', fontsize=11, color='b')
        ax2.set_ylabel('Accretion/(Dilution) (%)', fontweight='bold', fontsize=11, color='r')
        ax.set_title(title, fontweight='bold', fontsize=14, pad=20)

        # Add legend
        lines = line1 + line2
        labels = [l.get_label() for l in lines]
        ax.legend(lines, labels, loc='upper left')

        ax.grid(True, alpha=0.3)
        ax.spines['top'].set_visible(False)
        ax2.spines['top'].set_visible(False)

        plt.tight_layout()

        # Save chart
        chart_path = f"/Users/averyromain/Scraper/{filename}"
        plt.savefig(chart_path, dpi=300, bbox_inches='tight', facecolor=fig.get_facecolor())
        plt.close()

        return chart_path

    def _create_excel_output(self, financial_inputs, deal_structure, synergies,
                           purchase_price_calc, financing_calc, proforma_calc,
                           accretion_dilution, sensitivity_analysis, chart_files):

        """Create professional Excel output with multiple tabs"""

        wb = Workbook()

        # Create worksheets
        ws_summary = wb.active
        ws_summary.title = "Accretion Dilution Summary"

        ws_inputs = wb.create_sheet("Inputs & Assumptions")
        ws_calculations = wb.create_sheet("Calculations")
        ws_sensitivity = wb.create_sheet("Sensitivity Analysis")

        # Apply styles to workbook
        for style_name, style in self.styles.items():
            if style_name not in wb.named_styles:
                wb.add_named_style(style)

        # Create each tab
        self._create_summary_tab(ws_summary, accretion_dilution, chart_files)
        self._create_inputs_tab(ws_inputs, financial_inputs, deal_structure, synergies)
        self._create_calculations_tab(ws_calculations, purchase_price_calc, financing_calc, proforma_calc)
        self._create_sensitivity_tab(ws_sensitivity, sensitivity_analysis)

        # Save workbook
        filename = f"Accretion_Dilution_{self.buyer_ticker}_{self.seller_ticker}_{self.model_date}.xlsx"
        wb.save(filename)

        return filename

    def _create_summary_tab(self, ws, accretion_dilution, chart_files):

        """Create Accretion/Dilution Summary tab"""

        # Title
        ws['A1'] = f"Accretion/(Dilution) Analysis: {self.buyer_company} Acquiring {self.seller_company}"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:G1')

        current_row = 3

        # Executive Summary Box
        ws[f'A{current_row}'] = 'EXECUTIVE SUMMARY'
        ws[f'A{current_row}'].style = 'summary_box'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        # Key metrics
        summary_data = [
            ("Buyer Standalone EPS", f"${accretion_dilution['buyer_standalone_eps']:.2f}"),
            ("Pro Forma EPS", f"${accretion_dilution['proforma_eps']:.2f}"),
            ("EPS Impact", f"${accretion_dilution['eps_impact']:.2f}"),
            ("Accretion/(Dilution)", f"{accretion_dilution['accretion_dilution_pct']:.2%}"),
            ("Status", accretion_dilution['status'])
        ]

        for label, value in summary_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            if label == "Accretion/(Dilution)":
                ws[f'B{current_row}'].style = 'accretion' if accretion_dilution['is_accretive'] else 'dilution'
            elif label == "Status":
                ws[f'B{current_row}'].style = 'accretion' if accretion_dilution['is_accretive'] else 'dilution'
            else:
                ws[f'B{current_row}'].style = 'output'
            current_row += 1

        # Add main chart if available
        if 'main_chart' in chart_files and HAS_MATPLOTLIB:
            current_row += 2
            try:
                from openpyxl.drawing.image import Image
                img = Image(chart_files['main_chart'])
                img.width = 600
                img.height = 400
                ws.add_image(img, f'A{current_row}')
            except Exception as e:
                print(f"Warning: Could not insert main chart: {e}")

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15

    def _create_inputs_tab(self, ws, financial_inputs, deal_structure, synergies):

        """Create Inputs & Assumptions tab"""

        # Title
        ws['A1'] = f"Inputs & Assumptions: {self.buyer_company} + {self.seller_company}"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:E1')

        current_row = 3

        # Buyer Financials
        ws[f'A{current_row}'] = "BUYER COMPANY FINANCIALS"
        ws[f'A{current_row}'].style = 'company_header'
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 2

        buyer_data = [
            ("Company Name", financial_inputs['buyer']['company']),
            ("Net Income ($M)", financial_inputs['buyer']['net_income']),
            ("EPS ($)", financial_inputs['buyer']['eps']),
            ("Shares Outstanding (M)", financial_inputs['buyer']['shares_outstanding']),
            ("P/E Ratio", financial_inputs['buyer']['pe_ratio']),
            ("Market Cap ($M)", financial_inputs['buyer']['market_cap'])
        ]

        for label, value in buyer_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].style = 'input'
            current_row += 1

        current_row += 2

        # Seller Financials
        ws[f'A{current_row}'] = "SELLER COMPANY FINANCIALS"
        ws[f'A{current_row}'].style = 'company_header'
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 2

        seller_data = [
            ("Company Name", financial_inputs['seller']['company']),
            ("Net Income ($M)", financial_inputs['seller']['net_income']),
            ("EPS ($)", financial_inputs['seller']['eps']),
            ("Shares Outstanding (M)", financial_inputs['seller']['shares_outstanding']),
            ("P/E Ratio", financial_inputs['seller']['pe_ratio']),
            ("Market Cap ($M)", financial_inputs['seller']['market_cap'])
        ]

        for label, value in seller_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].style = 'input'
            current_row += 1

        current_row += 2

        # Deal Structure
        ws[f'A{current_row}'] = "DEAL STRUCTURE"
        ws[f'A{current_row}'].style = 'company_header'
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 2

        deal_data = [
            ("Purchase Premium (%)", deal_structure['purchase_premium_pct']),
            ("Cash Consideration (%)", deal_structure['consideration_mix']['cash_pct']),
            ("Stock Consideration (%)", deal_structure['consideration_mix']['stock_pct']),
            ("Debt Financing (%)", deal_structure['consideration_mix']['debt_pct']),
            ("Interest Rate on Debt (%)", deal_structure['financing']['interest_rate']),
            ("Tax Rate (%)", deal_structure['financing']['tax_rate'])
        ]

        for label, value in deal_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].style = 'input'
            current_row += 1

        current_row += 2

        # Synergies
        ws[f'A{current_row}'] = "SYNERGIES & COST SAVINGS"
        ws[f'A{current_row}'].style = 'company_header'
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 2

        synergies_data = [
            ("Revenue Synergies ($M)", synergies['revenue_synergies']),
            ("Cost Synergies ($M)", synergies['cost_synergies']),
            ("One-time Integration Costs ($M)", synergies['one_time_costs']),
            ("Total Annual Synergies ($M)", synergies['total_annual_synergies'])
        ]

        for label, value in synergies_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].style = 'input'
            current_row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15

    def _create_calculations_tab(self, ws, purchase_price_calc, financing_calc, proforma_calc):

        """Create Calculations tab"""

        # Title
        ws['A1'] = f"Pro Forma Calculations: {self.buyer_company} + {self.seller_company}"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:G1')

        current_row = 3

        # Purchase Price Calculations
        ws[f'A{current_row}'] = "PURCHASE PRICE CALCULATIONS"
        ws[f'A{current_row}'].style = 'company_header'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        purchase_data = [
            ("Seller Market Cap ($M)", purchase_price_calc['seller_market_cap']),
            ("Purchase Premium (%)", purchase_price_calc['purchase_premium_pct']),
            ("Equity Purchase Price ($M)", purchase_price_calc['equity_purchase_price']),
            ("Enterprise Purchase Price ($M)", purchase_price_calc['enterprise_purchase_price']),
            ("Cash Consideration ($M)", purchase_price_calc['consideration_breakdown']['cash']),
            ("Stock Consideration ($M)", purchase_price_calc['consideration_breakdown']['stock']),
            ("Debt Consideration ($M)", purchase_price_calc['consideration_breakdown']['debt']),
            ("Shares to Issue (M)", purchase_price_calc['shares_to_issue'])
        ]

        for label, value in purchase_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].style = 'calculation'
            current_row += 1

        current_row += 2

        # Financing Calculations
        ws[f'A{current_row}'] = "FINANCING CALCULATIONS"
        ws[f'A{current_row}'].style = 'company_header'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        financing_data = [
            ("New Debt Amount ($M)", financing_calc['new_debt_amount']),
            ("Annual Interest Expense ($M)", financing_calc['annual_interest_expense']),
            ("Tax Shield ($M)", financing_calc['tax_shield']),
            ("Net Interest Expense ($M)", financing_calc['net_interest_expense']),
            ("Cash Used ($M)", financing_calc['cash_used']),
            ("Foregone Interest Income ($M)", financing_calc['foregone_interest_income']),
            ("Net Financing Impact ($M)", financing_calc['net_financing_impact'])
        ]

        for label, value in financing_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].style = 'calculation'
            current_row += 1

        current_row += 2

        # Pro Forma Calculations
        ws[f'A{current_row}'] = "PRO FORMA FINANCIAL CALCULATIONS"
        ws[f'A{current_row}'].style = 'company_header'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        proforma_data = [
            ("Buyer Standalone Net Income ($M)", proforma_calc['buyer_standalone_net_income']),
            ("Seller Standalone Net Income ($M)", proforma_calc['seller_standalone_net_income']),
            ("Synergies Impact ($M)", proforma_calc['synergies_impact']),
            ("Financing Impact ($M)", proforma_calc['financing_impact']),
            ("One-time Costs Impact ($M)", proforma_calc['one_time_impact']),
            ("Pro Forma Net Income ($M)", proforma_calc['proforma_net_income']),
            ("Buyer Shares (M)", proforma_calc['buyer_shares']),
            ("Shares Issued (M)", proforma_calc['shares_issued']),
            ("Pro Forma Shares (M)", proforma_calc['proforma_shares']),
            ("Buyer Standalone EPS ($)", proforma_calc['buyer_standalone_eps']),
            ("Pro Forma EPS ($)", proforma_calc['proforma_eps'])
        ]

        for label, value in proforma_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].style = 'calculation'
            current_row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20

    def _create_sensitivity_tab(self, ws, sensitivity_analysis):

        """Create Sensitivity Analysis tab"""

        # Title
        ws['A1'] = f"Sensitivity Analysis: {self.buyer_company} Acquiring {self.seller_company}"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:H1')

        current_row = 3

        # Premium Sensitivity
        if sensitivity_analysis['premium_sensitivity']:
            ws[f'A{current_row}'] = "PREMIUM SENSITIVITY ANALYSIS"
            ws[f'A{current_row}'].style = 'company_header'
            ws.merge_cells(f'A{current_row}:H{current_row}')
            current_row += 2

            # Headers
            headers = ['Premium (%)', 'Purchase Price ($M)', 'Pro Forma EPS ($)', 'Accretion/(Dilution) (%)', 'Status']
            for col, header in enumerate(headers, 1):
                ws.cell(row=current_row, column=col, value=header).style = 'header'
            current_row += 1

            # Data
            for item in sensitivity_analysis['premium_sensitivity']:
                ws.cell(row=current_row, column=1, value=item['premium_pct']).style = 'calculation'
                ws.cell(row=current_row, column=2, value=item['purchase_price']).style = 'calculation'
                ws.cell(row=current_row, column=3, value=item['proforma_eps']).style = 'calculation'
                ws.cell(row=current_row, column=4, value=item['accretion_dilution_pct']).style = 'accretion' if item['is_accretive'] else 'dilution'
                ws.cell(row=current_row, column=5, value="ACCRETIVE" if item['is_accretive'] else "DILUTIVE").style = 'accretion' if item['is_accretive'] else 'dilution'
                current_row += 1

        current_row += 2

        # Synergies Sensitivity
        if sensitivity_analysis['synergies_sensitivity']:
            ws[f'A{current_row}'] = "SYNERGIES SENSITIVITY ANALYSIS"
            ws[f'A{current_row}'].style = 'company_header'
            ws.merge_cells(f'A{current_row}:H{current_row}')
            current_row += 2

            # Headers
            headers = ['Synergies ($M)', 'Pro Forma EPS ($)', 'Accretion/(Dilution) (%)', 'Status']
            for col, header in enumerate(headers, 1):
                ws.cell(row=current_row, column=col, value=header).style = 'header'
            current_row += 1

            # Data
            for item in sensitivity_analysis['synergies_sensitivity']:
                ws.cell(row=current_row, column=1, value=item['synergies_amount']).style = 'calculation'
                ws.cell(row=current_row, column=2, value=item['proforma_eps']).style = 'calculation'
                ws.cell(row=current_row, column=3, value=item['accretion_dilution_pct']).style = 'accretion' if item['is_accretive'] else 'dilution'
                ws.cell(row=current_row, column=4, value="ACCRETIVE" if item['is_accretive'] else "DILUTIVE").style = 'accretion' if item['is_accretive'] else 'dilution'
                current_row += 1

        # Set column widths
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _display_accretion_results(self, accretion_dilution, sensitivity_analysis):

        """Display key accretion/dilution results"""

        print("\n🎯 Accretion/Dilution Results:")
        print("=" * 80)

        # Main results
        print(f"Buyer Standalone EPS: ${accretion_dilution['buyer_standalone_eps']:.2f}")
        print(f"Pro Forma EPS: ${accretion_dilution['proforma_eps']:.2f}")
        print(f"EPS Impact: ${accretion_dilution['eps_impact']:.2f}")
        print(f"Accretion/(Dilution): {accretion_dilution['accretion_dilution_pct']:.2%}")
        print(f"Status: {accretion_dilution['status']}")

        # Premium sensitivity summary
        if sensitivity_analysis['premium_sensitivity']:
            print("\n📊 Premium Sensitivity:")
            premium_range = sensitivity_analysis['premium_sensitivity']
            accretive_count = sum(1 for item in premium_range if item['is_accretive'])
            print(f"   • Accretive scenarios: {accretive_count}/{len(premium_range)}")
            if premium_range:
                min_premium = min(item['accretion_dilution_pct'] for item in premium_range)
                max_premium = max(item['accretion_dilution_pct'] for item in premium_range)
                print(f"   • Range: {min_premium:.1%} to {max_premium:.1%}")

        # Synergies sensitivity summary
        if sensitivity_analysis['synergies_sensitivity']:
            print("\n💰 Synergies Sensitivity:")
            synergies_range = sensitivity_analysis['synergies_sensitivity']
            accretive_count = sum(1 for item in synergies_range if item['is_accretive'])
            print(f"   • Accretive scenarios: {accretive_count}/{len(synergies_range)}")
            if synergies_range:
                min_synergies = min(item['accretion_dilution_pct'] for item in synergies_range)
                max_synergies = max(item['accretion_dilution_pct'] for item in synergies_range)
                print(f"   • Range: {min_synergies:.1%} to {max_synergies:.1%}")

        print("\n💡 Key Insights:")
        print("=" * 80)
        if accretion_dilution['is_accretive']:
            print("✅ DEAL IS ACCRETIVE - Should be favorable for buyer shareholders")
        else:
            print("❌ DEAL IS DILUTIVE - May require synergies or premium adjustment")

        print(f"📈 Deal creates {abs(accretion_dilution['accretion_dilution_pct'])*100:.1f}¢ {'increase' if accretion_dilution['is_accretive'] else 'decrease'} in EPS")
        print("🎯 Focus on synergies and purchase price to optimize accretion")


def run_sample_accretion_dilution_model():
    """Run a sample accretion/dilution analysis"""

    print("📊 Running Professional Accretion/Dilution Model Sample")
    print("=" * 70)

    # Create model instance
    model = ProfessionalAccretionDilutionModel("TechCorp Inc.", "StartUp Ltd.", "TECHCORP", "STARTUP")

    # Run the model with sample data
    results = model.run_accretion_dilution_model(
        # Buyer financials (large tech company)
        buyer_net_income=1500.0,      # $1.5B NI
        buyer_eps=4.50,              # $4.50 EPS
        buyer_shares_outstanding=333.3,  # 333M shares
        buyer_pe_ratio=20.0,          # 20x P/E

        # Seller financials (smaller startup)
        seller_net_income=50.0,       # $50M NI
        seller_shares_outstanding=20.0,  # 20M shares
        seller_pe_ratio=25.0,         # 25x P/E

        # Deal structure (25% premium, mixed consideration)
        purchase_premium_pct=0.25,    # 25% premium
        cash_pct=0.60,                # 60% cash
        stock_pct=0.30,               # 30% stock
        debt_pct=0.10,                # 10% debt

        # Financing assumptions
        interest_rate=0.05,           # 5% interest rate
        tax_rate=0.25,                # 25% tax rate

        # Synergies and costs
        revenue_synergies=75.0,       # $75M revenue synergies
        cost_synergies=125.0,         # $125M cost synergies
        one_time_costs=25.0           # $25M integration costs
    )

    return results


if __name__ == "__main__":
    # Run sample accretion/dilution model
    results = run_sample_accretion_dilution_model()

    print("\n📋 Accretion/Dilution Model Complete!")
    print(f"📁 Excel file saved as: {results['excel_file']}")

    # Display key results
    print("\n🔑 Key Results Summary:")
    print("=" * 40)

    accretion_dilution = results['accretion_dilution']
    print(f"Buyer Standalone EPS: ${accretion_dilution['buyer_standalone_eps']:.2f}")
    print(f"Pro Forma EPS: ${accretion_dilution['proforma_eps']:.2f}")
    print(f"EPS Impact: ${accretion_dilution['eps_impact']:.2f}")
    print(f"Accretion/(Dilution): {accretion_dilution['accretion_dilution_pct']:.2%}")
    print(f"Status: {accretion_dilution['status']}")

    # Show pro forma calculations
    proforma_calc = results['proforma_calc']
    print(f"Pro Forma Net Income: ${proforma_calc['proforma_net_income']:.0f}M")
    print(f"Pro Forma Shares: {proforma_calc['proforma_shares']:.1f}M")
