#!/usr/bin/env python3
"""
Simple DCF Model Builder - Excel Only Version
No Google Sheets dependencies, just creates Excel files.
"""

import sys
import subprocess

def install_and_import(package, pip_name=None):
    pip_name = pip_name or package
    try:
        __import__(package)
    except ImportError:
        print(f"Installing missing package: {pip_name}")
        subprocess.check_call([sys.executable, "-m", "pip", "install", pip_name])
        __import__(package)

# Install required packages
REQUIRED_MODULES = [
    ("pandas", "pandas"),
    ("openpyxl", "openpyxl"),
    ("requests", "requests"),
    ("bs4", "beautifulsoup4"),
    ("yfinance", "yfinance"),
    ("openai", "openai"),
]
for mod, pip_name in REQUIRED_MODULES:
    install_and_import(mod, pip_name)

import os
import re
import requests
import pandas as pd
from bs4 import BeautifulSoup, Tag
from datetime import datetime
from openpyxl import Workbook
from urllib.parse import urljoin

# Optional imports
try:
    import openai
except ImportError:
    openai = None
try:
    import yfinance as yf
except ImportError:
    yf = None

# Load environment variables
from dotenv import load_dotenv
load_dotenv()
OPENAI_API_KEY = os.getenv('OPENAI_API_KEY')
if openai and OPENAI_API_KEY:
    openai.api_key = OPENAI_API_KEY

# Configuration
YEARS = 5
DISCOUNT_RATE = 0.10
TERMINAL_GROWTH = 0.03

def get_yfinance_financials(ticker):
    """Get financial data from yfinance for public companies."""
    if not yf:
        print("yfinance not available")
        return {}
    try:
        ticker_obj = yf.Ticker(ticker)
        fin = ticker_obj.financials.T
        years = fin.index[-YEARS:]
        data = {}
        if 'Total Revenue' in fin.columns:
            data['Revenue'] = [float(fin.loc[y, 'Total Revenue']) for y in years]
        if 'EBITDA' in fin.columns:
            data['EBITDA'] = [float(fin.loc[y, 'EBITDA']) for y in years]
        if 'Operating Income' in fin.columns:
            data['EBIT'] = [float(fin.loc[y, 'Operating Income']) for y in years]
        return data
    except Exception as e:
        print(f"yfinance extraction failed: {e}")
        return {}

def extract_financials_with_llm(company_name):
    """Extract financials using OpenAI."""
    if not openai or not OPENAI_API_KEY:
        print("OpenAI not available")
        return {}
    prompt = f"""
Extract the most recent annual financials for {company_name}. Return a JSON object with keys: Revenue, EBITDA, EBIT, Net Income, Depreciation, CapEx, Working Capital. Use 'Unknown' if not found.
"""
    try:
        response = openai.chat.completions.create(
            model="gpt-3.5-turbo-16k",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.0,
            max_tokens=600,
        )
        import json
        content = response.choices[0].message.content
        if content:
            return json.loads(content)
        else:
            return {}
    except Exception as e:
        print(f"LLM extraction failed: {e}")
        return {}

def project_financials(base, growth=0.08, years=YEARS):
    """Project financials into the future."""
    return [round(base * ((1 + growth) ** i), 2) for i in range(years)]

def calculate_fcf(revenue, ebitda, ebit, taxes, depreciation, capex, delta_wc):
    """Calculate Free Cash Flow."""
    tax_rate = 0.25
    fcf = []
    for i in range(len(revenue)):
        EBIT = ebit[i] if ebit else revenue[i] * 0.15
        Taxes = EBIT * tax_rate
        Dep = depreciation[i] if depreciation else revenue[i] * 0.05
        Cap = capex[i] if capex else revenue[i] * 0.06
        DWC = delta_wc[i] if delta_wc else revenue[i] * 0.02
        fcf.append(round(EBIT - Taxes + Dep - Cap - DWC, 2))
    return fcf

def calculate_terminal_value(last_fcf, discount_rate=DISCOUNT_RATE, growth=TERMINAL_GROWTH):
    """Calculate terminal value."""
    return last_fcf * (1 + growth) / (discount_rate - growth)

def discount_cash_flows(fcfs, terminal_value, discount_rate=DISCOUNT_RATE):
    """Discount cash flows to present value."""
    discounted = [fcf / ((1 + discount_rate) ** (i + 1)) for i, fcf in enumerate(fcfs)]
    tv_discounted = terminal_value / ((1 + discount_rate) ** len(fcfs))
    return discounted, tv_discounted

def write_to_excel(company, years, revenue, ebitda, ebit, fcf, terminal_value, enterprise_value, filename):
    """Write DCF model to Excel file."""
    wb = Workbook()
    ws = wb.active
    if ws:
        ws.title = "DCF Model"
        
        # Add data
        ws.append(["Company", company])
        ws.append([])
        ws.append(["Year"] + years)
        ws.append(["Revenue"] + revenue)
        ws.append(["EBITDA"] + ebitda)
        ws.append(["EBIT"] + ebit)
        ws.append(["FCF"] + fcf)
        ws.append([])
        ws.append(["Terminal Value", terminal_value])
        ws.append(["Enterprise Value", enterprise_value])
        
        # Add some formatting
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Format headers
        for cell in ws[3]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Save file
        wb.save(filename)
        print(f"‚úÖ Saved DCF model to {filename}")
    else:
        print("‚ùå Could not create Excel worksheet")

def build_dcf_model(company_name, ticker=None):
    """Build DCF model for a company."""
    print(f"Building DCF model for {company_name}")
    
    # Get financial data
    data = get_yfinance_financials(ticker) if ticker else {}
    if not data or not data.get('Revenue'):
        print("Falling back to LLM extraction...")
        data = extract_financials_with_llm(company_name)
    
    # Use default values if no data found
    if not data or not data.get('Revenue'):
        print("Using default financial assumptions...")
        base_revenue = 1000000  # $1M default
    else:
        base_revenue = float(str(data.get('Revenue', [0])[0] if isinstance(data.get('Revenue'), list) else data.get('Revenue', 0)).replace(',', '').replace('$', ''))
    
    # Project financials
    revenue_proj = project_financials(base_revenue)
    ebitda_proj = [r * 0.25 for r in revenue_proj]
    ebit_proj = [r * 0.15 for r in revenue_proj]
    fcf_proj = calculate_fcf(revenue_proj, ebitda_proj, ebit_proj, None, None, None, None)
    
    # Calculate terminal value and enterprise value
    terminal_value = calculate_terminal_value(fcf_proj[-1])
    discounted_fcfs, discounted_tv = discount_cash_flows(fcf_proj, terminal_value)
    enterprise_value = round(sum(discounted_fcfs) + discounted_tv, 2)
    
    # Create years
    years = [datetime.now().year + i for i in range(YEARS)]
    
    # Save to Excel
    filename = f"dcf_{company_name.replace(' ', '_')}.xlsx"
    write_to_excel(company_name, years, revenue_proj, ebitda_proj, ebit_proj, fcf_proj, terminal_value, enterprise_value, filename)
    
    # Print summary
    print(f"\nüìä DCF Model Summary for {company_name}")
    print(f"Base Revenue: ${base_revenue:,.0f}")
    print(f"Terminal Value: ${terminal_value:,.0f}")
    print(f"Enterprise Value: ${enterprise_value:,.0f}")
    print(f"Discount Rate: {DISCOUNT_RATE*100}%")
    print(f"Terminal Growth: {TERMINAL_GROWTH*100}%")

if __name__ == "__main__":
    print("üöÄ Simple DCF Model Builder")
    print("=" * 40)
    
    company_name = input("Enter company name: ").strip()
    ticker = input("Enter ticker symbol (optional): ").strip() or None
    
    build_dcf_model(company_name, ticker) 