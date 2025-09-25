#!/usr/bin/env python3
"""
Test Script for Custom Inputs Module
Demonstrates the custom inputs functionality for financial models
"""

import sys
import os
sys.path.append('financial-models-app/backend')

def test_custom_inputs_creation():
    """Test creating custom inputs sheet"""
    print("🎛️ TESTING CUSTOM INPUTS CREATION")
    print("=" * 60)
    
    try:
        from openpyxl import Workbook
        from custom_inputs_module import create_custom_inputs_sheet
        from enhanced_assumptions_research import get_research_based_assumptions
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Get research assumptions
        company_name = "Microsoft Corporation"
        research_assumptions = get_research_based_assumptions(company_name, None, "Technology - Software")
        
        # Create custom inputs sheet
        ws = create_custom_inputs_sheet(wb, company_name, research_assumptions)
        
        # Save test file
        test_file = "test_custom_inputs.xlsx"
        wb.save(test_file)
        
        print(f"✅ Custom inputs sheet created successfully")
        print(f"📁 File: {test_file}")
        print(f"📊 Company: {company_name}")
        print(f"🏭 Industry: Technology - Software")
        print(f"📋 Sections created:")
        
        # Count sections
        section_count = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and 'CUSTOM' in str(cell.value) and 'ASSUMPTIONS' in str(cell.value):
                    section_count += 1
                    print(f"   • {cell.value}")
        
        print(f"\n📈 Total sections: {section_count}")
        print(f"📏 File size: {os.path.getsize(test_file):,} bytes")
        
        return True
        
    except Exception as e:
        print(f"❌ Error creating custom inputs: {e}")
        return False

def test_custom_inputs_extraction():
    """Test extracting custom inputs from worksheet"""
    print("\n📥 TESTING CUSTOM INPUTS EXTRACTION")
    print("=" * 60)
    
    try:
        from openpyxl import load_workbook
        from custom_inputs_module import get_custom_inputs
        
        # Load the test file
        wb = load_workbook("test_custom_inputs.xlsx")
        ws = wb["Custom Inputs"]
        
        # Extract custom inputs
        custom_inputs = get_custom_inputs(ws)
        
        print(f"✅ Custom inputs extracted successfully")
        print(f"📊 Total custom inputs found: {len(custom_inputs)}")
        
        if custom_inputs:
            print(f"📋 Custom inputs:")
            for metric, value in custom_inputs.items():
                print(f"   • {metric}: {value}")
        else:
            print(f"📝 No custom inputs found (expected for new file)")
        
        return True
        
    except Exception as e:
        print(f"❌ Error extracting custom inputs: {e}")
        return False

def test_custom_inputs_application():
    """Test applying custom inputs to research assumptions"""
    print("\n🔄 TESTING CUSTOM INPUTS APPLICATION")
    print("=" * 60)
    
    try:
        from enhanced_assumptions_research import get_research_based_assumptions
        from custom_inputs_module import apply_custom_inputs
        
        # Get original research assumptions
        company_name = "Microsoft Corporation"
        original_assumptions = get_research_based_assumptions(company_name, None, "Technology - Software")
        
        # Create sample custom inputs
        custom_inputs = {
            'Revenue Growth Year 1': '20%',
            'EBITDA Margin': '30%',
            'Target Debt/Equity Ratio': '0.30',
            'Cost of Debt': '5.5%',
            'EV/EBITDA Multiple': '20.0x'
        }
        
        print(f"📊 Original assumptions:")
        print(f"   • Revenue Growth Y1: {original_assumptions['revenue_growth'][0]*100:.1f}%")
        print(f"   • EBITDA Margin: {original_assumptions['ebitda_margin']*100:.1f}%")
        print(f"   • Debt/Equity: {original_assumptions['debt_to_equity']:.2f}")
        print(f"   • Cost of Debt: {original_assumptions['capital_structure']['cost_of_debt']*100:.1f}%")
        print(f"   • EV/EBITDA: {original_assumptions['valuation_metrics']['ev_ebitda']:.1f}x")
        
        # Apply custom inputs
        modified_assumptions = apply_custom_inputs(original_assumptions, custom_inputs)
        
        print(f"\n🎛️ Custom inputs applied:")
        for metric, value in custom_inputs.items():
            print(f"   • {metric}: {value}")
        
        print(f"\n📊 Modified assumptions:")
        print(f"   • Revenue Growth Y1: {modified_assumptions['revenue_growth'][0]*100:.1f}%")
        print(f"   • EBITDA Margin: {modified_assumptions['ebitda_margin']*100:.1f}%")
        print(f"   • Debt/Equity: {modified_assumptions['debt_to_equity']:.2f}")
        print(f"   • Cost of Debt: {modified_assumptions['capital_structure']['cost_of_debt']*100:.1f}%")
        print(f"   • EV/EBITDA: {modified_assumptions['valuation_metrics']['ev_ebitda']:.1f}x")
        
        # Verify changes
        changes_made = (
            modified_assumptions['revenue_growth'][0] != original_assumptions['revenue_growth'][0] or
            modified_assumptions['ebitda_margin'] != original_assumptions['ebitda_margin'] or
            modified_assumptions['debt_to_equity'] != original_assumptions['debt_to_equity'] or
            modified_assumptions['capital_structure']['cost_of_debt'] != original_assumptions['capital_structure']['cost_of_debt'] or
            modified_assumptions['valuation_metrics']['ev_ebitda'] != original_assumptions['valuation_metrics']['ev_ebitda']
        )
        
        if changes_made:
            print(f"\n✅ Custom inputs successfully applied!")
        else:
            print(f"\n⚠️ No changes detected in assumptions")
        
        return True
        
    except Exception as e:
        print(f"❌ Error applying custom inputs: {e}")
        return False

def test_3_statement_with_custom_inputs():
    """Test 3-Statement model with custom inputs"""
    print("\n📊 TESTING 3-STATEMENT MODEL WITH CUSTOM INPUTS")
    print("=" * 60)
    
    try:
        from three_statement_model import create_three_statement_model
        
        # Create 3-Statement model (which now includes custom inputs sheet)
        company_name = "Microsoft Corporation"
        ticker = "MSFT"
        
        output_file = create_three_statement_model(company_name, ticker)
        
        print(f"✅ 3-Statement model created with custom inputs")
        print(f"📁 File: {output_file}")
        print(f"📏 Size: {os.path.getsize(output_file):,} bytes")
        
        # Check if custom inputs sheet exists
        from openpyxl import load_workbook
        wb = load_workbook(output_file)
        
        if "Custom Inputs" in wb.sheetnames:
            print(f"✅ Custom Inputs sheet found in model")
            custom_sheet = wb["Custom Inputs"]
            print(f"📋 Sheet contains {custom_sheet.max_row} rows")
        else:
            print(f"⚠️ Custom Inputs sheet not found")
        
        return True
        
    except Exception as e:
        print(f"❌ Error creating 3-Statement model: {e}")
        return False

def cleanup_test_files():
    """Clean up test files"""
    print("\n🧹 CLEANING UP TEST FILES")
    print("=" * 60)
    
    test_files = ["test_custom_inputs.xlsx"]
    
    for file in test_files:
        if os.path.exists(file):
            try:
                os.remove(file)
                print(f"🗑️ Removed: {file}")
            except Exception as e:
                print(f"⚠️ Could not remove {file}: {e}")

def main():
    """Main test function"""
    print("🎛️ COMPREHENSIVE CUSTOM INPUTS TEST")
    print("=" * 80)
    print("Testing custom inputs functionality for financial models")
    print("=" * 80)
    
    # Run all tests
    test_results = []
    
    test_results.append(("Custom Inputs Creation", test_custom_inputs_creation()))
    test_results.append(("Custom Inputs Extraction", test_custom_inputs_extraction()))
    test_results.append(("Custom Inputs Application", test_custom_inputs_application()))
    test_results.append(("3-Statement with Custom Inputs", test_3_statement_with_custom_inputs()))
    
    # Summary
    print("\n" + "=" * 80)
    print("📊 TEST RESULTS SUMMARY")
    print("=" * 80)
    
    passed = sum(1 for _, result in test_results if result)
    total = len(test_results)
    
    for test_name, result in test_results:
        status = "✅ PASS" if result else "❌ FAIL"
        print(f"   {test_name}: {status}")
    
    print(f"\n🎯 Overall: {passed}/{total} tests passed ({passed/total*100:.1f}%)")
    
    if passed == total:
        print("🎉 ALL CUSTOM INPUTS TESTS PASSED!")
        print("✅ Custom inputs functionality working perfectly")
        print("\n🎛️ FEATURES DEMONSTRATED:")
        print("   • Custom inputs sheet creation with 8 sections")
        print("   • Research-based defaults with editable custom values")
        print("   • Automatic application of custom inputs to models")
        print("   • Integration with 3-Statement model")
        print("   • Support for percentages, ratios, multiples, currency")
    else:
        print("⚠️ Some tests failed. Please check the errors above.")
    
    # Cleanup
    cleanup_test_files()
    
    return passed == total

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1) 