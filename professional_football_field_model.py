#!/usr/bin/env python3
"""
Professional Football Field Valuation Model
Synthesis tool that aggregates valuation ranges from multiple methodologies

Author: Investment Banking Modeler
Date: 2024

Features:
- Aggregates DCF, LBO, Trading Comps, Precedent Transactions
- Professional horizontal bar chart (Football Field)
- Enterprise Value and Equity Value views
- Banker-standard formatting and color coding
- Excel output with multiple tabs
- Integration with existing valuation models
"""

import numpy as np
import pandas as pd
from datetime import datetime
try:
    import matplotlib.pyplot as plt
    import matplotlib.patches as patches
    HAS_MATPLOTLIB = True
except ImportError:
    HAS_MATPLOTLIB = False
    print("⚠️  Warning: matplotlib not available. Charts will not be generated.")
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import warnings
warnings.filterwarnings('ignore')

# Professional color scheme for Football Field
FOOTBALL_FIELD_COLORS = {
    'header_blue': '1F4E79',
    'input_light_blue': 'DDEBF7',
    'calculation_black': '000000',
    'output_green': 'D8E4BC',
    'warning_red': 'F8CECC',
    'border_dark': '2C3E50',
    'dcf_orange': 'FF7F50',        # DCF methodology
    'trading_blue': '4682B4',     # Trading Comps
    'precedent_green': '32CD32',  # Precedent Transactions
    'lbo_purple': '9370DB',       # LBO Analysis
    'other_gray': '708090',       # Other methodologies
    'text_dark': '000000',
    'text_white': 'FFFFFF',
    'chart_background': 'F5F5F5',
    'grid_light': 'E0E0E0'
}

class ProfessionalFootballFieldModel:
    """
    Comprehensive Football Field Valuation Model with Professional Charting
    """

    def __init__(self, target_company="Target Company", target_ticker="TARGET"):
        self.target_company = target_company
        self.target_ticker = target_ticker
        self.model_date = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Initialize styles
        self.styles = self._create_styles()

    def _create_styles(self):
        """Create professional Excel styles"""
        styles = {}

        # Header style
        styles['header'] = NamedStyle(name='header')
        styles['header'].font = Font(name='Calibri', size=12, bold=True, color=FOOTBALL_FIELD_COLORS['text_white'])
        styles['header'].fill = PatternFill(start_color=FOOTBALL_FIELD_COLORS['header_blue'], end_color=FOOTBALL_FIELD_COLORS['header_blue'], fill_type='solid')
        styles['header'].alignment = Alignment(horizontal='center', vertical='center')

        # Input style
        styles['input'] = NamedStyle(name='input')
        styles['input'].font = Font(name='Calibri', size=10, color=FOOTBALL_FIELD_COLORS['text_dark'])
        styles['input'].fill = PatternFill(start_color=FOOTBALL_FIELD_COLORS['input_light_blue'], end_color=FOOTBALL_FIELD_COLORS['input_light_blue'], fill_type='solid')
        styles['input'].alignment = Alignment(horizontal='right', vertical='center')
        styles['input'].number_format = '#,##0.00'

        # Calculation style
        styles['calculation'] = NamedStyle(name='calculation')
        styles['calculation'].font = Font(name='Calibri', size=10, color=FOOTBALL_FIELD_COLORS['calculation_black'])
        styles['calculation'].alignment = Alignment(horizontal='right', vertical='center')
        styles['calculation'].number_format = '#,##0.00'

        # Output style
        styles['output'] = NamedStyle(name='output')
        styles['output'].font = Font(name='Calibri', size=10, bold=True, color=FOOTBALL_FIELD_COLORS['text_dark'])
        styles['output'].fill = PatternFill(start_color=FOOTBALL_FIELD_COLORS['output_green'], end_color=FOOTBALL_FIELD_COLORS['output_green'], fill_type='solid')
        styles['output'].alignment = Alignment(horizontal='right', vertical='center')
        styles['output'].number_format = '#,##0.00'

        # Methodology styles
        styles['dcf'] = NamedStyle(name='dcf')
        styles['dcf'].font = Font(name='Calibri', size=10, color=FOOTBALL_FIELD_COLORS['text_white'])
        styles['dcf'].fill = PatternFill(start_color=FOOTBALL_FIELD_COLORS['dcf_orange'], end_color=FOOTBALL_FIELD_COLORS['dcf_orange'], fill_type='solid')
        styles['dcf'].alignment = Alignment(horizontal='center', vertical='center')

        styles['trading'] = NamedStyle(name='trading')
        styles['trading'].font = Font(name='Calibri', size=10, color=FOOTBALL_FIELD_COLORS['text_white'])
        styles['trading'].fill = PatternFill(start_color=FOOTBALL_FIELD_COLORS['trading_blue'], end_color=FOOTBALL_FIELD_COLORS['trading_blue'], fill_type='solid')
        styles['trading'].alignment = Alignment(horizontal='center', vertical='center')

        styles['precedent'] = NamedStyle(name='precedent')
        styles['precedent'].font = Font(name='Calibri', size=10, color=FOOTBALL_FIELD_COLORS['text_white'])
        styles['precedent'].fill = PatternFill(start_color=FOOTBALL_FIELD_COLORS['precedent_green'], end_color=FOOTBALL_FIELD_COLORS['precedent_green'], fill_type='solid')
        styles['precedent'].alignment = Alignment(horizontal='center', vertical='center')

        styles['lbo'] = NamedStyle(name='lbo')
        styles['lbo'].font = Font(name='Calibri', size=10, color=FOOTBALL_FIELD_COLORS['text_white'])
        styles['lbo'].fill = PatternFill(start_color=FOOTBALL_FIELD_COLORS['lbo_purple'], end_color=FOOTBALL_FIELD_COLORS['lbo_purple'], fill_type='solid')
        styles['lbo'].alignment = Alignment(horizontal='center', vertical='center')

        # Warning style
        styles['warning'] = NamedStyle(name='warning')
        styles['warning'].font = Font(name='Calibri', size=10, color=FOOTBALL_FIELD_COLORS['text_dark'])
        styles['warning'].fill = PatternFill(start_color=FOOTBALL_FIELD_COLORS['warning_red'], end_color=FOOTBALL_FIELD_COLORS['warning_red'], fill_type='solid')
        styles['warning'].alignment = Alignment(horizontal='center', vertical='center')

        # Normal label style
        styles['label'] = NamedStyle(name='label')
        styles['label'].font = Font(name='Calibri', size=10, color=FOOTBALL_FIELD_COLORS['text_dark'])
        styles['label'].alignment = Alignment(horizontal='left', vertical='center')

        # Bold label style
        styles['label_bold'] = NamedStyle(name='label_bold')
        styles['label_bold'].font = Font(name='Calibri', size=10, bold=True, color=FOOTBALL_FIELD_COLORS['text_dark'])
        styles['label_bold'].alignment = Alignment(horizontal='left', vertical='center')

        # Summary box style
        styles['summary_box'] = NamedStyle(name='summary_box')
        styles['summary_box'].font = Font(name='Calibri', size=11, bold=True, color=FOOTBALL_FIELD_COLORS['text_white'])
        styles['summary_box'].fill = PatternFill(start_color=FOOTBALL_FIELD_COLORS['header_blue'], end_color=FOOTBALL_FIELD_COLORS['header_blue'], fill_type='solid')
        styles['summary_box'].alignment = Alignment(horizontal='center', vertical='center')
        styles['summary_box'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        return styles

    def run_football_field_model(self,
                               # Target Company Info
                               target_company=None,
                               target_ticker=None,

                               # DCF Valuation Ranges ($M)
                               dcf_ev_low=None,
                               dcf_ev_median=None,
                               dcf_ev_high=None,
                               dcf_equity_low=None,
                               dcf_equity_median=None,
                               dcf_equity_high=None,

                               # Trading Comps Valuation Ranges ($M)
                               trading_ev_low=None,
                               trading_ev_median=None,
                               trading_ev_high=None,
                               trading_equity_low=None,
                               trading_equity_median=None,
                               trading_equity_high=None,

                               # Precedent Transactions Valuation Ranges ($M)
                               precedent_ev_low=None,
                               precedent_ev_median=None,
                               precedent_ev_high=None,
                               precedent_equity_low=None,
                               precedent_equity_median=None,
                               precedent_equity_high=None,

                               # LBO Valuation Ranges ($M)
                               lbo_equity_low=None,
                               lbo_equity_median=None,
                               lbo_equity_high=None,

                               # Other Methodologies (optional)
                               other_methods=None,  # dict with method_name: {'ev_low', 'ev_high', 'equity_low', 'equity_high'}

                               # Current Share Price and Shares Outstanding
                               current_share_price=None,
                               shares_outstanding=None):

        """
        Run complete Football Field valuation model
        """

        if target_company:
            self.target_company = target_company
        if target_ticker:
            self.target_ticker = target_ticker

        print(f"🏟️  Building Professional Football Field Valuation Model for {self.target_company} ({self.target_ticker})")
        print("=" * 90)

        # Set default values if not provided
        if dcf_ev_low is None:
            # Default sample data
            dcf_ev_low, dcf_ev_median, dcf_ev_high = 2500.0, 3200.0, 4000.0
            dcf_equity_low, dcf_equity_median, dcf_equity_high = 2200.0, 2900.0, 3700.0

            trading_ev_low, trading_ev_median, trading_ev_high = 2800.0, 3500.0, 4200.0
            trading_equity_low, trading_equity_median, trading_equity_high = 2500.0, 3200.0, 3900.0

            precedent_ev_low, precedent_ev_median, precedent_ev_high = 3200.0, 3800.0, 4500.0
            precedent_equity_low, precedent_equity_median, precedent_equity_high = 2900.0, 3500.0, 4200.0

            lbo_equity_low, lbo_equity_median, lbo_equity_high = 2600.0, 3300.0, 4100.0

            current_share_price = 45.0
            shares_outstanding = 80.0

        # Step 1: Create Valuation Input Data
        valuation_inputs = self._create_valuation_inputs(
            dcf_ev_low, dcf_ev_median, dcf_ev_high, dcf_equity_low, dcf_equity_median, dcf_equity_high,
            trading_ev_low, trading_ev_median, trading_ev_high, trading_equity_low, trading_equity_median, trading_equity_high,
            precedent_ev_low, precedent_ev_median, precedent_ev_high, precedent_equity_low, precedent_equity_median, precedent_equity_high,
            lbo_equity_low, lbo_equity_median, lbo_equity_high,
            other_methods, current_share_price, shares_outstanding
        )

        # Step 2: Organize Valuation Ranges
        valuation_ranges = self._organize_valuation_ranges(valuation_inputs)

        # Step 3: Create Football Field Chart
        chart_files = self._create_football_field_chart(valuation_ranges)

        # Step 4: Create Excel Output
        excel_file = self._create_excel_output(valuation_ranges, chart_files)

        print("\n✅ Football Field Valuation Model Complete!")
        print("📊 Key Valuation Synthesis:")
        print(f"   • Target: {self.target_company} ({self.target_ticker})")
        print(f"   • Valuation Methods: {len(valuation_ranges['methods'])} methodologies")
        print(f"   • EV Range: ${valuation_ranges['ev_overall']['min']/1000:.1f}B - ${valuation_ranges['ev_overall']['max']/1000:.1f}B")
        print(f"   • Equity Range: ${valuation_ranges['equity_overall']['min']/1000:.1f}B - ${valuation_ranges['equity_overall']['max']/1000:.1f}B")
        print(f"   • Current Market Cap: ${(current_share_price * shares_outstanding)/1000:.1f}B")
        print(f"📁 Excel Output: {excel_file}")

        # Display methodology comparison
        self._display_methodology_comparison(valuation_ranges)

        return valuation_ranges, excel_file

    def _create_valuation_inputs(self, dcf_ev_low, dcf_ev_median, dcf_ev_high, dcf_equity_low, dcf_equity_median, dcf_equity_high,
                               trading_ev_low, trading_ev_median, trading_ev_high, trading_equity_low, trading_equity_median, trading_equity_high,
                               precedent_ev_low, precedent_ev_median, precedent_ev_high, precedent_equity_low, precedent_equity_median, precedent_equity_high,
                               lbo_equity_low, lbo_equity_median, lbo_equity_high,
                               other_methods, current_share_price, shares_outstanding):

        """Create comprehensive valuation inputs from multiple methodologies"""

        valuation_inputs = {
            'dcf': {
                'ev_low': dcf_ev_low,
                'ev_median': dcf_ev_median,
                'ev_high': dcf_ev_high,
                'equity_low': dcf_equity_low,
                'equity_median': dcf_equity_median,
                'equity_high': dcf_equity_high,
                'methodology': 'DCF',
                'description': 'Discounted Cash Flow Analysis'
            },
            'trading_comps': {
                'ev_low': trading_ev_low,
                'ev_median': trading_ev_median,
                'ev_high': trading_ev_high,
                'equity_low': trading_equity_low,
                'equity_median': trading_equity_median,
                'equity_high': trading_equity_high,
                'methodology': 'Trading Comps',
                'description': 'Public Market Comparables'
            },
            'precedent_transactions': {
                'ev_low': precedent_ev_low,
                'ev_median': precedent_ev_median,
                'ev_high': precedent_ev_high,
                'equity_low': precedent_equity_low,
                'equity_median': precedent_equity_median,
                'equity_high': precedent_equity_high,
                'methodology': 'Precedent Transactions',
                'description': 'Historical M&A Deals'
            },
            'lbo': {
                'ev_low': None,  # LBO typically focuses on equity value
                'ev_median': None,
                'ev_high': None,
                'equity_low': lbo_equity_low,
                'equity_median': lbo_equity_median,
                'equity_high': lbo_equity_high,
                'methodology': 'LBO',
                'description': 'Leveraged Buyout Analysis'
            },
            'current_market': {
                'ev_low': None,
                'ev_median': None,
                'ev_high': None,
                'equity_low': current_share_price * shares_outstanding,
                'equity_median': current_share_price * shares_outstanding,
                'equity_high': current_share_price * shares_outstanding,
                'methodology': 'Current Market',
                'description': 'Current Trading Price'
            }
        }

        # Add other methodologies if provided
        if other_methods:
            for method_name, method_data in other_methods.items():
                valuation_inputs[method_name] = method_data

        print("📋 Valuation Inputs Created:")
        print(f"   • Target Company: {self.target_company} ({self.target_ticker})")
        print(f"   • Current Share Price: ${current_share_price:.2f}")
        print(f"   • Shares Outstanding: {shares_outstanding:.1f}M")
        print(f"   • Current Market Cap: ${(current_share_price * shares_outstanding)/1000:.1f}B")
        print(f"   • Valuation Methodologies: {len(valuation_inputs)} methods")

        return valuation_inputs

    def _organize_valuation_ranges(self, valuation_inputs):
        """Organize valuation data into structured ranges for analysis"""

        methods = []
        ev_ranges = []
        equity_ranges = []

        # Extract data for each methodology
        for method_key, method_data in valuation_inputs.items():
            methods.append(method_data['methodology'])

            # EV ranges (skip if None)
            if method_data['ev_low'] is not None and method_data['ev_high'] is not None:
                ev_ranges.append({
                    'method': method_data['methodology'],
                    'low': method_data['ev_low'],
                    'median': method_data['ev_median'] or (method_data['ev_low'] + method_data['ev_high']) / 2,
                    'high': method_data['ev_high']
                })

            # Equity ranges
            if method_data['equity_low'] is not None and method_data['equity_high'] is not None:
                equity_ranges.append({
                    'method': method_data['methodology'],
                    'low': method_data['equity_low'],
                    'median': method_data['equity_median'] or (method_data['equity_low'] + method_data['equity_high']) / 2,
                    'high': method_data['equity_high']
                })

        # Calculate overall ranges
        ev_overall = self._calculate_overall_ranges(ev_ranges)
        equity_overall = self._calculate_overall_ranges(equity_ranges)

        valuation_ranges = {
            'methods': methods,
            'ev_ranges': ev_ranges,
            'equity_ranges': equity_ranges,
            'ev_overall': ev_overall,
            'equity_overall': equity_overall,
            'valuation_inputs': valuation_inputs
        }

        print("📊 Valuation Ranges Organized:")
        print(f"   • EV methodologies: {len(ev_ranges)} methods")
        print(f"   • Equity methodologies: {len(equity_ranges)} methods")
        print(f"   • Overall EV range: ${ev_overall['min']/1000:.1f}B - ${ev_overall['max']/1000:.1f}B")
        print(f"   • Overall Equity range: ${equity_overall['min']/1000:.1f}B - ${equity_overall['max']/1000:.1f}B")

        return valuation_ranges

    def _calculate_overall_ranges(self, ranges_list):
        """Calculate overall min/max ranges across all methodologies"""

        if not ranges_list:
            return {'min': 0, 'max': 0, 'range': 0}

        all_lows = [r['low'] for r in ranges_list if r['low'] is not None]
        all_highs = [r['high'] for r in ranges_list if r['high'] is not None]

        if not all_lows or not all_highs:
            return {'min': 0, 'max': 0, 'range': 0}

        overall_min = min(all_lows)
        overall_max = max(all_highs)
        overall_range = overall_max - overall_min

        return {
            'min': overall_min,
            'max': overall_max,
            'range': overall_range
        }

    def _create_football_field_chart(self, valuation_ranges):
        """Create professional Football Field charts for EV and Equity valuations"""

        chart_files = {}

        if not HAS_MATPLOTLIB:
            print("📈 Football Field Charts Skipped (matplotlib not available)")
            return chart_files

        # Create EV Football Field Chart
        if valuation_ranges['ev_ranges']:
            chart_files['ev_chart'] = self._create_single_football_field(
                valuation_ranges['ev_ranges'],
                "Enterprise Value Football Field",
                "Enterprise Value ($M)",
                "ev_chart.png"
            )

        # Create Equity Football Field Chart
        if valuation_ranges['equity_ranges']:
            chart_files['equity_chart'] = self._create_single_football_field(
                valuation_ranges['equity_ranges'],
                "Equity Value Football Field",
                "Equity Value ($M)",
                "equity_chart.png"
            )

        print("📈 Football Field Charts Created:")
        if 'ev_chart' in chart_files:
            print(f"   • Enterprise Value chart: {chart_files['ev_chart']}")
        if 'equity_chart' in chart_files:
            print(f"   • Equity Value chart: {chart_files['equity_chart']}")

        return chart_files

    def _create_single_football_field(self, ranges_data, title, x_label, filename):
        """Create a single football field chart"""

        # Set up the plot
        fig, ax = plt.subplots(figsize=(12, 8))
        fig.patch.set_facecolor('#' + FOOTBALL_FIELD_COLORS['chart_background'])

        # Define colors for different methodologies
        method_colors = {
            'DCF': '#' + FOOTBALL_FIELD_COLORS['dcf_orange'],
            'Trading Comps': '#' + FOOTBALL_FIELD_COLORS['trading_blue'],
            'Precedent Transactions': '#' + FOOTBALL_FIELD_COLORS['precedent_green'],
            'LBO': '#' + FOOTBALL_FIELD_COLORS['lbo_purple'],
            'Current Market': '#' + FOOTBALL_FIELD_COLORS['other_gray']
        }

        # Prepare data
        methods = [r['method'] for r in ranges_data]
        lows = [r['low'] for r in ranges_data]
        medians = [r['median'] for r in ranges_data]
        highs = [r['high'] for r in ranges_data]

        # Create horizontal bars
        y_positions = np.arange(len(methods))

        # Plot ranges as horizontal bars
        for i, method in enumerate(methods):
            color = method_colors.get(method, '#' + FOOTBALL_FIELD_COLORS['other_gray'])

            # Main range bar (low to high)
            ax.barh(y_positions[i], highs[i] - lows[i], left=lows[i],
                   height=0.6, color=color, alpha=0.7, edgecolor='black', linewidth=0.5)

            # Median marker
            ax.plot(medians[i], y_positions[i], 'ko', markersize=8, markerfacecolor='white',
                   markeredgecolor='black', markeredgewidth=2)

            # Add value labels
            ax.text(lows[i], y_positions[i], f'${lows[i]/1000:.1f}B',
                   ha='right', va='center', fontweight='bold', fontsize=9)
            ax.text(highs[i], y_positions[i], f'${highs[i]/1000:.1f}B',
                   ha='left', va='center', fontweight='bold', fontsize=9)
            ax.text(medians[i], y_positions[i] + 0.25, f'${medians[i]/1000:.1f}B',
                   ha='center', va='bottom', fontweight='bold', fontsize=8, color='red')

        # Customize the plot
        ax.set_yticks(y_positions)
        ax.set_yticklabels(methods, fontweight='bold', fontsize=10)
        ax.set_xlabel(x_label, fontweight='bold', fontsize=11)
        ax.set_title(title, fontweight='bold', fontsize=14, pad=20)
        ax.grid(True, alpha=0.3, color='#' + FOOTBALL_FIELD_COLORS['grid_light'])

        # Remove top and right spines
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)

        # Add a subtle background for the football field
        for i in range(len(methods)):
            if i % 2 == 0:
                ax.axhspan(i - 0.5, i + 0.5, alpha=0.05, color='gray')

        plt.tight_layout()

        # Save the chart
        chart_path = f"/Users/averyromain/Scraper/{filename}"
        plt.savefig(chart_path, dpi=300, bbox_inches='tight', facecolor=fig.get_facecolor())
        plt.close()

        return chart_path

    def _create_excel_output(self, valuation_ranges, chart_files):
        """Create professional Excel output with multiple tabs"""

        wb = Workbook()

        # Create worksheets
        ws_summary = wb.active
        ws_summary.title = "Valuation Summary"

        ws_inputs = wb.create_sheet("Valuation Inputs")
        ws_ev_ranges = wb.create_sheet("EV Valuation Ranges")
        ws_equity_ranges = wb.create_sheet("Equity Valuation Ranges")

        # Apply styles to workbook
        for style_name, style in self.styles.items():
            if style_name not in wb.named_styles:
                wb.add_named_style(style)

        # Create each tab
        self._create_summary_tab(ws_summary, valuation_ranges, chart_files)
        self._create_inputs_tab(ws_inputs, valuation_ranges)
        self._create_ev_ranges_tab(ws_ev_ranges, valuation_ranges)
        self._create_equity_ranges_tab(ws_equity_ranges, valuation_ranges)

        # Save workbook
        filename = f"Football_Field_{self.target_ticker}_{self.model_date}.xlsx"
        wb.save(filename)

        return filename

    def _create_summary_tab(self, ws, valuation_ranges, chart_files):
        """Create Valuation Summary tab with charts and key metrics"""

        # Title
        ws['A1'] = f"Valuation Summary – Football Field: {self.target_company} ({self.target_ticker})"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:I1')

        current_row = 3

        # Key Metrics Summary Box
        ws[f'A{current_row}'] = 'VALUATION SUMMARY'
        ws[f'A{current_row}'].style = 'summary_box'
        ws.merge_cells(f'A{current_row}:I{current_row}')
        current_row += 2

        # Overall valuation ranges
        summary_data = [
            ("Overall Enterprise Value Range", f"${valuation_ranges['ev_overall']['min']/1000:.1f}B", f"${valuation_ranges['ev_overall']['max']/1000:.1f}B", f"${(valuation_ranges['ev_overall']['max']-valuation_ranges['ev_overall']['min'])/1000:.1f}B"),
            ("Overall Equity Value Range", f"${valuation_ranges['equity_overall']['min']/1000:.1f}B", f"${valuation_ranges['equity_overall']['max']/1000:.1f}B", f"${(valuation_ranges['equity_overall']['max']-valuation_ranges['equity_overall']['min'])/1000:.1f}B"),
            ("Number of Methodologies", f"{len(valuation_ranges['methods'])}", "", "")
        ]

        # Headers
        ws.cell(row=current_row, column=1, value="Metric").style = 'header'
        ws.cell(row=current_row, column=2, value="Low").style = 'header'
        ws.cell(row=current_row, column=3, value="High").style = 'header'
        ws.cell(row=current_row, column=4, value="Range").style = 'header'
        current_row += 1

        # Data
        for metric, low, high, range_val in summary_data:
            ws.cell(row=current_row, column=1, value=metric).style = 'label_bold'
            ws.cell(row=current_row, column=2, value=low).style = 'output'
            ws.cell(row=current_row, column=3, value=high).style = 'output'
            ws.cell(row=current_row, column=4, value=range_val).style = 'output'
            current_row += 1

        # Add charts if available
        if 'ev_chart' in chart_files or 'equity_chart' in chart_files:
            current_row += 2
            ws[f'A{current_row}'] = 'FOOTBALL FIELD CHARTS'
            ws[f'A{current_row}'].style = 'header'
            ws.merge_cells(f'A{current_row}:I{current_row}')
            current_row += 1

            # Insert EV chart
            if 'ev_chart' in chart_files:
                try:
                    img = Image(chart_files['ev_chart'])
                    img.width = 600
                    img.height = 400
                    ws.add_image(img, f'A{current_row}')
                    current_row += 25
                except Exception as e:
                    print(f"Warning: Could not insert EV chart: {e}")

            # Insert Equity chart
            if 'equity_chart' in chart_files:
                try:
                    img = Image(chart_files['equity_chart'])
                    img.width = 600
                    img.height = 400
                    ws.add_image(img, f'A{current_row}')
                except Exception as e:
                    print(f"Warning: Could not insert Equity chart: {e}")

        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    def _create_inputs_tab(self, ws, valuation_ranges):
        """Create Valuation Inputs tab"""

        valuation_inputs = valuation_ranges['valuation_inputs']

        # Title
        ws['A1'] = f"Valuation Methodology Inputs: {self.target_company} ({self.target_ticker})"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:G1')

        current_row = 3

        # Methodology inputs
        for method_key, method_data in valuation_inputs.items():
            ws[f'A{current_row}'] = method_data['methodology']
            ws[f'A{current_row}'].style = self._get_methodology_style(method_key)
            ws.merge_cells(f'A{current_row}:G{current_row}')
            current_row += 1

            # Description
            ws[f'A{current_row}'] = method_data['description']
            ws[f'A{current_row}'].style = 'label'
            current_row += 1

            # EV Range
            if method_data['ev_low'] is not None:
                ws[f'A{current_row}'] = "Enterprise Value Range:"
                ws[f'A{current_row}'].style = 'label_bold'
                ws[f'B{current_row}'] = method_data['ev_low']
                ws[f'B{current_row}'].style = 'input'
                ws[f'C{current_row}'] = method_data['ev_median'] or (method_data['ev_low'] + method_data['ev_high']) / 2
                ws[f'C{current_row}'].style = 'input'
                ws[f'D{current_row}'] = method_data['ev_high']
                ws[f'D{current_row}'].style = 'input'
                current_row += 1

            # Equity Range
            if method_data['equity_low'] is not None:
                ws[f'A{current_row}'] = "Equity Value Range:"
                ws[f'A{current_row}'].style = 'label_bold'
                ws[f'B{current_row}'] = method_data['equity_low']
                ws[f'B{current_row}'].style = 'input'
                ws[f'C{current_row}'] = method_data['equity_median'] or (method_data['equity_low'] + method_data['equity_high']) / 2
                ws[f'C{current_row}'].style = 'input'
                ws[f'D{current_row}'] = method_data['equity_high']
                ws[f'D{current_row}'].style = 'input'
                current_row += 1

            current_row += 2

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    def _create_ev_ranges_tab(self, ws, valuation_ranges):
        """Create EV Valuation Ranges tab"""

        ev_ranges = valuation_ranges['ev_ranges']

        # Title
        ws['A1'] = f"Enterprise Value Valuation Ranges: {self.target_company} ({self.target_ticker})"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:D1')

        current_row = 3

        # EV Ranges Table
        ws[f'A{current_row}'] = "ENTERPRISE VALUE RANGES"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 2

        # Headers
        headers = ['Methodology', 'Low ($M)', 'Median ($M)', 'High ($M)']
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header).style = 'header'
        current_row += 1

        # Data
        for range_data in ev_ranges:
            method_key = self._get_method_key(range_data['method'])
            ws.cell(row=current_row, column=1, value=range_data['method']).style = self._get_methodology_style(method_key)
            ws.cell(row=current_row, column=2, value=range_data['low']).style = 'output'
            ws.cell(row=current_row, column=3, value=range_data['median']).style = 'output'
            ws.cell(row=current_row, column=4, value=range_data['high']).style = 'output'
            current_row += 1

        # Overall range
        current_row += 2
        ws[f'A{current_row}'] = "Overall Range:"
        ws[f'A{current_row}'].style = 'label_bold'
        ws[f'B{current_row}'] = valuation_ranges['ev_overall']['min']
        ws[f'B{current_row}'].style = 'output'
        ws[f'C{current_row}'] = valuation_ranges['ev_overall']['max']
        ws[f'C{current_row}'].style = 'output'
        ws[f'D{current_row}'] = valuation_ranges['ev_overall']['range']
        ws[f'D{current_row}'].style = 'output'

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    def _create_equity_ranges_tab(self, ws, valuation_ranges):
        """Create Equity Valuation Ranges tab"""

        equity_ranges = valuation_ranges['equity_ranges']

        # Title
        ws['A1'] = f"Equity Value Valuation Ranges: {self.target_company} ({self.target_ticker})"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:D1')

        current_row = 3

        # Equity Ranges Table
        ws[f'A{current_row}'] = "EQUITY VALUE RANGES"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 2

        # Headers
        headers = ['Methodology', 'Low ($M)', 'Median ($M)', 'High ($M)']
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header).style = 'header'
        current_row += 1

        # Data
        for range_data in equity_ranges:
            method_key = self._get_method_key(range_data['method'])
            ws.cell(row=current_row, column=1, value=range_data['method']).style = self._get_methodology_style(method_key)
            ws.cell(row=current_row, column=2, value=range_data['low']).style = 'output'
            ws.cell(row=current_row, column=3, value=range_data['median']).style = 'output'
            ws.cell(row=current_row, column=4, value=range_data['high']).style = 'output'
            current_row += 1

        # Overall range
        current_row += 2
        ws[f'A{current_row}'] = "Overall Range:"
        ws[f'A{current_row}'].style = 'label_bold'
        ws[f'B{current_row}'] = valuation_ranges['equity_overall']['min']
        ws[f'B{current_row}'].style = 'output'
        ws[f'C{current_row}'] = valuation_ranges['equity_overall']['max']
        ws[f'C{current_row}'].style = 'output'
        ws[f'D{current_row}'] = valuation_ranges['equity_overall']['range']
        ws[f'D{current_row}'].style = 'output'

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    def _get_method_key(self, method_name):
        """Get method key from method name"""
        method_map = {
            'DCF': 'dcf',
            'Trading Comps': 'trading_comps',
            'Precedent Transactions': 'precedent_transactions',
            'LBO': 'lbo',
            'Current Market': 'current_market'
        }
        return method_map.get(method_name, 'other')

    def _get_methodology_style(self, method_key):
        """Get the appropriate style for a methodology"""
        if method_key == 'dcf':
            return 'dcf'
        elif method_key == 'trading_comps' or method_key == 'trading':
            return 'trading'
        elif method_key == 'precedent_transactions' or method_key == 'precedent':
            return 'precedent'
        elif method_key == 'lbo':
            return 'lbo'
        else:
            return 'calculation'

    def _display_methodology_comparison(self, valuation_ranges):
        """Display a comparison of methodologies in the console"""

        print("\n📊 Methodology Comparison:")
        print("=" * 80)

        # EV comparison
        if valuation_ranges['ev_ranges']:
            print("\nEnterprise Value Ranges:")
            print("<20")
            print("-" * 90)

            for range_data in valuation_ranges['ev_ranges']:
                print(f"{range_data.get('method', 'Unknown'):<20}"
                      f"${range_data.get('min', 0)/1000:<12.1f}B"
                      f"${range_data.get('max', 0)/1000:<12.1f}B"
                      f"${range_data.get('median', 0)/1000:<12.1f}B")

        # Equity comparison
        if valuation_ranges['equity_ranges']:
            print("\nEquity Value Ranges:")
            print("<20")
            print("-" * 90)

            for range_data in valuation_ranges['equity_ranges']:
                print(f"{range_data.get('method', 'Unknown'):<20}"
                      f"${range_data.get('min', 0)/1000:<12.1f}B"
                      f"${range_data.get('max', 0)/1000:<12.1f}B"
                      f"${range_data.get('median', 0)/1000:<12.1f}B")

        print("\n💡 Football Field Analysis Insights:")
        print("=" * 80)
        print("• DCF: Intrinsic value based on cash flow generation")
        print("• Trading Comps: Market-based valuation multiples")
        print("• Precedent Transactions: Control premium analysis")
        print("• LBO: Private equity perspective on exit value")
        print("• Current Market: Current trading valuation")

        # Valuation conclusions
        ev_range = valuation_ranges['ev_overall']['range']
        equity_range = valuation_ranges['equity_overall']['range']

        print("\n🎯 Valuation Conclusions:")
        print(f"   • EV Valuation Range: ${(valuation_ranges['ev_overall']['min'])/1000:.1f}B - ${(valuation_ranges['ev_overall']['max'])/1000:.1f}B")
        print(f"   • Equity Valuation Range: ${(valuation_ranges['equity_overall']['min'])/1000:.1f}B - ${(valuation_ranges['equity_overall']['max'])/1000:.1f}B")
        print(f"   • EV Range Width: ${(ev_range)/1000:.1f}B ({ev_range/valuation_ranges['ev_overall']['min']*100:.1f}% of low end)")
        print(f"   • Equity Range Width: ${(equity_range)/1000:.1f}B ({equity_range/valuation_ranges['equity_overall']['min']*100:.1f}% of low end)")

        # Determine primary valuation anchor
        if valuation_ranges['ev_ranges']:
            ev_medians = [r['median'] for r in valuation_ranges['ev_ranges']]
            avg_ev = sum(ev_medians) / len(ev_medians)
            print(f"   • Average EV Valuation: ${(avg_ev)/1000:.1f}B")

        if valuation_ranges['equity_ranges']:
            equity_medians = [r['median'] for r in valuation_ranges['equity_ranges']]
            avg_equity = sum(equity_medians) / len(equity_medians)
            print(f"   • Average Equity Valuation: ${(avg_equity)/1000:.1f}B")


def run_sample_football_field_model():
    """Run a sample football field valuation model"""

    print("🏟️  Running Professional Football Field Valuation Model Sample")
    print("=" * 70)

    # Create model instance
    model = ProfessionalFootballFieldModel("TechTarget Inc.", "TECHTARGET")

    # Run the model with sample data from multiple methodologies
    valuation_ranges, excel_file = model.run_football_field_model(
        # Target Company Info
        target_company="TechTarget Inc.",
        target_ticker="TECHTARGET",

        # Current market data
        current_share_price=45.0,
        shares_outstanding=80.0
    )

    return valuation_ranges, excel_file


if __name__ == "__main__":
    # Run sample football field model
    valuation_ranges, excel_file = run_sample_football_field_model()

    print("\n📋 Football Field Model Complete!")
    print(f"📁 Excel file saved as: {excel_file}")

    # Display key results
    print("\n🔑 Key Results Summary:")
    print("=" * 40)

    print(f"Overall EV Range: ${valuation_ranges['ev_overall']['min']/1000:.1f}B - ${valuation_ranges['ev_overall']['max']/1000:.1f}B")
    print(f"Overall Equity Range: ${valuation_ranges['equity_overall']['min']/1000:.1f}B - ${valuation_ranges['equity_overall']['max']/1000:.1f}B")
    print(f"Number of Methodologies: {len(valuation_ranges['methods'])}")

    # Calculate average valuations
    if valuation_ranges['ev_ranges']:
        ev_medians = [r['median'] for r in valuation_ranges['ev_ranges']]
        avg_ev = sum(ev_medians) / len(ev_medians)
        print(f"Average EV Valuation: ${avg_ev/1000:.1f}B")

    if valuation_ranges['equity_ranges']:
        equity_medians = [r['median'] for r in valuation_ranges['equity_ranges']]
        avg_equity = sum(equity_medians) / len(equity_medians)
        print(f"Average Equity Valuation: ${avg_equity/1000:.1f}B")
