#!/usr/bin/env python3
"""
Professional Sum-of-the-Parts (SOTP) Valuation Model
Values each business unit separately using appropriate methodologies, then aggregates

Author: Investment Banking Modeler
Date: 2024

Features:
- Multi-segment business valuation
- Flexible valuation methods per segment
- Corporate consolidation and adjustments
- Equity value and per-share calculations
- Professional banker-standard formatting
- Excel output with multiple tabs
- Charts and visualizations
"""

import numpy as np
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
try:
    import matplotlib.pyplot as plt
    HAS_MATPLOTLIB = True
except ImportError:
    HAS_MATPLOTLIB = False
    print("⚠️  Warning: matplotlib not available. Charts will not be generated.")
import warnings
warnings.filterwarnings('ignore')

# Professional color scheme for Sum-of-the-Parts
SOTP_COLORS = {
    'header_blue': '1F4E79',
    'input_light_blue': 'DDEBF7',
    'calculation_black': '000000',
    'output_green': 'D8E4BC',
    'segment_orange': 'FF7F50',
    'segment_blue': '4682B4',
    'segment_green': '32CD32',
    'segment_purple': '9370DB',
    'segment_gray': '708090',
    'warning_red': 'F8CECC',
    'border_dark': '2C3E50',
    'text_dark': '000000',
    'text_white': 'FFFFFF',
    'chart_background': 'F5F5F5',
    'grid_light': 'E0E0E0',
    'highlight_gold': 'FFD700'
}

class ProfessionalSOTPModel:
    """
    Comprehensive Sum-of-the-Parts Valuation Model with Professional Formatting
    """

    def __init__(self, company_name="Company Name", ticker="TICKER"):
        self.company_name = company_name
        self.ticker = ticker
        self.model_date = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Initialize styles
        self.styles = self._create_styles()

    def _create_styles(self):
        """Create professional Excel styles"""
        styles = {}

        # Header style
        styles['header'] = NamedStyle(name='header')
        styles['header'].font = Font(name='Calibri', size=12, bold=True, color=SOTP_COLORS['text_white'])
        styles['header'].fill = PatternFill(start_color=SOTP_COLORS['header_blue'], end_color=SOTP_COLORS['header_blue'], fill_type='solid')
        styles['header'].alignment = Alignment(horizontal='center', vertical='center')

        # Input style
        styles['input'] = NamedStyle(name='input')
        styles['input'].font = Font(name='Calibri', size=10, color=SOTP_COLORS['text_dark'])
        styles['input'].fill = PatternFill(start_color=SOTP_COLORS['input_light_blue'], end_color=SOTP_COLORS['input_light_blue'], fill_type='solid')
        styles['input'].alignment = Alignment(horizontal='right', vertical='center')
        styles['input'].number_format = '#,##0.00'

        # Calculation style
        styles['calculation'] = NamedStyle(name='calculation')
        styles['calculation'].font = Font(name='Calibri', size=10, color=SOTP_COLORS['calculation_black'])
        styles['calculation'].alignment = Alignment(horizontal='right', vertical='center')
        styles['calculation'].number_format = '#,##0.00'

        # Output style
        styles['output'] = NamedStyle(name='output')
        styles['output'].font = Font(name='Calibri', size=10, bold=True, color=SOTP_COLORS['text_dark'])
        styles['output'].fill = PatternFill(start_color=SOTP_COLORS['output_green'], end_color=SOTP_COLORS['output_green'], fill_type='solid')
        styles['output'].alignment = Alignment(horizontal='right', vertical='center')
        styles['output'].number_format = '#,##0.00'

        # Segment styles
        styles['segment1'] = NamedStyle(name='segment1')
        styles['segment1'].font = Font(name='Calibri', size=10, color=SOTP_COLORS['text_white'])
        styles['segment1'].fill = PatternFill(start_color=SOTP_COLORS['segment_orange'], end_color=SOTP_COLORS['segment_orange'], fill_type='solid')
        styles['segment1'].alignment = Alignment(horizontal='center', vertical='center')

        styles['segment2'] = NamedStyle(name='segment2')
        styles['segment2'].font = Font(name='Calibri', size=10, color=SOTP_COLORS['text_white'])
        styles['segment2'].fill = PatternFill(start_color=SOTP_COLORS['segment_blue'], end_color=SOTP_COLORS['segment_blue'], fill_type='solid')
        styles['segment2'].alignment = Alignment(horizontal='center', vertical='center')

        styles['segment3'] = NamedStyle(name='segment3')
        styles['segment3'].font = Font(name='Calibri', size=10, color=SOTP_COLORS['text_white'])
        styles['segment3'].fill = PatternFill(start_color=SOTP_COLORS['segment_green'], end_color=SOTP_COLORS['segment_green'], fill_type='solid')
        styles['segment3'].alignment = Alignment(horizontal='center', vertical='center')

        styles['segment4'] = NamedStyle(name='segment4')
        styles['segment4'].font = Font(name='Calibri', size=10, color=SOTP_COLORS['text_white'])
        styles['segment4'].fill = PatternFill(start_color=SOTP_COLORS['segment_purple'], end_color=SOTP_COLORS['segment_purple'], fill_type='solid')
        styles['segment4'].alignment = Alignment(horizontal='center', vertical='center')

        # Corporate style
        styles['corporate'] = NamedStyle(name='corporate')
        styles['corporate'].font = Font(name='Calibri', size=10, color=SOTP_COLORS['text_white'])
        styles['corporate'].fill = PatternFill(start_color=SOTP_COLORS['border_dark'], end_color=SOTP_COLORS['border_dark'], fill_type='solid')
        styles['corporate'].alignment = Alignment(horizontal='center', vertical='center')

        # Warning style
        styles['warning'] = NamedStyle(name='warning')
        styles['warning'].font = Font(name='Calibri', size=10, color=SOTP_COLORS['text_dark'])
        styles['warning'].fill = PatternFill(start_color=SOTP_COLORS['warning_red'], end_color=SOTP_COLORS['warning_red'], fill_type='solid')
        styles['warning'].alignment = Alignment(horizontal='center', vertical='center')

        # Bold label style
        styles['label_bold'] = NamedStyle(name='label_bold')
        styles['label_bold'].font = Font(name='Calibri', size=10, bold=True, color=SOTP_COLORS['text_dark'])
        styles['label_bold'].alignment = Alignment(horizontal='left', vertical='center')

        # Normal label style
        styles['label'] = NamedStyle(name='label')
        styles['label'].font = Font(name='Calibri', size=10, color=SOTP_COLORS['text_dark'])
        styles['label'].alignment = Alignment(horizontal='left', vertical='center')

        # Summary box style
        styles['summary_box'] = NamedStyle(name='summary_box')
        styles['summary_box'].font = Font(name='Calibri', size=11, bold=True, color=SOTP_COLORS['text_white'])
        styles['summary_box'].fill = PatternFill(start_color=SOTP_COLORS['header_blue'], end_color=SOTP_COLORS['header_blue'], fill_type='solid')
        styles['summary_box'].alignment = Alignment(horizontal='center', vertical='center')
        styles['summary_box'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        return styles

    def run_sotp_model(self,
                      # Company Overview
                      company_name=None,
                      ticker=None,

                      # Business Segments Configuration
                      segments=None,  # List of segment dictionaries

                      # Corporate-Level Adjustments
                      net_debt=0.0,                    # $M
                      minority_interests=0.0,          # $M
                      cash_and_investments=0.0,        # $M
                      other_assets=0.0,                # $M
                      other_liabilities=0.0,           # $M
                      shares_outstanding=100.0,        # Million shares
                      current_share_price=0.0):        # $ per share

        """
        Run complete Sum-of-the-Parts valuation model
        """

        if company_name:
            self.company_name = company_name
        if ticker:
            self.ticker = ticker

        print(f"🧩 Building Professional Sum-of-the-Parts Valuation Model for {self.company_name} ({self.ticker})")
        print("=" * 95)

        # Set default segments if not provided
        if segments is None:
            segments = self._get_default_segments()

        # Step 1: Validate and Process Segments
        processed_segments = self._validate_segments(segments)

        # Step 2: Value Each Segment
        segment_valuations = self._value_segments(processed_segments)

        # Step 3: Corporate Consolidation
        consolidation = self._corporate_consolidation(
            segment_valuations, net_debt, minority_interests,
            cash_and_investments, other_assets, other_liabilities,
            shares_outstanding, current_share_price
        )

        # Step 4: Generate Outputs and Analysis
        analysis = self._generate_sotp_analysis(segment_valuations, consolidation)

        # Step 5: Create Charts
        chart_files = self._create_sotp_charts(segment_valuations, consolidation)

        # Step 6: Create Excel Output
        excel_file = self._create_excel_output(
            processed_segments, segment_valuations, consolidation, analysis, chart_files
        )

        print("\n✅ Sum-of-the-Parts Model Complete!")
        print("📊 Key Valuation Metrics:")
        print(f"   • Company: {self.company_name} ({self.ticker})")
        print(f"   • Segments: {len(processed_segments)} business units")
        print(f"   • Total EV: ${consolidation['total_ev']/1000:.1f}B")
        print(f"   • Equity Value: ${consolidation['equity_value']/1000:.1f}B")
        print(f"   • Implied Share Price: ${consolidation['implied_share_price']:.2f}")
        print(f"   • Current Share Price: ${current_share_price:.2f}")
        print(f"   • Premium/(Discount): {consolidation['premium_discount_pct']:.1%}")
        print(f"📁 Excel Output: {excel_file}")

        # Display segment breakdown
        self._display_segment_breakdown(segment_valuations, consolidation)

        return {
            'segments': processed_segments,
            'segment_valuations': segment_valuations,
            'consolidation': consolidation,
            'analysis': analysis,
            'excel_file': excel_file
        }

    def _get_default_segments(self):
        """Get default segment configuration for demonstration"""

        return [
            {
                'name': 'Technology Division',
                'description': 'Software, Cloud Services, Digital Platforms',
                'revenue': 5000.0,          # $M
                'ebitda': 1500.0,           # $M
                'net_income': 1200.0,       # $M
                'valuation_method': 'EV/EBITDA Multiple',
                'multiple': 18.0,           # Premium tech multiple
                'growth_rate': 0.08,        # 8% growth
                'beta': 1.2,                # Tech beta
                'risk_premium': 0.05        # 5% risk premium
            },
            {
                'name': 'Industrial Division',
                'description': 'Manufacturing, Aerospace, Energy Equipment',
                'revenue': 8000.0,          # $M
                'ebitda': 1200.0,           # $M
                'net_income': 800.0,        # $M
                'valuation_method': 'EV/EBITDA Multiple',
                'multiple': 10.0,           # Industrial average
                'growth_rate': 0.04,        # 4% growth
                'beta': 1.0,                # Market beta
                'risk_premium': 0.055       # 5.5% risk premium
            },
            {
                'name': 'Healthcare Division',
                'description': 'Medical Devices, Pharmaceuticals, Healthcare Services',
                'revenue': 6000.0,          # $M
                'ebitda': 1400.0,           # $M
                'net_income': 1000.0,       # $M
                'valuation_method': 'P/E Multiple',
                'multiple': 22.0,           # Healthcare premium
                'growth_rate': 0.06,        # 6% growth
                'beta': 0.9,                # Healthcare beta
                'risk_premium': 0.045       # 4.5% risk premium
            },
            {
                'name': 'Financial Services',
                'description': 'Banking, Insurance, Investment Management',
                'revenue': 3000.0,          # $M
                'ebitda': 900.0,            # $M
                'net_income': 600.0,        # $M
                'valuation_method': 'P/E Multiple',
                'multiple': 14.0,           # Financial services average
                'growth_rate': 0.035,       # 3.5% growth
                'beta': 1.1,                # Financial beta
                'risk_premium': 0.06        # 6% risk premium
            },
            {
                'name': 'Consumer Division',
                'description': 'Retail, Consumer Goods, Media',
                'revenue': 4000.0,          # $M
                'ebitda': 600.0,            # $M
                'net_income': 400.0,        # $M
                'valuation_method': 'EV/EBITDA Multiple',
                'multiple': 8.0,            # Consumer average
                'growth_rate': 0.03,        # 3% growth
                'beta': 0.95,               # Consumer beta
                'risk_premium': 0.05        # 5% risk premium
            }
        ]

    def _validate_segments(self, segments):
        """Validate and process segment data"""

        processed_segments = []

        for i, segment in enumerate(segments):
            # Ensure required fields
            required_fields = ['name', 'description', 'valuation_method']
            for field in required_fields:
                if field not in segment:
                    raise ValueError(f"Segment {i+1} missing required field: {field}")

            # Set defaults for optional fields
            segment.setdefault('revenue', 0.0)
            segment.setdefault('ebitda', 0.0)
            segment.setdefault('net_income', 0.0)
            segment.setdefault('multiple', 10.0)
            segment.setdefault('growth_rate', 0.03)
            segment.setdefault('beta', 1.0)
            segment.setdefault('risk_premium', 0.05)

            # Validate valuation method
            valid_methods = ['EV/EBITDA Multiple', 'P/E Multiple', 'EV/Revenue Multiple', 'DCF']
            if segment['valuation_method'] not in valid_methods:
                raise ValueError(f"Invalid valuation method: {segment['valuation_method']}. Must be one of {valid_methods}")

            processed_segments.append(segment)

        print("📋 Segments Validated:")
        print(f"   • Total Segments: {len(processed_segments)}")
        for i, segment in enumerate(processed_segments, 1):
            print(f"   • Segment {i}: {segment['name']} - {segment['valuation_method']}")

        return processed_segments

    def _value_segments(self, segments):
        """Value each segment using appropriate methodology"""

        segment_valuations = []

        for segment in segments:
            valuation = self._value_single_segment(segment)
            segment_valuations.append(valuation)

        print("📊 Segment Valuations Completed:")
        total_ev = sum(val['enterprise_value'] for val in segment_valuations)
        print(f"   • Total Enterprise Value: ${total_ev/1000:.1f}B")
        for val in segment_valuations:
            pct = val['enterprise_value'] / total_ev * 100 if total_ev > 0 else 0
            print(f"   • {val['name']}: ${val['enterprise_value']/1000:.1f}B ({pct:.1f}%)")

        return segment_valuations

    def _value_single_segment(self, segment):
        """Value a single segment using specified methodology"""

        method = segment['valuation_method']
        enterprise_value = 0.0

        if method == 'EV/EBITDA Multiple':
            enterprise_value = segment['ebitda'] * segment['multiple']
        elif method == 'P/E Multiple':
            enterprise_value = segment['net_income'] * segment['multiple']
        elif method == 'EV/Revenue Multiple':
            enterprise_value = segment['revenue'] * segment['multiple']
        elif method == 'DCF':
            enterprise_value = self._dcf_valuation(segment)

        # Calculate equity value (simplified - assuming no segment-specific debt)
        equity_value = enterprise_value

        return {
            'name': segment['name'],
            'description': segment['description'],
            'valuation_method': method,
            'metric_value': segment.get('ebitda', segment.get('net_income', segment.get('revenue', 0))),
            'multiple': segment['multiple'],
            'enterprise_value': enterprise_value,
            'equity_value': equity_value,
            'segment_data': segment
        }

    def _dcf_valuation(self, segment):
        """Perform DCF valuation for a segment"""

        # Simplified DCF calculation
        ebitda = segment['ebitda']
        growth_rate = segment['growth_rate']
        beta = segment['beta']
        risk_premium = segment['risk_premium']

        # Estimate WACC (simplified)
        risk_free_rate = 0.03  # 3% risk-free rate
        market_risk_premium = 0.055  # 5.5% market risk premium
        tax_rate = 0.25

        cost_of_equity = risk_free_rate + beta * market_risk_premium
        wacc = cost_of_equity * (1 - tax_rate)  # Simplified

        # Terminal value calculation (simplified)
        terminal_growth = 0.025  # 2.5% terminal growth
        terminal_value = ebitda * (1 + growth_rate) * (1 - tax_rate) / (wacc - terminal_growth)

        # PV of FCF (simplified - using current EBITDA as proxy)
        pv_fcf = ebitda * (1 - tax_rate) / wacc

        return pv_fcf + terminal_value

    def _corporate_consolidation(self, segment_valuations, net_debt, minority_interests,
                               cash_and_investments, other_assets, other_liabilities,
                               shares_outstanding, current_share_price):

        """Perform corporate-level consolidation and adjustments"""

        # Sum segment enterprise values
        total_segment_ev = sum(val['enterprise_value'] for val in segment_valuations)

        # Corporate adjustments
        total_ev = total_segment_ev
        total_ev -= net_debt  # Subtract net debt
        total_ev -= minority_interests  # Subtract minority interests
        total_ev += cash_and_investments  # Add cash and investments
        total_ev += other_assets  # Add other assets
        total_ev -= other_liabilities  # Subtract other liabilities

        # Calculate equity value
        equity_value = total_ev

        # Calculate per-share value
        implied_share_price = equity_value / shares_outstanding if shares_outstanding > 0 else 0

        # Calculate premium/discount to current share price
        premium_discount = implied_share_price - current_share_price
        premium_discount_pct = premium_discount / current_share_price if current_share_price > 0 else 0

        consolidation = {
            'total_segment_ev': total_segment_ev,
            'total_ev': total_ev,
            'equity_value': equity_value,
            'implied_share_price': implied_share_price,
            'current_share_price': current_share_price,
            'premium_discount': premium_discount,
            'premium_discount_pct': premium_discount_pct,
            'adjustments': {
                'net_debt': net_debt,
                'minority_interests': minority_interests,
                'cash_and_investments': cash_and_investments,
                'other_assets': other_assets,
                'other_liabilities': other_liabilities
            },
            'shares_outstanding': shares_outstanding
        }

        print("🏢 Corporate Consolidation Completed:")
        print(f"   • Total Segment EV: ${total_segment_ev/1000:.1f}B")
        print(f"   • Net Debt: ${net_debt/1000:.1f}B")
        print(f"   • Equity Value: ${equity_value/1000:.1f}B")
        print(f"   • Implied Share Price: ${implied_share_price:.2f}")
        print(f"   • Current Share Price: ${current_share_price:.2f}")
        print(f"   • Premium/(Discount): {premium_discount_pct:.1%}")

        return consolidation

    def _generate_sotp_analysis(self, segment_valuations, consolidation):

        """Generate SOTP analysis and insights"""

        analysis = {}

        # Segment contribution analysis
        total_ev = consolidation['total_ev']
        segment_contributions = []

        for val in segment_valuations:
            contribution_pct = val['enterprise_value'] / consolidation['total_segment_ev'] * 100 if consolidation['total_segment_ev'] > 0 else 0
            segment_contributions.append({
                'name': val['name'],
                'enterprise_value': val['enterprise_value'],
                'contribution_pct': contribution_pct
            })

        analysis['segment_contributions'] = segment_contributions

        # Conglomerate discount/premium analysis
        current_market_cap = consolidation['current_share_price'] * consolidation['shares_outstanding']
        implied_market_cap = consolidation['equity_value']
        conglomerate_effect = implied_market_cap - current_market_cap
        conglomerate_effect_pct = conglomerate_effect / current_market_cap if current_market_cap > 0 else 0

        analysis['conglomerate_analysis'] = {
            'current_market_cap': current_market_cap,
            'implied_market_cap': implied_market_cap,
            'conglomerate_effect': conglomerate_effect,
            'conglomerate_effect_pct': conglomerate_effect_pct,
            'status': "PREMIUM" if conglomerate_effect > 0 else "DISCOUNT"
        }

        print("📈 SOTP Analysis Generated:")
        print(f"   • Largest Segment: {max(segment_contributions, key=lambda x: x['enterprise_value'])['name']}")
        print(f"   • Conglomerate Effect: {analysis['conglomerate_analysis']['status']} of {abs(conglomerate_effect_pct):.1%}")

        return analysis

    def _create_sotp_charts(self, segment_valuations, consolidation):

        """Create professional charts for SOTP analysis"""

        chart_files = {}

        if not HAS_MATPLOTLIB:
            print("📈 Charts Skipped (matplotlib not available)")
            return chart_files

        # Create segment contribution pie chart
        chart_files['segment_pie'] = self._create_segment_pie_chart(segment_valuations, consolidation)

        # Create valuation breakdown bar chart
        chart_files['valuation_bar'] = self._create_valuation_bar_chart(segment_valuations, consolidation)

        print("📊 SOTP Charts Created:")
        for chart_name, chart_file in chart_files.items():
            print(f"   • {chart_name}: {chart_file}")

        return chart_files

    def _create_segment_pie_chart(self, segment_valuations, consolidation):

        """Create segment contribution pie chart"""

        fig, ax = plt.subplots(figsize=(10, 8))
        fig.patch.set_facecolor('#' + SOTP_COLORS['chart_background'])

        # Data
        segment_names = [val['name'] for val in segment_valuations]
        segment_values = [val['enterprise_value'] for val in segment_valuations]

        # Colors for segments
        colors = ['#' + SOTP_COLORS['segment_orange'],
                 '#' + SOTP_COLORS['segment_blue'],
                 '#' + SOTP_COLORS['segment_green'],
                 '#' + SOTP_COLORS['segment_purple'],
                 '#' + SOTP_COLORS['segment_gray']]

        # Create pie chart
        wedges, texts, autotexts = ax.pie(segment_values, labels=segment_names, autopct='%1.1f%%',
                                        colors=colors[:len(segment_names)], startangle=90)

        # Style the text
        for text in texts:
            text.set_fontsize(10)
            text.set_fontweight('bold')
        for autotext in autotexts:
            autotext.set_fontsize(9)
            autotext.set_fontweight('bold')
            autotext.set_color('white')

        ax.set_title(f'{self.company_name} - SOTP Segment Contribution', fontweight='bold', fontsize=14, pad=20)

        plt.tight_layout()

        # Save chart
        chart_path = "/Users/averyromain/Scraper/sotp_segment_pie.png"
        plt.savefig(chart_path, dpi=300, bbox_inches='tight', facecolor=fig.get_facecolor())
        plt.close()

        return chart_path

    def _create_valuation_bar_chart(self, segment_valuations, consolidation):

        """Create valuation breakdown bar chart"""

        fig, ax = plt.subplots(figsize=(12, 8))
        fig.patch.set_facecolor('#' + SOTP_COLORS['chart_background'])

        # Data
        segment_names = [val['name'] for val in segment_valuations]
        segment_evs = [val['enterprise_value'] / 1000 for val in segment_valuations]  # Convert to billions

        # Colors
        colors = ['#' + SOTP_COLORS['segment_orange'],
                 '#' + SOTP_COLORS['segment_blue'],
                 '#' + SOTP_COLORS['segment_green'],
                 '#' + SOTP_COLORS['segment_purple'],
                 '#' + SOTP_COLORS['segment_gray']]

        # Create bars
        bars = ax.bar(segment_names, segment_evs, color=colors[:len(segment_names)], alpha=0.8, width=0.6)

        # Add value labels on bars
        for bar, value in zip(bars, segment_evs):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                   f'${value:.1f}B', ha='center', va='bottom', fontweight='bold', fontsize=10)

        # Add total line
        total_ev = sum(segment_evs)
        ax.axhline(y=total_ev, color='red', linestyle='--', linewidth=2, label=f'Total EV: ${total_ev:.1f}B')

        # Styling
        ax.set_ylabel('Enterprise Value ($B)', fontweight='bold', fontsize=11)
        ax.set_title(f'{self.company_name} - SOTP Valuation Breakdown', fontweight='bold', fontsize=14, pad=20)
        ax.legend()
        ax.grid(True, alpha=0.3, axis='y')

        # Rotate x-axis labels
        plt.xticks(rotation=45, ha='right')

        plt.tight_layout()

        # Save chart
        chart_path = "/Users/averyromain/Scraper/sotp_valuation_bar.png"
        plt.savefig(chart_path, dpi=300, bbox_inches='tight', facecolor=fig.get_facecolor())
        plt.close()

        return chart_path

    def _create_excel_output(self, segments, segment_valuations, consolidation, analysis, chart_files):

        """Create professional Excel output with multiple tabs"""

        wb = Workbook()

        # Create worksheets
        ws_summary = wb.active
        ws_summary.title = "SOTP Valuation Summary"

        ws_inputs = wb.create_sheet("Segment Inputs")
        ws_valuations = wb.create_sheet("Segment Valuations")
        ws_consolidation = wb.create_sheet("Corporate Consolidation")

        # Apply styles to workbook
        for style_name, style in self.styles.items():
            if style_name not in wb.named_styles:
                wb.add_named_style(style)

        # Create each tab
        self._create_summary_tab(ws_summary, segment_valuations, consolidation, analysis, chart_files)
        self._create_inputs_tab(ws_inputs, segments)
        self._create_valuations_tab(ws_valuations, segment_valuations)
        self._create_consolidation_tab(ws_consolidation, consolidation, analysis)

        # Save workbook
        filename = f"SOTP_Valuation_{self.ticker}_{self.model_date}.xlsx"
        wb.save(filename)

        return filename

    def _create_summary_tab(self, ws, segment_valuations, consolidation, analysis, chart_files):

        """Create SOTP Valuation Summary tab"""

        # Title
        ws['A1'] = f"Sum-of-the-Parts Valuation: {self.company_name} ({self.ticker})"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:G1')

        current_row = 3

        # Executive Summary Box
        ws[f'A{current_row}'] = 'EXECUTIVE SUMMARY'
        ws[f'A{current_row}'].style = 'summary_box'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        # Key metrics
        summary_data = [
            ("Total Enterprise Value", f"${consolidation['total_ev']/1000:.1f}B"),
            ("Equity Value", f"${consolidation['equity_value']/1000:.1f}B"),
            ("Implied Share Price", f"${consolidation['implied_share_price']:.2f}"),
            ("Current Share Price", f"${consolidation['current_share_price']:.2f}"),
            ("Premium/(Discount)", f"{consolidation['premium_discount_pct']:.1%}"),
            ("Conglomerate Effect", analysis['conglomerate_analysis']['status'])
        ]

        for label, value in summary_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            if label == "Premium/(Discount)":
                ws[f'B{current_row}'].style = 'output' if consolidation['premium_discount'] > 0 else 'warning'
            else:
                ws[f'B{current_row}'].style = 'output'
            current_row += 1

        current_row += 2

        # Segment Summary
        ws[f'A{current_row}'] = 'SEGMENT SUMMARY'
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        # Headers
        ws.cell(row=current_row, column=1, value="Segment").style = 'header'
        ws.cell(row=current_row, column=2, value="Enterprise Value ($M)").style = 'header'
        ws.cell(row=current_row, column=3, value="Contribution (%)").style = 'header'
        ws.cell(row=current_row, column=4, value="Valuation Method").style = 'header'
        current_row += 1

        # Segment data
        for i, val in enumerate(segment_valuations):
            segment_style = self._get_segment_style(i)
            contribution_pct = val['enterprise_value'] / consolidation['total_segment_ev'] * 100 if consolidation['total_segment_ev'] > 0 else 0

            ws.cell(row=current_row, column=1, value=val['name']).style = segment_style
            ws.cell(row=current_row, column=2, value=val['enterprise_value']).style = 'output'
            ws.cell(row=current_row, column=3, value=contribution_pct).style = 'output'
            ws.cell(row=current_row, column=4, value=val['valuation_method']).style = 'output'
            current_row += 1

        # Add charts if available
        if chart_files and HAS_MATPLOTLIB:
            current_row += 2
            try:
                from openpyxl.drawing.image import Image

                # Add pie chart
                if 'segment_pie' in chart_files:
                    img = Image(chart_files['segment_pie'])
                    img.width = 500
                    img.height = 400
                    ws.add_image(img, f'A{current_row}')

                # Add bar chart
                if 'valuation_bar' in chart_files:
                    img = Image(chart_files['valuation_bar'])
                    img.width = 600
                    img.height = 400
                    ws.add_image(img, f'H{current_row}')

            except Exception as e:
                print(f"Warning: Could not insert charts: {e}")

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20

    def _create_inputs_tab(self, ws, segments):

        """Create Segment Inputs tab"""

        # Title
        ws['A1'] = f"Segment Inputs & Assumptions: {self.company_name} ({self.ticker})"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:H1')

        current_row = 3

        # Segment inputs
        for i, segment in enumerate(segments):
            segment_style = self._get_segment_style(i)

            ws[f'A{current_row}'] = segment['name']
            ws[f'A{current_row}'].style = segment_style
            ws.merge_cells(f'A{current_row}:H{current_row}')
            current_row += 1

            ws[f'A{current_row}'] = segment['description']
            ws[f'A{current_row}'].style = 'label'
            ws.merge_cells(f'A{current_row}:H{current_row}')
            current_row += 1

            # Financial inputs
            input_data = [
                ("Revenue ($M)", segment['revenue']),
                ("EBITDA ($M)", segment['ebitda']),
                ("Net Income ($M)", segment['net_income']),
                ("Valuation Method", segment['valuation_method']),
                ("Multiple", segment['multiple']),
                ("Growth Rate", segment['growth_rate']),
                ("Beta", segment['beta']),
                ("Risk Premium", segment['risk_premium'])
            ]

            for label, value in input_data:
                ws[f'A{current_row}'] = label
                ws[f'A{current_row}'].style = 'label_bold'
                ws[f'B{current_row}'] = value
                ws[f'B{current_row}'].style = 'input'
                current_row += 1

            current_row += 2

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15

    def _create_valuations_tab(self, ws, segment_valuations):

        """Create Segment Valuations tab"""

        # Title
        ws['A1'] = f"Segment Valuations: {self.company_name} ({self.ticker})"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:F1')

        current_row = 3

        # Valuation Table
        ws[f'A{current_row}'] = 'SEGMENT VALUATIONS'
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:F{current_row}')
        current_row += 2

        # Headers
        headers = ['Segment', 'Valuation Method', 'Metric', 'Multiple', 'Enterprise Value ($M)', 'Equity Value ($M)']
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header).style = 'header'
        current_row += 1

        # Valuation data
        for i, val in enumerate(segment_valuations):
            segment_style = self._get_segment_style(i)

            ws.cell(row=current_row, column=1, value=val['name']).style = segment_style
            ws.cell(row=current_row, column=2, value=val['valuation_method']).style = 'calculation'
            ws.cell(row=current_row, column=3, value=val['metric_value']).style = 'calculation'
            ws.cell(row=current_row, column=4, value=val['multiple']).style = 'calculation'
            ws.cell(row=current_row, column=5, value=val['enterprise_value']).style = 'output'
            ws.cell(row=current_row, column=6, value=val['equity_value']).style = 'output'
            current_row += 1

        # Total row
        current_row += 1
        ws.cell(row=current_row, column=1, value="TOTAL").style = 'corporate'
        ws.cell(row=current_row, column=5, value=sum(val['enterprise_value'] for val in segment_valuations)).style = 'output'
        ws.cell(row=current_row, column=6, value=sum(val['equity_value'] for val in segment_valuations)).style = 'output'

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20

    def _create_consolidation_tab(self, ws, consolidation, analysis):

        """Create Corporate Consolidation tab"""

        # Title
        ws['A1'] = f"Corporate Consolidation: {self.company_name} ({self.ticker})"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:F1')

        current_row = 3

        # Consolidation Table
        ws[f'A{current_row}'] = 'CORPORATE CONSOLIDATION'
        ws[f'A{current_row}'].style = 'corporate'
        ws.merge_cells(f'A{current_row}:F{current_row}')
        current_row += 2

        # Segment EV
        ws[f'A{current_row}'] = "Segment Enterprise Values"
        ws[f'A{current_row}'].style = 'label_bold'
        ws[f'B{current_row}'] = consolidation['total_segment_ev']
        ws[f'B{current_row}'].style = 'calculation'
        current_row += 1

        # Adjustments
        adjustments = consolidation['adjustments']
        adjustment_items = [
            ("Less: Net Debt", -adjustments['net_debt']),
            ("Less: Minority Interests", -adjustments['minority_interests']),
            ("Add: Cash & Investments", adjustments['cash_and_investments']),
            ("Add: Other Assets", adjustments['other_assets']),
            ("Less: Other Liabilities", -adjustments['other_liabilities'])
        ]

        for label, value in adjustment_items:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label'
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].style = 'calculation'
            current_row += 1

        # Total EV
        current_row += 1
        ws[f'A{current_row}'] = "Total Enterprise Value"
        ws[f'A{current_row}'].style = 'label_bold'
        ws[f'B{current_row}'] = consolidation['total_ev']
        ws[f'B{current_row}'].style = 'output'
        current_row += 2

        # Equity Value
        ws[f'A{current_row}'] = "Equity Value"
        ws[f'A{current_row}'].style = 'label_bold'
        ws[f'B{current_row}'] = consolidation['equity_value']
        ws[f'B{current_row}'].style = 'output'
        current_row += 1

        # Share information
        share_data = [
            ("Shares Outstanding (M)", consolidation['shares_outstanding']),
            ("Implied Share Price ($)", consolidation['implied_share_price']),
            ("Current Share Price ($)", consolidation['current_share_price']),
            ("Premium/(Discount) ($)", consolidation['premium_discount']),
            ("Premium/(Discount) (%)", consolidation['premium_discount_pct'])
        ]

        for label, value in share_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            if "Premium" in label and "%" in label:
                ws[f'B{current_row}'].style = 'output' if consolidation['premium_discount_pct'] > 0 else 'warning'
            else:
                ws[f'B{current_row}'].style = 'output'
            current_row += 1

        current_row += 2

        # Conglomerate Analysis
        ws[f'A{current_row}'] = 'CONGLOMERATE ANALYSIS'
        ws[f'A{current_row}'].style = 'corporate'
        ws.merge_cells(f'A{current_row}:F{current_row}')
        current_row += 2

        conglomerate = analysis['conglomerate_analysis']
        conglomerate_data = [
            ("Current Market Cap", conglomerate['current_market_cap']),
            ("Implied Market Cap", conglomerate['implied_market_cap']),
            ("Conglomerate Effect", conglomerate['conglomerate_effect']),
            ("Conglomerate Effect (%)", conglomerate['conglomerate_effect_pct']),
            ("Status", conglomerate['status'])
        ]

        for label, value in conglomerate_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            if label == "Status":
                ws[f'B{current_row}'].style = 'output' if conglomerate['status'] == "PREMIUM" else 'warning'
            else:
                ws[f'B{current_row}'].style = 'output'
            current_row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

    def _get_segment_style(self, index):

        """Get appropriate style for segment"""
        styles = ['segment1', 'segment2', 'segment3', 'segment4']
        return styles[index % len(styles)]

    def _display_segment_breakdown(self, segment_valuations, consolidation):

        """Display segment breakdown in console"""

        print("\n🏗️  Segment Breakdown:")
        print("=" * 80)

        total_ev = consolidation['total_segment_ev']
        print("<25")
        print("-" * 90)

        for val in segment_valuations:
            contribution_pct = val['enterprise_value'] / total_ev * 100 if total_ev > 0 else 0
            print(f"{val['segment_name']:<25}"
                  f"{contribution_pct:<20.1f}%")

        print("-" * 90)
        print(f"{'Total Enterprise Value':<25}"
              f"{100.0:<20.1f}%")

        print("\n🏢 Corporate Adjustments:")
        print("=" * 80)

        adjustments = consolidation['adjustments']
        print(f"{'Cash & Equivalents':<30}"
              f"${adjustments['cash']/1000:<20.1f}B")
        print(f"{'Total Debt':<30}"
              f"${adjustments['debt']/1000:<20.1f}B")
        print(f"{'Minority Interests':<30}"
              f"${adjustments['minority_interests']/1000:<20.1f}B")
        print(f"{'Other Adjustments':<30}"
              f"${adjustments['other']/1000:<20.1f}B")
        print(f"{'Net Corporate Adjustments':<30}"
              f"${adjustments['net_adjustments']/1000:<20.1f}B")
        print("-" * 60)
        print(f"{'Implied Equity Value':<30}"
              f"${consolidation['equity_value']/1000:<20.1f}B")

        print("\n💰 Final Valuation:")
        print("=" * 80)
        print(f"{'Implied Share Price':<30}"
              f"${consolidation['implied_share_price']:<20.2f}")
        print(f"{'Shares Outstanding':<30}"
              f"{consolidation['shares_outstanding']/1000000:<20.1f}M")
        print(f"{'Current Market Cap':<30}"
              f"${consolidation['current_market_cap']/1000:<20.1f}B")
        print(f"{'Upside/(Downside)':<30}"
              f"{consolidation['upside_downside']:<20.1f}%")
        print(f"{'Valuation Date':<30}"
              f"{consolidation.get('valuation_date', 'N/A'):<20}")

        conglomerate = consolidation.get('conglomerate_analysis', {})
        if conglomerate:
            print(f"{'Conglomerate Discount':<30}"
                  f"{conglomerate.get('discount', 0):<20.1f}%")
            print(f"{'Discounted Equity Value':<30}"
                  f"${conglomerate.get('discounted_equity', 0)/1000:<25.1f}B")


def run_sample_sotp_model():
    """Run a sample SOTP valuation model"""

    print("🧩 Running Professional Sum-of-the-Parts Valuation Model Sample")
    print("=" * 75)

    # Create model instance
    model = ProfessionalSOTPModel("Global Industries Inc.", "GLOBAL")

    # Run the model with default segments
    results = model.run_sotp_model(
        company_name="Global Industries Inc.",
        ticker="GLOBAL",

        # Corporate adjustments
        net_debt=2000.0,                  # $2B net debt
        minority_interests=500.0,         # $500M minority interests
        cash_and_investments=1500.0,      # $1.5B cash
        other_assets=300.0,               # $300M other assets
        other_liabilities=200.0,          # $200M other liabilities
        shares_outstanding=150.0,         # 150M shares
        current_share_price=85.0          # $85 current price
    )

    return results


if __name__ == "__main__":
    # Run sample SOTP model
    results = run_sample_sotp_model()

    print("\n📋 Sum-of-the-Parts Model Complete!")
    print(f"📁 Excel file saved as: {results['excel_file']}")

    # Display key results
    print("\n🔑 Key Results Summary:")
    print("=" * 40)

    consolidation = results['consolidation']
    print(f"Total Enterprise Value: ${consolidation['total_ev']/1000:.1f}B")
    print(f"Equity Value: ${consolidation['equity_value']/1000:.1f}B")
    print(f"Implied Share Price: ${consolidation['implied_share_price']:.2f}")
    print(f"Current Share Price: ${consolidation['current_share_price']:.2f}")
    print(f"Premium/(Discount): {consolidation['premium_discount_pct']:.1%}")

    analysis = results['analysis']
    conglomerate = analysis['conglomerate_analysis']
    print(f"Conglomerate Effect: {conglomerate['status']} of {abs(conglomerate['conglomerate_effect_pct']):.1%}")

    # Show largest segment
    largest_segment = max(results['segment_valuations'], key=lambda x: x['enterprise_value'])
    print(f"Largest Segment: {largest_segment['name']} (${largest_segment['enterprise_value']/1000:.1f}B)")
