#!/usr/bin/env python3
"""
Professional Free Cash Flow (FCF) Model Builder
Creates focused FCF analysis with professional formatting and Excel output

Author: Investment Banking Modeler
Date: 2024

Features:
- Revenue to UFCF calculation chain
- Multiple scenario analysis (Base, Bull, Bear)
- Sensitivity analysis tables
- Terminal value calculations
- Professional Excel Output with Multiple Tabs
- Investment Banker-Style Formatting
"""

import numpy as np
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# Professional color scheme
FCF_MODEL_COLORS = {
    'header_blue': '1F4E79',
    'input_light_blue': 'DDEBF7',
    'calculation_yellow': 'FFF2CC',
    'result_green': 'D8E4BC',
    'warning_red': 'F8CECC',
    'border_dark': '2C3E50',
    'text_dark': '000000',
    'text_white': 'FFFFFF',
    'scenario_blue': 'B8CCE4',
    'scenario_green': 'C4D79B',
    'scenario_red': 'F2DCDB',
    'base_case_highlight': 'FFFF00'  # Yellow for base case
}

class ProfessionalFCFModel:
    """
    Comprehensive Free Cash Flow Model with Professional Formatting
    """

    def __init__(self, company_name="Sample Company", ticker="TICK"):
        self.company_name = company_name
        self.ticker = ticker
        self.model_date = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Initialize styles
        self.styles = self._create_styles()

    def _create_styles(self):
        """Create professional Excel styles"""
        styles = {}

        # Header style
        styles['header'] = NamedStyle(name='header')
        styles['header'].font = Font(name='Calibri', size=12, bold=True, color=FCF_MODEL_COLORS['text_white'])
        styles['header'].fill = PatternFill(start_color=FCF_MODEL_COLORS['header_blue'], end_color=FCF_MODEL_COLORS['header_blue'], fill_type='solid')
        styles['header'].alignment = Alignment(horizontal='center', vertical='center')

        # Input style
        styles['input'] = NamedStyle(name='input')
        styles['input'].font = Font(name='Calibri', size=10, color=FCF_MODEL_COLORS['text_dark'])
        styles['input'].fill = PatternFill(start_color=FCF_MODEL_COLORS['input_light_blue'], end_color=FCF_MODEL_COLORS['input_light_blue'], fill_type='solid')
        styles['input'].alignment = Alignment(horizontal='right', vertical='center')
        styles['input'].number_format = '#,##0'

        # Calculation style
        styles['calculation'] = NamedStyle(name='calculation')
        styles['calculation'].font = Font(name='Calibri', size=10, color=FCF_MODEL_COLORS['text_dark'])
        styles['calculation'].fill = PatternFill(start_color=FCF_MODEL_COLORS['calculation_yellow'], end_color=FCF_MODEL_COLORS['calculation_yellow'], fill_type='solid')
        styles['calculation'].alignment = Alignment(horizontal='right', vertical='center')
        styles['calculation'].number_format = '#,##0'

        # Result style
        styles['result'] = NamedStyle(name='result')
        styles['result'].font = Font(name='Calibri', size=10, bold=True, color=FCF_MODEL_COLORS['text_dark'])
        styles['result'].fill = PatternFill(start_color=FCF_MODEL_COLORS['result_green'], end_color=FCF_MODEL_COLORS['result_green'], fill_type='solid')
        styles['result'].alignment = Alignment(horizontal='right', vertical='center')
        styles['result'].number_format = '#,##0'

        # Warning style
        styles['warning'] = NamedStyle(name='warning')
        styles['warning'].font = Font(name='Calibri', size=10, color=FCF_MODEL_COLORS['text_dark'])
        styles['warning'].fill = PatternFill(start_color=FCF_MODEL_COLORS['warning_red'], end_color=FCF_MODEL_COLORS['warning_red'], fill_type='solid')
        styles['warning'].alignment = Alignment(horizontal='center', vertical='center')

        # Scenario styles
        styles['scenario_base'] = NamedStyle(name='scenario_base')
        styles['scenario_base'].font = Font(name='Calibri', size=10, color=FCF_MODEL_COLORS['text_dark'])
        styles['scenario_base'].fill = PatternFill(start_color=FCF_MODEL_COLORS['scenario_blue'], end_color=FCF_MODEL_COLORS['scenario_blue'], fill_type='solid')

        styles['scenario_bull'] = NamedStyle(name='scenario_bull')
        styles['scenario_bull'].font = Font(name='Calibri', size=10, color=FCF_MODEL_COLORS['text_dark'])
        styles['scenario_bull'].fill = PatternFill(start_color=FCF_MODEL_COLORS['scenario_green'], end_color=FCF_MODEL_COLORS['scenario_green'], fill_type='solid')

        styles['scenario_bear'] = NamedStyle(name='scenario_bear')
        styles['scenario_bear'].font = Font(name='Calibri', size=10, color=FCF_MODEL_COLORS['text_dark'])
        styles['scenario_bear'].fill = PatternFill(start_color=FCF_MODEL_COLORS['scenario_red'], end_color=FCF_MODEL_COLORS['scenario_red'], fill_type='solid')

        # Base case highlight
        styles['base_case_highlight'] = NamedStyle(name='base_case_highlight')
        styles['base_case_highlight'].font = Font(name='Calibri', size=10, bold=True, color=FCF_MODEL_COLORS['text_dark'])
        styles['base_case_highlight'].fill = PatternFill(start_color=FCF_MODEL_COLORS['base_case_highlight'], end_color=FCF_MODEL_COLORS['base_case_highlight'], fill_type='solid')
        styles['base_case_highlight'].alignment = Alignment(horizontal='right', vertical='center')
        styles['base_case_highlight'].number_format = '#,##0'

        # Subtotal style
        styles['subtotal'] = NamedStyle(name='subtotal')
        styles['subtotal'].font = Font(name='Calibri', size=10, bold=True, color=FCF_MODEL_COLORS['text_dark'])
        styles['subtotal'].fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')  # Light grey
        styles['subtotal'].alignment = Alignment(horizontal='right', vertical='center')
        styles['subtotal'].number_format = '#,##0'

        # UFCF style (special highlighting)
        styles['ufcf'] = NamedStyle(name='ufcf')
        styles['ufcf'].font = Font(name='Calibri', size=10, bold=True, color=FCF_MODEL_COLORS['text_dark'])
        styles['ufcf'].fill = PatternFill(start_color='D5F4E6', end_color='D5F4E6', fill_type='solid')  # Light green
        styles['ufcf'].alignment = Alignment(horizontal='right', vertical='center')
        styles['ufcf'].number_format = '#,##0'

        # Bold label style
        styles['label_bold'] = NamedStyle(name='label_bold')
        styles['label_bold'].font = Font(name='Calibri', size=10, bold=True, color=FCF_MODEL_COLORS['text_dark'])
        styles['label_bold'].alignment = Alignment(horizontal='left', vertical='center')

        # Normal label style
        styles['label'] = NamedStyle(name='label')
        styles['label'].font = Font(name='Calibri', size=10, color=FCF_MODEL_COLORS['text_dark'])
        styles['label'].alignment = Alignment(horizontal='left', vertical='center')

        return styles

    def run_fcf_model(self,
                     # Company Basics
                     starting_revenue=1000.0,  # $M

                     # Revenue Growth Scenarios (%)
                     growth_base=[0.08, 0.07, 0.06, 0.05, 0.04],  # Year 1-5
                     growth_bull=[0.12, 0.10, 0.08, 0.06, 0.05],
                     growth_bear=[0.04, 0.03, 0.02, 0.02, 0.01],

                     # Operating Margins (%)
                     ebitda_margin=0.30,  # 30% EBITDA margin
                     ebit_margin=0.25,    # 25% EBIT margin

                     # Depreciation & Amortization (% of revenue)
                     depreciation_pct=0.05,  # 5% of revenue

                     # Capital Expenditures (% of revenue)
                     capex_pct=0.04,  # 4% of revenue

                     # Net Working Capital (% of revenue)
                     nwc_pct=0.10,  # 10% of revenue

                     # Tax Rate
                     tax_rate=0.25,  # 25%

                     # Terminal Value Assumptions
                     terminal_growth_rate=0.025,  # 2.5% perpetual growth
                     wacc=0.09,  # 9% WACC for terminal value
                     terminal_multiple=None,  # Alternative: use EBITDA multiple

                     forecast_years=5):

        """
        Run complete FCF model with scenarios and sensitivities
        """

        print(f"💰 Building Professional FCF Model for {self.company_name} ({self.ticker})")
        print("=" * 80)

        # Step 1: Create Assumptions & Drivers
        assumptions = self._create_assumptions(
            starting_revenue, growth_base, growth_bull, growth_bear,
            ebitda_margin, ebit_margin, depreciation_pct, capex_pct,
            nwc_pct, tax_rate, terminal_growth_rate, wacc, terminal_multiple, forecast_years
        )

        # Step 2: Generate FCF Forecasts for all scenarios
        fcf_forecasts = self._create_fcf_forecasts(assumptions)

        # Step 3: Calculate Terminal Values
        terminal_values = self._calculate_terminal_values(assumptions, fcf_forecasts)

        # Step 4: Create Sensitivity Analysis
        sensitivity_analysis = self._create_sensitivity_analysis(assumptions)

        # Step 5: Generate Summary Metrics
        summary_metrics = self._generate_summary_metrics(fcf_forecasts, terminal_values)

        # Compile results
        fcf_results = {
            'assumptions': assumptions,
            'fcf_forecasts': fcf_forecasts,
            'terminal_values': terminal_values,
            'sensitivity_analysis': sensitivity_analysis,
            'summary_metrics': summary_metrics
        }

        # Create Excel output
        excel_file = self._create_excel_output(fcf_results)

        print("\n✅ FCF Model Complete!")
        print("📊 Key Metrics (Base Case):")
        print(f"   • Year 5 UFCF: ${fcf_forecasts['base']['ufcf'][-1]:.0f}M")
        print(f"   • Average UFCF Margin: {np.mean([ufcf/rev for ufcf, rev in zip(fcf_forecasts['base']['ufcf'], fcf_forecasts['base']['revenue'])])*100:.1f}%")
        print(f"   • Total UFCF (Years 1-5): ${np.sum(fcf_forecasts['base']['ufcf']):.0f}M")
        if terminal_values['base']['terminal_value'] is not None:
            print(f"   • Terminal Value: ${terminal_values['base']['terminal_value']/1000:.1f}B")
            print(f"   • Total PV: ${terminal_values['base']['present_value']/1000:.1f}B")
        print(f"📁 Excel Output: {excel_file}")

        return fcf_results, excel_file

    def _create_assumptions(self, starting_revenue, growth_base, growth_bull, growth_bear,
                           ebitda_margin, ebit_margin, depreciation_pct, capex_pct,
                           nwc_pct, tax_rate, terminal_growth_rate, wacc, terminal_multiple, forecast_years):

        """Create comprehensive assumptions for all scenarios"""

        assumptions = {
            # Company Basics
            'starting_revenue': starting_revenue,
            'forecast_years': forecast_years,
            'years': list(range(2025, 2025 + forecast_years)),

            # Growth Scenarios (%)
            'growth_base': growth_base,
            'growth_bull': growth_bull,
            'growth_bear': growth_bear,

            # Operating Margins (%)
            'ebitda_margin': ebitda_margin,
            'ebit_margin': ebit_margin,

            # Depreciation & Amortization (% of revenue)
            'depreciation_pct': depreciation_pct,

            # Capital Expenditures (% of revenue)
            'capex_pct': capex_pct,

            # Net Working Capital (% of revenue)
            'nwc_pct': nwc_pct,

            # Tax Rate
            'tax_rate': tax_rate,

            # Terminal Value Assumptions
            'terminal_growth_rate': terminal_growth_rate,
            'wacc': wacc,
            'terminal_multiple': terminal_multiple
        }

        print("📋 Assumptions Created:")
        print(f"   • Starting Revenue: ${starting_revenue:.0f}M")
        print(f"   • Forecast Period: {forecast_years} years")
        print(f"   • Base Case Growth: {growth_base[0]*100:.1f}% → {growth_base[-1]*100:.1f}%")
        print(f"   • EBITDA Margin: {ebitda_margin*100:.1f}%")
        print(f"   • CapEx as % of Revenue: {capex_pct*100:.1f}%")
        print(f"   • NWC as % of Revenue: {nwc_pct*100:.1f}%")

        return assumptions

    def _create_fcf_forecasts(self, assumptions):
        """Create FCF forecasts for all scenarios"""

        fcf_forecasts = {}

        for scenario in ['base', 'bull', 'bear']:
            growth_rates = assumptions[f'growth_{scenario}']

            # Calculate revenue
            revenue = [assumptions['starting_revenue']]
            for i in range(assumptions['forecast_years'] - 1):
                if i < len(growth_rates):
                    growth = growth_rates[i]
                else:
                    growth = growth_rates[-1]  # Use terminal growth
                revenue.append(revenue[-1] * (1 + growth))

            # Calculate EBITDA
            ebitda = [rev * assumptions['ebitda_margin'] for rev in revenue]

            # Calculate EBIT (alternative calculation)
            ebit_alt = [rev * assumptions['ebit_margin'] for rev in revenue]

            # Use EBITDA margin for primary calculation (more common)
            ebit = ebitda.copy()

            # Depreciation & Amortization
            depreciation = [rev * assumptions['depreciation_pct'] for rev in revenue]

            # EBIT = EBITDA - D&A
            ebit = [ebitda[i] - depreciation[i] for i in range(len(ebitda))]

            # NOPAT = EBIT * (1 - tax rate)
            nopat = [ebit[i] * (1 - assumptions['tax_rate']) for i in range(len(ebit))]

            # Capital Expenditures
            capex = [rev * assumptions['capex_pct'] for rev in revenue]

            # Net Working Capital
            nwc = [rev * assumptions['nwc_pct'] for rev in revenue]

            # Change in NWC
            delta_nwc = [0]  # First year has no prior year
            for i in range(1, len(nwc)):
                delta_nwc.append(nwc[i] - nwc[i-1])

            # Unlevered Free Cash Flow (UFCF)
            # UFCF = NOPAT + D&A - CapEx - ∆NWC
            ufcf = [nopat[i] + depreciation[i] - capex[i] - delta_nwc[i] for i in range(len(nopat))]

            # Additional metrics
            ufcf_margin = [ufcf[i] / revenue[i] if revenue[i] != 0 else 0 for i in range(len(ufcf))]
            cumulative_ufcf = [sum(ufcf[:i+1]) for i in range(len(ufcf))]

            fcf_forecasts[scenario] = {
                'years': assumptions['years'],
                'revenue': revenue,
                'ebitda': ebitda,
                'depreciation': depreciation,
                'ebit': ebit,
                'nopat': nopat,
                'capex': capex,
                'nwc': nwc,
                'delta_nwc': delta_nwc,
                'ufcf': ufcf,
                'ufcf_margin': ufcf_margin,
                'cumulative_ufcf': cumulative_ufcf
            }

        print("💰 FCF Forecasts Created:")
        print(f"   • Base Case Year 5 Revenue: ${fcf_forecasts['base']['revenue'][-1]/1000:.1f}B")
        print(f"   • Base Case Year 5 UFCF: ${fcf_forecasts['base']['ufcf'][-1]:.0f}M")
        print(f"   • Base Case Average UFCF Margin: {np.mean(fcf_forecasts['base']['ufcf_margin'])*100:.1f}%")
        print(f"   • Base Case Total UFCF (Years 1-5): ${np.sum(fcf_forecasts['base']['ufcf']):.0f}M")

        return fcf_forecasts

    def _calculate_terminal_values(self, assumptions, fcf_forecasts):
        """Calculate terminal values for all scenarios"""

        terminal_values = {}

        for scenario in ['base', 'bull', 'bear']:
            fcf_data = fcf_forecasts[scenario]

            # Terminal year calculations
            terminal_year = assumptions['forecast_years']

            # Terminal revenue (using terminal growth rate)
            terminal_revenue = fcf_data['revenue'][-1] * (1 + assumptions['terminal_growth_rate'])

            # Terminal EBITDA
            terminal_ebitda = terminal_revenue * assumptions['ebitda_margin']

            # Terminal EBIT
            terminal_ebit = terminal_ebitda - (terminal_revenue * assumptions['depreciation_pct'])

            # Terminal NOPAT
            terminal_nopat = terminal_ebit * (1 - assumptions['tax_rate'])

            # Terminal UFCF
            terminal_ufcf = (terminal_nopat +
                           (terminal_revenue * assumptions['depreciation_pct']) -
                           (terminal_revenue * assumptions['capex_pct']) -
                           (terminal_revenue * assumptions['nwc_pct'] * assumptions['terminal_growth_rate']))

            terminal_value = None
            present_value = None

            if assumptions['terminal_multiple'] is not None:
                # Terminal value using EBITDA multiple
                terminal_value = terminal_ebitda * assumptions['terminal_multiple']
            else:
                # Terminal value using perpetuity growth formula
                terminal_value = terminal_ufcf / (assumptions['wacc'] - assumptions['terminal_growth_rate'])

            # Present value of terminal value (discounted back to end of forecast period)
            if terminal_value is not None:
                present_value = terminal_value / ((1 + assumptions['wacc']) ** terminal_year)

            terminal_values[scenario] = {
                'terminal_year': terminal_year,
                'terminal_revenue': terminal_revenue,
                'terminal_ebitda': terminal_ebitda,
                'terminal_ufcf': terminal_ufcf,
                'terminal_value': terminal_value,
                'present_value': present_value,
                'terminal_multiple': assumptions['terminal_multiple'],
                'terminal_growth_rate': assumptions['terminal_growth_rate'],
                'wacc': assumptions['wacc']
            }

        print("🎯 Terminal Values Calculated:")
        if terminal_values['base']['terminal_value'] is not None:
            print(f"   • Base Case Terminal Value: ${terminal_values['base']['terminal_value']/1000:.1f}B")
            print(f"   • Base Case PV of Terminal: ${terminal_values['base']['present_value']/1000:.1f}B")

        return terminal_values

    def _create_sensitivity_analysis(self, assumptions):
        """Create sensitivity analysis tables"""

        # Sensitivity: Growth Rate vs EBITDA Margin
        growth_rates = np.arange(0.02, 0.16, 0.02)  # 2% to 14%
        ebitda_margins = np.arange(0.15, 0.36, 0.05)  # 15% to 35%

        growth_margin_sensitivity = []
        for growth in growth_rates:
            row = []
            for margin in ebitda_margins:
                # Calculate UFCF for Year 5 with these assumptions
                revenue_yr5 = assumptions['starting_revenue'] * ((1 + growth) ** (assumptions['forecast_years'] - 1))
                ebitda_yr5 = revenue_yr5 * margin
                ebit_yr5 = ebitda_yr5 - (revenue_yr5 * assumptions['depreciation_pct'])
                nopat_yr5 = ebit_yr5 * (1 - assumptions['tax_rate'])
                ufcf_yr5 = (nopat_yr5 +
                          (revenue_yr5 * assumptions['depreciation_pct']) -
                          (revenue_yr5 * assumptions['capex_pct']) -
                          (revenue_yr5 * assumptions['nwc_pct'] * growth))
                row.append(ufcf_yr5)
            growth_margin_sensitivity.append(row)

        # Sensitivity: CapEx % vs NWC %
        capex_rates = np.arange(0.02, 0.11, 0.02)  # 2% to 10%
        nwc_rates = np.arange(0.05, 0.21, 0.05)    # 5% to 20%

        capex_nwc_sensitivity = []
        for capex_pct in capex_rates:
            row = []
            for nwc_pct in nwc_rates:
                # Use base case assumptions for other variables
                revenue_yr5 = assumptions['starting_revenue'] * ((1 + assumptions['growth_base'][-1]) ** (assumptions['forecast_years'] - 1))
                ebitda_yr5 = revenue_yr5 * assumptions['ebitda_margin']
                ebit_yr5 = ebitda_yr5 - (revenue_yr5 * assumptions['depreciation_pct'])
                nopat_yr5 = ebit_yr5 * (1 - assumptions['tax_rate'])
                ufcf_yr5 = (nopat_yr5 +
                          (revenue_yr5 * assumptions['depreciation_pct']) -
                          (revenue_yr5 * capex_pct) -
                          (revenue_yr5 * nwc_pct * assumptions['growth_base'][-1]))
                row.append(ufcf_yr5)
            capex_nwc_sensitivity.append(row)

        sensitivity_analysis = {
            'growth_rates': growth_rates,
            'ebitda_margins': ebitda_margins,
            'growth_margin_sensitivity': growth_margin_sensitivity,
            'capex_rates': capex_rates,
            'nwc_rates': nwc_rates,
            'capex_nwc_sensitivity': capex_nwc_sensitivity
        }

        print("📊 Sensitivity Analysis Created:")
        print("   • Growth Rate vs EBITDA Margin sensitivity table")
        print("   • CapEx % vs NWC % sensitivity table")
        return sensitivity_analysis

    def _generate_summary_metrics(self, fcf_forecasts, terminal_values):
        """Generate summary metrics for all scenarios"""

        summary_metrics = {}

        for scenario in ['base', 'bull', 'bear']:
            fcf_data = fcf_forecasts[scenario]
            terminal_data = terminal_values[scenario]

            # Calculate key metrics
            total_ufcf = np.sum(fcf_data['ufcf'])
            avg_ufcf_margin = np.mean(fcf_data['ufcf_margin'])
            max_ufcf = np.max(fcf_data['ufcf'])
            min_ufcf = np.min(fcf_data['ufcf'])
            ufcf_cagr = ((fcf_data['ufcf'][-1] / fcf_data['ufcf'][0]) ** (1 / (len(fcf_data['ufcf']) - 1))) - 1 if fcf_data['ufcf'][0] != 0 else 0

            # Revenue CAGR
            revenue_cagr = ((fcf_data['revenue'][-1] / fcf_data['revenue'][0]) ** (1 / (len(fcf_data['revenue']) - 1))) - 1

            summary_metrics[scenario] = {
                'total_ufcf': total_ufcf,
                'avg_ufcf_margin': avg_ufcf_margin,
                'max_ufcf': max_ufcf,
                'min_ufcf': min_ufcf,
                'ufcf_cagr': ufcf_cagr,
                'revenue_cagr': revenue_cagr,
                'terminal_value': terminal_data['terminal_value'],
                'present_value_terminal': terminal_data['present_value'],
                'total_present_value': total_ufcf + (terminal_data['present_value'] or 0)
            }

        print("📈 Summary Metrics Generated:")
        print(f"   • Base Case Total UFCF: ${summary_metrics['base']['total_ufcf']:.0f}M")
        print(f"   • Base Case UFCF CAGR: {summary_metrics['base']['ufcf_cagr']*100:.1f}%")
        print(f"   • Base Case Revenue CAGR: {summary_metrics['base']['revenue_cagr']*100:.1f}%")

        return summary_metrics

    def _create_excel_output(self, fcf_results):
        """Create professional Excel output with multiple tabs"""

        wb = Workbook()

        # Create worksheets
        ws_assumptions = wb.active
        ws_assumptions.title = "Assumptions"

        ws_fcf_base = wb.create_sheet("FCF Forecast - Base")
        ws_fcf_bull = wb.create_sheet("FCF Forecast - Bull")
        ws_fcf_bear = wb.create_sheet("FCF Forecast - Bear")

        ws_sensitivity = wb.create_sheet("Sensitivity Analysis")
        ws_summary = wb.create_sheet("Summary Snapshot")

        # Apply styles to workbook
        for style_name, style in self.styles.items():
            if style_name not in wb.named_styles:
                wb.add_named_style(style)

        # Create each tab
        self._create_assumptions_tab(ws_assumptions, fcf_results)
        self._create_fcf_forecast_tabs(ws_fcf_base, ws_fcf_bull, ws_fcf_bear, fcf_results)
        self._create_sensitivity_tab(ws_sensitivity, fcf_results)
        self._create_summary_tab(ws_summary, fcf_results)

        # Save workbook
        filename = f"FCF_Model_{self.ticker}_{self.model_date}.xlsx"
        wb.save(filename)

        return filename

    def _create_assumptions_tab(self, ws, results):
        """Create Assumptions tab"""

        assumptions = results['assumptions']

        # Title
        ws['A1'] = f"{self.company_name} ({self.ticker}) - FCF Model Assumptions"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:G1')

        current_row = 3

        # Company Basics
        ws[f'A{current_row}'] = "COMPANY BASICS"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        basics_data = [
            ("Starting Revenue ($M)", assumptions['starting_revenue']),
            ("Forecast Years", assumptions['forecast_years'])
        ]

        for label, value in basics_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].style = 'input'
            current_row += 1

        current_row += 2

        # Growth Scenarios
        ws[f'A{current_row}'] = "REVENUE GROWTH SCENARIOS (%)"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        growth_headers = ['Year'] + ['Base Case', 'Bull Case', 'Bear Case']

        for col, header in enumerate(growth_headers, 1):
            ws.cell(row=current_row, column=col, value=header).style = 'header'

        current_row += 1

        for i in range(max(len(assumptions['growth_base']), len(assumptions['growth_bull']), len(assumptions['growth_bear']))):
            ws.cell(row=current_row, column=1, value=f'Year {i+1}').style = 'label'

            if i < len(assumptions['growth_base']):
                ws.cell(row=current_row, column=2, value=f"{assumptions['growth_base'][i]*100:.1f}%").style = 'input'
            if i < len(assumptions['growth_bull']):
                ws.cell(row=current_row, column=3, value=f"{assumptions['growth_bull'][i]*100:.1f}%").style = 'input'
            if i < len(assumptions['growth_bear']):
                ws.cell(row=current_row, column=4, value=f"{assumptions['growth_bear'][i]*100:.1f}%").style = 'input'

            current_row += 1

        current_row += 2

        # Operating Assumptions
        ws[f'A{current_row}'] = "OPERATING ASSUMPTIONS"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        operating_data = [
            ("EBITDA Margin (%)", f"{assumptions['ebitda_margin']*100:.1f}%"),
            ("EBIT Margin (%)", f"{assumptions['ebit_margin']*100:.1f}%"),
            ("Depreciation (% of Revenue)", f"{assumptions['depreciation_pct']*100:.1f}%"),
            ("CapEx (% of Revenue)", f"{assumptions['capex_pct']*100:.1f}%"),
            ("NWC (% of Revenue)", f"{assumptions['nwc_pct']*100:.1f}%"),
            ("Tax Rate", f"{assumptions['tax_rate']*100:.1f}%")
        ]

        for label, value in operating_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].style = 'input'
            current_row += 1

        current_row += 2

        # Terminal Value Assumptions
        ws[f'A{current_row}'] = "TERMINAL VALUE ASSUMPTIONS"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 2

        terminal_data = [
            ("Terminal Growth Rate", f"{assumptions['terminal_growth_rate']*100:.1f}%"),
            ("WACC", f"{assumptions['wacc']*100:.1f}%"),
            ("Terminal EBITDA Multiple", assumptions['terminal_multiple'] or "N/A")
        ]

        for label, value in terminal_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].style = 'label_bold'
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].style = 'input'
            current_row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    def _create_fcf_forecast_tabs(self, ws_base, ws_bull, ws_bear, results):
        """Create FCF Forecast tabs for all scenarios"""

        scenarios = [('base', ws_base), ('bull', ws_bull), ('bear', ws_bear)]

        for scenario, ws in scenarios:
            fcf_data = results['fcf_forecasts'][scenario]
            terminal_data = results['terminal_values'][scenario]

            # Title
            ws['A1'] = f"{self.company_name} - FCF Forecast ({scenario.title()} Case)"
            ws['A1'].style = 'header'
            ws.merge_cells('A1:K1')

            # Headers
            headers = ['Year', 'Revenue', 'EBITDA', 'D&A', 'EBIT', 'Taxes', 'NOPAT', 'CapEx', 'ΔNWC', 'UFCF', 'UFCF Margin']

            for col, header in enumerate(headers, 1):
                ws.cell(row=3, column=col, value=header).style = 'header'

            # Data rows
            for i, year in enumerate(fcf_data['years']):
                row = i + 4
                ws.cell(row=row, column=1, value=year).style = 'label'

                values = [
                    fcf_data['revenue'][i],
                    fcf_data['ebitda'][i],
                    fcf_data['depreciation'][i],
                    fcf_data['ebit'][i],
                    fcf_data['ebit'][i] * results['assumptions']['tax_rate'],  # Taxes
                    fcf_data['nopat'][i],
                    fcf_data['capex'][i],
                    fcf_data['delta_nwc'][i],
                    fcf_data['ufcf'][i],
                    fcf_data['ufcf_margin'][i]
                ]

                for col, value in enumerate(values, 2):
                    cell = ws.cell(row=row, column=col, value=value)
                    if col in [5, 7, 11]:  # EBIT, NOPAT, UFCF
                        cell.style = 'subtotal'
                    elif col == 11:  # UFCF
                        cell.style = 'ufcf'
                    else:
                        cell.style = 'calculation'

                    if col == 11:  # UFCF Margin
                        cell.number_format = '0.0%'
                    else:
                        cell.number_format = '#,##0'

            # Add double underline to UFCF row
            from openpyxl.styles import Border, Side
            double_border = Border(bottom=Side(style='double'))

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=len(fcf_data['years']) + 4, column=col)
                if col == 11:  # UFCF column
                    cell.border = double_border

            # Terminal Value section
            terminal_row = len(fcf_data['years']) + 6

            ws.cell(row=terminal_row, column=1, value=f"Terminal Year ({terminal_data['terminal_year'] + 1})").style = 'label_bold'

            if terminal_data['terminal_value'] is not None:
                ws.cell(row=terminal_row, column=10, value=terminal_data['terminal_ufcf']).style = 'calculation'
                ws.cell(row=terminal_row, column=10).number_format = '#,##0'

                ws.cell(row=terminal_row + 1, column=1, value="Terminal Value").style = 'label_bold'
                ws.cell(row=terminal_row + 1, column=10, value=terminal_data['terminal_value']).style = 'result'
                ws.cell(row=terminal_row + 1, column=10).number_format = '#,##0'

                ws.cell(row=terminal_row + 2, column=1, value="PV of Terminal Value").style = 'label_bold'
                ws.cell(row=terminal_row + 2, column=10, value=terminal_data['present_value']).style = 'result'
                ws.cell(row=terminal_row + 2, column=10).number_format = '#,##0'

            # Set column widths
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 12

    def _create_sensitivity_tab(self, ws, results):
        """Create Sensitivity Analysis tab"""

        sensitivity = results['sensitivity_analysis']

        # Title
        ws['A1'] = f"{self.company_name} - Sensitivity Analysis"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:K1')

        # Growth Rate vs EBITDA Margin Sensitivity
        ws['A3'] = 'SENSITIVITY: Growth Rate vs EBITDA Margin (Year 5 UFCF, $M)'
        ws['A3'].style = 'header'
        ws.merge_cells('A3:K3')

        # Growth vs EBITDA table
        growth_headers = ['EBITDA Margin \\ Growth'] + [f'{rate*100:.0f}%' for rate in sensitivity['growth_rates']]

        for col, header in enumerate(growth_headers, 1):
            ws.cell(row=5, column=col, value=header).style = 'header'

        for i, margin in enumerate(sensitivity['ebitda_margins']):
            row = i + 6
            ws.cell(row=row, column=1, value=f'{margin*100:.0f}%').style = 'label_bold'

            for j, ufcf_value in enumerate(sensitivity['growth_margin_sensitivity'][i]):
                cell = ws.cell(row=row, column=j+2, value=ufcf_value)
                cell.style = 'calculation'
                cell.number_format = '#,##0'

        # CapEx % vs NWC % Sensitivity
        current_row = len(sensitivity['ebitda_margins']) + 10

        ws[f'A{current_row}'] = 'SENSITIVITY: CapEx % vs NWC % (Year 5 UFCF, $M)'
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:K{current_row}')

        # CapEx vs NWC table
        capex_headers = ['NWC % \\ CapEx %'] + [f'{rate*100:.0f}%' for rate in sensitivity['capex_rates']]

        for col, header in enumerate(capex_headers, 1):
            ws.cell(row=current_row+2, column=col, value=header).style = 'header'

        for i, nwc_rate in enumerate(sensitivity['nwc_rates']):
            row = current_row + 3 + i
            ws.cell(row=row, column=1, value=f'{nwc_rate*100:.0f}%').style = 'label_bold'

            for j, ufcf_value in enumerate(sensitivity['capex_nwc_sensitivity'][i]):
                cell = ws.cell(row=row, column=j+2, value=ufcf_value)
                cell.style = 'calculation'
                cell.number_format = '#,##0'

        # Set column widths
        for col in range(1, max(len(growth_headers), len(capex_headers)) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12

    def _create_summary_tab(self, ws, results):
        """Create Summary Snapshot tab"""

        # Title
        ws['A1'] = f"{self.company_name} ({self.ticker}) - FCF Model Summary"
        ws['A1'].style = 'header'
        ws.merge_cells('A1:F1')

        current_row = 3

        # Key Metrics Summary
        ws[f'A{current_row}'] = "KEY METRICS SUMMARY"
        ws[f'A{current_row}'].style = 'header'
        ws.merge_cells(f'A{current_row}:F{current_row}')
        current_row += 2

        scenarios = ['base', 'bull', 'bear']

        # Headers
        ws.cell(row=current_row, column=1, value="Metric").style = 'header'
        for i, scenario in enumerate(scenarios):
            ws.cell(row=current_row, column=i+2, value=f"{scenario.title()} Case").style = 'header'
        current_row += 1

        # Summary data
        summary_data = [
            ("Total UFCF (Years 1-5, $M)", [results['summary_metrics'][s]['total_ufcf'] for s in scenarios]),
            ("Average UFCF Margin", [results['summary_metrics'][s]['avg_ufcf_margin'] for s in scenarios]),
            ("UFCF CAGR", [results['summary_metrics'][s]['ufcf_cagr'] for s in scenarios]),
            ("Revenue CAGR", [results['summary_metrics'][s]['revenue_cagr'] for s in scenarios]),
            ("Max Annual UFCF ($M)", [results['summary_metrics'][s]['max_ufcf'] for s in scenarios]),
            ("Min Annual UFCF ($M)", [results['summary_metrics'][s]['min_ufcf'] for s in scenarios]),
            ("Terminal Value ($M)", [results['summary_metrics'][s]['terminal_value'] or 0 for s in scenarios]),
            ("PV of Terminal ($M)", [results['summary_metrics'][s]['present_value_terminal'] or 0 for s in scenarios]),
            ("Total PV ($M)", [results['summary_metrics'][s]['total_present_value'] or results['summary_metrics'][s]['total_ufcf'] for s in scenarios])
        ]

        for label, values in summary_data:
            ws.cell(row=current_row, column=1, value=label).style = 'label_bold'
            for i, value in enumerate(values):
                cell = ws.cell(row=current_row, column=i+2, value=value)
                cell.style = 'result'
                if 'Margin' in label or 'CAGR' in label:
                    cell.number_format = '0.0%'
                elif 'UFCF' in label or 'Value' in label or 'PV' in label:
                    cell.number_format = '#,##0'
            current_row += 1

            # Add spacing
            if "Total UFCF" in label or "Terminal Value" in label:
                current_row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 25
        for col in range(2, len(scenarios) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 15


def run_sample_fcf_model():
    """Run a sample FCF model"""

    print("💰 Running Professional FCF Model Sample")
    print("=" * 60)

    # Create model instance
    model = ProfessionalFCFModel("TechFlow Inc.", "TECHFLOW")

    # Run the model with sample assumptions
    results, excel_file = model.run_fcf_model(
        # Company Basics
        starting_revenue=1000.0,  # $1B starting revenue

        # Revenue Growth Scenarios (%)
        growth_base=[0.08, 0.07, 0.06, 0.05, 0.04],  # Year 1-5
        growth_bull=[0.12, 0.10, 0.08, 0.06, 0.05],
        growth_bear=[0.04, 0.03, 0.02, 0.02, 0.01],

        # Operating Margins (%)
        ebitda_margin=0.30,  # 30% EBITDA margin
        ebit_margin=0.25,    # 25% EBIT margin

        # Depreciation & Amortization (% of revenue)
        depreciation_pct=0.05,  # 5% of revenue

        # Capital Expenditures (% of revenue)
        capex_pct=0.04,  # 4% of revenue

        # Net Working Capital (% of revenue)
        nwc_pct=0.10,  # 10% of revenue

        # Tax Rate
        tax_rate=0.25,  # 25%

        # Terminal Value Assumptions
        terminal_growth_rate=0.025,  # 2.5% perpetual growth
        wacc=0.09,  # 9% WACC

        forecast_years=5
    )

    return results, excel_file


if __name__ == "__main__":
    # Run sample FCF model
    results, excel_file = run_sample_fcf_model()

    print("\n📋 Model Complete!")
    print(f"📁 Excel file saved as: {excel_file}")

    # Display key results
    print("\n🔑 Key Results Summary:")
    print("=" * 40)

    scenarios = ['base', 'bull', 'bear']
    summary = results['summary_metrics']

    print("Scenario Comparison:")
    print("<25")
    print("-" * 70)

    for scenario in scenarios:
        print("<25"
              "<15.0f"
              "<12.1f"
              "<12.1f"
              "<15.0f")
